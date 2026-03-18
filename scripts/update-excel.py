"""
스포짐 기획문서 v3 → v4 업데이트 스크립트
- IA(정보구조): 7개 페이지 추가
- 기능명세서: 11개 기능 추가
- 데이터 정의서: 14개 테이블 컬럼 추가
- API명세서: 10개 API 추가
- QA테스트케이스: QA 리포트 3건 추가
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

SRC = "excel/스포짐_기획문서_v3.xlsx"
DST = "excel/스포짐_기획문서_v4.xlsx"

wb = openpyxl.load_workbook(SRC)

# ──────────────────────────────────────────
# 헬퍼: 마지막 데이터 행 이후에 행 추가
# ──────────────────────────────────────────
def append_rows(ws, rows):
    """rows: list of list (순서는 헤더 컬럼 순서와 동일)"""
    start = ws.max_row + 1
    for i, row in enumerate(rows):
        r = start + i
        for j, val in enumerate(row, 1):
            ws.cell(r, j, val)
    return start, start + len(rows) - 1


# ══════════════════════════════════════════
# 1. IA(정보구조) — 7개 페이지 추가
# 컬럼: Depth1|Depth2|Depth3|화면ID|화면명|플랫폼|URL|로그인필요|권한|설명|상태|비고
# ══════════════════════════════════════════
ws_ia = wb["IA(정보구조)"]

ia_rows = [
    ["수업관리", "수업", "수업 관리",   "SCR-070", "수업 관리",   "Web", "/lessons",      "Y", "매니저 이상", "수업 등록/수정/삭제, 강사 배정, 정원 관리",   "개발완료", "BROJ CRM 이식 (2026-03-18)"],
    ["수업관리", "수업", "횟수 관리",   "SCR-071", "횟수 관리",   "Web", "/lesson-counts", "Y", "매니저 이상", "수강권 횟수 차감 관리",                       "개발완료", "BROJ CRM 이식"],
    ["수업관리", "수업", "페널티 관리", "SCR-072", "페널티 관리", "Web", "/penalties",     "Y", "매니저 이상", "노쇼/지각 페널티 관리",                       "개발완료", "BROJ CRM 이식"],
    ["매출관리", "매출", "환불 관리",   "SCR-073", "환불 관리",   "Web", "/refunds",       "Y", "관리자 이상", "환불 내역 조회/관리",                         "개발완료", "BROJ CRM 이식"],
    ["매출관리", "매출", "미수금 관리", "SCR-074", "미수금 관리", "Web", "/unpaid",        "Y", "매니저 이상", "미결제 내역 추적/관리",                       "개발완료", "BROJ CRM 이식"],
    ["매출관리", "매출", "매출 통계",   "SCR-075", "매출 통계",   "Web", "/sales/stats",   "Y", "매니저 이상", "상품별/결제수단별/종목별 매출 분석",           "개발완료", "BROJ CRM 이식"],
    ["시설관리", "시설", "운동복 관리", "SCR-076", "운동복 관리", "Web", "/clothing",      "Y", "스태프 이상", "운동복 재고/대여/세탁/파손 관리",             "개발완료", "BROJ CRM 이식"],
]

s1, e1 = append_rows(ws_ia, ia_rows)
print(f"[IA] {e1 - s1 + 1}개 행 추가 (행 {s1}~{e1}), 총 {ws_ia.max_row}행")


# ══════════════════════════════════════════
# 2. 기능명세서 — 11개 기능 추가
# 컬럼(21개): 기능ID|기능명|기능분류|플랫폼|우선순위|기능설명|상세요구사항|화면ID|
#              입력항목|출력항목|비즈니스규칙|예외처리|Web구현사항|App구현사항|
#              관련API|상태|담당자|[자동]화면명|[자동]API엔드포인트|비고|[자동]관련UI요소
# ══════════════════════════════════════════
ws_fn = wb["기능명세서"]

fn_rows = [
    # FN-070
    ["FN-070","수업 CRUD","수업관리","Web","High",
     "수업 등록/수정/삭제 + 강사 배정",
     "1)수업명/유형/강사/정원/시간/색상 입력 2)수업 목록 테이블",
     "SCR-070","수업명,유형,강사ID,정원,시간,색상","수업 목록",
     "정원은 1 이상","수업 삭제 시 연결된 일정 확인",
     "Supabase lessons CRUD","","API-070","개발완료","개발자1",None,None,None,None],
    # FN-071
    ["FN-071","수업 일정 관리","수업관리","Web","High",
     "수업 일정 생성/예약/출석 처리",
     "1)일정 생성(수업/시간/정원) 2)회원 예약 3)출석/노쇼 처리",
     "SCR-070","lessonId,시작시간,종료시간","일정 목록,예약 목록",
     "예약은 정원 초과 불가","정원 초과 시 대기열 안내",
     "FullCalendar 연동","","API-071","개발완료","개발자1",None,None,None,None],
    # FN-072
    ["FN-072","횟수 차감","수업관리","Web","High",
     "수강권 횟수 차감 + 이력 관리",
     "1)1회 차감 버튼 2)차감 이력 조회 3)잔여횟수 표시",
     "SCR-071","lessonCountId","차감 이력",
     "잔여 0이면 차감 불가","잔여 횟수 0: 차감 불가 안내",
     "usedCount 업데이트","","API-072","개발완료","개발자1",None,None,None,None],
    # FN-073
    ["FN-073","페널티 관리","수업관리","Web","Medium",
     "노쇼/지각 페널티 등록/삭제",
     "1)회원검색 2)유형선택 3)차감횟수 입력",
     "SCR-072","memberId,유형,차감횟수,사유","페널티 목록",
     "페널티 시 횟수 자동 차감","회원 미존재 에러",
     "penalties CRUD","","API-073","개발완료","개발자1",None,None,None,None],
    # FN-074
    ["FN-074","환불 처리","매출관리","Web","High",
     "환불 등록/조회/통계",
     "1)매출건 선택 2)환불금액/사유 입력 3)환불 통계",
     "SCR-073","saleId,환불금액,환불방법,사유","환불 목록,통계",
     "환불액은 결제액 초과 불가","원 결제건 없음 에러",
     "refunds CRUD","","API-074","개발완료","개발자1",None,None,None,None],
    # FN-075
    ["FN-075","미수금 관리","매출관리","Web","High",
     "미수금 등록/상태변경/추적",
     "1)미수금 목록 2)상태변경(미결제→완료) 3)메모 편집",
     "SCR-074","unpaidId,상태,메모","미수금 목록,통계",
     "30일 초과 시 연체 처리","이미 완료된 건 변경 불가",
     "unpaid CRUD","","API-075","개발완료","개발자1",None,None,None,None],
    # FN-076
    ["FN-076","매출 통계","매출관리","Web","High",
     "상품별/결제수단별/종목별 분석",
     "1)기간 필터 2)상품별 바차트 3)결제수단 파이차트",
     "SCR-075","기간,브랜치ID","통계 데이터,차트",
     "기간 필수","데이터 없음 시 빈 차트",
     "SVG 도넛+Tailwind 바 차트","","API-076","개발완료","개발자1",None,None,None,None],
    # FN-077
    ["FN-077","운동복 관리","시설관리","Web","Medium",
     "운동복 재고/대여/반납/세탁",
     "1)운동복 등록 2)상태변경 3)대여/반납",
     "SCR-076","번호,사이즈,타입","운동복 목록",
     "대여 시 회원ID 필수","중복 번호 등록 방지",
     "clothing CRUD","","API-077","개발완료","개발자1",None,None,None,None],
    # FN-078
    ["FN-078","뉴스피드/알림센터","공통","Web","Medium",
     "CRM 활동 알림 타임라인",
     "1)최근 알림 조회 2)읽음 처리 3)미읽 배지",
     "공통","branchId","알림 목록,미읽수",
     "최근 50건 표시","알림 없음 시 빈 상태",
     "우측 퀵메뉴 패널","","API-078","개발완료","개발자1",None,None,None,None],
    # FN-079
    ["FN-079","홀딩/연장/양도","회원관리","Web","High",
     "이용권 홀딩/연장/양도 처리",
     "1)홀딩(시작일/종료일/사유) 2)연장(추가일수) 3)양도(대상회원)",
     "SCR-011","memberId,productId,사유","이력 목록",
     "홀딩 중 재홀딩 불가","활성 이용권 없으면 불가",
     "member_holdings/extensions/transfer_logs","","API-079","개발완료","개발자1",None,None,None,None],
    # FN-080
    ["FN-080","상품타입 분리","상품관리","Web","High",
     "회원권/수강권/대여권/일반 타입 분리",
     "1)타입탭 2)현금가/카드가 분리 3)종목/태그/키오스크",
     "SCR-050","productType,cashPrice,cardPrice","상품목록",
     "타입 필수선택","기존 상품 마이그레이션",
     "productType 컬럼","","API-050","개발완료","개발자1",None,None,None,None],
]

s2, e2 = append_rows(ws_fn, fn_rows)
print(f"[기능명세서] {e2 - s2 + 1}개 행 추가 (행 {s2}~{e2}), 총 {ws_fn.max_row}행")


# ══════════════════════════════════════════
# 3. 데이터 정의서 — 14개 테이블 컬럼 추가
# 컬럼(13개): 테이블명|컬럼명|데이터타입|길이|NULL허용|기본값|PK|FK|인덱스|설명|관련API|상태|비고
# ══════════════════════════════════════════
ws_dd = wb["데이터 정의서"]

NOTE = "BROJ CRM 이식 (2026-03-18)"

# 각 테이블 정의: (table_name, [(col_name, dtype, length, nullable, default, pk, fk, idx, desc, api)])
tables = [
    ("news_feed", [
        ("id",         "SERIAL",          None, "N", None,  "Y", "",          "Y", "PK",                      "API-078"),
        ("branchId",   "INTEGER",         None, "N", None,  "",  "branches",  "Y", "지점 FK",                 "API-078"),
        ("userId",     "INTEGER",         None, "N", None,  "",  "users",     "Y", "작성자 FK",               "API-078"),
        ("userName",   "TEXT",            None, "Y", None,  "",  "",          "",  "작성자 이름",              "API-078"),
        ("type",       "TEXT",            None, "N", None,  "",  "",          "",  "알림 유형",                "API-078"),
        ("action",     "TEXT",            None, "N", None,  "",  "",          "",  "액션명",                   "API-078"),
        ("message",    "TEXT",            None, "N", None,  "",  "",          "",  "알림 메시지",              "API-078"),
        ("targetType", "TEXT",            None, "Y", None,  "",  "",          "",  "대상 유형",                "API-078"),
        ("targetId",   "INTEGER",         None, "Y", None,  "",  "",          "",  "대상 ID",                  "API-078"),
        ("isRead",     "BOOLEAN",         None, "N", "false","", "",          "",  "읽음 여부",                "API-078"),
        ("createdAt",  "TIMESTAMPTZ",     None, "N", "now()","","",          "",  "생성일시",                 "API-078"),
    ]),
    ("lessons", [
        ("id",             "SERIAL",      None, "N", None,   "Y", "",         "Y", "PK",                      "API-070"),
        ("branchId",       "INTEGER",     None, "N", None,   "",  "branches", "Y", "지점 FK",                 "API-070"),
        ("name",           "TEXT",        None, "N", None,   "",  "",         "",  "수업명",                   "API-070"),
        ("type",           "TEXT",        None, "N", None,   "",  "",         "",  "수업 유형",                "API-070"),
        ("instructorId",   "INTEGER",     None, "Y", None,   "",  "staff",    "Y", "강사 FK",                 "API-070"),
        ("instructorName", "TEXT",        None, "Y", None,   "",  "",         "",  "강사 이름",                "API-070"),
        ("capacity",       "INTEGER",     None, "N", None,   "",  "",         "",  "정원",                     "API-070"),
        ("duration",       "INTEGER",     None, "N", None,   "",  "",         "",  "수업 시간(분)",            "API-070"),
        ("color",          "TEXT",        None, "Y", None,   "",  "",         "",  "캘린더 색상",              "API-070"),
        ("isActive",       "BOOLEAN",     None, "N", "true", "",  "",         "",  "활성 여부",                "API-070"),
        ("createdAt",      "TIMESTAMPTZ", None, "N", "now()","",  "",         "",  "생성일시",                 "API-070"),
        ("updatedAt",      "TIMESTAMPTZ", None, "N", "now()","",  "",         "",  "수정일시",                 "API-070"),
    ]),
    ("lesson_schedules", [
        ("id",           "SERIAL",      None, "N", None,   "Y", "",          "Y", "PK",                     "API-071"),
        ("lessonId",     "INTEGER",     None, "N", None,   "",  "lessons",   "Y", "수업 FK",                "API-071"),
        ("branchId",     "INTEGER",     None, "N", None,   "",  "branches",  "Y", "지점 FK",                "API-071"),
        ("instructorId", "INTEGER",     None, "Y", None,   "",  "staff",     "Y", "강사 FK",                "API-071"),
        ("startAt",      "TIMESTAMPTZ", None, "N", None,   "",  "",          "",  "시작 일시",               "API-071"),
        ("endAt",        "TIMESTAMPTZ", None, "N", None,   "",  "",          "",  "종료 일시",               "API-071"),
        ("currentCount", "INTEGER",     None, "N", "0",    "",  "",          "",  "현재 예약자 수",          "API-071"),
        ("capacity",     "INTEGER",     None, "N", None,   "",  "",          "",  "정원",                    "API-071"),
        ("status",       "TEXT",        None, "N", None,   "",  "",          "",  "일정 상태",               "API-071"),
        ("memo",         "TEXT",        None, "Y", None,   "",  "",          "",  "메모",                    "API-071"),
        ("createdAt",    "TIMESTAMPTZ", None, "N", "now()","",  "",          "",  "생성일시",                "API-071"),
    ]),
    ("lesson_bookings", [
        ("id",          "SERIAL",      None, "N", None,  "Y", "",                "Y", "PK",             "API-071"),
        ("scheduleId",  "INTEGER",     None, "N", None,  "",  "lesson_schedules","Y", "일정 FK",        "API-071"),
        ("memberId",    "INTEGER",     None, "N", None,  "",  "members",         "Y", "회원 FK",        "API-071"),
        ("memberName",  "TEXT",        None, "Y", None,  "",  "",                "",  "회원 이름",       "API-071"),
        ("status",      "TEXT",        None, "N", None,  "",  "",                "",  "예약 상태",       "API-071"),
        ("bookedAt",    "TIMESTAMPTZ", None, "N", "now()","", "",                "",  "예약 일시",       "API-071"),
        ("attendedAt",  "TIMESTAMPTZ", None, "Y", None,  "",  "",                "",  "출석 일시",       "API-071"),
        ("cancelledAt", "TIMESTAMPTZ", None, "Y", None,  "",  "",                "",  "취소 일시",       "API-071"),
        ("cancelReason","TEXT",        None, "Y", None,  "",  "",                "",  "취소 사유",       "API-071"),
    ]),
    ("lesson_counts", [
        ("id",          "SERIAL",      None, "N", None,   "Y", "",         "Y", "PK",               "API-072"),
        ("memberId",    "INTEGER",     None, "N", None,   "",  "members",  "Y", "회원 FK",          "API-072"),
        ("productId",   "INTEGER",     None, "N", None,   "",  "products", "Y", "상품 FK",          "API-072"),
        ("productName", "TEXT",        None, "Y", None,   "",  "",         "",  "상품명",            "API-072"),
        ("totalCount",  "INTEGER",     None, "N", None,   "",  "",         "",  "총 횟수",           "API-072"),
        ("usedCount",   "INTEGER",     None, "N", "0",    "",  "",         "",  "사용 횟수",         "API-072"),
        ("startDate",   "DATE",        None, "Y", None,   "",  "",         "",  "시작일",            "API-072"),
        ("endDate",     "DATE",        None, "Y", None,   "",  "",         "",  "종료일",            "API-072"),
        ("status",      "TEXT",        None, "N", None,   "",  "",         "",  "상태",              "API-072"),
        ("createdAt",   "TIMESTAMPTZ", None, "N", "now()","",  "",         "",  "생성일시",          "API-072"),
        ("updatedAt",   "TIMESTAMPTZ", None, "N", "now()","",  "",         "",  "수정일시",          "API-072"),
    ]),
    ("penalties", [
        ("id",          "SERIAL",      None, "N", None,   "Y", "",                "Y", "PK",          "API-073"),
        ("branchId",    "INTEGER",     None, "N", None,   "",  "branches",        "Y", "지점 FK",     "API-073"),
        ("memberId",    "INTEGER",     None, "N", None,   "",  "members",         "Y", "회원 FK",     "API-073"),
        ("memberName",  "TEXT",        None, "Y", None,   "",  "",                "",  "회원 이름",   "API-073"),
        ("scheduleId",  "INTEGER",     None, "Y", None,   "",  "lesson_schedules","Y", "일정 FK",     "API-073"),
        ("type",        "TEXT",        None, "N", None,   "",  "",                "",  "페널티 유형", "API-073"),
        ("deductCount", "INTEGER",     None, "N", None,   "",  "",                "",  "차감 횟수",   "API-073"),
        ("reason",      "TEXT",        None, "Y", None,   "",  "",                "",  "사유",        "API-073"),
        ("appliedAt",   "TIMESTAMPTZ", None, "N", "now()","",  "",                "",  "적용 일시",   "API-073"),
        ("appliedBy",   "INTEGER",     None, "Y", None,   "",  "users",           "Y", "처리자 FK",  "API-073"),
    ]),
    ("refunds", [
        ("id",              "SERIAL",      None, "N", None,   "Y", "",       "Y", "PK",               "API-074"),
        ("saleId",          "INTEGER",     None, "N", None,   "",  "sales",  "Y", "매출 FK",          "API-074"),
        ("branchId",        "INTEGER",     None, "N", None,   "",  "branches","Y", "지점 FK",         "API-074"),
        ("memberId",        "INTEGER",     None, "N", None,   "",  "members","Y", "회원 FK",          "API-074"),
        ("memberName",      "TEXT",        None, "Y", None,   "",  "",       "",  "회원 이름",         "API-074"),
        ("productName",     "TEXT",        None, "Y", None,   "",  "",       "",  "상품명",            "API-074"),
        ("refundAmount",    "DECIMAL(12,2)",None,"N", None,   "",  "",       "",  "환불 금액",         "API-074"),
        ("refundMethod",    "TEXT",        None, "N", None,   "",  "",       "",  "환불 방법",         "API-074"),
        ("reason",          "TEXT",        None, "Y", None,   "",  "",       "",  "환불 사유",         "API-074"),
        ("status",          "TEXT",        None, "N", None,   "",  "",       "",  "처리 상태",         "API-074"),
        ("processedBy",     "INTEGER",     None, "Y", None,   "",  "users",  "Y", "처리자 FK",        "API-074"),
        ("processedByName", "TEXT",        None, "Y", None,   "",  "",       "",  "처리자 이름",       "API-074"),
        ("processedAt",     "TIMESTAMPTZ", None, "Y", None,   "",  "",       "",  "처리 일시",         "API-074"),
        ("createdAt",       "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "생성일시",          "API-074"),
    ]),
    ("unpaid", [
        ("id",           "SERIAL",      None, "N", None,   "Y", "",       "Y", "PK",               "API-075"),
        ("saleId",       "INTEGER",     None, "Y", None,   "",  "sales",  "Y", "매출 FK",          "API-075"),
        ("branchId",     "INTEGER",     None, "N", None,   "",  "branches","Y","지점 FK",          "API-075"),
        ("memberId",     "INTEGER",     None, "N", None,   "",  "members","Y", "회원 FK",          "API-075"),
        ("memberName",   "TEXT",        None, "Y", None,   "",  "",       "",  "회원 이름",         "API-075"),
        ("productName",  "TEXT",        None, "Y", None,   "",  "",       "",  "상품명",            "API-075"),
        ("unpaidAmount", "DECIMAL(12,2)",None,"N", None,   "",  "",       "",  "미수금 금액",       "API-075"),
        ("dueDate",      "DATE",        None, "Y", None,   "",  "",       "",  "납부 기한",         "API-075"),
        ("status",       "TEXT",        None, "N", None,   "",  "",       "",  "처리 상태",         "API-075"),
        ("memo",         "TEXT",        None, "Y", None,   "",  "",       "",  "메모",              "API-075"),
        ("createdAt",    "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "생성일시",          "API-075"),
        ("updatedAt",    "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "수정일시",          "API-075"),
    ]),
    ("visit_logs", [
        ("id",         "SERIAL",      None, "N", None,   "Y", "",       "Y", "PK",               "API-078"),
        ("branchId",   "INTEGER",     None, "N", None,   "",  "branches","Y","지점 FK",          "API-078"),
        ("memberId",   "INTEGER",     None, "N", None,   "",  "members","Y", "회원 FK",          "API-078"),
        ("memberName", "TEXT",        None, "Y", None,   "",  "",       "",  "회원 이름",         "API-078"),
        ("checkInAt",  "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "입장 일시",         "API-078"),
        ("checkOutAt", "TIMESTAMPTZ", None, "Y", None,   "",  "",       "",  "퇴장 일시",         "API-078"),
        ("method",     "TEXT",        None, "Y", None,   "",  "",       "",  "인증 방법",         "API-078"),
        ("gateId",     "TEXT",        None, "Y", None,   "",  "",       "",  "게이트 ID",         "API-078"),
    ]),
    ("clothing", [
        ("id",         "SERIAL",      None, "N", None,   "Y", "",       "Y", "PK",               "API-077"),
        ("branchId",   "INTEGER",     None, "N", None,   "",  "branches","Y","지점 FK",          "API-077"),
        ("number",     "TEXT",        None, "N", None,   "",  "",       "",  "운동복 번호",       "API-077"),
        ("size",       "TEXT",        None, "N", None,   "",  "",       "",  "사이즈",            "API-077"),
        ("type",       "TEXT",        None, "N", None,   "",  "",       "",  "유형(상의/하의)",  "API-077"),
        ("status",     "TEXT",        None, "N", None,   "",  "",       "",  "상태",              "API-077"),
        ("memberId",   "INTEGER",     None, "Y", None,   "",  "members","Y", "대여 회원 FK",     "API-077"),
        ("memberName", "TEXT",        None, "Y", None,   "",  "",       "",  "대여 회원 이름",   "API-077"),
        ("rentedAt",   "TIMESTAMPTZ", None, "Y", None,   "",  "",       "",  "대여 일시",         "API-077"),
        ("returnDue",  "DATE",        None, "Y", None,   "",  "",       "",  "반납 예정일",       "API-077"),
        ("memo",       "TEXT",        None, "Y", None,   "",  "",       "",  "메모",              "API-077"),
        ("createdAt",  "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "생성일시",          "API-077"),
        ("updatedAt",  "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "수정일시",          "API-077"),
    ]),
    ("member_holdings", [
        ("id",            "SERIAL",      None, "N", None,   "Y", "",         "Y", "PK",             "API-079"),
        ("memberId",      "INTEGER",     None, "N", None,   "",  "members",  "Y", "회원 FK",        "API-079"),
        ("productId",     "INTEGER",     None, "N", None,   "",  "products", "Y", "상품 FK",        "API-079"),
        ("productName",   "TEXT",        None, "Y", None,   "",  "",         "",  "상품명",          "API-079"),
        ("startDate",     "DATE",        None, "N", None,   "",  "",         "",  "홀딩 시작일",     "API-079"),
        ("endDate",       "DATE",        None, "N", None,   "",  "",         "",  "홀딩 종료일",     "API-079"),
        ("reason",        "TEXT",        None, "Y", None,   "",  "",         "",  "홀딩 사유",       "API-079"),
        ("status",        "TEXT",        None, "N", None,   "",  "",         "",  "처리 상태",       "API-079"),
        ("createdBy",     "INTEGER",     None, "Y", None,   "",  "users",    "Y", "처리자 FK",      "API-079"),
        ("createdByName", "TEXT",        None, "Y", None,   "",  "",         "",  "처리자 이름",     "API-079"),
        ("createdAt",     "TIMESTAMPTZ", None, "N", "now()","",  "",         "",  "생성일시",        "API-079"),
    ]),
    ("member_extensions", [
        ("id",              "SERIAL",      None, "N", None,   "Y", "",         "Y", "PK",           "API-079"),
        ("memberId",        "INTEGER",     None, "N", None,   "",  "members",  "Y", "회원 FK",      "API-079"),
        ("productId",       "INTEGER",     None, "N", None,   "",  "products", "Y", "상품 FK",      "API-079"),
        ("productName",     "TEXT",        None, "Y", None,   "",  "",         "",  "상품명",        "API-079"),
        ("extraDays",       "INTEGER",     None, "N", None,   "",  "",         "",  "연장 일수",     "API-079"),
        ("originalEndDate", "DATE",        None, "N", None,   "",  "",         "",  "원래 종료일",   "API-079"),
        ("newEndDate",      "DATE",        None, "N", None,   "",  "",         "",  "새 종료일",     "API-079"),
        ("reason",          "TEXT",        None, "Y", None,   "",  "",         "",  "연장 사유",     "API-079"),
        ("createdBy",       "INTEGER",     None, "Y", None,   "",  "users",    "Y", "처리자 FK",    "API-079"),
        ("createdByName",   "TEXT",        None, "Y", None,   "",  "",         "",  "처리자 이름",   "API-079"),
        ("createdAt",       "TIMESTAMPTZ", None, "N", "now()","",  "",         "",  "생성일시",      "API-079"),
    ]),
    ("member_transfer_logs", [
        ("id",              "SERIAL",      None, "N", None,   "Y", "",         "Y", "PK",             "API-079"),
        ("fromMemberId",    "INTEGER",     None, "N", None,   "",  "members",  "Y", "양도인 FK",      "API-079"),
        ("fromMemberName",  "TEXT",        None, "Y", None,   "",  "",         "",  "양도인 이름",     "API-079"),
        ("toMemberId",      "INTEGER",     None, "N", None,   "",  "members",  "Y", "양수인 FK",      "API-079"),
        ("toMemberName",    "TEXT",        None, "Y", None,   "",  "",         "",  "양수인 이름",     "API-079"),
        ("productId",       "INTEGER",     None, "N", None,   "",  "products", "Y", "상품 FK",        "API-079"),
        ("productName",     "TEXT",        None, "Y", None,   "",  "",         "",  "상품명",          "API-079"),
        ("reason",          "TEXT",        None, "Y", None,   "",  "",         "",  "양도 사유",       "API-079"),
        ("createdBy",       "INTEGER",     None, "Y", None,   "",  "users",    "Y", "처리자 FK",      "API-079"),
        ("createdByName",   "TEXT",        None, "Y", None,   "",  "",         "",  "처리자 이름",     "API-079"),
        ("createdAt",       "TIMESTAMPTZ", None, "N", "now()","",  "",         "",  "생성일시",        "API-079"),
    ]),
    ("member_memos", [
        ("id",            "SERIAL",      None, "N", None,   "Y", "",       "Y", "PK",               ""),
        ("memberId",      "INTEGER",     None, "N", None,   "",  "members","Y", "회원 FK",          ""),
        ("branchId",      "INTEGER",     None, "N", None,   "",  "branches","Y","지점 FK",          ""),
        ("content",       "TEXT",        None, "N", None,   "",  "",       "",  "메모 내용",         ""),
        ("createdBy",     "INTEGER",     None, "Y", None,   "",  "users",  "Y", "작성자 FK",        ""),
        ("createdByName", "TEXT",        None, "Y", None,   "",  "",       "",  "작성자 이름",       ""),
        ("createdAt",     "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "생성일시",          ""),
        ("updatedAt",     "TIMESTAMPTZ", None, "N", "now()","",  "",       "",  "수정일시",          ""),
    ]),
]

dd_rows = []
for table_name, cols in tables:
    for col in cols:
        col_name, dtype, length, nullable, default, pk, fk, idx, desc, api = col
        dd_rows.append([
            table_name, col_name, dtype,
            length if length else "",
            nullable, default if default else "", pk, fk, idx, desc, api,
            "확정", NOTE
        ])

s3, e3 = append_rows(ws_dd, dd_rows)
print(f"[데이터 정의서] {e3 - s3 + 1}개 행 추가 (행 {s3}~{e3}), 총 {ws_dd.max_row}행")


# ══════════════════════════════════════════
# 4. API명세서 — 10개 API 추가
# 컬럼(18개): API ID|API명|Method|엔드포인트|설명|Request Headers|Request Body|
#              Response(성공)|Response(실패)|인증필요|화면ID|기능ID|상태|담당자|
#              [자동]화면명|[자동]기능명|비고|[자동]관련UI요소
# ══════════════════════════════════════════
ws_api = wb["API명세서"]

HDR = "Authorization: Bearer {token}"
ERR = '{ "error": "SERVER_ERROR" }'

api_rows = [
    ["API-070","수업 목록 조회",      "GET",         "/api/v1/lessons",           "지점별 수업 목록",              HDR, "",                                   '{ "lessons": [...] }',   ERR, "Y", "SCR-070", "FN-070", "개발완료", "", None, None, None, None],
    ["API-071","수업 일정 조회",      "GET",         "/api/v1/lesson-schedules",  "기간별 수업 일정 조회",          HDR, "",                                   '{ "schedules": [...] }', ERR, "Y", "SCR-070", "FN-071", "개발완료", "", None, None, None, None],
    ["API-072","횟수 조회/차감",      "GET/PUT",     "/api/v1/lesson-counts",     "수강권 횟수 조회 및 차감",       HDR, "",                                   '{ "counts": [...] }',    ERR, "Y", "SCR-071", "FN-072", "개발완료", "", None, None, None, None],
    ["API-073","페널티 CRUD",         "GET/POST/DELETE","/api/v1/penalties",      "페널티 조회/등록/삭제",          HDR, '{ "memberId": 1, "type": "NOSHOW" }','{ "penalties": [...] }', ERR, "Y", "SCR-072", "FN-073", "개발완료", "", None, None, None, None],
    ["API-074","환불 관리",           "GET/POST",    "/api/v1/refunds",           "환불 조회/등록/통계",            HDR, '{ "saleId": 1, "amount": 50000 }',   '{ "refunds": [...] }',   ERR, "Y", "SCR-073", "FN-074", "개발완료", "", None, None, None, None],
    ["API-075","미수금 관리",         "GET/POST/PUT","/api/v1/unpaid",            "미수금 조회/등록/상태변경",      HDR, '{ "status": "PAID" }',               '{ "unpaid": [...] }',    ERR, "Y", "SCR-074", "FN-075", "개발완료", "", None, None, None, None],
    ["API-076","매출 통계",           "GET",         "/api/v1/sales/stats",       "상품별/결제수단별 매출 통계",    HDR, "",                                   '{ "stats": {...} }',     ERR, "Y", "SCR-075", "FN-076", "개발완료", "", None, None, None, None],
    ["API-077","운동복 CRUD",         "GET/POST/PUT","/api/v1/clothing",          "운동복 조회/등록/상태변경",      HDR, '{ "number": "001", "size": "M" }',   '{ "items": [...] }',     ERR, "Y", "SCR-076", "FN-077", "개발완료", "", None, None, None, None],
    ["API-078","뉴스피드",            "GET/POST/PUT","/api/v1/news-feed",         "알림 조회/생성/읽음처리",        HDR, "",                                   '{ "feed": [...] }',      ERR, "Y", "공통",    "FN-078", "개발완료", "", None, None, None, None],
    ["API-079","홀딩/연장/양도",      "GET/POST",    "/api/v1/member-holdings",   "홀딩/연장/양도 이력 관리",       HDR, '{ "memberId": 1, "startDate": "..." }','{ "holdings": [...] }',ERR, "Y", "SCR-011", "FN-079", "개발완료", "", None, None, None, None],
]

s4, e4 = append_rows(ws_api, api_rows)
print(f"[API명세서] {e4 - s4 + 1}개 행 추가 (행 {s4}~{e4}), 총 {ws_api.max_row}행")


# ══════════════════════════════════════════
# 5. QA테스트케이스 — 3개 리포트 결과 추가
# 컬럼(29개): TC-ID|플랫폼|테스트영역|하위영역|테스트유형|테스트시나리오|사전조건|
#              테스트절차|기대결과|실제결과|테스트상태|우선순위|심각도|담당자|
#              테스트환경|브라우저/앱|디바이스|OS|테스트사이클|자동화여부|
#              화면ID|기능ID|API ID|버그티켓|스크린샷|[자동]화면명|[자동]기능명|[자동]API엔드포인트|비고
# ══════════════════════════════════════════
ws_qa = wb["QA테스트케이스"]

ENV  = "Staging"
BROW = "Chrome"
DEV  = "Desktop"
OS_  = "macOS"
CYCO = "2026-03-17"
TESTER = "Claude"

def qa_row(tc_id, area, sub, scenario, procedure, expected, actual, status, priority, severity,
           screen_id="", fn_id="", api_id="", bug="", note=""):
    return [
        tc_id, "Web", area, sub, "기능테스트",
        scenario, "로그인 완료 / 해당 계정 접속", procedure, expected, actual,
        status, priority, severity, TESTER,
        ENV, BROW, DEV, OS_, CYCO, "수동",
        screen_id, fn_id, api_id, bug, "",
        None, None, None, note
    ]

qa_rows = [
    # ── QA-VERIFY-2026-03-17.md (6건) ──
    qa_row("TC-V-001","RBAC","권한제어",
           "readonly 사용자 헤더 버튼 미표시 확인",
           "1)readonly1 로그인 2)/payroll/statements 이동 3)헤더 버튼 영역 확인",
           "'+회원등록' 버튼 미표시",
           "버튼 미표시 확인",
           "Pass","Medium","Medium","","","","BUG-13","QA-VERIFY-2026-03-17"),
    qa_row("TC-V-002","RBAC","권한제어",
           "fc(trainer1) 엑셀 다운로드 버튼 미표시",
           "1)trainer1 로그인 2)/members 이동 3)헤더 버튼 확인",
           "엑셀 다운로드 버튼 미표시",
           "버튼 미표시 확인",
           "Pass","Medium","Medium","","","","BUG-15","QA-VERIFY-2026-03-17"),
    qa_row("TC-V-003","RBAC","권한제어",
           "fc(trainer1) 회원추가 버튼 미표시",
           "1)trainer1 로그인 2)/members 이동 3)헤더 버튼 확인",
           "'회원 추가' 버튼 미표시",
           "버튼 미표시 확인",
           "Pass","Medium","Medium","","","","BUG-16","QA-VERIFY-2026-03-17"),
    qa_row("TC-V-004","RBAC","권한제어",
           "staff 엑셀 다운로드 버튼 미표시",
           "1)staff1 로그인 2)/members 이동 3)헤더 버튼 확인",
           "엑셀 다운로드 버튼 미표시",
           "버튼 미표시 확인",
           "Pass","Medium","Medium","","","","BUG-17","QA-VERIFY-2026-03-17"),
    qa_row("TC-V-005","전자계약","배지",
           "전자계약 회원 상태 배지 정상 표시",
           "1)/contracts/new 이동 2)회원 상태 배지 확인",
           "활성/만료/홀딩 배지 정상 표시",
           "활성(초록)/만료(빨강)/홀딩(주황) 정상",
           "Pass","Low","Low","","","","UX-11","QA-VERIFY-2026-03-17"),
    qa_row("TC-V-006","RBAC","403페이지",
           "readonly 403 페이지 무한 루프 미발생",
           "1)readonly1 로그인 2)403 페이지 확인 3)버튼 클릭 4)이동 확인",
           "'급여 명세서로 이동' 버튼 표시, 무한 루프 없음",
           "무한 루프 없이 /payroll/statements 이동",
           "Pass","Medium","Medium","","","","UX-13","QA-VERIFY-2026-03-17"),

    # ── QA-LOG-2026-03-17-슈퍼관리자.md (5탭) ──
    qa_row("TC-SA-001","슈퍼관리자","통합대시보드",
           "슈퍼관리자 통합 대시보드 접근 및 렌더링",
           "1)superadmin 로그인 2)/super-dashboard 이동 3)KPI 카드 확인 4)지점별 현황 확인",
           "페이지 렌더링 정상, KPI 카드 4개 표시",
           "렌더링 정상, KPI 초기값 0 (데이터 없음)",
           "Pass","High","High","","","","BUG-SA-01","QA-LOG-2026-03-17-슈퍼관리자"),
    qa_row("TC-SA-002","슈퍼관리자","지점관리",
           "슈퍼관리자 지점 관리 페이지 접근",
           "1)/branches 이동 2)통계카드 확인 3)지점 목록 확인 4)상태 배지 확인",
           "3개 지점 목록 표시, 상태 배지 정상",
           "3개 지점 표시, 상태 배지 '폐점' 오류 발견",
           "Fail","High","High","","","","BUG-SA-02","QA-LOG-2026-03-17-슈퍼관리자"),
    qa_row("TC-SA-003","슈퍼관리자","지점비교리포트",
           "슈퍼관리자 지점 비교 리포트 정상 동작",
           "1)/branch-report 이동 2)KPI 카드 확인 3)바 차트 확인 4)상세 테이블 확인",
           "3지점 비교 데이터 정상 표시",
           "강남25명, 서초10명, 송파10명 정상",
           "Pass","High","Medium","","","","","QA-LOG-2026-03-17-슈퍼관리자"),
    qa_row("TC-SA-004","슈퍼관리자","감사로그",
           "슈퍼관리자 감사 로그 페이지 접근",
           "1)/audit-log 이동 2)필터 UI 확인 3)검색 실행",
           "에러 없이 빈 상태 표시",
           "에러 없이 '감사 로그가 없습니다' 표시",
           "Pass","Medium","Medium","","","","BUG-SA-04","QA-LOG-2026-03-17-슈퍼관리자"),
    qa_row("TC-SA-005","슈퍼관리자","구독관리",
           "슈퍼관리자 구독 관리 페이지 정상 동작",
           "1)/subscription 이동 2)플랜 정보 확인 3)사용량 확인",
           "Fit Pro Plan, 사용량 프로그레스 정상 표시",
           "Fit Pro Plan, 회원 850/1000 등 정상",
           "Pass","Medium","Low","","","","","QA-LOG-2026-03-17-슈퍼관리자"),

    # ── QA-VERIFY-SA-2026-03-17.md (4건 수정 확인) ──
    qa_row("TC-SA-V-001","슈퍼관리자","통합대시보드",
           "BUG-SA-01 수정 후 KPI 정상 표시 확인",
           "1)superadmin 로그인 2)/super-dashboard 이동 3)KPI 카드 값 확인",
           "전체 회원/매출/출석/직원 실제 데이터 표시",
           "전체 회원 26명, 매출 784만원, 출석 1명, 직원 9명",
           "Pass","High","High","","","","BUG-SA-01","QA-VERIFY-SA-2026-03-17"),
    qa_row("TC-SA-V-002","슈퍼관리자","지점관리",
           "BUG-SA-02 수정 후 지점 상태 배지 정상 확인",
           "1)/branches 이동 2)각 지점 상태 배지 확인",
           "강남/서초 '운영중', 송파 '임시휴업' 정상 표시",
           "강남/서초 운영중(초록), 송파 임시휴업(주황)",
           "Pass","High","High","","","","BUG-SA-02","QA-VERIFY-SA-2026-03-17"),
    qa_row("TC-SA-V-003","슈퍼관리자","지점관리",
           "BUG-SA-03 수정 후 지점별 회원/직원 수 정상 확인",
           "1)/branches 이동 2)각 지점 회원/직원 수 확인",
           "각 지점 실제 회원/직원 수 표시",
           "강남 13/7, 서초 8/2, 송파 5/0",
           "Pass","High","Medium","","","","BUG-SA-03","QA-VERIFY-SA-2026-03-17"),
    qa_row("TC-SA-V-004","슈퍼관리자","감사로그",
           "BUG-SA-04 수정 후 audit_logs 테이블 생성 확인",
           "1)/audit-log 이동 2)에러 여부 확인 3)빈 상태 확인",
           "테이블 생성, 에러 없이 빈 상태 표시",
           "테이블 생성 완료, 데이터 0건 (신규 테이블)",
           "Pass","Medium","Medium","","","","BUG-SA-04","QA-VERIFY-SA-2026-03-17"),
]

s5, e5 = append_rows(ws_qa, qa_rows)
print(f"[QA테스트케이스] {e5 - s5 + 1}개 행 추가 (행 {s5}~{e5}), 총 {ws_qa.max_row}행")


# ══════════════════════════════════════════
# 저장
# ══════════════════════════════════════════
wb.save(DST)
print(f"\n저장 완료: {DST}")

# 행 수 비교 요약
print("\n=== 시트별 행 수 비교 ===")
wb_old = openpyxl.load_workbook(SRC)
wb_new = openpyxl.load_workbook(DST)
targets = ["IA(정보구조)", "기능명세서", "데이터 정의서", "API명세서", "QA테스트케이스"]
for name in targets:
    old_r = wb_old[name].max_row
    new_r = wb_new[name].max_row
    print(f"  [{name}] {old_r}행 → {new_r}행 (+{new_r - old_r})")
