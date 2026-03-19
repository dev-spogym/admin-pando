"""
기획문서 v6 업데이트: 레슨북/온핏/브로제이 경쟁사 기능 반영
- 상품 레슨북 컬럼 (홀딩/양도/포인트/판매유형)
- 상품 동적 카테고리
- 노쇼/취소 정책 설정 모달
- 대기열 자동 승격
- 얼굴 인식 설정
- 동선 분석 (구역별 현황)
- QR 운동복
"""
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "excel/스포짐_기획문서_v5_온핏반영.xlsx"
DST = "excel/스포짐_기획문서_v6_레슨북반영.xlsx"

# 복사
shutil.copy2(SRC, DST)
wb = load_workbook(DST)

# 스타일
header_font = Font(bold=True, size=10)
normal_font = Font(size=10)
fill_new = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 연녹색
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def append_row(ws, values, is_header=False):
    """시트에 행 추가 + 스타일 적용"""
    row_num = ws.max_row + 1
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = header_font if is_header else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if not is_header:
            cell.fill = fill_new
    return row_num


# ═══════════════════════════════════════════════════════════════
# 1. 기능명세서 시트 — 신규 기능 추가
# ═══════════════════════════════════════════════════════════════
ws = wb['기능명세서']

new_features = [
    # (기능ID, 기능명, 분류, 플랫폼, 우선순위, 설명, 상세요구사항, 화면ID, 입력, 출력, 비즈니스규칙, 예외처리, Web구현, App, API, 상태, 담당자, 화면명, API엔드포인트, 비고, UI요소)
    ("FN-110", "상품 동적 카테고리", "상품관리", "Web", "High",
     "고정 카테고리(이용권/PT/GX/기타) 외에 골프/식품/보충제 등 자유롭게 카테고리 추가 가능",
     "1) product_groups 테이블에서 동적 분류 로드\n2) ComboBox로 기존 선택 또는 직접 입력\n3) ProductList에 동적 카테고리 필터 탭 표시\n4) 카테고리별 상품 수 카운트",
     "SCR-040", "카테고리명", "동적 카테고리 목록", "분류 관리 탭에서 CRUD 가능", "중복 분류명 방지", "ProductList.tsx + ProductDetailPanel.tsx ComboBox", "", "", "개발완료", "", "상품 목록", "", "레슨북 핵심 장점: 한 화면 자유도", ""),

    ("FN-111", "상품 레슨북 컬럼 (홀딩/양도/포인트/판매유형)", "상품관리", "Web", "High",
     "레슨북 Group 3.png 기준 누락 컬럼 4개 추가: 홀딩 가능, 양도 가능, 포인트 적립, 판매유형",
     "1) holdingEnabled: 기간정지 허용 여부 (토글)\n2) transferEnabled: 양도 가능 여부 (토글)\n3) pointAccrual: 포인트 적립 여부 (토글)\n4) salesChannel: ALL/KIOSK/COUNTER/ONLINE\n5) 테이블에 ✓/✗ 표시",
     "SCR-040", "홀딩/양도/포인트/판매유형 값", "상품 속성 업데이트", "기본값: 홀딩=off, 양도=off, 포인트=on, 판매=전체", "DB 컬럼 추가 필요", "ProductDetailPanel.tsx 옵션 섹션 + ProductList.tsx 테이블 컬럼", "", "", "개발완료", "", "상품 목록", "", "레슨북 상품관리 컬럼 100% 반영", ""),

    ("FN-112", "상품 요일/시간 제한 표시", "상품관리", "Web", "High",
     "상품 목록 테이블에 레슨북 스타일 월~일 요일별 ✓/✗ + 이용시간 컬럼 표시",
     "1) usage_restrictions JSON 필드에서 availableDays 읽기\n2) 월/화/수/목/금/토/일 7컬럼에 ✓(초록)/✗(빨강) 표시\n3) 이용시간: 06:00~22:00 또는 '전체'\n4) 패널 열리면 요일 컬럼 숨김",
     "SCR-040", "usage_restrictions JSON", "요일별 ✓/✗ 렌더링", "미설정 시 전체 ✓ 표시", "", "ProductList.tsx 테이블 + ProductDetailPanel.tsx 요일 원형 버튼", "", "", "개발완료", "", "상품 목록", "", "레슨북 최대 장점", ""),

    ("FN-113", "노쇼/취소 자동 정책 설정", "수업관리", "Web", "Critical",
     "수업 노쇼/취소 관련 정책을 설정 모달에서 통합 관리",
     "1) 취소 마감: 수업 N시간 전 (기본 3시간)\n2) 마감 후 취소 패널티: 횟수 차감 토글\n3) 노쇼 시 세션 차감: 토글\n4) 연속 노쇼 경고: N회 기준 (기본 3회)\n5) 미처리 수업 자동 완료: N시간 후 (기본 24시간)\n6) settings 테이블 lesson_policy 키에 JSON 저장",
     "SCR-070", "정책 설정값", "정책 JSON 저장", "기본값 적용, 지점별 독립 설정", "settings 테이블 upsert", "LessonManagement.tsx 정책 모달 + businessLogic.ts", "", "", "개발완료", "", "수업 관리", "", "레슨북+온핏 통합", ""),

    ("FN-114", "GX 예약 N시간 전 자동 오픈", "수업관리", "Web", "High",
     "GX 수업 예약을 수업 시작 N시간 전에 자동으로 오픈 (PENDING→OPEN)",
     "1) lesson_policy에 reservationAutoOpenHours 설정 (기본 48시간)\n2) runDailySync에서 syncAutoOpenReservations 호출\n3) lesson_schedules.status를 PENDING→OPEN 전환",
     "SCR-070", "자동 오픈 시간 설정", "예약 오픈 상태 전환", "오픈 시간 전까지는 PENDING 유지", "", "businessLogic.ts syncAutoOpenReservations", "", "", "개발완료", "", "수업 관리", "", "온핏 요청사항", ""),

    ("FN-115", "대기열 시스템 (Waitlist)", "수업관리", "Web", "High",
     "정원 초과 시 WAITLIST 상태로 예약 등록, 취소 발생 시 대기 1순위 자동 승격",
     "1) joinWaitlist: WAITLIST 상태로 booking 생성\n2) promoteFromWaitlist: 취소 시 대기 1순위→BOOKED 전환\n3) cancelBookingWithWaitlistPromotion: 취소+승격 통합\n4) 정책 모달에서 대기열 활성/자동승격 토글",
     "SCR-070", "회원ID, 스케줄ID", "대기 순번, 승격 결과", "중복 대기 방지, 정원 체크", "이미 예약/대기 중이면 거부", "businessLogic.ts + LessonManagement.tsx 정책 모달", "", "", "개발완료", "", "수업 관리", "", "브로제이 대기열 참고", ""),

    ("FN-116", "즐겨찾기 회원 FC 실시간 알림", "회원관리", "Web", "High",
     "즐겨찾기 회원 입장 시 Supabase Realtime으로 실시간 알림 (벨+toast+브라우저)",
     "1) AppHeader에서 attendance 테이블 INSERT 구독\n2) favorites 목록과 대조\n3) 매칭 시: 알림 벨 목록 추가 + toast + 브라우저 Notification\n4) 브라우저 알림 권한 요청",
     "SCR-002", "출석 이벤트", "알림 표시", "즐겨찾기 회원만 대상", "알림 권한 거부 시 toast만", "AppHeader.tsx Supabase Realtime subscription", "", "", "개발완료", "", "대시보드", "", "온핏 핵심 장점: FC 핸드폰 알림", ""),

    ("FN-117B", "얼굴 인식 출입 설정", "설정", "Web", "Medium",
     "키오스크 체크인에 얼굴 인식 방식 추가 + 설정 UI",
     "1) checkInMethods에 'face' 추가 (기본 활성)\n2) 인식 모드: 단일/다중(10명 동시)\n3) 민감도: 높음/보통/낮음\n4) 병행 인증: 전화번호/QR\n5) 자동 등록: 수동/첫 방문 시",
     "SCR-090", "얼굴 인식 설정값", "설정 저장", "카메라 장치 등록 필요 안내", "", "KioskSettings.tsx 얼굴 인식 설정 섹션", "", "", "개발완료", "", "키오스크 설정", "", "브로제이: 10명 동시 인식", ""),

    ("FN-118", "키오스크/프로젝터 자동 전원 스케줄", "설정", "Web", "Medium",
     "IoT 설정에 자동 전원 탭 추가, 장비별 ON/OFF 시간 스케줄 설정",
     "1) '자동 전원' 탭 추가\n2) 장비별 power ON/OFF 시간 설정\n3) 토글로 스케줄 활성/비활성\n4) WoL/스마트플러그/RS-232 프로토콜 안내",
     "SCR-091", "전원 스케줄 목록", "스케줄 저장", "IoT 기기 탭에서 장비 먼저 등록", "", "IotSettings.tsx 자동 전원 탭", "", "", "개발완료", "", "IoT 설정", "", "레슨북 고척점: 오픈/마감 자동 제어", ""),

    ("FN-119", "동선 분석 (구역별 이용 현황)", "출석관리", "Web", "Medium",
     "출석 페이지에 구역별 실시간 이용 현황 + 시간대별 방문 밀도 히트맵 표시",
     "1) 구역: 웨이트/유산소/GX룸/탈의실\n2) 정원 대비 사용률 프로그레스 바\n3) 혼잡도 색상 (초록/노랑/빨강)\n4) 시간대별(06~23시) 히트맵 차트",
     "SCR-020", "출석/구역 데이터", "구역별 현황 + 히트맵", "실시간 표시 (새로고침 기반)", "", "Attendance.tsx 동선 분석 섹션", "", "", "개발완료", "", "출석 기록", "", "대치점 제안: 동선 파악 + 데이터 분석", ""),

    ("FN-120", "수건/운동복 QR 관리", "시설관리", "Web", "Medium",
     "운동복에 QR 코드 부여 + QR 스캔으로 빠른 대여/반납/세탁 처리",
     "1) ClothingItem에 qrCode 필드 추가\n2) QR 스캔 모달: 코드/번호 입력 → 즉시 조회\n3) 대여중이면 반납/세탁 버튼, 대기면 대여 버튼\n4) 키오스크 연동 안내",
     "SCR-056", "QR 코드/운동복 번호", "운동복 상태 조회/변경", "미등록 QR: 에러 메시지", "", "ClothingManagement.tsx QR 스캔 모달", "", "", "개발완료", "", "수건/운동복 관리", "", "QR 기반 자동 관리", ""),
]

for feat in new_features:
    append_row(ws, feat)

print(f"기능명세서: {len(new_features)}개 기능 추가")


# ═══════════════════════════════════════════════════════════════
# 2. 비즈니스정책 시트 — 신규 정책 추가
# ═══════════════════════════════════════════════════════════════
ws = wb['비즈니스정책']

new_policies = [
    # (정책ID, 정책영역, 정책명, 정책설명, 관련화면, 관련기능, 우선순위, 상태)
    ("BP-200", "상품관리", "상품 동적 카테고리 정책",
     "1) 기본 카테고리: 이용권, PT, GX, 골프, 식품, 기타\n2) product_groups에서 자유롭게 추가 가능\n3) 분류 관리 탭에서 CRUD\n4) 미리 정의되지 않은 카테고리도 기본 설정(기간제) 적용\n5) 레슨북 스타일: 헬스장에서 식품도 팔 수 있음",
     "SCR-040, SCR-092", "FN-110, FN-106", "High", "확정"),

    ("BP-201", "상품관리", "상품 홀딩/양도/포인트 정책",
     "1) holdingEnabled: 기간정지 허용 여부 (상품별 개별 설정)\n2) transferEnabled: 양도 가능 여부 (상품별)\n3) pointAccrual: 포인트 적립 여부 (기본 true)\n4) salesChannel: 판매유형 (ALL/KIOSK/COUNTER/ONLINE)\n5) 테이블에 ✓/✗로 한눈에 확인",
     "SCR-040", "FN-111", "High", "확정"),

    ("BP-202", "상품관리", "상품 요일/시간 이용 제한 정책",
     "1) usage_restrictions JSON에 저장\n2) availableDays: 이용 가능 요일 (0=일~6=토)\n3) availableTimeStart/End: 이용 가능 시간대\n4) weekdayPrice/weekendPrice: 주중/주말 가격 분리\n5) 미설정 시 전체 이용 가능\n6) 주중권/주말권 등 가격 차별화 지원",
     "SCR-040", "FN-112", "High", "확정"),

    ("BP-203", "수업관리", "노쇼/취소 자동 정책",
     "1) 취소 마감: 수업 시작 N시간 전 (기본 3시간)\n2) 마감 후 취소: 수업 진행으로 간주 (횟수 차감)\n3) 노쇼 = 수업 진행: 세션 차감 (토글)\n4) 연속 N회 노쇼: 회원 메모에 경고 기록 (기본 3회)\n5) 미처리 수업 자동 완료: 종료 N시간 후 (기본 24시간)\n6) 지점별 독립 설정 (settings.lesson_policy)",
     "SCR-070", "FN-113", "Critical", "확정"),

    ("BP-204", "수업관리", "대기열(Waitlist) 정책",
     "1) 정원 초과 시 WAITLIST 상태로 등록 허용\n2) 대기 순번 자동 계산 (createdAt 기준)\n3) 예약 취소 발생 시 대기 1순위 자동 승격 (WAITLIST→BOOKED)\n4) 이미 예약/대기 중이면 중복 등록 거부\n5) 대기열 기능/자동승격 각각 토글 가능\n6) 브로제이 꼬리물기 참고",
     "SCR-070, SCR-054", "FN-115", "High", "확정"),

    ("BP-205", "수업관리", "GX 예약 자동 오픈 정책",
     "1) 수업 시작 N시간 전 예약 자동 오픈 (기본 48시간)\n2) lesson_schedules.status: PENDING→OPEN 전환\n3) runDailySync에서 일괄 처리\n4) 온핏 요청: N시간 전 자동 예약 오픈 기능",
     "SCR-070", "FN-114", "High", "확정"),

    ("BP-206", "출입관리", "얼굴 인식 출입 정책",
     "1) 키오스크 체크인 방식에 face 추가\n2) 인식 모드: 단일(1명)/다중(최대 10명 동시)\n3) 민감도: 높음(정확도)/보통/낮음(속도)\n4) 인식 실패 시 전화번호/QR 병행 인증\n5) 카메라 IoT 장치 등록 필수\n6) 브로제이 참고: 빠르고 정확한 인식",
     "SCR-090", "FN-117B", "Medium", "확정"),

    ("BP-207", "시설관리", "자동 전원 제어 정책",
     "1) 장비별 ON/OFF 시간 스케줄 독립 설정\n2) WoL(Wake on LAN) 또는 스마트 플러그 API\n3) 프로젝터: RS-232/IP 프로토콜\n4) 시간 종료 시 자동 종료 (에너지 절약)\n5) 레슨북 고척점: 오픈/마감 시 키오스크·프로젝터 자동",
     "SCR-091", "FN-118", "Medium", "확정"),

    ("BP-208", "출석관리", "동선 분석 정책",
     "1) 구역: 웨이트/유산소/GX룸/탈의실 (확장 가능)\n2) 구역별 정원 설정\n3) 사용률 80% 이상: 빨강(혼잡), 50~80%: 노랑, 50% 미만: 초록\n4) 시간대별 히트맵: 06~23시 밀도 시각화\n5) 대치점 제안: 동선 파악으로 운영 효율화",
     "SCR-020", "FN-119", "Medium", "확정"),

    ("BP-209", "시설관리", "QR 운동복 관리 정책",
     "1) 운동복 등록 시 QR 코드 자동 생성\n2) QR 스캔: 코드/번호 입력 → 즉시 조회\n3) 대여중 → 반납/세탁, 대기 → 대여 처리\n4) 키오스크에서 QR 스캔으로 무인 대여/반납\n5) 자동 재고 파악",
     "SCR-056", "FN-120", "Medium", "확정"),
]

for pol in new_policies:
    append_row(ws, pol)

print(f"비즈니스정책: {len(new_policies)}개 정책 추가")


# ═══════════════════════════════════════════════════════════════
# 3. 데이터 정의서 시트 — 신규 컬럼 추가
# ═══════════════════════════════════════════════════════════════
ws = wb['데이터 정의서']

new_columns = [
    # (테이블명, 컬럼명, 데이터타입, 길이, NULL, 기본값, PK, FK, 인덱스, 설명, API, 상태, 비고)
    ("products", "holdingEnabled", "BOOLEAN", "", "Y", "false", "N", "", "N", "홀딩(기간정지) 가능 여부", "", "확정", "레슨북 컬럼"),
    ("products", "transferEnabled", "BOOLEAN", "", "Y", "false", "N", "", "N", "양도 가능 여부", "", "확정", "레슨북 컬럼"),
    ("products", "pointAccrual", "BOOLEAN", "", "Y", "true", "N", "", "N", "포인트 적립 여부", "", "확정", "레슨북 컬럼"),
    ("products", "salesChannel", "TEXT", "", "Y", "ALL", "N", "", "N", "판매유형 (ALL/KIOSK/COUNTER/ONLINE)", "", "확정", "레슨북 컬럼"),
    ("products", "usage_restrictions", "JSONB", "", "Y", "NULL", "N", "", "N", "이용 제한: {availableDays, availableTimeStart/End, weekdayPrice, weekendPrice}", "", "확정", "요일/시간 제한"),
    ("clothing", "qrCode", "TEXT", "", "Y", "NULL", "N", "", "N", "운동복 QR 코드 (자동 생성)", "", "확정", "QR 관리"),
]

for col in new_columns:
    append_row(ws, col)

print(f"데이터 정의서: {len(new_columns)}개 컬럼 추가")


# ═══════════════════════════════════════════════════════════════
# 4. UI 요소 상세 시트 — 신규 UI 요소 추가
# ═══════════════════════════════════════════════════════════════
ws = wb['UI 요소 상세']

new_ui = [
    # (요소ID, 화면ID, 요소명, 유형, 설명, 인터랙션, 기능ID, APIID, Web, App, 유효성, 에러, 접근성, 상태, 담당자, 화면명, 기능명, API, 비고, 내부구성, 크기, 상태목록, 호버, 섹션, 정렬, 조건부, 토큰)
    ("UI-300", "SCR-040", "동적 카테고리 필터 탭", "TabGroup", "product_groups + 실제 상품에서 카테고리 추출하여 동적 필터 탭 표시", "탭 클릭 시 해당 카테고리 필터 적용, 카운트 표시", "FN-110", "", "ProductList.tsx 내 동적 탭", "", "", "", "", "확정", "", "상품 목록", "상품 동적 카테고리", "", "레슨북 스타일", "전체 탭 + 동적 카테고리 버튼들", "", "active/inactive", "active: bg-content, hover: bg-surface-secondary", "필터 영역", "", "", ""),
    ("UI-301", "SCR-040", "요일별 ✓/✗ 컬럼 (월~일)", "TableColumn", "상품 테이블에 요일별 이용 가능 여부를 ✓(초록)/✗(빨강)로 표시", "읽기 전용, 패널 열리면 숨김", "FN-112", "", "ProductList.tsx 7개 th/td", "", "", "", "", "확정", "", "상품 목록", "상품 요일/시간 제한", "", "레슨북 Group 3.png", "월/화/수/목/금/토/일 각 24px 컬럼", "w-24px", "", "", "테이블", "", "패널 닫힌 상태에서만", ""),
    ("UI-302", "SCR-040", "홀딩/양도/포인트 ✓/✗ 컬럼", "TableColumn", "상품 테이블에 홀딩/양도/포인트 가능 여부를 ✓/✗로 표시", "읽기 전용, 패널 열리면 숨김", "FN-111", "", "ProductList.tsx 3개 th/td", "", "", "", "", "확정", "", "상품 목록", "상품 레슨북 컬럼", "", "레슨북 컬럼 완전 반영", "", "", "", "", "테이블", "", "패널 닫힌 상태에서만", ""),
    ("UI-303", "SCR-040", "현금가/카드가 분리 컬럼", "TableColumn", "기존 '가격' 단일 컬럼을 현금가/카드가 2개로 분리 표시", "읽기 전용", "FN-111", "", "ProductList.tsx 2개 th/td", "", "", "", "", "확정", "", "상품 목록", "상품 레슨북 컬럼", "", "", "", "", "", "", "테이블", "", "항상", ""),
    ("UI-304", "SCR-070", "노쇼/취소 정책 설정 모달", "Modal", "수업 관리 헤더에 '노쇼/취소 정책' 버튼 클릭 시 열리는 설정 모달", "취소 마감, 노쇼 차감, 자동 완료, 대기열 등 정책 설정 후 저장", "FN-113, FN-114, FN-115", "", "LessonManagement.tsx Modal", "", "", "", "", "확정", "", "수업 관리", "노쇼/취소 정책", "", "6개 섹션: 취소마감/노쇼/자동처리/GX자동오픈/대기열/요약", "", "", "", "", "", "", "", ""),
    ("UI-305", "SCR-020", "구역별 이용 현황 카드", "Card", "웨이트/유산소/GX룸/탈의실 실시간 현황 (정원 대비 사용률 프로그레스 바)", "읽기 전용, 혼잡도 색상 자동", "FN-119", "", "Attendance.tsx 구역 카드 4개", "", "", "", "", "확정", "", "출석 기록", "동선 분석", "", "80%↑=빨강, 50~80%=노랑, ~50%=초록", "", "", "", "", "", "", "", ""),
    ("UI-306", "SCR-020", "시간대별 방문 밀도 히트맵", "Chart", "06~23시 시간대별 방문 밀도를 파란색 농도로 시각화", "호버 시 시간/밀도 tooltip", "FN-119", "", "Attendance.tsx 히트맵 바", "", "", "", "", "확정", "", "출석 기록", "동선 분석", "", "18개 타임슬롯, rgba 색상", "", "", "", "", "", "", "", ""),
    ("UI-307", "SCR-056", "QR 스캔 모달", "Modal", "운동복 QR 코드/번호 입력 → 즉시 조회 → 반납/세탁/대여 처리", "Enter키 조회, 상태에 따라 액션 버튼 변경", "FN-120", "", "ClothingManagement.tsx Modal", "", "", "QR 코드 미등록: 에러 toast", "", "확정", "", "수건/운동복 관리", "QR 운동복", "", "키오스크 연동 안내 포함", "", "", "", "", "", "", "", ""),
    ("UI-308", "SCR-090", "얼굴 인식 설정 섹션", "FormSection", "키오스크 설정에서 face 체크인 활성 시 표시되는 얼굴 인식 상세 설정", "인식모드/민감도/병행인증/자동등록 select", "FN-117B", "", "KioskSettings.tsx FormSection", "", "", "", "", "확정", "", "키오스크 설정", "얼굴 인식 설정", "", "face 체크인 활성 시에만 표시", "", "", "", "", "", "", "checkInMethods.includes('face')", ""),
    ("UI-309", "SCR-091", "자동 전원 스케줄 리스트", "List", "장비별 ON/OFF 시간 설정 + 토글 + 추가/삭제", "시간 입력, 토글, 삭제 버튼", "FN-118", "", "IotSettings.tsx 자동 전원 탭", "", "", "", "", "확정", "", "IoT 설정", "자동 전원", "", "activeTab=power에서만 표시", "", "", "", "", "", "", "", ""),
]

for ui in new_ui:
    append_row(ws, ui)

print(f"UI 요소 상세: {len(new_ui)}개 요소 추가")


# ═══════════════════════════════════════════════════════════════
# 저장
# ═══════════════════════════════════════════════════════════════
wb.save(DST)
wb.close()
print(f"\n✅ 저장 완료: {DST}")
print("  - 기능명세서: 11개 신규 기능")
print("  - 비즈니스정책: 10개 신규 정책")
print("  - 데이터 정의서: 6개 신규 컬럼")
print("  - UI 요소 상세: 10개 신규 요소")
