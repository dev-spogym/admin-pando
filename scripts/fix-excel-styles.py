#!/usr/bin/env python3
"""
FitGenie CRM 기획문서 v4.xlsx — 새로 추가된 행에 기존 서식 적용
"""

from copy import copy
import openpyxl

FILE_PATH = "/Users/simjaehyeong/Desktop/pando/pando-beta/excel/FitGenie CRM_기획문서_v4.xlsx"

# 각 시트별 설정: (시트 이름 일부, 새 행 시작 행번호, 참조 행번호)
SHEET_CONFIG = [
    ("IA",          67, 66),
    ("사용자 스토리", 62, 61),
    ("화면설계서",   150, 149),
    ("UI 요소 상세", 194, 193),
    ("기능명세서",   148, 147),
    ("API명세서",     69,  68),
    ("데이터 정의서", 371, 370),
    ("QA테스트케이스", 281, 280),
    ("비즈니스정책",  139, 138),
    ("화면흐름도",    56,  55),
]


def copy_style(src, dst):
    """소스 셀의 스타일을 타겟 셀에 복사"""
    if src.font:
        dst.font = copy(src.font)
    if src.border:
        dst.border = copy(src.border)
    if src.fill and src.fill.fgColor and src.fill.fgColor.rgb != "00000000":
        dst.fill = copy(src.fill)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.number_format:
        dst.number_format = src.number_format


def find_sheet(wb, name_part):
    """시트 이름에 name_part가 포함된 시트를 반환"""
    for name in wb.sheetnames:
        if name_part in name:
            return wb[name]
    return None


def apply_styles(wb, sheet_name_part, new_start_row, ref_row_num):
    ws = find_sheet(wb, sheet_name_part)
    if ws is None:
        print(f"  [SKIP] '{sheet_name_part}' 시트를 찾을 수 없음")
        return

    max_col = ws.max_column
    max_row = ws.max_row

    if new_start_row > max_row:
        print(f"  [SKIP] '{ws.title}' — 새 행 없음 (max_row={max_row}, new_start={new_start_row})")
        return

    # 참조 행 스타일 캐시
    ref_styles = {}
    for col in range(1, max_col + 1):
        ref_styles[col] = ws.cell(row=ref_row_num, column=col)

    # 참조 행 높이
    ref_height = ws.row_dimensions[ref_row_num].height

    applied = 0
    for row in range(new_start_row, max_row + 1):
        for col in range(1, max_col + 1):
            src = ref_styles[col]
            dst = ws.cell(row=row, column=col)
            copy_style(src, dst)
        # 행 높이 동일하게
        if ref_height is not None:
            ws.row_dimensions[row].height = ref_height
        applied += 1

    print(f"  [OK] '{ws.title}' — {new_start_row}~{max_row}행 ({applied}행) 서식 적용 완료")


def main():
    print(f"파일 로드: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH)
    print(f"시트 목록: {wb.sheetnames}\n")

    for name_part, new_start, ref_row in SHEET_CONFIG:
        print(f"처리 중: '{name_part}' (참조행={ref_row}, 새행시작={new_start})")
        apply_styles(wb, name_part, new_start, ref_row)

    wb.save(FILE_PATH)
    print(f"\n저장 완료: {FILE_PATH}")


if __name__ == "__main__":
    main()

