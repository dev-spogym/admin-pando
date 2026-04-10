"""
FitGenie CRM_기획문서_v4.xlsx 나머지 5개 시트 데이터 추가 스크립트
- 사용자 스토리: US-070~080 (11개)
- 화면설계서: SCR-070~076 (7개)
- UI 요소 상세: UI-700~776 (신규 페이지 요소)
- 비즈니스정책: BP-070~078 (9개)
- 화면흐름도: FLW-070~081 (12개)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

FILE_PATH = "excel/FitGenie CRM_기획문서_v4.xlsx"

wb = openpyxl.load_workbook(FILE_PATH)

# ─────────────────────────────────────────────
# 유틸: 마지막 행 스타일 복사하여 새 행 추가
# ─────────────────────────────────────────────
def append_row_with_style(ws, values):
    """마지막 데이터 행 스타일을 복사해 새 행 삽입"""
    last_row = ws.max_row
    new_row = last_row + 1

    # 스타일 복사 원본 행 (마지막 데이터 행)
    src_row = last_row

    for col_idx, value in enumerate(values, start=1):
        src_cell = ws.cell(row=src_row, column=col_idx)
        dst_cell = ws.cell(row=new_row, column=col_idx)

        dst_cell.value = value

        # 스타일 복사
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

    return new_row


# ─────────────────────────────────────────────
# 1. 사용자 스토리 (13컬럼)
# 컬럼: 스토리ID|에픽|사용자유형|사용자스토리|인수조건|화면ID|우선순위|사이즈|담당자|상태|[자동]화면명|[자동]기능수|비고
# ─────────────────────────────────────────────
ws_story = wb["사용자 스토리"]
before_story = ws_story.max_row

user_stories = [
    ["US-070", "수업관리", "매니저",
     "매니저로서 수업을 등록하고 강사를 배정하여 수업 스케줄을 관리하고 싶다",
     "1)수업명/유형/강사/정원/시간 입력 가능 2)수업 목록 조회/검색 3)수업 수정/삭제",
     "SCR-070", "High", "L", "개발자1", "Done", "수업 관리", 1, "BROJ CRM 이식"],
    ["US-071", "수업관리", "매니저",
     "매니저로서 수강권 횟수를 관리하여 회원의 남은 수업 횟수를 추적하고 싶다",
     "1)횟수 목록 조회 2)1회 차감 가능 3)잔여횟수 실시간 표시",
     "SCR-071", "High", "M", "개발자1", "Done", "횟수 관리", 1, "BROJ CRM 이식"],
    ["US-072", "수업관리", "매니저",
     "매니저로서 노쇼/지각 페널티를 관리하여 수업 출석률을 높이고 싶다",
     "1)페널티 등록(회원/유형/차감횟수) 2)페널티 목록 조회 3)페널티 삭제(취소)",
     "SCR-072", "Medium", "M", "개발자1", "Done", "페널티 관리", 1, "BROJ CRM 이식"],
    ["US-073", "매출관리", "관리자",
     "관리자로서 환불 내역을 관리하여 환불 처리를 체계화하고 싶다",
     "1)환불 목록 조회 2)기간별 필터 3)환불 통계(총액/건수/평균)",
     "SCR-073", "High", "M", "개발자1", "Done", "환불 관리", 1, "BROJ CRM 이식"],
    ["US-074", "매출관리", "매니저",
     "매니저로서 미수금을 추적하여 미결제 건을 관리하고 싶다",
     "1)미수금 목록 조회 2)상태변경(미결제→완료) 3)연체 자동표시 4)메모 편집",
     "SCR-074", "High", "M", "개발자1", "Done", "미수금 관리", 1, "BROJ CRM 이식"],
    ["US-075", "매출관리", "매니저",
     "매니저로서 매출 통계를 분석하여 상품별/결제수단별 매출을 파악하고 싶다",
     "1)상품별 매출 바차트 2)결제수단별 파이차트 3)종목별 분석 4)기간 필터",
     "SCR-075", "High", "L", "개발자1", "Done", "매출 통계", 1, "BROJ CRM 이식"],
    ["US-076", "시설관리", "스태프",
     "스태프로서 운동복을 관리하여 재고와 대여 현황을 추적하고 싶다",
     "1)운동복 등록(번호/사이즈/타입) 2)상태변경(대기/대여/세탁/파손) 3)반납 처리",
     "SCR-076", "Medium", "M", "개발자1", "Done", "운동복 관리", 1, "BROJ CRM 이식"],
    ["US-077", "공통", "매니저",
     "매니저로서 우측 알림센터에서 CRM 활동을 실시간으로 확인하고 싶다",
     "1)최근 알림 타임라인 2)미읽 배지 3)전체 읽음 처리",
     "공통", "Medium", "M", "개발자1", "Done", "알림센터", 1, "BROJ CRM 이식"],
    ["US-078", "회원관리", "매니저",
     "매니저로서 회원 이용권을 홀딩/연장/양도하여 유연하게 관리하고 싶다",
     "1)홀딩(시작일/종료일/사유) 2)연장(추가일수) 3)양도(대상회원 선택)",
     "SCR-011", "High", "L", "개발자1", "Done", "홀딩/연장/양도", 1, "BROJ CRM 이식"],
    ["US-079", "상품관리", "관리자",
     "관리자로서 상품을 타입별(회원권/수강권/대여권/일반)로 분류하여 관리하고 싶다",
     "1)상품 타입 탭 4개 2)현금가/카드가 분리 3)종목/태그 필터",
     "SCR-050", "High", "M", "개발자1", "Done", "상품타입 분리", 1, "BROJ CRM 이식"],
    ["US-080", "수업관리", "매니저",
     "매니저로서 캘린더에서 수업 일정을 필터링하고 우측 패널에서 상세를 확인하고 싶다",
     "1)강사/수업명/정원 필터 2)우측 일별 일정 리스트 3)유효 수업 탭",
     "SCR-020", "High", "L", "개발자1", "Done", "캘린더 수업 통합", 1, "BROJ CRM 이식"],
]

for row_data in user_stories:
    append_row_with_style(ws_story, row_data)

after_story = ws_story.max_row
print(f"[사용자 스토리] {before_story}행 → {after_story}행 (+{after_story - before_story}행)")


# ─────────────────────────────────────────────
# 2. 화면설계서 (31컬럼)
# ─────────────────────────────────────────────
ws_screen = wb["화면설계서"]
before_screen = ws_screen.max_row

screens = [
    # 화면ID|화면명|플랫폼|화면경로|화면설명|[자동]UI요소목록|인터랙션정의|Web특이사항|App특이사항|진입경로
    # |이동가능화면|에러화면정의|접근성체크|반응형체크|가로세로모드|디자인시안링크|상태|담당자
    # |[자동]관련기능수|[자동]관련TC수|[자동]관련API수|비고|[자동]UI요소수|레이아웃구조|섹션순서
    # |화면유형|진입트리거|Loading상태|Empty상태|반응형상세|고정요소
    ["SCR-070", "수업 관리", "Web", "/lessons",
     "수업 등록/수정/삭제 + 강사 배정 + 정원 관리",
     None, None, None, None, None,
     None, None, None, None, None, None,
     "확정", "개발팀", 1, None, 1, "BROJ CRM 이식",
     4, "[헤더]+[통계카드4]+[테이블]+[등록모달]",
     "1:헤더→2:통계카드→3:테이블", "page",
     "사이드바 '수업 관리' 클릭", "테이블 스켈레톤", "수업이 없습니다", None, "GNB 고정"],

    ["SCR-071", "횟수 관리", "Web", "/lesson-counts",
     "수강권 횟수 차감 관리",
     None, None, None, None, None,
     None, None, None, None, None, None,
     "확정", "개발팀", 1, None, 1, "BROJ CRM 이식",
     3, "[헤더]+[상태탭]+[테이블]",
     "1:헤더→2:탭→3:테이블", "page",
     "사이드바 '횟수 관리' 클릭", "테이블 스켈레톤", "횟수 데이터가 없습니다", None, "GNB 고정"],

    ["SCR-072", "페널티 관리", "Web", "/penalties",
     "노쇼/지각 페널티 등록/관리",
     None, None, None, None, None,
     None, None, None, None, None, None,
     "확정", "개발팀", 1, None, 1, "BROJ CRM 이식",
     4, "[헤더]+[통계카드4]+[테이블]+[등록모달]",
     "1:헤더→2:통계카드→3:테이블", "page",
     "사이드바 '페널티 관리' 클릭", "테이블 스켈레톤", "페널티가 없습니다", None, "GNB 고정"],

    ["SCR-073", "환불 관리", "Web", "/refunds",
     "환불 내역 조회/관리 + 통계",
     None, None, None, None, None,
     None, None, None, None, None, None,
     "확정", "개발팀", 1, None, 1, "BROJ CRM 이식",
     3, "[헤더]+[통계카드3]+[날짜필터]+[테이블]",
     "1:헤더→2:통계카드→3:필터→4:테이블", "page",
     "사이드바 '환불 관리' 클릭", "테이블 스켈레톤", "환불 내역이 없습니다", None, "GNB 고정"],

    ["SCR-074", "미수금 관리", "Web", "/unpaid",
     "미결제 내역 추적 + 상태변경",
     None, None, None, None, None,
     None, None, None, None, None, None,
     "확정", "개발팀", 1, None, 1, "BROJ CRM 이식",
     4, "[헤더]+[통계카드4]+[상태탭]+[테이블]",
     "1:헤더→2:통계카드→3:탭→4:테이블", "page",
     "사이드바 '미수금 관리' 클릭", "테이블 스켈레톤", "미수금이 없습니다", None, "GNB 고정"],

    ["SCR-075", "매출 통계", "Web", "/sales/stats",
     "상품별/결제수단별/종목별 매출 분석",
     None, None, None, None, None,
     None, None, None, None, None, None,
     "확정", "개발팀", 1, None, 1, "BROJ CRM 이식",
     5, "[헤더]+[날짜필터]+[탭4]+[차트]+[테이블]",
     "1:헤더→2:필터→3:탭→4:차트→5:테이블", "page",
     "사이드바 '매출 통계' 클릭", "차트 로딩 스피너", "매출 데이터가 없습니다", None, "GNB 고정"],

    ["SCR-076", "운동복 관리", "Web", "/clothing",
     "운동복 재고/대여/세탁/파손 관리",
     None, None, None, None, None,
     None, None, None, None, None, None,
     "확정", "개발팀", 1, None, 1, "BROJ CRM 이식",
     4, "[헤더]+[통계카드4]+[상태탭]+[테이블]+[등록모달]",
     "1:헤더→2:통계카드→3:탭→4:테이블", "page",
     "사이드바 '운동복' 클릭", "테이블 스켈레톤", "운동복 데이터가 없습니다", None, "GNB 고정"],
]

for row_data in screens:
    append_row_with_style(ws_screen, row_data)

after_screen = ws_screen.max_row
print(f"[화면설계서] {before_screen}행 → {after_screen}행 (+{after_screen - before_screen}행)")


# ─────────────────────────────────────────────
# 3. UI 요소 상세 (27컬럼)
# 컬럼: 요소ID|화면ID|요소명|요소유형|요소설명|인터랙션정의|기능ID|API ID|Web구현|App구현
#       |유효성검사|에러메시지|접근성|상태|담당자|[자동]화면명|[자동]기능명|[자동]API엔드포인트|비고
#       |내부구성|크기/간격|상태목록|호버/클릭피드백|소속섹션|정렬순서|조건부표시|관련디자인토큰
# ─────────────────────────────────────────────
ws_ui = wb["UI 요소 상세"]
before_ui = ws_ui.max_row

ui_elements = [
    # ── SCR-070 수업 관리 ──
    ["UI-700", "SCR-070", "수업 등록 버튼", "Button", "새 수업을 등록하는 기본 액션 버튼",
     "클릭 시 등록 모달 열림", "FN-070", None, "Y", "N",
     None, None, "aria-label='수업 등록'", "완료", "개발팀",
     "수업 관리", None, None, "BROJ CRM 이식",
     None, "h-10 px-4", "default/hover/disabled",
     "배경색 darken", "헤더 우측", 1, None, "color.primary"],

    ["UI-701", "SCR-070", "수업 목록 테이블", "Table", "등록된 수업 목록 표시 (수업명/유형/강사/정원/시간/색상)",
     "행 클릭 시 수정 모달", "FN-070", "API-070", "Y", "N",
     None, None, "role='table'", "완료", "개발팀",
     "수업 관리", None, None, "BROJ CRM 이식",
     "수업명|유형|강사|정원|시간|색상|액션", "w-full", "loaded/empty/loading",
     None, "메인 컨텐츠", 2, None, None],

    ["UI-702", "SCR-070", "등록 모달", "Modal", "수업 등록/수정 폼 모달 (수업명/유형/강사/정원/시간/색상)",
     "저장 클릭 시 API 호출 후 목록 갱신, ESC/배경 클릭 시 닫힘", "FN-070", "API-070", "Y", "N",
     "수업명 필수, 정원 1 이상 정수", "수업명을 입력해주세요", "role='dialog'", "완료", "개발팀",
     "수업 관리", None, None, "BROJ CRM 이식",
     "수업명Input|유형Select|강사Select|정원Input|시간Input|색상Picker", "w-[480px]", "open/closed",
     None, "오버레이", 3, None, None],

    ["UI-703", "SCR-070", "수정/삭제 버튼", "ButtonGroup", "테이블 행별 수정·삭제 액션",
     "수정: 모달 열기, 삭제: confirm 다이얼로그 후 API 호출", "FN-070", "API-070", "Y", "N",
     None, "삭제하시겠습니까?", None, "완료", "개발팀",
     "수업 관리", None, None, "BROJ CRM 이식",
     "수정Button|삭제Button", "gap-2", "default/hover",
     "아이콘 색상 변경", "테이블 액션 컬럼", 4, None, None],

    # ── SCR-071 횟수 관리 ──
    ["UI-710", "SCR-071", "상태 탭", "Tabs", "전체/이용중/만료/일시정지 탭 전환",
     "탭 클릭 시 테이블 필터링", "FN-071", None, "Y", "N",
     None, None, "role='tablist'", "완료", "개발팀",
     "횟수 관리", None, None, "BROJ CRM 이식",
     "전체|이용중|만료|일시정지", "h-10", "active/inactive",
     "하단 보더 표시", "테이블 상단", 1, None, None],

    ["UI-711", "SCR-071", "횟수 테이블", "Table", "회원별 수강권 잔여 횟수 목록 (회원명/상품/잔여/총횟수/상태)",
     None, "FN-071", "API-071", "Y", "N",
     None, None, "role='table'", "완료", "개발팀",
     "횟수 관리", None, None, "BROJ CRM 이식",
     "회원명|상품|잔여횟수|총횟수|상태|액션", "w-full", "loaded/empty/loading",
     None, "메인 컨텐츠", 2, None, None],

    ["UI-712", "SCR-071", "1회 차감 버튼", "Button", "해당 회원의 수강권에서 1회 차감",
     "클릭 시 confirm 후 API 호출, 잔여횟수 즉시 업데이트", "FN-071", "API-071", "Y", "N",
     "잔여횟수 > 0", "잔여 횟수가 없습니다", "aria-label='1회 차감'", "완료", "개발팀",
     "횟수 관리", None, None, "BROJ CRM 이식",
     None, "h-8 px-3 text-sm", "default/disabled",
     "배경색 darken", "테이블 액션 컬럼", 3, "잔여횟수 > 0", None],

    ["UI-713", "SCR-071", "차감 이력 모달", "Modal", "해당 회원의 횟수 차감 이력 타임라인",
     "이력 클릭 시 모달 열기", "FN-071", "API-071", "Y", "N",
     None, None, "role='dialog'", "완료", "개발팀",
     "횟수 관리", None, None, "BROJ CRM 이식",
     "날짜|차감수|사유|처리자", "w-[400px]", "open/closed",
     None, "오버레이", 4, None, None],

    # ── SCR-072 페널티 관리 ──
    ["UI-720", "SCR-072", "페널티 등록 모달", "Modal", "노쇼/지각 페널티 등록 폼 (회원검색/유형/차감횟수/사유)",
     "저장 시 API 호출 후 목록 갱신", "FN-073", "API-072", "Y", "N",
     "회원 필수, 차감횟수 1 이상", "회원을 선택해주세요", "role='dialog'", "완료", "개발팀",
     "페널티 관리", None, None, "BROJ CRM 이식",
     "회원검색Input|유형Select|차감횟수Input|사유Textarea", "w-[480px]", "open/closed",
     None, "오버레이", 1, None, None],

    ["UI-721", "SCR-072", "페널티 테이블", "Table", "등록된 페널티 목록 (회원/유형/차감횟수/사유/등록일)",
     None, "FN-073", "API-072", "Y", "N",
     None, None, "role='table'", "완료", "개발팀",
     "페널티 관리", None, None, "BROJ CRM 이식",
     "회원명|페널티유형|차감횟수|사유|등록일|액션", "w-full", "loaded/empty/loading",
     None, "메인 컨텐츠", 2, None, None],

    ["UI-722", "SCR-072", "삭제(취소) 버튼", "Button", "페널티 삭제(페널티 취소) 액션",
     "클릭 시 confirm 후 API 삭제, 목록 갱신", "FN-073", "API-072", "Y", "N",
     None, "페널티를 삭제하시겠습니까?", None, "완료", "개발팀",
     "페널티 관리", None, None, "BROJ CRM 이식",
     None, "h-8 px-3 text-sm", "default/hover",
     "색상 red", "테이블 액션 컬럼", 3, None, "color.danger"],

    # ── SCR-073 환불 관리 ──
    ["UI-730", "SCR-073", "날짜 프리셋 버튼", "ButtonGroup", "오늘/이번주/이번달 날짜 빠른 선택",
     "클릭 시 날짜 범위 자동 설정 후 테이블 갱신", "FN-074", None, "Y", "N",
     None, None, "role='group'", "완료", "개발팀",
     "환불 관리", None, None, "BROJ CRM 이식",
     "오늘|이번주|이번달", "h-9 gap-2", "active/inactive",
     "배경색 filled", "필터 영역", 1, None, None],

    ["UI-731", "SCR-073", "환불 테이블", "Table", "환불 내역 목록 (회원/상품/환불금액/처리일/처리자)",
     None, "FN-074", "API-073", "Y", "N",
     None, None, "role='table'", "완료", "개발팀",
     "환불 관리", None, None, "BROJ CRM 이식",
     "회원명|상품|환불금액|처리일|처리자", "w-full", "loaded/empty/loading",
     None, "메인 컨텐츠", 2, None, None],

    ["UI-732", "SCR-073", "통계 카드", "StatCard", "총 환불액/건수/평균 환불액 3개 카드",
     None, "FN-074", "API-073", "Y", "N",
     None, None, None, "완료", "개발팀",
     "환불 관리", None, None, "BROJ CRM 이식",
     "총환불액|건수|평균환불액", "grid-cols-3 gap-4", "loaded/loading",
     None, "헤더 하단", 3, None, None],

    # ── SCR-074 미수금 관리 ──
    ["UI-740", "SCR-074", "상태 탭", "Tabs", "미결제/일부/연체/완료 탭 전환",
     "탭 클릭 시 테이블 필터링", "FN-075", None, "Y", "N",
     None, None, "role='tablist'", "완료", "개발팀",
     "미수금 관리", None, None, "BROJ CRM 이식",
     "미결제|일부납부|연체|완료", "h-10", "active/inactive",
     "하단 보더", "테이블 상단", 1, None, None],

    ["UI-741", "SCR-074", "결제완료 버튼", "Button", "미수금 건을 완료 처리하는 액션",
     "클릭 시 상태 '완료' 변경 API 호출", "FN-075", "API-074", "Y", "N",
     None, None, "aria-label='결제완료 처리'", "완료", "개발팀",
     "미수금 관리", None, None, "BROJ CRM 이식",
     None, "h-8 px-3 text-sm", "default/disabled",
     "배경색 green", "테이블 액션 컬럼", 2, "상태 != 완료", "color.success"],

    ["UI-742", "SCR-074", "메모 편집 모달", "Modal", "미수금 건 메모 입력/수정",
     "저장 시 메모 업데이트 API 호출", "FN-075", "API-074", "Y", "N",
     None, None, "role='dialog'", "완료", "개발팀",
     "미수금 관리", None, None, "BROJ CRM 이식",
     "메모Textarea", "w-[400px]", "open/closed",
     None, "오버레이", 3, None, None],

    # ── SCR-075 매출 통계 ──
    ["UI-750", "SCR-075", "분석 탭", "Tabs", "상품별/타입별/결제수단별/종목별 분석 탭",
     "탭 클릭 시 차트/테이블 전환", "FN-076", None, "Y", "N",
     None, None, "role='tablist'", "완료", "개발팀",
     "매출 통계", None, None, "BROJ CRM 이식",
     "상품별|타입별|결제수단별|종목별", "h-10", "active/inactive",
     "하단 보더", "차트 상단", 1, None, None],

    ["UI-751", "SCR-075", "바 차트", "Chart", "상품별/종목별 매출 수직 바 차트",
     "바 hover 시 툴팁 표시", "FN-076", "API-075", "Y", "N",
     None, None, "aria-label='매출 바 차트'", "완료", "개발팀",
     "매출 통계", None, None, "BROJ CRM 이식",
     "X축:상품명, Y축:매출액, 툴팁", "h-[300px]", "loaded/loading/empty",
     "바 hover darken", "차트 영역", 2, None, None],

    ["UI-752", "SCR-075", "도넛 차트", "Chart", "결제수단별 매출 비율 도넛 차트",
     "세그먼트 hover 시 툴팁 표시", "FN-076", "API-075", "Y", "N",
     None, None, "aria-label='결제수단별 차트'", "완료", "개발팀",
     "매출 통계", None, None, "BROJ CRM 이식",
     "레이블:결제수단, 값:비율, 툴팁", "h-[300px]", "loaded/loading/empty",
     "세그먼트 hover opacity", "차트 영역", 3, "결제수단별 탭 선택 시", None],

    # ── SCR-076 운동복 관리 ──
    ["UI-760", "SCR-076", "운동복 등록 모달", "Modal", "운동복 등록 폼 (번호/사이즈/타입)",
     "저장 시 API 호출 후 목록 갱신", "FN-077", "API-076", "Y", "N",
     "번호 필수, 사이즈 필수", "운동복 번호를 입력해주세요", "role='dialog'", "완료", "개발팀",
     "운동복 관리", None, None, "BROJ CRM 이식",
     "번호Input|사이즈Select|타입Select", "w-[400px]", "open/closed",
     None, "오버레이", 1, None, None],

    ["UI-761", "SCR-076", "상태변경 버튼", "ButtonGroup", "대기/대여/세탁/파손 상태 변경 (반납 포함)",
     "클릭 시 상태 변경 API 호출 후 행 업데이트", "FN-077", "API-076", "Y", "N",
     None, None, "aria-label='상태 변경'", "완료", "개발팀",
     "운동복 관리", None, None, "BROJ CRM 이식",
     "반납Button|세탁Button|파손Button", "gap-1", "default/active",
     "상태별 색상 뱃지", "테이블 액션 컬럼", 2, None, None],
]

for row_data in ui_elements:
    append_row_with_style(ws_ui, row_data)

after_ui = ws_ui.max_row
print(f"[UI 요소 상세] {before_ui}행 → {after_ui}행 (+{after_ui - before_ui}행)")


# ─────────────────────────────────────────────
# 4. 비즈니스정책 (8컬럼)
# 컬럼: 정책ID|정책영역|정책명|정책설명|관련화면|관련기능|우선순위|상태
# ─────────────────────────────────────────────
ws_bp = wb["비즈니스정책"]
before_bp = ws_bp.max_row

policies = [
    ["BP-070", "수업관리", "수업 정원 정책",
     "정원 초과 예약 불가. 대기열 미지원 (v1)",
     "SCR-070", "FN-070", "High", "확정"],
    ["BP-071", "수업관리", "페널티 정책",
     "노쇼 시 1회 차감, 지각취소(수업 1시간 전) 시 1회 차감, 지각은 경고만",
     "SCR-072", "FN-073", "High", "확정"],
    ["BP-072", "매출관리", "환불 정책",
     "결제금액 초과 환불 불가. 환불 시 원결제 상태 '환불' 변경. 부분환불 가능",
     "SCR-073", "FN-074", "Critical", "확정"],
    ["BP-073", "매출관리", "미수금 정책",
     "결제기한 30일 초과 시 자동 '연체' 처리. 연체 시 알림 발송",
     "SCR-074", "FN-075", "High", "확정"],
    ["BP-074", "회원관리", "이용권 홀딩 정책",
     "홀딩 중 재홀딩 불가. 홀딩 기간만큼 이용권 종료일 자동 연장. 홀딩 최대 30일",
     "SCR-011", "FN-079", "High", "확정"],
    ["BP-075", "회원관리", "이용권 연장 정책",
     "연장 사유 필수 입력. 연장일수만큼 종료일 변경. 연장 이력 기록",
     "SCR-011", "FN-079", "High", "확정"],
    ["BP-076", "회원관리", "이용권 양도 정책",
     "양도 시 원회원 이용권 비활성, 대상회원에 동일 상품 새로 생성. 양도 수수료 없음(v1)",
     "SCR-011", "FN-079", "High", "확정"],
    ["BP-077", "시설관리", "운동복 대여 정책",
     "대여 시 회원ID 필수. 반납 시 자동 세탁중 전환 옵션. 파손 시 별도 관리",
     "SCR-076", "FN-077", "Medium", "확정"],
    ["BP-078", "상품관리", "상품 타입 정책",
     "모든 상품은 회원권/수강권/대여권/일반 중 하나의 타입 필수. 수강권은 총횟수 필수 입력",
     "SCR-050", "FN-080", "High", "확정"],
]

for row_data in policies:
    append_row_with_style(ws_bp, row_data)

after_bp = ws_bp.max_row
print(f"[비즈니스정책] {before_bp}행 → {after_bp}행 (+{after_bp - before_bp}행)")


# ─────────────────────────────────────────────
# 5. 화면흐름도 (10컬럼)
# 컬럼: 흐름ID|출발화면|트리거요소|트리거액션|도착화면|전환방식|전달데이터|뒤로가기동작|조건|비고
# ─────────────────────────────────────────────
ws_flow = wb["화면흐름도"]
before_flow = ws_flow.max_row

flows = [
    ["FLW-070", "AppSidebar", "수업 관리 메뉴", "click",
     "SCR-070 (수업 관리)", "page", "없음", "사이드바 이전 페이지", "매니저 이상 권한", None],
    ["FLW-071", "AppSidebar", "횟수 관리 메뉴", "click",
     "SCR-071 (횟수 관리)", "page", "없음", "사이드바 이전 페이지", "매니저 이상 권한", None],
    ["FLW-072", "AppSidebar", "페널티 관리 메뉴", "click",
     "SCR-072 (페널티 관리)", "page", "없음", "사이드바 이전 페이지", "매니저 이상 권한", None],
    ["FLW-073", "AppSidebar", "환불 관리 메뉴", "click",
     "SCR-073 (환불 관리)", "page", "없음", "사이드바 이전 페이지", "관리자 이상 권한", None],
    ["FLW-074", "AppSidebar", "미수금 관리 메뉴", "click",
     "SCR-074 (미수금 관리)", "page", "없음", "사이드바 이전 페이지", "매니저 이상 권한", None],
    ["FLW-075", "AppSidebar", "매출 통계 메뉴", "click",
     "SCR-075 (매출 통계)", "page", "없음", "사이드바 이전 페이지", "매니저 이상 권한", None],
    ["FLW-076", "AppSidebar", "운동복 메뉴", "click",
     "SCR-076 (운동복 관리)", "page", "없음", "사이드바 이전 페이지", "스태프 이상 권한", None],
    ["FLW-077", "SCR-011 (회원상세)", "상세내역 탭 > 홀딩하기", "click",
     "SCR-011 (홀딩 모달)", "modal", "memberId, productId", "모달 닫기", "활성 이용권 존재 시", None],
    ["FLW-078", "SCR-011 (회원상세)", "결제내역 탭", "click",
     "SCR-011 (결제내역)", "tab", "memberId", "탭 전환", "항상", None],
    ["FLW-079", "SCR-011 (회원상세)", "예약내역 탭", "click",
     "SCR-011 (예약내역)", "tab", "memberId", "탭 전환", "항상", None],
    ["FLW-080", "RightQuickPanel", "알림센터 버튼", "click",
     "뉴스피드 패널", "slide-panel", "branchId", "패널 닫기", "항상", None],
    ["FLW-081", "RightQuickPanel", "방문회원 버튼", "click",
     "방문회원 패널", "slide-panel", "branchId", "패널 닫기", "항상", None],
]

for row_data in flows:
    append_row_with_style(ws_flow, row_data)

after_flow = ws_flow.max_row
print(f"[화면흐름도] {before_flow}행 → {after_flow}행 (+{after_flow - before_flow}행)")


# ─────────────────────────────────────────────
# 저장
# ─────────────────────────────────────────────
wb.save(FILE_PATH)
print(f"\n저장 완료: {FILE_PATH}")
print("\n=== 최종 요약 ===")
print(f"사용자 스토리: {before_story} → {after_story} (+{after_story - before_story}행)")
print(f"화면설계서:   {before_screen} → {after_screen} (+{after_screen - before_screen}행)")
print(f"UI 요소 상세: {before_ui} → {after_ui} (+{after_ui - before_ui}행)")
print(f"비즈니스정책: {before_bp} → {after_bp} (+{after_bp - before_bp}행)")
print(f"화면흐름도:   {before_flow} → {after_flow} (+{after_flow - before_flow}행)")

