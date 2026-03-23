"""
FitGenie CRM 기획문서 v4 엑셀에 QA 리포트 결과를 반영하는 스크립트
- QA테스트케이스 시트: BUG/UX 항목 TC 행 추가
- 기능명세서 시트: 수정된 기능 비고 컬럼 업데이트
- 화면설계서 시트: 수정된 화면 비고 업데이트
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

EXCEL_PATH = "/Users/simjaehyeong/Desktop/pando/pando-beta/excel/FitGenie CRM_기획문서_v4.xlsx"

# ─────────────────────────────────────────────
# 1. QA 데이터 정의
# ─────────────────────────────────────────────

# QA테스트케이스 시트에 추가할 TC 행들
# (tc_id, platform, area, sub_area, test_type, scenario, precondition, steps, expected, actual, status, priority, severity, env)

QA_TC_ROWS = [
    # ── QA-REPORT-2026-03-17: BUG-01~05, UX-01~05 ──
    (
        "TC-BUG-01", "Web", "지점관리", "지점비교리포트", "권한테스트",
        "admin(primary) 역할로 지점비교리포트 접근",
        "admin 계정 로그인 완료",
        "1) admin으로 로그인\n2) /branch-report 경로 접근",
        "primary 역할은 모든 페이지 접근 가능해야 함",
        "403 '접근 권한이 없습니다' 페이지 표시",
        "수정완료", "P1", "High", "localhost:5173",
    ),
    (
        "TC-BUG-02", "Web", "슈퍼관리자", "슈퍼대시보드", "권한테스트",
        "admin 역할로 슈퍼대시보드 접근 시 403 차단 여부 확인",
        "admin 계정 로그인 완료",
        "1) admin으로 로그인\n2) /super-dashboard 경로 접근",
        "isSuperAdmin 정책에 따라 접근 차단 또는 허용",
        "admin 역할로 접근 시 403 차단 (비즈니스 확인 필요)",
        "수정완료", "P2", "High", "localhost:5173",
    ),
    (
        "TC-BUG-03", "Web", "감사로그", "감사로그 조회", "기능테스트",
        "감사로그 페이지 진입 시 정상 데이터 표시 확인",
        "admin 계정 로그인 완료",
        "1) admin으로 로그인\n2) /audit-log 접근\n3) 데이터 로드 확인",
        "감사 로그 목록 정상 표시",
        "페이지 진입 시 '감사 로그 조회에 실패했습니다' toast 에러, 데이터 0건",
        "수정완료", "P0", "High", "localhost:5173",
    ),
    (
        "TC-BUG-04", "Web", "지점관리", "지점 상태", "기능테스트",
        "지점 상태(운영중/임시휴업/폐점) 매핑 정확성 확인",
        "admin 계정 로그인, 시드 데이터(강남본점/서초점/송파점) 존재",
        "1) admin으로 로그인\n2) /branches 접근\n3) 각 지점 상태 레이블 확인",
        "강남본점·서초점은 '운영중', 송파점은 '임시휴업'으로 표시",
        "3개 지점 모두 '폐점' 상태로 표시, DB ACTIVE → UI '폐점' 매핑 오류",
        "수정완료", "P0", "High", "localhost:5173",
    ),
    (
        "TC-BUG-05", "Web", "시설관리", "락커관리", "기능테스트",
        "락커관리 페이지에서 락커 목록 정상 표시 확인",
        "admin 계정 로그인, 시드 데이터 50개 락커 존재",
        "1) admin으로 로그인\n2) /locker 접근\n3) 락커 목록 수 확인",
        "시드 데이터 기준 50개 락커 표시",
        "전체 락커 0개, 검색 결과 없음 (사물함관리 /locker/management는 50개 정상)",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
    (
        "TC-UX-01", "Web", "대시보드", "트레이너 데이터 격리", "UX검증",
        "trainer 역할 로그인 시 담당 회원만 대시보드에 표시되는지 확인",
        "trainer1 계정 로그인 완료",
        "1) trainer1으로 로그인\n2) 대시보드 진입\n3) 회원 수·매출 데이터 확인",
        "코치 담당 회원만 표시 (전체 데이터 노출 금지)",
        "전체 회원 23명, 매출 586만 표시 — 담당 회원 필터링 미적용",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
    (
        "TC-UX-02", "Web", "출석관리", "빈 화면 UX", "UX검증",
        "출석 데이터 없을 때 Empty State 및 CTA 안내 표시 확인",
        "출석 데이터 0건 상태",
        "1) admin으로 로그인\n2) /attendance 접근\n3) 빈 상태 UI 확인",
        "'조회된 데이터가 없습니다' 메시지 + 수동 출석 CTA 강조 표시",
        "'조회된 데이터가 없습니다' 메시지만 표시, 빈 공간 과도함",
        "수정완료", "P3", "Low", "localhost:5173",
    ),
    (
        "TC-UX-03", "Web", "마일리지", "요약 카드 일관성", "UX검증",
        "마일리지 요약 카드 발행/사용/잔여 집계 일관성 확인",
        "마일리지 데이터 존재",
        "1) admin으로 로그인\n2) /mileage 접근\n3) 통계 카드 수치 확인",
        "전체 발행 = 전체 사용 + 잔여 마일리지",
        "'전체 발행: 0', '전체 사용: 0'이지만 '잔여: 395,599' — 집계 불일치",
        "수정완료", "P2", "Low", "localhost:5173",
    ),
    (
        "TC-UX-04", "Web", "회원관리", "회원이전 직접 접근", "UX검증",
        "회원 ID 없이 /members/transfer 직접 접근 시 안내 처리 확인",
        "admin 계정 로그인",
        "1) admin으로 로그인\n2) URL에 /members/transfer 직접 입력\n3) 메시지 확인",
        "명확한 안내 메시지 표시 또는 사이드바 메뉴 미노출로 접근 차단",
        "'잘못된 접근입니다. 회원 ID가 없습니다' toast — URL 직접 입력 시 사용자 혼란",
        "수정완료", "P3", "Low", "localhost:5173",
    ),
    (
        "TC-UX-05", "Web", "전자계약", "회원 상태 배지", "UX검증",
        "전자계약 위저드 회원 선택 단계에서 활성 회원 상태 배지 정확성 확인",
        "admin 계정 로그인, 활성 회원 데이터 존재",
        "1) admin으로 로그인\n2) /contracts/new 접근\n3) 회원 목록 상태 배지 확인",
        "활성 회원은 '활성', 만료 회원은 '만료' 배지로 정확히 표시",
        "모든 회원이 '만료' 상태 배지로 표시",
        "수정완료", "P3", "Low", "localhost:5173",
    ),

    # ── QA-LOG-대시보드: BUG-13, BUG-14, UX-13 ──
    (
        "TC-BUG-13", "Web", "대시보드", "헤더 권한 체크", "권한테스트",
        "readonly 사용자에게 헤더 '+ 회원등록' 버튼 미표시 확인",
        "readonly1 계정 로그인 완료",
        "1) readonly1으로 로그인\n2) /payroll/statements 접근\n3) 헤더 '+ 회원등록' 버튼 표시 여부 확인",
        "readonly 역할에게 '+ 회원등록' 버튼 미표시",
        "헤더 '+ 회원등록' 버튼이 readonly 사용자에게 표시됨",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
    (
        "TC-BUG-14", "Web", "대시보드", "급여명세서 권한", "권한테스트",
        "readonly1이 '직원 선택' 드롭다운으로 다른 직원 급여 조회 불가 확인",
        "readonly1 계정 로그인 완료",
        "1) readonly1으로 로그인\n2) /payroll/statements 접근\n3) 직원 선택 드롭다운 변경 시도",
        "readonly 역할은 본인 급여명세서만 조회 가능",
        "readonly1이 '직원 선택' 드롭다운으로 다른 직원 급여 조회 가능",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
    (
        "TC-UX-13", "Web", "대시보드", "403 페이지 CTA", "UX검증",
        "403 페이지 '대시보드로 이동' 버튼이 readonly 역할에 적합한지 확인",
        "readonly1 계정 로그인",
        "1) readonly1으로 로그인\n2) 403 페이지(/forbidden) 이동\n3) '대시보드로 이동' 버튼 클릭",
        "readonly 역할에게는 접근 가능한 페이지(급여명세서)로 안내",
        "readonly는 대시보드 접근 불가 — '대시보드로 이동' 클릭 시 무한 403 루프",
        "수정완료", "P2", "Medium", "localhost:5173",
    ),

    # ── QA-LOG-회원탭: BUG-09~12, UX-10~12 ──
    (
        "TC-BUG-09", "Web", "출석관리", "출석유형 필터", "기능테스트",
        "출석유형/체크인방식 필터 선택 시 테이블 정상 필터링 확인",
        "admin 계정 로그인, 출석 데이터 존재",
        "1) admin으로 로그인\n2) /attendance 접근\n3) 출석유형 필터 'PT' 선택\n4) 테이블 결과 확인",
        "PT 유형 출석 건만 테이블에 표시",
        "SearchFilter 선택해도 테이블 필터링 안 됨 (전체 데이터 그대로 표시)",
        "수정완료", "P1", "High", "localhost:5173",
    ),
    (
        "TC-BUG-10", "Web", "전자계약", "Step2 상품 목록", "기능테스트",
        "전자계약 위저드 Step2에서 카테고리별 상품 정상 표시 확인",
        "admin 계정 로그인, 상품 데이터 존재",
        "1) admin으로 로그인\n2) /contracts/new → 회원 선택\n3) Step2 각 탭(시설이용/1:1수업/그룹수업/옵션) 확인",
        "각 탭에 해당 카테고리 상품 목록 표시",
        "모든 탭 비어있음 — DB 카테고리(PT/MEMBERSHIP)와 탭 매핑 불일치",
        "수정완료", "P1", "High", "localhost:5173",
    ),
    (
        "TC-BUG-11", "Web", "마일리지", "통계카드 집계", "기능테스트",
        "마일리지 통계카드 발행/사용 합계 집계 정확성 확인",
        "admin 계정 로그인, 마일리지 데이터 존재",
        "1) admin으로 로그인\n2) /mileage 접근\n3) 통계카드 수치 확인",
        "전체 발행 마일리지 = 회원별 발행 합계 정확 표시",
        "'전체 발행: 0', '전체 사용: 0'이나 '잔여: 395,599' — reduce 집계 미적용",
        "수정완료", "P2", "Medium", "localhost:5173",
    ),
    (
        "TC-BUG-12", "Web", "마일리지", "아이콘 렌더링", "UI테스트",
        "마일리지 테이블 연락처 컬럼 아이콘 정상 렌더링 확인",
        "admin 계정 로그인",
        "1) admin으로 로그인\n2) /mileage 접근\n3) 연락처 컬럼 아이콘 확인",
        "Smartphone 아이콘 정상 표시",
        "연락처 셀에 □ 문자 표시 — Smartphone 아이콘 렌더링 깨짐",
        "수정완료", "P2", "Medium", "localhost:5173",
    ),
    (
        "TC-UX-10", "Web", "출석관리", "날짜 레이블", "UX검증",
        "과거 날짜 선택 시 출석 통계카드 날짜 레이블 정확성 확인",
        "admin 계정 로그인",
        "1) admin으로 로그인\n2) /attendance 접근\n3) 과거 날짜(예: 3/11)로 이동\n4) 통계카드 레이블 확인",
        "선택된 날짜 기준 'MM/DD 출석'으로 표시",
        "과거 날짜 선택해도 '오늘 출석'으로 하드코딩 표시",
        "수정완료", "P2", "Medium", "localhost:5173",
    ),
    (
        "TC-UX-11", "Web", "전자계약", "회원 상태 배지(재확인)", "UX검증",
        "전자계약 위저드 Step1 회원 상태 배지 정확성 재확인",
        "admin 계정 로그인, 활성/만료 회원 혼재",
        "1) admin으로 로그인\n2) /contracts/new 접근\n3) 회원 목록 상태 배지 확인",
        "회원별 실제 상태 배지 표시",
        "모든 회원 '만료' 배지 (UX-05 재확인)",
        "수정완료", "P3", "Low", "localhost:5173",
    ),
    (
        "TC-UX-12", "Web", "전자계약", "회원 선택 후 스크롤", "UX검증",
        "전자계약 위저드에서 회원 선택 후 '다음 단계' 버튼 가시성 확인",
        "admin 계정 로그인",
        "1) admin으로 로그인\n2) /contracts/new 접근\n3) 회원 선택\n4) '다음 단계' 버튼 위치 확인",
        "회원 선택 후 자동 스크롤로 '다음 단계' 버튼 화면에 표시",
        "회원 선택 후 '다음 단계' 버튼이 화면 밖에 위치, 수동 스크롤 필요",
        "수정완료", "P3", "Low", "localhost:5173",
    ),

    # ── QA-LOG-회원목록: BUG-15~17 ──
    (
        "TC-BUG-15", "Web", "회원관리", "엑셀 다운로드 권한(fc)", "권한테스트",
        "fc(trainer1) 역할에게 엑셀 다운로드 버튼 미표시 확인",
        "trainer1 계정 로그인 완료",
        "1) trainer1으로 로그인\n2) /members 접근\n3) '엑셀 다운로드' 버튼 표시 여부 확인",
        "fc 역할에게 '엑셀 다운로드' 버튼 미표시 (primary/owner/manager만 가능)",
        "'엑셀 다운로드' 버튼이 fc(trainer1)에게 표시됨",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
    (
        "TC-BUG-16", "Web", "회원관리", "회원 추가 권한(fc)", "권한테스트",
        "fc(trainer1) 역할에게 '회원 추가' 버튼 미표시 확인",
        "trainer1 계정 로그인 완료",
        "1) trainer1으로 로그인\n2) /members 접근\n3) '회원 추가' 버튼 표시 여부 확인",
        "fc 역할에게 '회원 추가' 버튼 미표시 (/members/new 접근 불가)",
        "'회원 추가' 버튼이 fc(trainer1)에게 표시됨",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
    (
        "TC-BUG-17", "Web", "회원관리", "엑셀 다운로드 권한(staff)", "권한테스트",
        "staff 역할에게 엑셀 다운로드 버튼 미표시 확인",
        "staff1 계정 로그인 완료",
        "1) staff1으로 로그인\n2) /members 접근\n3) '엑셀 다운로드' 버튼 표시 여부 확인",
        "staff 역할에게 '엑셀 다운로드' 버튼 미표시 (primary/owner/manager만 가능)",
        "'엑셀 다운로드' 버튼이 staff에게 표시됨",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),

    # ── QA-LOG-매출: BUG-21 ──
    (
        "TC-BUG-21", "Web", "매출관리", "POS 결제 저장", "기능테스트",
        "POS 결제 완료 시 sales 테이블 정상 저장 확인",
        "admin 계정 로그인, 상품 및 회원 데이터 존재",
        "1) admin으로 로그인\n2) /pos 접근\n3) 상품 선택 → 장바구니 추가\n4) 결제하기 → 회원 선택\n5) 결제 완료 클릭",
        "결제 데이터 sales 테이블 정상 저장, 매출현황 반영",
        "POS 결제 완료 시 500 에러 — sales 테이블에 description 컬럼 미존재",
        "수정완료", "P0", "High", "localhost:5173",
    ),

    # ── QA-LOG-상품: BUG-22, BUG-23 ──
    (
        "TC-BUG-22", "Web", "상품관리", "상품 등록 버튼 권한(manager)", "권한테스트",
        "manager 역할에게 상품 등록 버튼 미표시 확인",
        "manager1 계정 로그인 완료",
        "1) manager1으로 로그인\n2) /products 접근\n3) '상품 등록' 버튼 표시 여부 확인",
        "manager 역할에게 '상품 등록' 버튼 미표시 (productCreate 권한 없음)",
        "'상품 등록' 버튼이 manager에게 표시됨 — productCreate 권한 체크 누락",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
    (
        "TC-BUG-23", "Web", "상품관리", "수정/삭제 아이콘 권한(manager)", "권한테스트",
        "manager 역할에게 상품 수정/삭제 아이콘 미표시 확인",
        "manager1 계정 로그인 완료",
        "1) manager1으로 로그인\n2) /products 접근\n3) 상품 행의 수정/삭제 아이콘 표시 여부 확인",
        "manager 역할에게 수정/삭제 아이콘 미표시 (productEdit 권한 없음)",
        "수정/삭제 아이콘이 manager에게 표시됨 — productEdit 권한 체크 누락",
        "수정완료", "P1", "Medium", "localhost:5173",
    ),
]

# ─────────────────────────────────────────────
# 2. 기능명세서 업데이트 정의
# ─────────────────────────────────────────────
# {fn_id: 추가할 비고 텍스트}
FN_NOTE_UPDATES = {
    "FN-002": "BUG-13,14 수정: readonly 권한체크 (2026-03-17 QA 수정완료)",
    "FN-003": "BUG-15,16,17 수정: RBAC 권한별 버튼 표시 (2026-03-17 QA 수정완료)",
    "FN-004": "BUG-08~17 수정: RBAC 권한, 필터, 아이콘 (2026-03-17 QA 수정완료)",
    "FN-008": "BUG-09 수정: 출석유형 필터 정상화 (2026-03-17 QA 수정완료)",
    "FN-010": "BUG-21 수정: POS 결제 매핑 (2026-03-17 QA 수정완료)",
    "FN-011": "BUG-21 수정: POS sales 컬럼 매핑 수정 (2026-03-17 QA 수정완료)",
    "FN-012": "BUG-21 수정: description→productName 컬럼 수정 (2026-03-17 QA 수정완료)",
    "FN-013": "BUG-22,23 수정: 권한별 버튼 표시 (2026-03-17 QA 수정완료)",
    "FN-014": "BUG-22,23 수정: productCreate/productEdit 권한 체크 추가 (2026-03-17 QA 수정완료)",
    "FN-026": "BUG-11,12 수정: 마일리지 집계 및 아이콘 렌더링 (2026-03-17 QA 수정완료)",
}

# ─────────────────────────────────────────────
# 3. 화면설계서 업데이트 정의
# ─────────────────────────────────────────────
# {scr_id: 추가할 비고 텍스트}  — 화면설계서는 col17=상태, col18=담당자, col19=[자동]관련기능수
# 비고 컬럼이 없으므로 상태(col17) 옆에 별도 처리
SCR_NOTE_UPDATES = {
    "SCR-002": "RBAC BUG-13,14 수정 (2026-03-17 QA 수정완료)",
}

# ─────────────────────────────────────────────
# 4. 스타일 헬퍼
# ─────────────────────────────────────────────
STATUS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 연두색
STATUS_FONT = Font(color="276221", bold=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def get_row_style(ws, ref_row):
    """참조 행의 스타일을 반환 (채우기 색상 참조용)"""
    return ws.row_dimensions[ref_row].height


def apply_cell_style(cell, value, bold=False, fill=None, font_color=None, wrap=True, align="left"):
    cell.value = value
    cell.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    cell.border = THIN_BORDER
    f = Font(bold=bold)
    if font_color:
        f = Font(bold=bold, color=font_color)
    cell.font = f
    if fill:
        cell.fill = fill


def copy_row_style(ws, src_row, dst_row):
    """src_row의 스타일을 dst_row에 복사"""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.alignment = copy(src.alignment)
            dst.border = copy(src.border)
            dst.number_format = src.number_format


# ─────────────────────────────────────────────
# 5. 메인 실행
# ─────────────────────────────────────────────
def main():
    print(f"엑셀 파일 로드 중: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # ── 5-1. QA테스트케이스 시트 업데이트 ──
    ws_qa = wb["QA테스트케이스"]
    before_qa_rows = ws_qa.max_row
    print(f"\n[QA테스트케이스] 현재 행 수: {before_qa_rows}")

    # 기존 TC-BUG/TC-UX ID 목록 수집 (중복 방지)
    existing_ids = set()
    for r in range(2, ws_qa.max_row + 1):
        v = ws_qa.cell(r, 1).value
        if v:
            existing_ids.add(str(v))

    # 헤더는 1행, 2행부터 데이터 — 마지막 행 다음에 추가
    start_row = ws_qa.max_row + 1
    added = 0

    for tc in QA_TC_ROWS:
        tc_id = tc[0]
        if tc_id in existing_ids:
            print(f"  SKIP (이미 존재): {tc_id}")
            continue

        r = start_row + added

        # 컬럼 매핑:
        # 1=TC-ID, 2=플랫폼, 3=테스트영역, 4=하위영역, 5=테스트유형
        # 6=시나리오, 7=사전조건, 8=테스트절차, 9=기대결과, 10=실제결과
        # 11=테스트상태, 12=우선순위, 13=심각도, 14=담당자, 15=테스트환경
        (tc_id_, platform, area, sub, t_type,
         scenario, precond, steps, expected, actual,
         status, priority, severity, env) = tc

        vals = [
            tc_id_, platform, area, sub, t_type,
            scenario, precond, steps, expected, actual,
            status, priority, severity, "QA팀", env,
        ]

        for c_idx, val in enumerate(vals, 1):
            cell = ws_qa.cell(r, c_idx)
            cell.value = val
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            # 테스트 상태 열(11) 강조
            if c_idx == 11:
                cell.fill = STATUS_FILL
                cell.font = STATUS_FONT
            else:
                cell.font = Font()

        ws_qa.row_dimensions[r].height = 60
        existing_ids.add(tc_id_)
        added += 1
        print(f"  추가: {tc_id_} | {area} | {sub}")

    after_qa_rows = ws_qa.max_row
    print(f"[QA테스트케이스] 추가 후 행 수: {after_qa_rows} (+{after_qa_rows - before_qa_rows}행)")

    # ── 5-2. 기능명세서 시트 비고 업데이트 ──
    ws_fn = wb["기능명세서"]
    NOTE_COL = 20  # col20 = 비고
    print(f"\n[기능명세서] 비고 업데이트")

    updated_fn = 0
    for r in range(2, ws_fn.max_row + 1):
        fn_id = str(ws_fn.cell(r, 1).value)
        if fn_id in FN_NOTE_UPDATES:
            cell = ws_fn.cell(r, NOTE_COL)
            existing = cell.value or ""
            append_text = FN_NOTE_UPDATES[fn_id]
            # 이미 반영된 경우 스킵
            if "QA 수정완료" in str(existing):
                print(f"  SKIP (이미 반영): {fn_id}")
                continue
            sep = "\n" if existing else ""
            cell.value = f"{existing}{sep}{append_text}"
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            print(f"  업데이트: {fn_id} | {ws_fn.cell(r, 2).value}")
            updated_fn += 1

    print(f"[기능명세서] {updated_fn}개 행 업데이트")

    # ── 5-3. 화면설계서 시트 비고 업데이트 ──
    ws_scr = wb["화면설계서"]
    # 화면설계서 컬럼 확인: 비고 컬럼이 없으면 상태(col17) 옆 추가 정보로 처리
    # col17=상태, col18=담당자, col19=[자동]관련기능수 → col20 이후 사용 가능 여부 확인
    scr_max_col = ws_scr.max_column
    # 비고 컬럼 찾기
    scr_note_col = None
    for c in range(1, scr_max_col + 1):
        h = str(ws_scr.cell(1, c).value)
        if "비고" in h:
            scr_note_col = c
            break
    if scr_note_col is None:
        # 비고 컬럼이 없으면 마지막 컬럼 다음에 추가
        scr_note_col = scr_max_col + 1
        ws_scr.cell(1, scr_note_col).value = "비고"
        ws_scr.cell(1, scr_note_col).font = Font(bold=True)
        ws_scr.cell(1, scr_note_col).border = THIN_BORDER

    print(f"\n[화면설계서] 비고 컬럼: col{scr_note_col}, 업데이트")

    updated_scr = 0
    for r in range(2, ws_scr.max_row + 1):
        scr_id = str(ws_scr.cell(r, 1).value)
        if scr_id in SCR_NOTE_UPDATES:
            cell = ws_scr.cell(r, scr_note_col)
            existing = cell.value or ""
            append_text = SCR_NOTE_UPDATES[scr_id]
            if "QA 수정완료" in str(existing):
                print(f"  SKIP (이미 반영): {scr_id}")
                continue
            sep = "\n" if existing else ""
            cell.value = f"{existing}{sep}{append_text}"
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            print(f"  업데이트: {scr_id} | {ws_scr.cell(r, 2).value}")
            updated_scr += 1

    print(f"[화면설계서] {updated_scr}개 행 업데이트")

    # ── 5-4. 저장 ──
    print(f"\n저장 중: {EXCEL_PATH}")
    wb.save(EXCEL_PATH)
    print("저장 완료!")

    print(f"""
=== 업데이트 요약 ===
QA테스트케이스: {before_qa_rows}행 → {after_qa_rows}행 (+{after_qa_rows - before_qa_rows}행 추가)
기능명세서    : {updated_fn}개 행 비고 업데이트
화면설계서    : {updated_scr}개 행 비고 업데이트
""")


if __name__ == "__main__":
    main()

