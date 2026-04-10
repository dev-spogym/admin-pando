from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "excel" / "기획자_체크리스트_점검_2026-04-05.xlsx"


CHECK_ROWS = [
    {
        "category": "1. 범위 잠금 체크리스트",
        "status": "부분 적용",
        "status_type": "partial",
        "applied": (
            "우선순위/Phase 정리, 경쟁사 비교, GAP 분석 문서는 존재함. "
            "기능명세서와 사용자스토리도 우선순위 기준으로 정리됨."
        ),
        "policy_only": (
            "기획/구현 간 차이를 별도 GAP 문서로 설명하고 있으며, 데모 성격에 가까운 흔적이 일부 존재함."
        ),
        "missing": (
            "1차/2차/보류 범위 확정서, 데모/실구현 구분, 항목별 기획 완료 기준, "
            "3개 솔루션 기준 기능 매핑표 단독본은 없음."
        ),
        "references": (
            "docs/05_기능명세서/README.md\n"
            "docs/02_사용자스토리/README.md\n"
            "docs/onfit-gap-analysis.md\n"
            "excel/GAP_분석_기획문서vs구현.md"
        ),
    },
    {
        "category": "2. 통합 원칙 정의 체크리스트",
        "status": "부분 적용",
        "status_type": "partial",
        "applied": (
            "온핏/브로제이/레슨북을 합친 방향은 명확하며, 경쟁사 장단점 분석과 반영 우선순위가 정리돼 있음."
        ),
        "policy_only": (
            "상품은 레슨북, 출입/얼굴인식은 브로제이, 회원/매출은 온핏 참고라는 흐름이 문서와 코드 주석에 드러남."
        ),
        "missing": (
            "모듈별 기준 솔루션 확정표, 기능 충돌 시 UI/정책/데이터 우선순위 규칙, "
            "신규 정책/향후 고도화 분리표가 독립 문서로 정리되진 않음."
        ),
        "references": (
            "docs/competitor-analysis.md\n"
            "docs/onfit-gap-analysis.md\n"
            "docs/BROJ-CRM-기능분석.md"
        ),
    },
    {
        "category": "3. 공통 정책 정의 체크리스트",
        "status": "문서 기준 적용",
        "status_type": "done",
        "applied": (
            "회원/상품/출입/예약/알림 정책이 비즈니스 정책서에 상세 정의되어 있음."
        ),
        "policy_only": (
            "일부 정책은 화면에 완전 강제되기보다 문서와 일부 로직 수준으로 존재함."
        ),
        "missing": (
            "체크리스트 기준의 상태값 세트(신규/휴면/이탈위험 등)와 현재 enum이 완전히 일치하지 않으며, "
            "실제 화면/DB 강제 적용은 아직 불완전함."
        ),
        "references": "docs/09_비즈니스정책/README.md",
    },
    {
        "category": "4. 데이터 정의 체크리스트",
        "status": "부분 적용",
        "status_type": "partial",
        "applied": (
            "핵심 엔티티, 필드, enum, FK, 멀티테넌트 격리 규칙이 정의되어 있음."
        ),
        "policy_only": (
            "데이터 정의는 잘 되어 있으나 KPI 원천 데이터 매핑, 수기 입력/연동 수집 구분은 약함."
        ),
        "missing": (
            "리드/상담/예약/수업세션/출입로그/알림/락커까지 체크리스트 수준으로 모두 명확히 분리된 표는 부족하고, "
            "KPI 산출 연계도 약함."
        ),
        "references": (
            "docs/07_데이터정의서/README.md\n"
            "docs/BROJ-CRM-데이터모델.md"
        ),
    },
    {
        "category": "5. 화면 기획 체크리스트",
        "status": "적용",
        "status_type": "done",
        "applied": (
            "IA, 권한 매트릭스, 화면설계서, 기능명세서, 회원앱 문서까지 다층 구조로 정리되어 있음."
        ),
        "policy_only": (
            "일부 화면은 설계 대비 구현이 시연 수준이거나 mock 기반임."
        ),
        "missing": (
            "키오스크 실운영용 상세 시나리오와 빈상태/에러상태/권한없음 상태는 전 화면 공통으로 완비되진 않음."
        ),
        "references": (
            "docs/01_IA/README.md\n"
            "docs/03_화면설계서/README.md\n"
            "src/App.tsx"
        ),
    },
    {
        "category": "6. KPI / Dashboard 기획 체크리스트",
        "status": "부분 적용",
        "status_type": "partial",
        "applied": (
            "대시보드 화면, 일부 KPI 카드, 매출/회원 통계 방향은 구현 및 문서에 존재함."
        ),
        "policy_only": (
            "KPI 문맥은 있으나 대부분 화면 지표 수준이며, 계산 정의와 입력 출처 연결은 약함."
        ),
        "missing": (
            "KPI별 원천 데이터, 입력 화면, 노출 대상, 필터 기준을 체계적으로 엮은 정의서는 부족함."
        ),
        "references": (
            "src/pages/Dashboard.tsx\n"
            "excel/extracted/대시보드.json"
        ),
    },
    {
        "category": "7. 연동 기획 체크리스트",
        "status": "부분 적용",
        "status_type": "partial",
        "applied": (
            "API 명세서와 키오스크/IoT 설정 화면은 있으며, 얼굴인식 등 연동 대상도 문서화 흔적이 있음."
        ),
        "policy_only": (
            "실제 API/IoT 연동보다는 정의와 데모 성격의 설정 화면이 중심임."
        ),
        "missing": (
            "연동 정의서 단독본, one-way/two-way 구분, 실패 시 수동 플로우, "
            "매장 예외 시나리오, 데모/운영 단계 구분이 부족함."
        ),
        "references": (
            "docs/06_API명세서/README.md\n"
            "src/pages/KioskSettings.tsx\n"
            "src/pages/IotSettings.tsx"
        ),
    },
    {
        "category": "8. 산출물 패키지 체크리스트",
        "status": "부분 적용",
        "status_type": "partial",
        "applied": (
            "기능명세서, 정책서, 데이터정의서, 화면설계서, QA 테스트케이스, 이슈 문서가 존재함."
        ),
        "policy_only": (
            "실무 산출물 꾸러미 형태는 상당 부분 갖춰져 있음."
        ),
        "missing": (
            "범위 확정서, 기능 매핑표 단독본, 자동 알림 트리거 정의서 단독본, 연동 정의서 단독본은 미흡함."
        ),
        "references": (
            "docs/README.md\n"
            "docs/05_기능명세서/README.md\n"
            "docs/08_QA테스트케이스/README.md"
        ),
    },
    {
        "category": "9. 기획자 범위에서 빼야 하는 것",
        "status": "구분 불명확",
        "status_type": "missing",
        "applied": (
            "문서 폴더는 기획 산출물 체계를 따르고 있음."
        ),
        "policy_only": (
            "프로젝트 설명상 시연/정책 정리 중심이라는 정황은 있음."
        ),
        "missing": (
            "저장소에 DB 마이그레이션, API 코드, 실제 페이지 구현이 함께 있어 "
            "기획 범위와 개발 범위를 문서상 명확히 분리한 선언은 없음."
        ),
        "references": (
            "prisma/schema.prisma\n"
            "src/api/endpoints/products.ts\n"
            "src/pages/ProductList.tsx"
        ),
    },
]


DETAIL_ROWS = [
    ("상품관리", "실제 반영", "레슨북 스타일의 마스터-디테일, 요일/시간 설정, 홀딩/양도/키오스크 노출이 구현됨.", "src/pages/ProductList.tsx\nsrc/components/ProductDetailPanel.tsx"),
    ("회원상세", "실제 반영", "브로제이 스타일 확장 탭, 상담/체성분/평가/서명 관련 UI가 다수 들어가 있음.", "src/pages/MemberDetail.tsx"),
    ("출입/키오스크", "부분 반영", "실시간 입장 팝업, 얼굴 인식 설정 UI는 있음. 다만 운영 연동보다는 시연/설정 수준임.", "src/pages/Attendance.tsx\nsrc/pages/KioskSettings.tsx"),
    ("예약/노쇼/대기열", "정책+일부 로직", "노쇼/취소/대기열 자동승격 로직 파일은 존재하나 전면 운영 구현으로 보긴 어려움.", "src/lib/businessLogic.ts"),
    ("데이터/API", "혼재", "일부는 Supabase 직접 CRUD, 일부는 mock 데이터, 일부는 문서상 개발완료로 표기되어 일관성이 떨어짐.", "src/api/endpoints/products.ts\nsrc/server/mockData.ts\ndocs/06_API명세서/README.md"),
]


SOURCE_ROWS = [
    ("문서 인덱스", "docs/README.md"),
    ("기획문서 vs 구현 GAP", "excel/GAP_분석_기획문서vs구현.md"),
    ("경쟁사 분석", "docs/competitor-analysis.md"),
    ("온핏 GAP 분석", "docs/onfit-gap-analysis.md"),
    ("브로제이 기능 분석", "docs/BROJ-CRM-기능분석.md"),
    ("비즈니스 정책", "docs/09_비즈니스정책/README.md"),
    ("데이터 정의", "docs/07_데이터정의서/README.md"),
    ("API 명세", "docs/06_API명세서/README.md"),
    ("IA", "docs/01_IA/README.md"),
    ("라우팅", "src/App.tsx"),
    ("상품 목록", "src/pages/ProductList.tsx"),
    ("상품 상세 패널", "src/components/ProductDetailPanel.tsx"),
]


def apply_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_borders(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def color_status(cell, status_type):
    fills = {
        "done": "E2F0D9",
        "partial": "FFF2CC",
        "missing": "F4CCCC",
    }
    cell.fill = PatternFill("solid", fgColor=fills[status_type])
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


wb = Workbook()

ws = wb.active
ws.title = "체크결과"
ws.append(["체크리스트", "판정", "현재 적용된 것", "정책상만 반영된 것", "누락/보강 필요", "근거 파일"])
for row in CHECK_ROWS:
    ws.append([
        row["category"],
        row["status"],
        row["applied"],
        row["policy_only"],
        row["missing"],
        row["references"],
    ])
apply_header(ws)
for idx, row in enumerate(CHECK_ROWS, start=2):
    color_status(ws[f"B{idx}"], row["status_type"])
ws.freeze_panes = "A2"
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 34
ws.column_dimensions["E"].width = 44
ws.column_dimensions["F"].width = 34
set_borders(ws)

summary = wb.create_sheet("요약")
summary.append(["항목", "내용"])
summary_rows = [
    ("작성일", "2026-04-05"),
    ("판정 기준", "적용 / 부분 적용 / 문서 기준 적용 / 구분 불명확"),
    ("전체 평가", "화면·정책·명세는 많이 진행됐지만, 범위 잠금·연동 정의·KPI 정의는 보강 필요"),
    ("프로젝트 성격", "운영 완성형보다는 시연 가능한 화면 + 정책/정의 정리 + 일부 실제 CRUD 혼합 상태"),
    ("주의사항", "문서상 '개발완료'와 실제 mock/시연 상태가 일부 불일치함"),
]
for row in summary_rows:
    summary.append(row)
apply_header(summary)
summary.freeze_panes = "A2"
summary.column_dimensions["A"].width = 18
summary.column_dimensions["B"].width = 120
set_borders(summary)

detail = wb.create_sheet("반영포인트")
detail.append(["영역", "분류", "내용", "근거 파일"])
for row in DETAIL_ROWS:
    detail.append(list(row))
apply_header(detail)
detail.freeze_panes = "A2"
detail.column_dimensions["A"].width = 18
detail.column_dimensions["B"].width = 14
detail.column_dimensions["C"].width = 70
detail.column_dimensions["D"].width = 40
set_borders(detail)

sources = wb.create_sheet("근거목록")
sources.append(["구분", "파일"])
for row in SOURCE_ROWS:
    sources.append(list(row))
apply_header(sources)
sources.freeze_panes = "A2"
sources.column_dimensions["A"].width = 18
sources.column_dimensions["B"].width = 70
set_borders(sources)

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"saved: {OUT}")
