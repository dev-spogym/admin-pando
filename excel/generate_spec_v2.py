#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""스포짐 기획문서 v2 생성 스크립트"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# 상수
# ============================================================
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "스포짐_기획문서_v2.xlsx")

HEADER_FILL = PatternFill(start_color="FF2B3A67", end_color="FF2B3A67", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

GROUP_HEADER_FILL = PatternFill(start_color="FFE8EDF5", end_color="FFE8EDF5", fill_type="solid")
GROUP_HEADER_FONT = Font(bold=True, size=10, color="FF2B3A67")

SECTION_FILL = PatternFill(start_color="FFDEE2E6", end_color="FFDEE2E6", fill_type="solid")
SECTION_FONT = Font(bold=True, size=10)

DASH_TITLE_FONT = Font(bold=True, size=16, color="FF2B3A67")
DASH_SECTION_FILL = PatternFill(start_color="FFE8EDF5", end_color="FFE8EDF5", fill_type="solid")
DASH_SECTION_FONT = Font(bold=True, size=10, color="FF2B3A67")

THIN_BORDER = Border(
    left=Side(style="thin", color="FFD0D0D0"),
    right=Side(style="thin", color="FFD0D0D0"),
    top=Side(style="thin", color="FFD0D0D0"),
    bottom=Side(style="thin", color="FFD0D0D0"),
)

담당자_LIST = "기획자1,기획자2,디자이너1,개발자1,개발자2,QA담당자1,QA담당자2"

# 스포짐 화면 ID 목록
SCR_IDS = [
    "SCR-001","SCR-002","SCR-010","SCR-011","SCR-012","SCR-013","SCR-014",
    "SCR-020","SCR-021","SCR-030","SCR-031","SCR-032","SCR-040","SCR-041",
    "SCR-050","SCR-051","SCR-052","SCR-053","SCR-060","SCR-061","SCR-062",
    "SCR-063","SCR-070","SCR-071","SCR-072","SCR-073","SCR-080","SCR-081",
    "SCR-090","SCR-091","SCR-092","SCR-093","SCR-094",
]
SCR_IDS_STR = ",".join(SCR_IDS)

FN_IDS = [f"FN-{i:03d}" for i in range(1, 34)]
FN_IDS_STR = ",".join(FN_IDS)

API_IDS = [f"API-{i:03d}" for i in range(1, 28)]
API_IDS_STR = ",".join(API_IDS)


def apply_header(ws, headers, col_widths=None):
    """헤더 행 스타일 적용"""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER
    if col_widths:
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w


def add_dv(ws, col_letter, formula1, max_row=200):
    """드롭다운 DataValidation 추가"""
    dv = DataValidation(type="list", formula1=f'"{formula1}"', allow_blank=True)
    dv.sqref = f"{col_letter}2:{col_letter}{max_row}"
    ws.add_data_validation(dv)


def write_rows(ws, data, start_row=2):
    """데이터 행 쓰기 (리스트의 리스트)"""
    for ri, row_data in enumerate(data):
        for ci, val in enumerate(row_data):
            if val is not None:
                c = ws.cell(start_row + ri, ci + 1, val)
                c.border = THIN_BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)


# ============================================================
# 1. IA(정보구조) 시트
# ============================================================
def create_ia_sheet(wb):
    ws = wb.active
    ws.title = "IA(정보구조)"

    headers = ["Depth 1","Depth 2","Depth 3","화면 ID","화면명","플랫폼",
               "화면 URL/경로","로그인 필요","권한","설명","상태","비고"]
    col_widths = {"A":14,"B":16,"C":14,"D":12,"E":18,"F":14,"G":22,"H":10,"I":12,"J":30,"K":10,"L":20}
    apply_header(ws, headers, col_widths)

    data = [
        ["인증","로그인",None,"SCR-001","로그인","Web","/login","N","비회원","이메일/비밀번호 로그인 화면. 지점 선택 포함","확정",None],
        ["대시보드","메인",None,"SCR-002","대시보드","Web","/","Y","전체","지점별 핵심 KPI 요약 대시보드","확정",None],
        ["회원관리","회원",None,"SCR-010","회원 목록","Web","/members","Y","매니저 이상","회원 검색/필터/목록 조회","확정",None],
        ["회원관리","회원",None,"SCR-011","회원 상세","Web","/members/detail","Y","매니저 이상","회원 프로필, 이용권, 출석, 결제, 체성분, 메모 탭","확정",None],
        ["회원관리","회원",None,"SCR-012","회원 등록","Web","/members/new","Y","매니저 이상","신규 회원 등록 폼","확정",None],
        ["회원관리","회원",None,"SCR-013","회원 수정","Web","/members/edit","Y","매니저 이상","기존 회원 정보 수정 폼","확정",None],
        ["회원관리","체성분",None,"SCR-014","체성분 분석","Web","/body-composition","Y","FC 이상","체성분 측정 기록 및 변화 추이 차트","확정",None],
        ["출석/일정","출석",None,"SCR-020","출석 기록","Web","/attendance","Y","스태프 이상","회원 출석 현황 조회 (일별/주별/월별)","확정",None],
        ["출석/일정","일정",None,"SCR-021","일정 관리","Web","/calendar","Y","스태프 이상","PT/GX 일정 캘린더 뷰","확정",None],
        ["매출/결제","매출",None,"SCR-030","매출 조회","Web","/sales","Y","매니저 이상","매출 현황 조회, 필터, 통계 요약","확정",None],
        ["매출/결제","POS",None,"SCR-031","POS","Web","/pos","Y","스태프 이상","POS 상품 선택 및 장바구니","확정",None],
        ["매출/결제","POS",None,"SCR-032","POS 결제","Web","/pos/payment","Y","스태프 이상","결제 수단 선택 및 결제 처리","확정",None],
        ["상품관리","상품",None,"SCR-040","상품 목록","Web","/products","Y","매니저 이상","이용권/PT/기타 상품 목록","확정",None],
        ["상품관리","상품",None,"SCR-041","상품 등록","Web","/products/new","Y","센터장 이상","신규 상품 등록 폼","확정",None],
        ["시설관리","락커",None,"SCR-050","락커 현황","Web","/locker","Y","스태프 이상","락커 구역별 배정 현황 (색상 그리드)","확정",None],
        ["시설관리","락커",None,"SCR-051","락커 관리","Web","/locker/management","Y","매니저 이상","락커 배정/해제/일괄 관리","확정",None],
        ["시설관리","RFID",None,"SCR-052","RFID 관리","Web","/rfid","Y","매니저 이상","RFID 카드 등록/해제/이력 조회","확정",None],
        ["시설관리","룸",None,"SCR-053","룸 관리","Web","/rooms","Y","매니저 이상","GX룸/PT룸 관리 및 예약 설정","확정",None],
        ["직원관리","직원",None,"SCR-060","직원 목록","Web","/staff","Y","센터장 이상","직원 검색/필터/목록 조회","확정",None],
        ["직원관리","직원",None,"SCR-061","직원 등록","Web","/staff/new","Y","센터장 이상","신규 직원 등록 폼","확정",None],
        ["직원관리","급여",None,"SCR-062","급여 현황","Web","/payroll","Y","센터장 이상","월별 급여 현황 요약","확정",None],
        ["직원관리","급여",None,"SCR-063","급여명세서","Web","/payroll/statements","Y","센터장 이상","직원별 급여명세서 상세","확정",None],
        ["메시지/마케팅","메시지",None,"SCR-070","메시지 발송","Web","/message","Y","매니저 이상","SMS/카카오톡 메시지 발송","확정",None],
        ["메시지/마케팅","자동알림",None,"SCR-071","자동 알림","Web","/message/auto-alarm","Y","매니저 이상","자동 알림 규칙 설정 (만료/생일 등)","확정",None],
        ["메시지/마케팅","쿠폰",None,"SCR-072","쿠폰 관리","Web","/message/coupon","Y","매니저 이상","쿠폰 생성/발급/이력 관리","확정",None],
        ["메시지/마케팅","마일리지",None,"SCR-073","마일리지 관리","Web","/mileage","Y","매니저 이상","회원 마일리지 적립/사용 이력","확정",None],
        ["계약/구독","계약",None,"SCR-080","계약 마법사","Web","/contracts/new","Y","매니저 이상","단계별 계약 등록 위자드","확정",None],
        ["계약/구독","구독",None,"SCR-081","구독 관리","Web","/subscription","Y","센터장 이상","정기 구독 상품 관리","확정",None],
        ["설정","설정",None,"SCR-090","설정","Web","/settings","Y","센터장 이상","시스템 기본 설정","확정",None],
        ["설정","권한",None,"SCR-091","권한 설정","Web","/settings/permissions","Y","최고관리자","역할별 메뉴/기능 권한 매트릭스","확정",None],
        ["설정","키오스크",None,"SCR-092","키오스크 설정","Web","/settings/kiosk","Y","센터장 이상","키오스크 디스플레이/동작 설정","확정",None],
        ["설정","IoT",None,"SCR-093","IoT 설정","Web","/settings/iot","Y","센터장 이상","출입문/체성분 기기 연동 설정","확정",None],
        ["설정","지점",None,"SCR-094","지점 관리","Web","/branches","Y","최고관리자","지점 CRUD 및 멀티테넌트 관리","확정",None],
    ]
    write_rows(ws, data)

    # 드롭다운
    add_dv(ws, "F", "Web,App (iOS),App (Android),Web + App,공통")
    add_dv(ws, "H", "Y,N")
    add_dv(ws, "K", "작성중,검토중,확정,변경필요,삭제")

    # 자동필터
    ws.auto_filter.ref = "A1:L1"
    return ws


# ============================================================
# 2. 사용자 스토리 시트
# ============================================================
def create_story_sheet(wb):
    ws = wb.create_sheet("사용자 스토리")

    headers = ["스토리 ID","에픽","사용자 유형","사용자 스토리","인수 조건",
               "화면 ID","우선순위","사이즈","담당자","상태",
               "[자동] 화면명","[자동] 기능 수","비고"]
    col_widths = {"A":12,"B":14,"C":12,"D":40,"E":35,"F":12,"G":10,"H":8,"I":12,"J":12,"K":18,"L":12,"M":20}
    apply_header(ws, headers, col_widths)

    data = [
        ["US-001","인증","비회원","비회원으로서 이메일/비밀번호로 로그인하여 시스템에 접근하고 싶다",
         "1) 올바른 이메일/비밀번호 입력 시 대시보드로 이동\n2) 5회 실패 시 계정 잠금\n3) 지점 선택 후 해당 지점 데이터만 표시",
         "SCR-001","Critical","S","개발자1","To Do",
         '=IFERROR(VLOOKUP(F2,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F2),0)',None],
        ["US-002","대시보드","센터장","센터장으로서 대시보드에서 지점 핵심 KPI를 한눈에 파악하고 싶다",
         "1) 총회원/활성/만료임박/만료 회원 수 카드 표시\n2) 오늘 생일 회원 리스트\n3) 미납 회원 리스트\n4) 당월 매출 차트",
         "SCR-002","Critical","M","개발자1","To Do",
         '=IFERROR(VLOOKUP(F3,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F3),0)',None],
        ["US-003","회원관리","매니저","매니저로서 회원을 검색/필터하여 원하는 회원을 빠르게 찾고 싶다",
         "1) 이름/연락처/회원번호로 검색\n2) 상태/성별/담당자 필터\n3) 엑셀 다운로드 가능",
         "SCR-010","Critical","M","개발자1","To Do",
         '=IFERROR(VLOOKUP(F4,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F4),0)',None],
        ["US-004","회원관리","매니저","매니저로서 신규 회원을 등록하고 싶다",
         "1) 필수 항목(이름/연락처/성별) 입력\n2) 담당 FC 배정\n3) 등록 완료 시 회원 상세로 이동",
         "SCR-012","Critical","M","개발자1","To Do",
         '=IFERROR(VLOOKUP(F5,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F5),0)',None],
        ["US-005","회원관리","매니저","매니저로서 회원 상세 정보와 이용 이력을 확인하고 싶다",
         "1) 프로필/이용권/출석/결제/체성분/메모 탭 구성\n2) 이용권 만료일 강조 표시\n3) 출석 히트맵",
         "SCR-011","High","L","개발자1","To Do",
         '=IFERROR(VLOOKUP(F6,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F6),0)',None],
        ["US-006","회원관리","매니저","매니저로서 회원 정보를 수정하고 싶다",
         "1) 기존 정보 프리필\n2) 변경 이력 자동 기록\n3) 담당자 변경 가능",
         "SCR-013","High","S","개발자1","To Do",
         '=IFERROR(VLOOKUP(F7,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F7),0)',None],
        ["US-007","회원관리","FC","FC로서 회원의 체성분 측정 결과를 기록하고 변화 추이를 분석하고 싶다",
         "1) 체중/골격근/체지방률 등 입력\n2) 측정일 기준 변화 차트\n3) 목표 대비 달성률 표시",
         "SCR-014","High","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F8,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F8),0)',None],
        ["US-008","출석관리","스태프","스태프로서 회원의 출석 현황을 조회하고 싶다",
         "1) 일별/주별/월별 출석 조회\n2) 출석 유형(일반/PT/GX) 구분\n3) 타지점 출석 표시",
         "SCR-020","High","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F9,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F9),0)',None],
        ["US-009","출석관리","스태프","스태프로서 PT/GX 일정을 캘린더로 관리하고 싶다",
         "1) 월/주/일 캘린더 뷰\n2) 드래그앤드롭 일정 변경\n3) 트레이너별 필터",
         "SCR-021","High","L","개발자2","To Do",
         '=IFERROR(VLOOKUP(F10,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F10),0)',None],
        ["US-010","매출관리","매니저","매니저로서 매출 현황을 조회하고 통계를 확인하고 싶다",
         "1) 기간별 매출 필터\n2) 유형별(이용권/PT/상품) 분류\n3) 총매출/현금/카드/마일리지 요약\n4) 엑셀 다운로드",
         "SCR-030","Critical","L","개발자1","To Do",
         '=IFERROR(VLOOKUP(F11,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F11),0)',None],
        ["US-011","매출관리","스태프","스태프로서 POS에서 상품을 선택하고 결제를 처리하고 싶다",
         "1) 카테고리별 상품 그리드\n2) 장바구니 담기/수량 변경\n3) 현금/카드/마일리지 복합결제\n4) 영수증 출력",
         "SCR-031","Critical","XL","개발자1","To Do",
         '=IFERROR(VLOOKUP(F12,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F12),0)',None],
        ["US-012","상품관리","매니저","매니저로서 상품 목록을 조회하고 관리하고 싶다",
         "1) 카테고리별 필터\n2) 키오스크 노출 여부 토글\n3) 상품 상태(판매중/중지) 관리",
         "SCR-040","High","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F13,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F13),0)',None],
        ["US-013","상품관리","센터장","센터장으로서 신규 상품을 등록하고 싶다",
         "1) 카테고리/이름/가격/기간 입력\n2) 현금가/카드가 별도 설정\n3) 키오스크 노출 설정",
         "SCR-041","High","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F14,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F14),0)',None],
        ["US-014","시설관리","스태프","스태프로서 락커 배정 현황을 한눈에 파악하고 싶다",
         "1) 구역별 락커 그리드 (색상으로 상태 구분)\n2) 사용중/빈락커/만료임박 구분\n3) 클릭 시 상세 모달",
         "SCR-050","High","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F15,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F15),0)',None],
        ["US-015","시설관리","매니저","매니저로서 락커를 회원에게 배정/해제하고 싶다",
         "1) 회원 검색 후 빈 락커 배정\n2) 만료일 설정\n3) 일괄 해제 기능",
         "SCR-051","High","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F16,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F16),0)',None],
        ["US-016","시설관리","매니저","매니저로서 RFID 카드를 관리하고 싶다",
         "1) RFID 카드 등록/해제\n2) 회원-카드 매핑\n3) 분실 처리",
         "SCR-052","Medium","S","개발자2","To Do",
         '=IFERROR(VLOOKUP(F17,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F17),0)',None],
        ["US-017","시설관리","매니저","매니저로서 GX룸/PT룸을 관리하고 싶다",
         "1) 룸 등록/수정/삭제\n2) 수용 인원 설정\n3) 예약 가능 시간대 설정",
         "SCR-053","Medium","S","개발자2","To Do",
         '=IFERROR(VLOOKUP(F18,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F18),0)',None],
        ["US-018","직원관리","센터장","센터장으로서 직원 목록을 조회하고 싶다",
         "1) 역할/상태별 필터\n2) 검색 기능\n3) 직원 상세 이동",
         "SCR-060","High","M","개발자1","To Do",
         '=IFERROR(VLOOKUP(F19,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F19),0)',None],
        ["US-019","직원관리","센터장","센터장으로서 신규 직원을 등록하고 싶다",
         "1) 이름/역할/연락처/입사일 입력\n2) 관리자 계정 생성\n3) 근무 유형 설정",
         "SCR-061","High","M","개발자1","To Do",
         '=IFERROR(VLOOKUP(F20,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F20),0)',None],
        ["US-020","직원관리","센터장","센터장으로서 월별 급여 현황을 확인하고 싶다",
         "1) 월별 급여 요약\n2) 기본급/인센티브/공제 내역\n3) 지급 상태 관리",
         "SCR-062","High","L","개발자1","To Do",
         '=IFERROR(VLOOKUP(F21,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F21),0)',None],
        ["US-021","직원관리","센터장","센터장으로서 직원별 급여명세서를 확인하고 싶다",
         "1) 직원별 상세 급여명세서\n2) PDF 다운로드\n3) 이력 조회",
         "SCR-063","Medium","M","개발자1","To Do",
         '=IFERROR(VLOOKUP(F22,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F22),0)',None],
        ["US-022","메시지/마케팅","매니저","매니저로서 회원에게 SMS/카카오톡 메시지를 발송하고 싶다",
         "1) 수신자 선택 (개별/그룹/전체)\n2) 메시지 유형 선택\n3) 발송 이력 조회",
         "SCR-070","High","L","개발자2","To Do",
         '=IFERROR(VLOOKUP(F23,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F23),0)',None],
        ["US-023","메시지/마케팅","매니저","매니저로서 자동 알림 규칙을 설정하고 싶다",
         "1) 만료 D-7/D-3/D-1 알림\n2) 생일 축하 알림\n3) 장기 미출석 알림",
         "SCR-071","Medium","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F24,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F24),0)',None],
        ["US-024","메시지/마케팅","매니저","매니저로서 쿠폰을 생성하고 발급하고 싶다",
         "1) 할인율/금액 쿠폰 생성\n2) 유효기간 설정\n3) 사용 이력 추적",
         "SCR-072","Medium","M","개발자2","To Do",
         '=IFERROR(VLOOKUP(F25,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F25),0)',None],
        ["US-025","메시지/마케팅","매니저","매니저로서 회원 마일리지를 관리하고 싶다",
         "1) 마일리지 적립/차감\n2) 잔액 조회\n3) 이력 조회",
         "SCR-073","Medium","S","개발자2","To Do",
         '=IFERROR(VLOOKUP(F26,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F26),0)',None],
        ["US-026","계약/구독","매니저","매니저로서 단계별 계약 등록을 하고 싶다",
         "1) 회원 선택 → 상품 선택 → 기간/금액 설정 → 결제 → 확인\n2) 계약서 자동 생성",
         "SCR-080","Critical","XL","개발자1","To Do",
         '=IFERROR(VLOOKUP(F27,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F27),0)',None],
        ["US-027","계약/구독","센터장","센터장으로서 정기 구독 상품을 관리하고 싶다",
         "1) 구독 상품 CRUD\n2) 자동 결제 설정\n3) 구독 현황 대시보드",
         "SCR-081","Medium","L","개발자1","To Do",
         '=IFERROR(VLOOKUP(F28,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F28),0)',None],
        ["US-028","설정/권한","센터장","센터장으로서 시스템 기본 설정을 변경하고 싶다",
         "1) 센터명/영업시간/연락처 설정\n2) 알림 설정\n3) 테마 설정",
         "SCR-090","Medium","M","개발자1","To Do",
         '=IFERROR(VLOOKUP(F29,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F29),0)',None],
        ["US-029","설정/권한","최고관리자","최고관리자로서 역할별 권한을 세밀하게 설정하고 싶다",
         "1) 6단계 역할별 메뉴 접근 권한\n2) 기능별 읽기/쓰기/삭제 권한\n3) 변경 이력 로그",
         "SCR-091","Critical","L","개발자1","To Do",
         '=IFERROR(VLOOKUP(F30,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F30),0)',None],
        ["US-030","설정/권한","최고관리자","최고관리자로서 지점을 추가/관리하고 싶다",
         "1) 지점 CRUD\n2) 지점별 독립 데이터\n3) 지점 간 회원 이동 처리",
         "SCR-094","Critical","L","개발자1","To Do",
         '=IFERROR(VLOOKUP(F31,\'IA(정보구조)\'!D:E,2,FALSE()),"")',
         '=IFERROR(COUNTIF(기능명세서!H:H,F31),0)',None],
    ]
    write_rows(ws, data)

    # 드롭다운
    add_dv(ws, "F", SCR_IDS_STR)
    add_dv(ws, "G", "Critical,High,Medium,Low")
    add_dv(ws, "H", "XS,S,M,L,XL")
    add_dv(ws, "I", 담당자_LIST)
    add_dv(ws, "J", "Backlog,To Do,In Progress,In Review,Done")

    ws.auto_filter.ref = "A1:M1"
    return ws


# ============================================================
# 3. 화면설계서 시트
# ============================================================
def create_screen_spec_sheet(wb):
    ws = wb.create_sheet("화면설계서")

    headers = ["화면 ID","화면명","플랫폼","화면 경로","화면 설명",
               "[자동] UI 요소 목록","인터랙션 정의","Web 특이사항","App 특이사항",
               "진입 경로","이동 가능 화면","에러 화면 정의","접근성 체크",
               "반응형 체크 (Web)","가로/세로 모드 (App)","디자인 시안 링크",
               "상태","담당자","[자동] 관련 기능 수","[자동] 관련 TC 수",
               "[자동] 관련 API 수","비고","[자동] UI 요소 수"]
    col_widths = {"A":12,"B":16,"C":14,"D":20,"E":30,"F":35,"G":30,"H":25,"I":20,
                  "J":20,"K":20,"L":25,"M":20,"N":18,"O":14,"P":20,"Q":10,"R":12,
                  "S":14,"T":14,"U":14,"V":20,"W":14}
    apply_header(ws, headers, col_widths)

    # 33개 화면 데이터
    screens = [
        ("SCR-001","로그인","Web","/login","이메일/비밀번호 기반 로그인 화면. 지점 선택 드롭다운 포함",
         "비밀번호 마스킹 토글, Enter키 로그인, 지점 선택 시 로고 변경","JWT 토큰 기반 세션 관리",None,
         "URL 직접 접근, 세션 만료 시 리다이렉트","SCR-002 (대시보드)",
         "이메일 미입력: '이메일을 입력하세요'\n비밀번호 오류: '비밀번호가 올바르지 않습니다'\n계정 잠금: '5회 실패로 잠금됨'",
         "Tab 이동 순서, aria-label","1024px 이상 중앙 정렬",None),
        ("SCR-002","대시보드","Web","/","지점별 핵심 KPI 대시보드. 회원/매출/출석 통계 카드 + 알림 리스트",
         "통계 카드 클릭 시 상세 이동, 차트 호버 툴팁, 기간 필터","반응형 그리드 (2열→1열)",None,
         "SCR-001 (로그인 후)","SCR-010, SCR-030, SCR-020",
         "데이터 로딩 실패: '데이터를 불러올 수 없습니다' + 재시도 버튼","차트 alt-text","1280px 이상 4열 카드",None),
        ("SCR-010","회원 목록","Web","/members","회원 검색, 필터, 페이지네이션이 포함된 목록 화면",
         "실시간 검색, 필터 조합, 정렬, 페이지네이션, 엑셀 다운로드","테이블 가로스크롤 (좁은 화면)",None,
         "GNB 메뉴, SCR-002 바로가기","SCR-011, SCR-012",
         "검색 결과 없음: '조건에 맞는 회원이 없습니다'\n로딩 실패: 스켈레톤 + 재시도","테이블 aria-label, 스크린리더 페이지 안내","테이블 반응형 스크롤",None),
        ("SCR-011","회원 상세","Web","/members/detail","회원 프로필 + 탭(이용권/출석/결제/체성분/메모) 구성",
         "탭 전환, 이용권 만료일 강조, 출석 히트맵, 결제 내역 페이지네이션","탭 URL 파라미터 유지",None,
         "SCR-010 회원 클릭","SCR-013, SCR-014, SCR-010",
         "회원 미존재: 404 페이지\n데이터 로딩 실패: 탭별 에러 표시","탭 키보드 네비게이션","탭 세로 배치 (좁은 화면)",None),
        ("SCR-012","회원 등록","Web","/members/new","신규 회원 등록 폼. 필수/선택 항목 구분",
         "실시간 유효성 검사, 연락처 자동 포맷, 주소 검색 API 연동","폼 상태 유지 (뒤로가기 시)",None,
         "SCR-010 '회원등록' 버튼","SCR-011 (등록 완료 후)",
         "필수 항목 미입력: 필드 하단 에러 메시지\n중복 연락처: '이미 등록된 연락처입니다'\n서버 오류: 토스트 알림","필수 항목 aria-required","폼 1열 배치",None),
        ("SCR-013","회원 수정","Web","/members/edit","기존 회원 정보 수정 폼. 기존 데이터 프리필",
         "변경사항 비교 표시, 저장 전 확인 모달","변경 이력 자동 기록",None,
         "SCR-011 '수정' 버튼","SCR-011 (저장 후)",
         "변경 없이 저장: '변경된 내용이 없습니다'\n서버 오류: 토스트 알림","폼 접근성 동일","폼 1열 배치",None),
        ("SCR-014","체성분 분석","Web","/body-composition","회원별 체성분 측정 기록 입력 및 변화 추이 차트",
         "측정값 입력 폼, 날짜별 라인 차트, 목표 대비 게이지","차트 라이브러리 (Recharts)",None,
         "SCR-011 '체성분' 탭","SCR-011",
         "측정 데이터 없음: '아직 측정 기록이 없습니다'\n입력 범위 초과: 경고 토스트","차트 alt-text, 데이터 테이블 대체","차트 반응형",None),
        ("SCR-020","출석 기록","Web","/attendance","회원 출석 현황 조회. 일별/주별/월별 뷰 전환",
         "날짜 필터, 출석 유형 필터, 타지점 출석 표시","캘린더 컴포넌트",None,
         "GNB 메뉴","SCR-011 (회원 클릭)",
         "데이터 없음: '해당 기간 출석 기록이 없습니다'","날짜 선택 키보드 접근","테이블 반응형",None),
        ("SCR-021","일정 관리","Web","/calendar","PT/GX 일정 캘린더. 월/주/일 뷰 전환",
         "드래그앤드롭 일정 이동, 클릭 상세 모달, 트레이너 필터","FullCalendar 라이브러리",None,
         "GNB 메뉴","SCR-011 (회원 연결)",
         "일정 충돌: '해당 시간에 이미 예약이 있습니다'\n과거 일정 수정 불가","캘린더 키보드 네비게이션","뷰 자동 전환 (주→일)",None),
        ("SCR-030","매출 조회","Web","/sales","매출 현황 조회. 기간/유형/상태 필터 + 통계 요약 카드",
         "날짜 범위 선택, 유형 필터, 상태 필터, 통계 카드 갱신, 엑셀 다운로드","금액 포맷 (천단위 콤마)",None,
         "GNB 메뉴, SCR-002 매출 카드","SCR-031 (POS 이동)",
         "데이터 없음: '해당 기간 매출이 없습니다'\n다운로드 실패: 재시도 안내","금액 스크린리더 읽기","통계 카드 2열→1열",None),
        ("SCR-031","POS","Web","/pos","POS 상품 선택 화면. 카테고리 탭 + 상품 그리드 + 장바구니",
         "카테고리 탭 전환, 상품 클릭 시 장바구니 추가, 수량 조절, 삭제","2컬럼 레이아웃 (상품 | 장바구니)",None,
         "GNB 메뉴, SCR-030","SCR-032 (결제 진행)",
         "장바구니 비어있음: '상품을 선택해주세요'\n재고 없음: 상품 비활성화","상품 그리드 키보드 선택","장바구니 하단 고정",None),
        ("SCR-032","POS 결제","Web","/pos/payment","결제 수단 선택 및 결제 처리. 복합결제 지원",
         "결제수단 선택, 금액 입력, 마일리지 사용, 결제 확인 모달","결제 API 연동",None,
         "SCR-031 '결제' 버튼","SCR-030 (결제 완료 후)",
         "결제 실패: '결제에 실패했습니다. 다시 시도해주세요'\n잔액 부족: 안내 메시지","금액 입력 접근성","모달 팝업",None),
        ("SCR-040","상품 목록","Web","/products","이용권/PT/기타 상품 목록. 카테고리 필터",
         "카테고리 필터, 상태 토글, 정렬","테이블 레이아웃",None,
         "GNB 메뉴","SCR-041 (등록), SCR-040 (수정 모달)",
         "상품 없음: '등록된 상품이 없습니다'","테이블 접근성","테이블 반응형",None),
        ("SCR-041","상품 등록","Web","/products/new","신규 상품 등록 폼. 카테고리/이름/가격/기간 설정",
         "카테고리 선택 시 하위 필드 변경, 가격 입력 포맷","폼 유효성 검사",None,
         "SCR-040 '등록' 버튼","SCR-040 (저장 후)",
         "필수 미입력: 필드별 에러\n중복 상품명: 경고","폼 접근성","폼 1열",None),
        ("SCR-050","락커 현황","Web","/locker","구역별 락커 그리드. 색상으로 상태 구분 (사용중/빈/만료임박)",
         "구역 탭 전환, 락커 클릭 시 상세 모달, 범례 표시","CSS Grid 레이아웃",None,
         "GNB 메뉴","SCR-051 (관리 이동)",
         "락커 없음: '등록된 락커가 없습니다'","색상 + 패턴으로 상태 구분 (색약 대응)","그리드 반응형",None),
        ("SCR-051","락커 관리","Web","/locker/management","락커 배정/해제/일괄 관리",
         "회원 검색 → 빈 락커 선택 → 배정, 일괄 선택 해제","배정 모달",None,
         "SCR-050, GNB","SCR-050",
         "이미 배정된 락커: '이미 사용중인 락커입니다'","모달 접근성","모달 반응형",None),
        ("SCR-052","RFID 관리","Web","/rfid","RFID 카드 등록/해제/이력 조회",
         "카드 번호 스캔 입력, 회원 매핑, 분실 처리 버튼","RFID 리더기 연동",None,
         "GNB 메뉴","SCR-011 (회원 연결)",
         "미등록 카드: '등록되지 않은 카드입니다'\n이미 매핑: '다른 회원에게 배정된 카드'","폼 접근성","테이블 반응형",None),
        ("SCR-053","룸 관리","Web","/rooms","GX룸/PT룸 등록/수정/삭제. 예약 시간대 설정",
         "룸 카드 뷰, 수용인원 설정, 시간대 슬롯 설정","카드 레이아웃",None,
         "GNB 메뉴","SCR-021 (일정 연결)",
         "룸 삭제 시 예약 존재: '기존 예약이 있어 삭제할 수 없습니다'","카드 접근성","카드 반응형",None),
        ("SCR-060","직원 목록","Web","/staff","직원 검색/필터/목록 조회",
         "역할 필터, 상태 필터, 검색, 테이블 정렬","테이블 레이아웃",None,
         "GNB 메뉴","SCR-061 (등록), 직원 상세 모달",
         "검색 결과 없음: '조건에 맞는 직원이 없습니다'","테이블 접근성","테이블 반응형",None),
        ("SCR-061","직원 등록","Web","/staff/new","신규 직원 등록 폼",
         "역할 선택 시 권한 미리보기, 관리자 계정 자동 생성 옵션","폼 레이아웃",None,
         "SCR-060 '등록' 버튼","SCR-060 (저장 후)",
         "필수 미입력: 필드별 에러\n중복 이메일: 경고","폼 접근성","폼 1열",None),
        ("SCR-062","급여 현황","Web","/payroll","월별 급여 현황 테이블. 합계/평균 요약",
         "월 선택, 지급상태 필터, 합계 행","테이블 + 요약 카드",None,
         "GNB 메뉴","SCR-063 (명세서)",
         "데이터 없음: '해당 월 급여 데이터가 없습니다'","테이블 접근성","테이블 반응형",None),
        ("SCR-063","급여명세서","Web","/payroll/statements","직원별 급여명세서 상세. PDF 다운로드",
         "직원 선택, 월 선택, 명세서 상세 표시, PDF 다운로드","PDF 생성 (html2pdf)",None,
         "SCR-062 직원 클릭","SCR-062",
         "명세서 없음: '해당 월 명세서가 없습니다'","명세서 접근성","A4 레이아웃",None),
        ("SCR-070","메시지 발송","Web","/message","SMS/카카오톡 메시지 발송. 수신자 선택 + 본문 편집",
         "수신자 검색/그룹 선택, 메시지 유형 탭, 미리보기, 발송 이력 테이블","메시지 API 연동",None,
         "GNB 메뉴","발송 이력 상세 모달",
         "수신자 미선택: '수신자를 선택해주세요'\n발송 실패: 실패 건 재발송 옵션","폼 접근성","2컬럼→1컬럼",None),
        ("SCR-071","자동 알림","Web","/message/auto-alarm","자동 알림 규칙 설정 (만료/생일/미출석 등)",
         "규칙 토글 ON/OFF, 발송 시점 설정, 메시지 템플릿 편집","규칙 엔진 설정 UI",None,
         "SCR-070, GNB","SCR-070 (발송 이력)",
         "템플릿 미입력: '메시지 내용을 입력해주세요'","토글 접근성","폼 1열",None),
        ("SCR-072","쿠폰 관리","Web","/message/coupon","쿠폰 생성/발급/이력 관리",
         "쿠폰 생성 모달, 발급 대상 선택, 사용 이력 테이블","모달 + 테이블",None,
         "GNB 메뉴","쿠폰 상세 모달",
         "유효기간 만료: '만료된 쿠폰입니다'\n사용 한도 초과: 비활성화","모달 접근성","테이블 반응형",None),
        ("SCR-073","마일리지 관리","Web","/mileage","회원 마일리지 적립/사용 이력 조회 및 관리",
         "회원 검색, 적립/차감 모달, 이력 테이블","모달 + 테이블",None,
         "GNB 메뉴","SCR-011 (회원 연결)",
         "잔액 부족 차감: '마일리지가 부족합니다'","테이블 접근성","테이블 반응형",None),
        ("SCR-080","계약 마법사","Web","/contracts/new","단계별 계약 등록 위자드 (회원→상품→기간→결제→확인)",
         "스텝 인디케이터, 이전/다음 버튼, 단계별 유효성 검사","위자드 컴포넌트",None,
         "SCR-011, SCR-031","SCR-011 (완료 후)",
         "필수 단계 미완료: '이전 단계를 완료해주세요'\n결제 실패: 재시도 옵션","스텝 접근성","스텝 세로 배치",None),
        ("SCR-081","구독 관리","Web","/subscription","정기 구독 상품 관리. 자동 결제 현황",
         "구독 상품 목록, 결제 현황 대시보드, 구독 상세 모달","테이블 + 카드",None,
         "GNB 메뉴","구독 상세 모달",
         "결제 실패 구독: 경고 배지 표시","테이블 접근성","카드 반응형",None),
        ("SCR-090","설정","Web","/settings","시스템 기본 설정. 센터 정보/영업시간/알림 등",
         "탭 구성 (기본/알림/테마), 저장 버튼","탭 + 폼",None,
         "GNB 메뉴","SCR-091, SCR-092, SCR-093",
         "저장 실패: 토스트 에러","탭 접근성","폼 1열",None),
        ("SCR-091","권한 설정","Web","/settings/permissions","역할별 메뉴/기능 권한 매트릭스",
         "역할 탭, 권한 체크박스 매트릭스, 일괄 설정","매트릭스 테이블",None,
         "SCR-090, GNB","SCR-090",
         "권한 충돌: 경고 표시","체크박스 접근성","테이블 가로스크롤",None),
        ("SCR-092","키오스크 설정","Web","/settings/kiosk","키오스크 디스플레이/동작 설정",
         "미리보기, 로고 업로드, 동작 시나리오 설정","미리보기 컴포넌트",None,
         "SCR-090, GNB","SCR-090",
         "이미지 용량 초과: '5MB 이하 파일만 업로드 가능합니다'","폼 접근성","폼 1열",None),
        ("SCR-093","IoT 설정","Web","/settings/iot","출입문/체성분 기기 연동 설정",
         "기기 목록, 연결 상태 표시, 테스트 버튼","기기 연동 API",None,
         "SCR-090, GNB","SCR-090",
         "연결 실패: '기기와 연결할 수 없습니다' + 재시도","상태 아이콘 alt-text","테이블 반응형",None),
        ("SCR-094","지점 관리","Web","/branches","지점 CRUD. 멀티테넌트 관리",
         "지점 카드/테이블 뷰 전환, 지점 추가 모달, 데이터 초기화","카드/테이블 토글",None,
         "GNB 메뉴 (최고관리자만)","지점 상세 모달",
         "지점 삭제 시 데이터 존재: '소속 데이터가 있어 삭제 불가'","카드/테이블 접근성","카드 반응형",None),
    ]

    for ri, s in enumerate(screens):
        row = ri + 2
        ws.cell(row, 1, s[0])  # 화면 ID
        ws.cell(row, 2, s[1])  # 화면명
        ws.cell(row, 3, s[2])  # 플랫폼
        ws.cell(row, 4, s[3])  # 화면 경로
        ws.cell(row, 5, s[4])  # 화면 설명
        # F: [자동] UI 요소 목록 - 수식
        ws.cell(row, 6, f"=IFERROR(_xlfn.TEXTJOIN(CHAR(10),TRUE(),IF('UI 요소 상세'!B$1:B$500=A{row},\"□ \"&'UI 요소 상세'!C$1:C$500,\"\")),\"\")")
        ws.cell(row, 7, s[5])   # 인터랙션
        ws.cell(row, 8, s[6])   # Web 특이사항
        ws.cell(row, 9, s[7])   # App 특이사항
        ws.cell(row, 10, s[8])  # 진입 경로
        ws.cell(row, 11, s[9])  # 이동 가능 화면
        ws.cell(row, 12, s[10]) # 에러 화면 정의
        ws.cell(row, 13, s[11]) # 접근성
        ws.cell(row, 14, s[12]) # 반응형
        ws.cell(row, 15, s[13]) # 가로세로
        ws.cell(row, 16, None)  # 디자인 시안 링크
        ws.cell(row, 17, "확정") # 상태
        ws.cell(row, 18, None)  # 담당자
        # S: [자동] 관련 기능 수
        ws.cell(row, 19, f'=IFERROR(COUNTIF(기능명세서!H:H,A{row}),0)')
        # T: [자동] 관련 TC 수
        ws.cell(row, 20, f'=IFERROR(COUNTIF(QA테스트케이스!U:U,A{row}),0)')
        # U: [자동] 관련 API 수
        ws.cell(row, 21, f'=IFERROR(COUNTIF(API명세서!K:K,A{row}),0)')
        ws.cell(row, 22, None)  # 비고
        # W: [자동] UI 요소 수
        ws.cell(row, 23, f"=IFERROR(COUNTIF('UI 요소 상세'!B:B,A{row}),0)")

        # 스타일
        for ci in range(1, 24):
            c = ws.cell(row, ci)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # 드롭다운
    add_dv(ws, "A", SCR_IDS_STR)
    add_dv(ws, "C", "Web,App (iOS),App (Android),Web + App,공통")
    add_dv(ws, "Q", "작성중,검토중,확정,변경필요,삭제")
    add_dv(ws, "R", 담당자_LIST)

    ws.auto_filter.ref = "A1:V1"
    return ws


# ============================================================
# 4. UI 요소 상세 시트
# ============================================================
def create_ui_detail_sheet(wb):
    ws = wb.create_sheet("UI 요소 상세")

    headers = ["요소 ID","화면 ID","요소명","요소 유형","요소 설명","인터랙션 정의",
               "기능 ID","API ID","Web 구현","App 구현","유효성 검사","에러 메시지",
               "접근성","상태","담당자","[자동] 화면명","[자동] 기능명","[자동] API 엔드포인트","비고"]
    col_widths = {"A":10,"B":10,"C":16,"D":12,"E":25,"F":30,"G":10,"H":10,
                  "I":22,"J":14,"K":20,"L":20,"M":18,"N":10,"O":12,"P":16,"Q":16,"R":20,"S":18}
    apply_header(ws, headers, col_widths)

    # UI 요소 데이터: (화면ID, 화면명, [(요소ID, 요소명, 유형, 설명, 인터랙션, 기능ID, API_ID, Web구현, 유효성, 에러메시지, 접근성)])
    ui_groups = [
        ("SCR-001", "로그인", [
            ("UI-001","이메일 입력","Input","로그인 이메일 주소 입력 필드","포커스 시 라벨 상단 이동, 실시간 이메일 형식 검증","FN-001","API-001","<input type='email'> + 실시간 validation","이메일 형식 검증 (regex)","올바른 이메일 주소를 입력하세요","aria-label='이메일'"),
            ("UI-002","비밀번호 입력","Input","로그인 비밀번호 입력 필드","마스킹 토글 버튼, Enter키 로그인 트리거","FN-001","API-001","<input type='password'> + visibility toggle","최소 1자 이상","비밀번호를 입력하세요","aria-label='비밀번호'"),
            ("UI-003","지점 선택","Select","로그인 시 접근할 지점 선택 드롭다운","선택 시 지점 로고/색상 변경","FN-001","API-001","<select> 커스텀 드롭다운","필수 선택","지점을 선택하세요","aria-label='지점 선택'"),
            ("UI-004","로그인 버튼","Button","로그인 실행 버튼","클릭/Enter 시 로그인 API 호출, 로딩 스피너","FN-001","API-001","<button> + loading state","이메일/비밀번호/지점 모두 입력 시 활성화","로그인에 실패했습니다","aria-label='로그인'"),
            ("UI-005","비밀번호 찾기 링크","Link","비밀번호 재설정 페이지 이동 링크","클릭 시 비밀번호 찾기 모달 오픈",None,None,"<a> + 모달 트리거",None,None,"aria-label='비밀번호 찾기'"),
        ]),
        ("SCR-002", "대시보드", [
            ("UI-006","총회원 통계 카드","Card","전체 회원 수 표시 카드","클릭 시 SCR-010 이동, 전일 대비 증감 표시","FN-002","API-003","Card 컴포넌트 + 숫자 애니메이션",None,None,"aria-label='총 회원 수'"),
            ("UI-007","활성회원 통계 카드","Card","활성(이용중) 회원 수 카드","클릭 시 SCR-010 활성 필터 이동","FN-002","API-003","Card 컴포넌트",None,None,"aria-label='활성 회원 수'"),
            ("UI-008","만료임박 통계 카드","Card","7일 내 만료 예정 회원 수 카드","클릭 시 SCR-010 만료임박 필터 이동, 경고 색상","FN-002","API-006","Card 컴포넌트 + 경고 스타일",None,None,"aria-label='만료 임박 회원'"),
            ("UI-009","만료회원 통계 카드","Card","이용권 만료된 회원 수 카드","클릭 시 SCR-010 만료 필터 이동","FN-002","API-003","Card 컴포넌트",None,None,"aria-label='만료 회원 수'"),
            ("UI-010","생일회원 리스트","List","오늘 생일인 회원 목록","회원명 클릭 시 SCR-011 이동","FN-002","API-004","리스트 컴포넌트 (최대 5명)",None,None,"aria-label='오늘 생일 회원'"),
            ("UI-011","미납회원 리스트","List","미납 내역이 있는 회원 목록","회원명 클릭 시 SCR-011 이동, 미납 금액 표시","FN-002","API-005","리스트 컴포넌트 + 금액 강조",None,None,"aria-label='미납 회원'"),
            ("UI-012","만료예정 리스트","List","7일 내 만료 예정 회원 목록","회원명 클릭 시 SCR-011 이동, D-day 표시","FN-002","API-006","리스트 컴포넌트 + D-day 배지",None,None,"aria-label='만료 예정 회원'"),
            ("UI-013","매출 차트","Card","당월 매출 추이 라인 차트","호버 시 일별 금액 툴팁, 기간 필터","FN-002","API-003","Recharts LineChart",None,None,"차트 데이터 테이블 대체 제공"),
        ]),
        ("SCR-010", "회원 목록", [
            ("UI-014","검색바","Input","회원 이름/연락처/번호 검색 입력","실시간 검색 (300ms debounce), 검색 아이콘","FN-003","API-007","SearchInput 컴포넌트","최소 2자 이상",None,"aria-label='회원 검색'"),
            ("UI-015","상태 필터","Select","회원 상태(활성/만료/휴회/탈퇴) 필터","멀티 선택 가능, 선택 시 즉시 필터 적용","FN-003","API-007","MultiSelect 컴포넌트",None,None,"aria-label='상태 필터'"),
            ("UI-016","성별 필터","Select","성별(전체/남/여) 필터","단일 선택, 즉시 적용","FN-003","API-007","Select 컴포넌트",None,None,"aria-label='성별 필터'"),
            ("UI-017","담당자 필터","Select","담당 FC 필터","단일 선택, 즉시 적용","FN-003","API-007","Select 컴포넌트",None,None,"aria-label='담당자 필터'"),
            ("UI-018","회원 테이블","List","회원 목록 데이터 테이블","행 클릭 시 SCR-011 이동, 컬럼 정렬, 행 호버","FN-003","API-007","DataTable 컴포넌트 + 정렬/페이지네이션",None,None,"aria-label='회원 목록 테이블'"),
            ("UI-019","페이지네이션","Nav","테이블 하단 페이지 네비게이션","페이지 이동, 페이지당 행 수 선택","FN-003","API-007","Pagination 컴포넌트",None,None,"aria-label='페이지 이동'"),
            ("UI-020","엑셀 다운로드","Button","현재 필터 결과 엑셀 다운로드","클릭 시 xlsx 파일 다운로드, 로딩 표시","FN-003","API-007","다운로드 버튼 + 로딩 스피너",None,None,"aria-label='엑셀 다운로드'"),
            ("UI-021","회원등록 버튼","Button","신규 회원 등록 페이지 이동 버튼","클릭 시 SCR-012 이동","FN-004",None,"Primary Button",None,None,"aria-label='회원 등록'"),
        ]),
        ("SCR-011", "회원 상세", [
            ("UI-022","프로필 영역","Card","회원 기본 정보 카드 (사진/이름/연락처/상태)","수정 버튼 → SCR-013 이동","FN-006","API-008","프로필 Card 컴포넌트",None,None,"aria-label='회원 프로필'"),
            ("UI-023","이용권 탭","Tab","보유 이용권 목록 탭","만료일 강조, 잔여 횟수 표시","FN-006","API-008","Tab + 이용권 리스트",None,None,"aria-label='이용권 탭'"),
            ("UI-024","출석 탭","Tab","출석 이력 탭 (히트맵 + 리스트)","히트맵 날짜 클릭 시 상세, 타지점 출석 구분","FN-006","API-013","Tab + 히트맵 + 테이블",None,None,"aria-label='출석 탭'"),
            ("UI-025","결제 탭","Tab","결제/매출 이력 탭","결제 상세 모달, 환불 처리","FN-006","API-015","Tab + 결제 테이블",None,None,"aria-label='결제 탭'"),
            ("UI-026","체성분 탭","Tab","체성분 측정 이력 + 차트 탭","측정 추가, 변화 추이 차트","FN-007","API-008","Tab + 차트 + 입력 폼",None,None,"aria-label='체성분 탭'"),
            ("UI-027","메모 탭","Tab","상담/메모 기록 탭","메모 추가/수정/삭제, 시간순 정렬","FN-006","API-008","Tab + 메모 리스트 + 에디터",None,None,"aria-label='메모 탭'"),
        ]),
        ("SCR-012", "회원 등록", [
            ("UI-028","이름 입력","Input","회원 이름 입력 필드","필수 항목, 한글 2~20자","FN-004","API-009","<input> + validation","한글 2~20자","이름을 입력하세요","aria-required='true'"),
            ("UI-029","성별 선택","Radio","성별 선택 (남/여)","라디오 버튼 2개","FN-004","API-009","RadioGroup 컴포넌트","필수 선택","성별을 선택하세요","role='radiogroup'"),
            ("UI-030","생년월일","Input","생년월일 입력 (date picker)","달력 팝업 또는 직접 입력","FN-004","API-009","DatePicker 컴포넌트","미래 날짜 불가","올바른 날짜를 입력하세요","aria-label='생년월일'"),
            ("UI-031","연락처 입력","Input","연락처 입력 (자동 하이픈)","입력 시 자동 포맷 (010-xxxx-xxxx)","FN-004","API-009","PhoneInput 컴포넌트","전화번호 형식 검증, 중복 체크","올바른 연락처를 입력하세요","aria-label='연락처'"),
            ("UI-032","이메일 입력","Input","이메일 주소 입력","선택 항목","FN-004","API-009","<input type='email'>","이메일 형식 검증","올바른 이메일을 입력하세요","aria-label='이메일'"),
            ("UI-033","주소 입력","Input","주소 검색 및 입력","다음 주소 API 연동","FN-004","API-009","AddressSearch 컴포넌트",None,None,"aria-label='주소'"),
            ("UI-034","메모 입력","Textarea","관리자 메모 입력","최대 500자, 글자수 카운터","FN-004","API-009","<textarea> + counter","최대 500자",None,"aria-label='메모'"),
            ("UI-035","담당자 선택","Select","담당 FC 선택 드롭다운","직원 목록에서 FC 역할만 필터","FN-004","API-009","Select 컴포넌트",None,None,"aria-label='담당자 선택'"),
            ("UI-036","저장 버튼","Button","회원 등록 저장","모든 필수 항목 입력 시 활성화, 클릭 시 API 호출","FN-004","API-009","Primary Button + loading","필수 항목 검증","등록에 실패했습니다","aria-label='저장'"),
            ("UI-037","취소 버튼","Button","등록 취소, 목록으로 이동","변경사항 있을 시 확인 모달",None,None,"Secondary Button",None,None,"aria-label='취소'"),
        ]),
        ("SCR-030", "매출 조회", [
            ("UI-038","날짜 범위 선택","Input","시작일~종료일 범위 선택","DateRangePicker, 프리셋(오늘/이번주/이번달)","FN-010","API-015","DateRangePicker 컴포넌트","시작일 ≤ 종료일","올바른 기간을 선택하세요","aria-label='조회 기간'"),
            ("UI-039","유형 필터","Select","매출 유형(이용권/PT/상품/기타) 필터","멀티 선택, 즉시 적용","FN-010","API-015","MultiSelect 컴포넌트",None,None,"aria-label='매출 유형'"),
            ("UI-040","상태 필터","Select","결제 상태(완료/환불/미납) 필터","멀티 선택, 즉시 적용","FN-010","API-015","MultiSelect 컴포넌트",None,None,"aria-label='결제 상태'"),
            ("UI-041","검색바","Input","구매자/상품명 검색","실시간 검색 (debounce)","FN-010","API-015","SearchInput",None,None,"aria-label='매출 검색'"),
            ("UI-042","매출 테이블","List","매출 내역 데이터 테이블","행 클릭 시 상세 모달, 정렬, 페이지네이션","FN-010","API-015","DataTable 컴포넌트",None,None,"aria-label='매출 목록'"),
            ("UI-043","총매출 요약","Card","총매출/현금/카드/마일리지 요약 카드","필터 변경 시 자동 갱신","FN-010","API-017","요약 Card 4개",None,None,"aria-label='매출 요약'"),
            ("UI-044","엑셀 다운로드","Button","매출 데이터 엑셀 다운로드","클릭 시 xlsx 다운로드","FN-010","API-015","다운로드 버튼",None,None,"aria-label='엑셀 다운로드'"),
        ]),
        ("SCR-031", "POS", [
            ("UI-045","카테고리 탭","Tab","상품 카테고리 탭 (이용권/PT/GX/기타)","탭 전환 시 상품 그리드 갱신","FN-011","API-018","Tab 컴포넌트",None,None,"aria-label='상품 카테고리'"),
            ("UI-046","상품 그리드","List","카테고리별 상품 카드 그리드","카드 클릭 시 장바구니 추가","FN-011","API-018","CSS Grid + Card",None,None,"aria-label='상품 목록'"),
            ("UI-047","장바구니","List","선택한 상품 목록 + 합계","수량 조절, 삭제, 합계 자동 계산","FN-011",None,"장바구니 컴포넌트",None,None,"aria-label='장바구니'"),
            ("UI-048","결제 버튼","Button","결제 화면(SCR-032) 이동","장바구니 1개 이상 시 활성화","FN-011",None,"Primary Button","장바구니 비어있으면 비활성","상품을 선택해주세요","aria-label='결제하기'"),
        ]),
        ("SCR-032", "POS 결제", [
            ("UI-049","결제 요약","Card","결제 대상 상품 요약 및 총액","상품명, 수량, 금액 표시","FN-012","API-016","요약 Card",None,None,"aria-label='결제 요약'"),
            ("UI-050","결제수단 선택","Radio","현금/카드/마일리지/복합 선택","복합 선택 시 금액 분배 입력 필드 표시","FN-012","API-016","RadioGroup + 조건부 입력",None,None,"role='radiogroup'"),
            ("UI-051","회원 검색","Input","결제 대상 회원 검색","회원 선택 시 마일리지 잔액 표시","FN-012","API-007","AutoComplete 컴포넌트",None,None,"aria-label='회원 검색'"),
            ("UI-052","결제 확인 버튼","Button","결제 실행","결제 확인 모달 → API 호출 → 영수증","FN-012","API-016","Primary Button + 확인 모달","결제수단/금액 검증","결제에 실패했습니다","aria-label='결제 확인'"),
        ]),
        ("SCR-050", "락커 현황", [
            ("UI-053","구역 탭","Tab","락커 구역(A/B/C...) 탭","탭 전환 시 해당 구역 그리드 표시","FN-015","API-020","Tab 컴포넌트",None,None,"aria-label='락커 구역'"),
            ("UI-054","락커 그리드","List","락커 번호별 색상 그리드","사용중(파랑)/빈(회색)/만료임박(주황)/고장(빨강)","FN-015","API-020","CSS Grid + 색상 매핑",None,None,"색상+패턴으로 상태 구분"),
            ("UI-055","범례","Label","락커 상태별 색상 범례","4가지 상태 색상 설명",None,None,"범례 컴포넌트",None,None,"aria-label='범례'"),
            ("UI-056","락커 상세 모달","Modal","락커 클릭 시 상세 정보 모달","배정 회원, 만료일, 이력 표시","FN-015","API-020","Modal 컴포넌트",None,None,"aria-modal='true'"),
        ]),
        ("SCR-060", "직원 목록", [
            ("UI-057","검색바","Input","직원 이름/연락처 검색","실시간 검색","FN-019","API-022","SearchInput",None,None,"aria-label='직원 검색'"),
            ("UI-058","역할 필터","Select","역할(센터장/매니저/FC/스태프) 필터","단일 선택","FN-019","API-022","Select",None,None,"aria-label='역할 필터'"),
            ("UI-059","상태 필터","Select","재직 상태(재직/휴직/퇴사) 필터","단일 선택","FN-019","API-022","Select",None,None,"aria-label='상태 필터'"),
            ("UI-060","직원 테이블","List","직원 목록 데이터 테이블","행 클릭 시 상세 모달","FN-019","API-022","DataTable",None,None,"aria-label='직원 목록'"),
            ("UI-061","직원등록 버튼","Button","SCR-061 이동","클릭 시 등록 페이지 이동","FN-020",None,"Primary Button",None,None,"aria-label='직원 등록'"),
        ]),
        ("SCR-070", "메시지 발송", [
            ("UI-062","수신자 선택","Input","회원 검색/그룹 선택으로 수신자 지정","AutoComplete + 그룹 선택 토글","FN-023",None,"AutoComplete + ChipInput",None,None,"aria-label='수신자'"),
            ("UI-063","메시지 유형","Tab","SMS/카카오톡 탭","탭에 따라 본문 제한 변경","FN-023",None,"Tab 컴포넌트",None,None,"aria-label='메시지 유형'"),
            ("UI-064","본문 에디터","Textarea","메시지 본문 입력","글자수 카운터, 변수 삽입 (#{이름} 등)","FN-023",None,"Textarea + 변수 버튼","SMS 90자/LMS 2000자",None,"aria-label='메시지 본문'"),
            ("UI-065","발송 버튼","Button","메시지 발송 실행","발송 전 미리보기 확인 모달","FN-023",None,"Primary Button + 확인 모달","수신자/본문 필수","수신자를 선택해주세요","aria-label='발송'"),
            ("UI-066","발송 이력","List","발송 이력 테이블","발송일시/유형/수신자수/성공률 표시","FN-023",None,"DataTable",None,None,"aria-label='발송 이력'"),
        ]),
        ("SCR-080", "계약 마법사", [
            ("UI-067","스텝 인디케이터","Nav","계약 단계 표시 (1.회원→2.상품→3.기간→4.결제→5.확인)","현재 단계 강조, 완료 단계 체크마크","FN-027",None,"Stepper 컴포넌트",None,None,"aria-label='계약 단계'"),
            ("UI-068","회원 선택 스텝","Card","계약 대상 회원 검색/선택","AutoComplete 검색, 선택 시 프로필 카드 표시","FN-027","API-007","검색 + 프로필 Card","회원 필수 선택","회원을 선택하세요","aria-label='회원 선택'"),
            ("UI-069","상품 선택 스텝","Card","계약 상품 선택","카테고리별 상품 목록, 복수 선택 가능","FN-027","API-018","상품 Card Grid","상품 필수 선택","상품을 선택하세요","aria-label='상품 선택'"),
            ("UI-070","기간/금액 스텝","Card","계약 기간 및 결제 금액 설정","시작일/종료일, 할인 적용, 총액 자동 계산","FN-027",None,"DatePicker + 금액 입력","시작일 필수","기간을 설정하세요","aria-label='기간 설정'"),
            ("UI-071","결제 스텝","Card","결제 수단 선택 및 처리","현금/카드/마일리지/복합결제","FN-027","API-016","결제 컴포넌트","결제수단 필수","결제에 실패했습니다","aria-label='결제'"),
            ("UI-072","확인 스텝","Card","계약 내용 최종 확인 및 완료","계약 요약, 확인 버튼, 인쇄 옵션","FN-027",None,"요약 Card + 확인 버튼",None,None,"aria-label='계약 확인'"),
        ]),
        ("SCR-090", "설정", [
            ("UI-073","기본 설정 탭","Tab","센터명/영업시간/연락처 등 기본 정보","폼 입력, 저장 버튼","FN-029",None,"Tab + Form",None,None,"aria-label='기본 설정'"),
            ("UI-074","알림 설정 탭","Tab","푸시/이메일/SMS 알림 ON/OFF","토글 스위치","FN-029",None,"Tab + Toggle List",None,None,"aria-label='알림 설정'"),
            ("UI-075","저장 버튼","Button","설정 저장","변경사항 감지 시 활성화","FN-029",None,"Primary Button",None,None,"aria-label='설정 저장'"),
        ]),
        ("SCR-091", "권한 설정", [
            ("UI-076","역할 탭","Tab","6단계 역할별 탭","탭 전환 시 해당 역할 권한 매트릭스","FN-030",None,"Tab 컴포넌트",None,None,"aria-label='역할 선택'"),
            ("UI-077","권한 매트릭스","List","메뉴/기능별 읽기/쓰기/삭제 체크박스 테이블","체크/해제 시 즉시 반영, 변경 이력 기록","FN-030",None,"체크박스 매트릭스 테이블",None,None,"aria-label='권한 매트릭스'"),
            ("UI-078","일괄 설정 버튼","Button","전체 선택/해제","메뉴 단위 일괄 설정","FN-030",None,"Toggle Button",None,None,"aria-label='일괄 설정'"),
        ]),
        ("SCR-094", "지점 관리", [
            ("UI-079","지점 카드 뷰","Card","지점별 카드 (이름/주소/회원수/상태)","카드 클릭 시 상세 모달","FN-033",None,"Card Grid",None,None,"aria-label='지점 목록'"),
            ("UI-080","지점 추가 버튼","Button","신규 지점 추가 모달 오픈","모달 폼 (이름/코드/주소/연락처)","FN-033",None,"Primary Button + Modal","지점명/코드 필수","지점명을 입력하세요","aria-label='지점 추가'"),
            ("UI-081","데이터 초기화 버튼","Button","지점 데이터 초기화 (개발용)","2중 확인 모달 (비밀번호 입력)","FN-033","API-027","Danger Button + 확인 모달","관리자 비밀번호 필수","비밀번호가 올바르지 않습니다","aria-label='데이터 초기화'"),
        ]),
        ("SCR-040", "상품 목록", [
            ("UI-082","카테고리 필터","Select","상품 카테고리 필터","단일 선택, 즉시 적용","FN-013","API-018","Select 컴포넌트",None,None,"aria-label='카테고리'"),
            ("UI-083","상품 테이블","List","상품 목록 테이블","정렬, 키오스크 노출 토글","FN-013","API-018","DataTable + Toggle",None,None,"aria-label='상품 목록'"),
            ("UI-084","상품등록 버튼","Button","SCR-041 이동","클릭 시 등록 페이지","FN-014",None,"Primary Button",None,None,"aria-label='상품 등록'"),
        ]),
        ("SCR-041", "상품 등록", [
            ("UI-085","카테고리 선택","Select","상품 카테고리 선택","선택에 따라 하위 필드 동적 변경","FN-014","API-019","Select + 조건부 필드","필수 선택","카테고리를 선택하세요","aria-required"),
            ("UI-086","상품명 입력","Input","상품명 입력","중복 체크","FN-014","API-019","<input> + 중복 체크","필수, 중복 불가","상품명을 입력하세요","aria-required"),
            ("UI-087","가격 입력","Input","현금가/카드가 입력","숫자만 입력, 천단위 콤마 포맷","FN-014","API-019","CurrencyInput","0 이상 숫자","올바른 금액을 입력하세요","aria-label='가격'"),
            ("UI-088","저장 버튼","Button","상품 등록 저장","필수 항목 검증 후 API 호출","FN-014","API-019","Primary Button","필수 항목 검증","등록에 실패했습니다","aria-label='저장'"),
        ]),
        ("SCR-051", "락커 관리", [
            ("UI-089","회원 검색","Input","배정 대상 회원 검색","AutoComplete","FN-016","API-007","AutoComplete",None,None,"aria-label='회원 검색'"),
            ("UI-090","빈 락커 선택","List","배정 가능한 빈 락커 그리드","클릭 선택, 선택 시 강조","FN-016","API-020","Grid + 선택 상태",None,None,"aria-label='락커 선택'"),
            ("UI-091","배정 버튼","Button","락커 배정 실행","회원+락커 선택 시 활성화","FN-016","API-021","Primary Button","회원/락커 필수","배정에 실패했습니다","aria-label='배정'"),
            ("UI-092","일괄 해제 버튼","Button","만료 락커 일괄 해제","확인 모달 후 실행","FN-016","API-021","Secondary Button",None,None,"aria-label='일괄 해제'"),
        ]),
        ("SCR-061", "직원 등록", [
            ("UI-093","이름 입력","Input","직원 이름","필수","FN-020","API-023","<input>","필수","이름을 입력하세요","aria-required"),
            ("UI-094","역할 선택","Select","직원 역할 선택","선택 시 권한 미리보기 표시","FN-020","API-023","Select + 권한 미리보기","필수","역할을 선택하세요","aria-required"),
            ("UI-095","연락처 입력","Input","직원 연락처","자동 포맷","FN-020","API-023","PhoneInput","전화번호 형식","올바른 연락처를 입력하세요","aria-label='연락처'"),
            ("UI-096","입사일 선택","Input","입사일 date picker","달력 팝업","FN-020","API-023","DatePicker","필수","입사일을 선택하세요","aria-label='입사일'"),
            ("UI-097","저장 버튼","Button","직원 등록 저장","필수 검증 후 저장","FN-020","API-023","Primary Button","필수 항목","등록에 실패했습니다","aria-label='저장'"),
        ]),
        ("SCR-062", "급여 현황", [
            ("UI-098","월 선택","Select","급여 조회 월 선택","월 변경 시 테이블 갱신","FN-021",None,"MonthPicker",None,None,"aria-label='조회 월'"),
            ("UI-099","급여 테이블","List","직원별 급여 테이블","기본급/인센티브/공제/실지급액 컬럼","FN-021",None,"DataTable",None,None,"aria-label='급여 현황'"),
            ("UI-100","합계 행","Label","전체 합계/평균 표시 행","테이블 하단 고정","FN-021",None,"테이블 footer",None,None,"aria-label='합계'"),
        ]),
        ("SCR-071", "자동 알림", [
            ("UI-101","알림 규칙 리스트","List","자동 알림 규칙 목록","각 규칙에 ON/OFF 토글","FN-024",None,"List + Toggle",None,None,"aria-label='알림 규칙'"),
            ("UI-102","규칙 설정 모달","Modal","알림 규칙 상세 설정","발송 시점/대상/메시지 템플릿 설정","FN-024",None,"Modal + Form",None,None,"aria-modal='true'"),
        ]),
        ("SCR-072", "쿠폰 관리", [
            ("UI-103","쿠폰 리스트","List","쿠폰 목록 테이블","상태/유형별 필터","FN-025",None,"DataTable",None,None,"aria-label='쿠폰 목록'"),
            ("UI-104","쿠폰 생성 버튼","Button","쿠폰 생성 모달 오픈","모달 폼 (이름/유형/할인값/기간)","FN-025",None,"Primary Button + Modal",None,None,"aria-label='쿠폰 생성'"),
        ]),
        ("SCR-073", "마일리지 관리", [
            ("UI-105","회원 검색","Input","마일리지 조회 회원 검색","AutoComplete","FN-026",None,"AutoComplete",None,None,"aria-label='회원 검색'"),
            ("UI-106","마일리지 이력","List","적립/사용 이력 테이블","유형/날짜/금액/잔액 표시","FN-026",None,"DataTable",None,None,"aria-label='마일리지 이력'"),
            ("UI-107","적립/차감 버튼","Button","마일리지 적립/차감 모달","금액 입력, 사유 선택","FN-026",None,"Button + Modal","금액 필수","금액을 입력하세요","aria-label='마일리지 변경'"),
        ]),
        ("SCR-052", "RFID 관리", [
            ("UI-108","카드 번호 입력","Input","RFID 카드 번호 스캔/입력","스캔 시 자동 입력, 직접 입력 가능","FN-017",None,"<input> + 스캔 연동",None,None,"aria-label='카드 번호'"),
            ("UI-109","회원 매핑","Select","카드-회원 매핑","회원 검색 후 매핑","FN-017",None,"AutoComplete",None,None,"aria-label='회원 매핑'"),
            ("UI-110","카드 이력","List","카드 등록/해제 이력","시간순 정렬","FN-017",None,"DataTable",None,None,"aria-label='카드 이력'"),
        ]),
        ("SCR-053", "룸 관리", [
            ("UI-111","룸 카드","Card","GX/PT 룸 카드 뷰","룸명/수용인원/상태 표시","FN-018",None,"Card Grid",None,None,"aria-label='룸 목록'"),
            ("UI-112","룸 등록 모달","Modal","신규 룸 등록","룸명/유형/수용인원/시간대 설정","FN-018",None,"Modal + Form","룸명 필수","룸명을 입력하세요","aria-modal='true'"),
        ]),
        ("SCR-081", "구독 관리", [
            ("UI-113","구독 목록","List","정기 구독 상품 목록","상태/결제일/금액 표시","FN-028",None,"DataTable",None,None,"aria-label='구독 목록'"),
            ("UI-114","구독 상세 모달","Modal","구독 상세 정보","결제 이력, 해지 옵션","FN-028",None,"Modal",None,None,"aria-modal='true'"),
        ]),
        ("SCR-092", "키오스크 설정", [
            ("UI-115","미리보기","Card","키오스크 화면 미리보기","실시간 설정 반영","FN-031",None,"Preview 컴포넌트",None,None,"aria-label='미리보기'"),
            ("UI-116","로고 업로드","Button","센터 로고 업로드","이미지 파일 선택, 5MB 제한","FN-031",None,"FileUpload","5MB 이하, JPG/PNG","5MB 이하 파일만 가능","aria-label='로고 업로드'"),
        ]),
        ("SCR-093", "IoT 설정", [
            ("UI-117","기기 목록","List","연동 기기 목록","연결상태 표시 (온라인/오프라인)","FN-032",None,"DataTable + 상태 아이콘",None,None,"aria-label='기기 목록'"),
            ("UI-118","연결 테스트 버튼","Button","기기 연결 테스트","클릭 시 ping 테스트, 결과 표시","FN-032",None,"Button + 결과 표시",None,None,"aria-label='연결 테스트'"),
        ]),
        ("SCR-013", "회원 수정", [
            ("UI-119","기존 정보 프리필 폼","Card","기존 회원 정보가 채워진 폼","변경된 필드 강조 표시","FN-005","API-010","폼 + diff 표시",None,None,"aria-label='회원 정보 수정'"),
            ("UI-120","저장 버튼","Button","수정 저장","변경사항 확인 모달 후 저장","FN-005","API-010","Primary Button + 확인 모달","변경사항 존재","수정에 실패했습니다","aria-label='저장'"),
        ]),
        ("SCR-014", "체성분 분석 추가", [
            ("UI-121","측정값 입력 폼","Card","체중/골격근/체지방률 등 입력","숫자 입력, 단위 표시","FN-007",None,"입력 폼",None,None,"aria-label='체성분 입력'"),
            ("UI-122","변화 추이 차트","Card","측정일 기준 라인 차트","항목별 토글, 호버 툴팁","FN-007",None,"Recharts LineChart",None,None,"차트 데이터 테이블 대체"),
        ]),
        ("SCR-020", "출석 기록 추가", [
            ("UI-123","날짜 필터","Input","출석 조회 날짜 선택","일별/주별/월별 토글","FN-008","API-013","DatePicker + 뷰 토글",None,None,"aria-label='날짜 필터'"),
            ("UI-124","출석 테이블","List","출석 이력 데이터 테이블","유형/시간/타지점 여부 표시","FN-008","API-013","DataTable",None,None,"aria-label='출석 목록'"),
        ]),
        ("SCR-021", "일정 관리 추가", [
            ("UI-125","캘린더 뷰","Card","월/주/일 캘린더","드래그앤드롭 일정 이동","FN-009",None,"FullCalendar",None,None,"키보드 네비게이션 지원"),
            ("UI-126","일정 상세 모달","Modal","일정 클릭 시 상세","수정/삭제/회원 연결","FN-009",None,"Modal + Form",None,None,"aria-modal='true'"),
        ]),
        ("SCR-063", "급여명세서 추가", [
            ("UI-127","직원/월 선택","Select","급여명세서 조회 조건","직원 + 월 선택","FN-022",None,"Select 2개",None,None,"aria-label='조회 조건'"),
            ("UI-128","명세서 상세","Card","급여명세서 상세 내역","기본급/수당/공제/실지급 항목별 표시","FN-022",None,"명세서 Card",None,None,"aria-label='급여명세서'"),
            ("UI-129","PDF 다운로드","Button","명세서 PDF 다운로드","클릭 시 PDF 생성/다운로드","FN-022",None,"Button + html2pdf",None,None,"aria-label='PDF 다운로드'"),
        ]),
    ]

    row = 2
    ui_id_counter = 1
    for scr_id, scr_name, elements in ui_groups:
        # 그룹 헤더 행
        ws.cell(row, 1, f"▼ {scr_id}")
        ws.cell(row, 2, scr_name)
        ws.cell(row, 3, f"{len(elements)}개 요소")
        for ci in range(1, 20):
            c = ws.cell(row, ci)
            c.fill = GROUP_HEADER_FILL
            c.font = GROUP_HEADER_FONT
            c.border = THIN_BORDER
        row += 1

        # 하위 요소 행
        for elem in elements:
            eid, ename, etype, edesc, einter, efn, eapi, eweb, evalid, eerr, eacc = elem
            ws.cell(row, 1, eid)
            ws.cell(row, 2, scr_id)
            ws.cell(row, 3, ename)
            ws.cell(row, 4, etype)
            ws.cell(row, 5, edesc)
            ws.cell(row, 6, einter)
            ws.cell(row, 7, efn)
            ws.cell(row, 8, eapi)
            ws.cell(row, 9, eweb)
            ws.cell(row, 10, None)  # App 구현
            ws.cell(row, 11, evalid)
            ws.cell(row, 12, eerr)
            ws.cell(row, 13, eacc)
            ws.cell(row, 14, "설계중")  # 상태
            ws.cell(row, 15, None)  # 담당자
            # [자동] 수식
            ws.cell(row, 16, f"=IFERROR(VLOOKUP(B{row},'IA(정보구조)'!D:E,2,FALSE()),\"\")")
            ws.cell(row, 17, f'=IFERROR(VLOOKUP(G{row},기능명세서!A:B,2,FALSE()),"")')
            ws.cell(row, 18, f'=IFERROR(VLOOKUP(H{row},API명세서!A:D,4,FALSE()),"")')
            ws.cell(row, 19, None)  # 비고

            # 행 그룹핑
            ws.row_dimensions[row].outline_level = 1

            for ci in range(1, 20):
                c = ws.cell(row, ci)
                c.border = THIN_BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

    # 드롭다운
    add_dv(ws, "B", SCR_IDS_STR, 500)
    add_dv(ws, "D", "Button,Input,Textarea,Select,Checkbox,Radio,Toggle,Link,Nav,Tab,Slide,Card,List,Modal,Toast,Tooltip,Image,Icon,Badge,Label,Text,Divider,기타", 500)
    add_dv(ws, "G", FN_IDS_STR, 500)
    add_dv(ws, "H", API_IDS_STR, 500)
    add_dv(ws, "N", "설계중,개발중,개발완료,테스트중,완료", 500)
    add_dv(ws, "O", 담당자_LIST, 500)

    ws.auto_filter.ref = "A1:S1"
    return ws


# ============================================================
# 5. 기능명세서 시트
# ============================================================
def create_function_spec_sheet(wb):
    ws = wb.create_sheet("기능명세서")

    headers = ["기능 ID","기능명","기능 분류","플랫폼","우선순위","기능 설명","상세 요구사항",
               "화면 ID","입력 항목","출력 항목","비즈니스 규칙","예외 처리",
               "Web 구현 사항","App 구현 사항","관련 API ID","상태","담당자",
               "[자동] 화면명","[자동] API 엔드포인트","비고","[자동] 관련 UI 요소"]
    col_widths = {"A":12,"B":18,"C":14,"D":10,"E":10,"F":35,"G":35,"H":12,"I":30,"J":25,
                  "K":30,"L":25,"M":25,"N":14,"O":12,"P":10,"Q":12,"R":18,"S":20,"T":20,"U":25}
    apply_header(ws, headers, col_widths)

    # (화면ID, 화면설명, [(기능ID, 기능명, 분류, 플랫폼, 우선순위, 설명, 상세요구, 입력, 출력, 비즈니스규칙, 예외처리, Web구현, API_ID)])
    fn_groups = [
        ("SCR-001", "로그인 (1개 기능)", [
            ("FN-001","로그인","인증","Web","Critical",
             "이메일/비밀번호 기반 로그인 + 지점 선택",
             "1) 이메일/비밀번호 입력\n2) 지점 선택 필수\n3) JWT 토큰 발급\n4) 로그인 실패 횟수 추적",
             "이메일, 비밀번호, 지점ID","JWT 토큰, 사용자 정보, 권한 목록",
             "1) 5회 연속 실패 시 30분 계정 잠금\n2) 비밀번호 90일 경과 시 변경 안내\n3) 동시 세션 1개 제한",
             "계정 잠금: 잠금 해제 안내\n서버 오류: 재시도 안내\n탈퇴 계정: 접근 불가 안내",
             "React Hook Form + JWT localStorage 저장",
             "API-001"),
        ]),
        ("SCR-002", "대시보드 (1개 기능)", [
            ("FN-002","대시보드 통계 조회","대시보드","Web","Critical",
             "지점별 핵심 KPI 대시보드 (회원/매출/출석 요약)",
             "1) 총회원/활성/만료임박/만료 카운트\n2) 오늘 생일 회원 리스트\n3) 미납 회원 리스트\n4) 만료 예정 회원\n5) 당월 매출 차트",
             "지점ID (JWT에서 추출)","통계 데이터, 리스트, 차트 데이터",
             "1) branch_id 기준 데이터 필터\n2) 만료임박 = D-7 이내\n3) 매출 차트는 최근 30일",
             "데이터 로딩 실패: 섹션별 에러 표시 + 재시도",
             "React Query + Recharts 차트",
             "API-003"),
        ]),
        ("SCR-010", "회원 목록 (2개 기능)", [
            ("FN-003","회원 목록 조회/검색/필터","회원관리","Web","Critical",
             "회원 검색, 필터, 페이지네이션 목록 조회",
             "1) 이름/연락처/회원번호 검색 (300ms debounce)\n2) 상태/성별/담당자 필터 조합\n3) 컬럼 정렬\n4) 페이지당 20/50/100건\n5) 엑셀 다운로드",
             "검색어, 필터 조건, 페이지 번호, 정렬 기준","회원 목록 (이름/연락처/상태/이용권/담당자)",
             "1) branch_id 기준 데이터 격리\n2) 권한별 조회 범위 제한\n3) 삭제 회원 기본 미표시",
             "검색 결과 없음: 빈 상태 UI\n서버 오류: 재시도 안내",
             "TanStack Table + React Query + xlsx export",
             "API-007"),
            ("FN-004","회원 등록","회원관리","Web","Critical",
             "신규 회원 등록",
             "1) 필수: 이름, 성별, 연락처\n2) 선택: 이메일, 주소, 생년월일, 메모\n3) 담당 FC 배정\n4) 연락처 중복 체크",
             "이름, 성별, 생년월일, 연락처, 이메일, 주소, 메모, 담당자ID","등록된 회원 정보",
             "1) 연락처 중복 불가 (동일 지점 내)\n2) 등록 시 first_reg_date 자동 설정\n3) 기본 상태: 활성",
             "중복 연락처: 기존 회원 안내\n필수 미입력: 필드별 에러",
             "React Hook Form + Zod validation",
             "API-009"),
        ]),
        ("SCR-011", "회원 상세 (2개 기능)", [
            ("FN-005","회원 정보 수정","회원관리","Web","High",
             "기존 회원 정보 수정",
             "1) 기존 데이터 프리필\n2) 변경 필드 강조\n3) 변경 이력 자동 기록\n4) 담당자 변경 가능",
             "수정할 회원 정보","수정된 회원 정보",
             "1) 변경 이력 audit log 자동 생성\n2) 연락처 변경 시 중복 체크",
             "변경 없이 저장: 안내 메시지\n서버 오류: 롤백",
             "React Hook Form + diff 비교",
             "API-010"),
            ("FN-006","회원 상세 조회","회원관리","Web","High",
             "회원 프로필 + 탭별 상세 정보 조회",
             "1) 프로필: 기본 정보 + 사진\n2) 이용권 탭: 보유 이용권 목록\n3) 출석 탭: 히트맵 + 리스트\n4) 결제 탭: 결제 이력\n5) 체성분 탭: 측정 기록 + 차트\n6) 메모 탭: 상담/메모",
             "회원ID","회원 전체 정보 (프로필+이용권+출석+결제+체성분+메모)",
             "1) 탭별 lazy loading\n2) 이용권 만료일 강조\n3) 출석 히트맵 최근 6개월",
             "회원 미존재: 404 페이지\n데이터 로딩 실패: 탭별 에러",
             "React Tabs + React Query parallel queries",
             "API-008"),
        ]),
        ("SCR-014", "체성분 (1개 기능)", [
            ("FN-007","체성분 기록/분석","회원관리","Web","High",
             "회원 체성분 측정 기록 입력 및 변화 추이 분석",
             "1) 체중/골격근/체지방률/BMI 등 입력\n2) 측정일 기준 시계열 차트\n3) 목표 대비 달성률 게이지",
             "체중, 골격근, 체지방률, 측정일","체성분 이력, 변화 차트",
             "1) 측정값 유효 범위 검증 (체중 20~300kg)\n2) 동일 날짜 중복 측정 시 덮어쓰기 확인",
             "범위 초과: 경고 후 저장 가능\n측정 기록 없음: 빈 상태 UI",
             "Recharts + 게이지 컴포넌트",
             "API-008"),
        ]),
        ("SCR-020", "출석 (1개 기능)", [
            ("FN-008","출석 기록 조회","출석관리","Web","High",
             "회원 출석 현황 조회 (일별/주별/월별)",
             "1) 날짜 필터 (일/주/월 뷰)\n2) 출석 유형(일반/PT/GX) 필터\n3) 타지점 출석 표시\n4) 키오스크/앱 체크인 구분",
             "날짜 범위, 필터 조건","출석 목록, 통계",
             "1) branch_id 기준 + 타지점 출석 포함 옵션\n2) 출석 유형별 색상 구분",
             "데이터 없음: 빈 상태 UI",
             "DataTable + 캘린더 히트맵",
             "API-013"),
        ]),
        ("SCR-021", "일정 (1개 기능)", [
            ("FN-009","일정 관리","출석관리","Web","High",
             "PT/GX 일정 캘린더 관리",
             "1) 월/주/일 캘린더 뷰 전환\n2) 드래그앤드롭 일정 이동\n3) 트레이너별 색상 구분\n4) 일정 충돌 감지",
             "일정 제목, 시간, 트레이너, 회원, 룸","캘린더 일정 목록",
             "1) 동일 시간대 동일 트레이너 중복 불가\n2) 동일 룸 동일 시간 중복 불가\n3) 과거 일정 수정 불가",
             "일정 충돌: 경고 메시지\n과거 일정: 수정 비활성화",
             "FullCalendar React + React Query",
             None),
        ]),
        ("SCR-030", "매출 (2개 기능)", [
            ("FN-010","매출 조회/통계","매출관리","Web","Critical",
             "매출 현황 조회 및 통계 요약",
             "1) 기간별 필터 (프리셋: 오늘/이번주/이번달/커스텀)\n2) 유형별 필터 (이용권/PT/상품/기타)\n3) 상태별 필터 (완료/환불/미납)\n4) 통계 카드: 총매출/현금/카드/마일리지\n5) 엑셀 다운로드",
             "기간, 유형, 상태, 페이지","매출 목록, 통계 요약",
             "1) branch_id 기준 데이터 격리\n2) 환불 건 매출에서 차감\n3) 천단위 콤마 포맷",
             "데이터 없음: 빈 상태\n다운로드 실패: 재시도",
             "DataTable + 통계 Card + xlsx export",
             "API-015"),
            ("FN-011","POS 상품 선택/결제","매출관리","Web","Critical",
             "POS 상품 선택 및 장바구니 관리",
             "1) 카테고리별 상품 그리드\n2) 상품 클릭 → 장바구니 추가\n3) 수량 조절/삭제\n4) 합계 자동 계산",
             "상품 선택, 수량","장바구니 목록, 합계",
             "1) 판매중 상품만 표시\n2) 장바구니 비어있으면 결제 불가",
             "재고 없는 상품: 비활성화",
             "Zustand 장바구니 상태관리 + Grid",
             "API-018"),
        ]),
        ("SCR-032", "POS 결제 (1개 기능)", [
            ("FN-012","POS 결제 처리","매출관리","Web","Critical",
             "결제 수단 선택 및 결제 실행",
             "1) 현금/카드/마일리지/복합결제\n2) 회원 검색 → 마일리지 잔액 표시\n3) 결제 확인 모달\n4) 영수증 출력 옵션",
             "결제수단, 금액, 회원ID","결제 결과, 영수증",
             "1) 복합결제 시 총액 일치 검증\n2) 마일리지 잔액 초과 사용 불가\n3) 결제 완료 시 매출 자동 등록",
             "결제 실패: 재시도 옵션\n마일리지 부족: 잔액 안내",
             "결제 API + 영수증 컴포넌트",
             "API-016"),
        ]),
        ("SCR-040", "상품 (2개 기능)", [
            ("FN-013","상품 목록 조회","상품관리","Web","High",
             "이용권/PT/GX/기타 상품 목록 조회",
             "1) 카테고리 필터\n2) 상태(판매중/중지) 필터\n3) 키오스크 노출 토글",
             "카테고리, 상태","상품 목록",
             "1) branch_id 기준 데이터 격리",
             "상품 없음: 빈 상태",
             "DataTable + Toggle",
             "API-018"),
            ("FN-014","상품 등록/수정","상품관리","Web","High",
             "신규 상품 등록 및 기존 상품 수정",
             "1) 카테고리/이름/현금가/카드가/기간 입력\n2) 키오스크 노출 설정\n3) 상품명 중복 체크",
             "카테고리, 이름, 가격, 기간","등록/수정된 상품 정보",
             "1) 동일 카테고리 내 상품명 중복 불가\n2) 가격 0원 이상",
             "중복 상품명: 경고\n필수 미입력: 에러",
             "React Hook Form",
             "API-019"),
        ]),
        ("SCR-050", "락커 (2개 기능)", [
            ("FN-015","락커 현황 조회","시설관리","Web","High",
             "구역별 락커 배정 현황 조회 (색상 그리드)",
             "1) 구역별 탭\n2) 색상 상태 (사용중/빈/만료임박/고장)\n3) 클릭 시 상세 모달",
             "구역","락커 그리드 데이터",
             "1) branch_id 기준\n2) 만료임박 = D-3 이내",
             "데이터 없음: 락커 미등록 안내",
             "CSS Grid + 색상 매핑",
             "API-020"),
            ("FN-016","락커 배정/관리","시설관리","Web","High",
             "락커 배정/해제/일괄 관리",
             "1) 회원 검색 → 빈 락커 선택 → 배정\n2) 만료일 설정\n3) 일괄 해제 (만료 락커)\n4) 고장 등록",
             "회원ID, 락커번호, 만료일","배정 결과",
             "1) 이미 배정된 락커 재배정 불가\n2) 만료일 필수",
             "중복 배정: 에러\n배정 실패: 재시도",
             "모달 + API 호출",
             "API-021"),
        ]),
        ("SCR-052", "RFID (1개 기능)", [
            ("FN-017","RFID 관리","시설관리","Web","Medium",
             "RFID 카드 등록/해제/이력 조회",
             "1) 카드 번호 스캔/수동 입력\n2) 회원-카드 매핑\n3) 분실 처리\n4) 이력 조회",
             "카드 번호, 회원ID","매핑 결과, 이력",
             "1) 1회원 1카드 원칙\n2) 분실 처리 시 기존 매핑 해제",
             "미등록 카드: 에러\n이미 매핑: 경고",
             "RFID 리더 연동 + DataTable",
             None),
        ]),
        ("SCR-053", "룸 (1개 기능)", [
            ("FN-018","룸 관리","시설관리","Web","Medium",
             "GX룸/PT룸 등록/수정/삭제",
             "1) 룸 CRUD\n2) 수용 인원 설정\n3) 예약 가능 시간대 설정",
             "룸명, 유형, 수용인원, 시간대","룸 정보",
             "1) 룸명 중복 불가\n2) 기존 예약 있으면 삭제 불가",
             "예약 존재 삭제: 에러 메시지",
             "Card + Modal CRUD",
             None),
        ]),
        ("SCR-060", "직원 (2개 기능)", [
            ("FN-019","직원 목록 조회","직원관리","Web","High",
             "직원 검색/필터/목록 조회",
             "1) 이름/연락처 검색\n2) 역할/상태 필터\n3) 테이블 정렬",
             "검색어, 필터","직원 목록",
             "1) branch_id 기준\n2) 센터장 이상만 접근",
             "결과 없음: 빈 상태",
             "DataTable + React Query",
             "API-022"),
            ("FN-020","직원 등록/수정","직원관리","Web","High",
             "신규 직원 등록 및 기존 직원 수정",
             "1) 이름/역할/연락처/입사일 입력\n2) 관리자 계정 자동 생성\n3) 근무 유형 설정",
             "이름, 역할, 연락처, 입사일","등록된 직원 정보",
             "1) 역할별 기본 권한 자동 배정\n2) 관리자 계정 이메일 필수",
             "중복 이메일: 에러\n필수 미입력: 에러",
             "React Hook Form + 권한 미리보기",
             "API-023"),
        ]),
        ("SCR-062", "급여 (2개 기능)", [
            ("FN-021","급여 관리","직원관리","Web","High",
             "월별 급여 현황 조회 및 관리",
             "1) 월별 급여 테이블\n2) 기본급/인센티브/공제 내역\n3) 지급 상태 관리\n4) 합계/평균 요약",
             "연월","급여 목록, 합계",
             "1) 실지급액 = 기본급 + 인센티브 - 공제\n2) 지급 완료 시 수정 불가",
             "데이터 없음: 빈 상태",
             "DataTable + footer 합계",
             None),
            ("FN-022","급여명세서 조회","직원관리","Web","Medium",
             "직원별 급여명세서 상세 조회 및 PDF 다운로드",
             "1) 직원/월 선택\n2) 상세 항목별 표시\n3) PDF 다운로드",
             "직원ID, 연월","급여명세서 상세",
             "1) 본인 명세서만 조회 가능 (센터장 이상은 전체)",
             "명세서 없음: 안내",
             "명세서 컴포넌트 + html2pdf",
             None),
        ]),
        ("SCR-070", "메시지 (2개 기능)", [
            ("FN-023","메시지 발송","메시지/마케팅","Web","High",
             "SMS/카카오톡 메시지 발송",
             "1) 수신자 선택 (개별/그룹/전체)\n2) 메시지 유형(SMS/LMS/카카오)\n3) 변수 삽입 (#{이름})\n4) 발송 이력 조회",
             "수신자, 유형, 본문","발송 결과",
             "1) SMS 90자, LMS 2000자 제한\n2) 발송 비용 사전 표시\n3) 예약 발송 지원",
             "수신자 없음: 에러\n발송 실패: 실패 건 재발송",
             "메시지 API + 변수 치환",
             None),
            ("FN-024","자동 알림 설정","메시지/마케팅","Web","Medium",
             "자동 알림 규칙 설정",
             "1) 만료 D-7/D-3/D-1 알림\n2) 생일 축하 알림\n3) 장기 미출석(30일) 알림\n4) 규칙별 ON/OFF",
             "규칙 설정","알림 발송",
             "1) 매일 오전 9시 자동 실행\n2) 중복 발송 방지 (동일 회원 동일 규칙 1회/일)",
             "템플릿 미입력: 에러",
             "규칙 엔진 + 스케줄러",
             None),
        ]),
        ("SCR-072", "쿠폰/마일리지 (2개 기능)", [
            ("FN-025","쿠폰 관리","메시지/마케팅","Web","Medium",
             "쿠폰 생성/발급/이력 관리",
             "1) 할인율/금액 쿠폰 생성\n2) 유효기간 설정\n3) 사용 한도 설정\n4) 발급 대상 선택\n5) 사용 이력 추적",
             "쿠폰 정보","쿠폰 목록, 이력",
             "1) 만료 쿠폰 자동 비활성화\n2) 사용 한도 도달 시 자동 마감",
             "만료 쿠폰 사용: 에러\n한도 초과: 에러",
             "CRUD + DataTable",
             None),
            ("FN-026","마일리지 관리","메시지/마케팅","Web","Medium",
             "회원 마일리지 적립/사용 관리",
             "1) 마일리지 적립 (결제 시 자동/수동)\n2) 마일리지 차감\n3) 잔액 조회\n4) 이력 조회",
             "회원ID, 금액, 유형","마일리지 잔액, 이력",
             "1) 잔액 부족 시 차감 불가\n2) 결제 시 자동 적립률 설정 가능",
             "잔액 부족: 에러\n음수 입력: 에러",
             "모달 + DataTable",
             None),
        ]),
        ("SCR-080", "계약/구독 (2개 기능)", [
            ("FN-027","계약 등록","계약/구독","Web","Critical",
             "단계별 계약 등록 위자드",
             "1) 회원 선택 → 상품 선택 → 기간/금액 설정 → 결제 → 확인\n2) 할인 적용\n3) 계약서 자동 생성\n4) 결제 처리",
             "회원ID, 상품ID, 기간, 금액, 결제수단","계약 정보, 영수증",
             "1) 계약 시작일 기준 이용권 자동 생성\n2) 할인 최대 50%\n3) 결제 완료 시 매출 자동 등록",
             "결제 실패: 계약 보류\n이전 단계 미완료: 다음 진행 불가",
             "Stepper + React Hook Form + 결제 API",
             "API-016"),
            ("FN-028","구독 관리","계약/구독","Web","Medium",
             "정기 구독 상품 관리",
             "1) 구독 상품 CRUD\n2) 자동 결제 설정\n3) 구독 현황 대시보드\n4) 해지 처리",
             "구독 상품 정보","구독 목록, 결제 이력",
             "1) 자동 결제 실패 시 3일 내 재시도\n2) 3회 실패 시 구독 일시정지",
             "결제 실패: 회원 알림\n해지: 잔여 기간 환불 계산",
             "DataTable + Modal + 결제 스케줄러",
             None),
        ]),
        ("SCR-090", "설정/권한 (5개 기능)", [
            ("FN-029","시스템 설정","설정","Web","Medium",
             "시스템 기본 설정 관리",
             "1) 센터명/영업시간/연락처 설정\n2) 알림 기본값 설정\n3) 테마 설정",
             "설정 값","저장 결과",
             "1) 설정 변경 즉시 반영\n2) 변경 이력 기록",
             "저장 실패: 재시도",
             "Tab + Form",
             None),
            ("FN-030","권한 관리","설정","Web","Critical",
             "역할별 메뉴/기능 권한 설정",
             "1) 6단계 역할 (최고관리자~조회전용)\n2) 메뉴별 접근 권한\n3) 기능별 읽기/쓰기/삭제\n4) 변경 이력 로그",
             "역할, 권한 매트릭스","권한 설정 결과",
             "1) 최고관리자 권한 변경 불가\n2) 하위 역할에 상위 권한 부여 불가",
             "권한 충돌: 경고\n자기 자신 권한 축소: 확인 모달",
             "체크박스 매트릭스 + API",
             None),
            ("FN-031","키오스크 설정","설정","Web","Medium",
             "키오스크 디스플레이/동작 설정",
             "1) 로고/배경 업로드\n2) 동작 시나리오 설정\n3) 미리보기",
             "이미지 파일, 설정 값","키오스크 설정",
             "1) 이미지 5MB 이하\n2) JPG/PNG만 허용",
             "용량 초과: 에러\n형식 오류: 에러",
             "FileUpload + Preview",
             None),
            ("FN-032","IoT 설정","설정","Web","Medium",
             "출입문/체성분 기기 연동 설정",
             "1) 기기 등록/해제\n2) 연결 상태 모니터링\n3) 연결 테스트",
             "기기 정보","기기 목록, 연결 상태",
             "1) 기기 IP/포트 유효성 검증\n2) 연결 상태 5분 주기 갱신",
             "연결 실패: 재시도 안내",
             "DataTable + 상태 아이콘 + ping API",
             None),
            ("FN-033","지점 관리","설정","Web","Critical",
             "지점 CRUD 및 멀티테넌트 관리",
             "1) 지점 추가 (이름/코드/주소/연락처)\n2) 지점 수정\n3) 지점 비활성화\n4) 데이터 초기화 (개발용)",
             "지점 정보","지점 목록",
             "1) 지점 코드 고유\n2) 지점별 데이터 완전 격리\n3) 데이터 초기화 시 2중 확인",
             "중복 코드: 에러\n데이터 존재 삭제: 비활성화만 가능",
             "Card/Table + Modal CRUD",
             "API-027"),
        ]),
    ]

    row = 2
    for scr_id, group_label, functions in fn_groups:
        # 그룹 헤더
        ws.cell(row, 1, f"▼ {scr_id}")
        ws.cell(row, 2, group_label)
        for ci in range(1, 22):
            c = ws.cell(row, ci)
            c.fill = GROUP_HEADER_FILL
            c.font = GROUP_HEADER_FONT
            c.border = THIN_BORDER
        row += 1

        # 기능 행
        for fn in functions:
            fn_id, fn_name, fn_cat, fn_plat, fn_pri, fn_desc, fn_detail, fn_input, fn_output, fn_biz, fn_exc, fn_web, fn_api = fn
            ws.cell(row, 1, fn_id)
            ws.cell(row, 2, fn_name)
            ws.cell(row, 3, fn_cat)
            ws.cell(row, 4, fn_plat)
            ws.cell(row, 5, fn_pri)
            ws.cell(row, 6, fn_desc)
            ws.cell(row, 7, fn_detail)
            ws.cell(row, 8, scr_id)
            ws.cell(row, 9, fn_input)
            ws.cell(row, 10, fn_output)
            ws.cell(row, 11, fn_biz)
            ws.cell(row, 12, fn_exc)
            ws.cell(row, 13, fn_web)
            ws.cell(row, 14, None)  # App 구현
            ws.cell(row, 15, fn_api)
            ws.cell(row, 16, "확정")  # 상태
            ws.cell(row, 17, None)  # 담당자
            # [자동] 수식
            ws.cell(row, 18, f"=IFERROR(VLOOKUP(H{row},'IA(정보구조)'!D:E,2,FALSE()),\"\")")
            ws.cell(row, 19, f'=IFERROR(VLOOKUP(O{row},API명세서!A:D,4,FALSE()),"")')
            ws.cell(row, 20, None)  # 비고
            ws.cell(row, 21, f"=IFERROR(_xlfn.TEXTJOIN(\", \",TRUE(),IF('UI 요소 상세'!G$1:G$500=A{row},'UI 요소 상세'!C$1:C$500,\"\")),\"\")")

            ws.row_dimensions[row].outline_level = 1
            for ci in range(1, 22):
                c = ws.cell(row, ci)
                c.border = THIN_BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

    # 드롭다운
    add_dv(ws, "D", "Web,App (iOS),App (Android),Web + App,공통")
    add_dv(ws, "E", "Critical,High,Medium,Low")
    add_dv(ws, "H", SCR_IDS_STR)
    add_dv(ws, "P", "작성중,검토중,확정,변경필요,삭제")
    add_dv(ws, "Q", 담당자_LIST)

    ws.auto_filter.ref = "A1:T1"
    return ws


# ============================================================
# 6. API명세서 시트
# ============================================================
def create_api_spec_sheet(wb):
    ws = wb.create_sheet("API명세서")

    headers = ["API ID","API명","Method","엔드포인트","설명","Request Headers",
               "Request Body","Response (성공)","Response (실패)","인증 필요",
               "화면 ID","기능 ID","상태","담당자","[자동] 화면명","[자동] 기능명",
               "비고","[자동] 관련 UI 요소"]
    col_widths = {"A":12,"B":20,"C":8,"D":25,"E":30,"F":25,"G":35,"H":35,"I":30,
                  "J":10,"K":12,"L":10,"M":10,"N":12,"O":18,"P":16,"Q":20,"R":25}
    apply_header(ws, headers, col_widths)

    apis = [
        ("API-001","로그인","POST","/api/v1/auth/login","이메일/비밀번호 로그인 인증",
         "Content-Type: application/json",
         '{\n  "email": "string",\n  "password": "string",\n  "branch_id": "number"\n}',
         '{\n  "token": "jwt_string",\n  "user": {\n    "id": 1,\n    "name": "홍길동",\n    "role": "manager",\n    "branch_id": 1\n  }\n}',
         '{\n  "error": "INVALID_CREDENTIALS",\n  "message": "이메일 또는 비밀번호가 올바르지 않습니다"\n}',
         "N","SCR-001","FN-001","개발완료"),
        ("API-002","로그아웃","POST","/api/v1/auth/logout","세션 종료 및 토큰 무효화",
         "Authorization: Bearer {token}",
         None,
         '{\n  "message": "로그아웃 되었습니다"\n}',
         '{\n  "error": "UNAUTHORIZED",\n  "message": "인증이 필요합니다"\n}',
         "Y","SCR-002","FN-001","개발완료"),
        ("API-003","대시보드 통계","GET","/api/v1/dashboard/stats","지점별 대시보드 핵심 KPI 조회",
         "Authorization: Bearer {token}",
         None,
         '{\n  "totalMembers": 450,\n  "activeMembers": 320,\n  "expiringMembers": 25,\n  "expiredMembers": 105,\n  "monthlySales": 15000000,\n  "todayAttendance": 87\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "데이터를 불러올 수 없습니다"\n}',
         "Y","SCR-002","FN-002","개발완료"),
        ("API-004","생일 회원 조회","GET","/api/v1/dashboard/birthday-members","오늘 생일인 회원 목록 조회",
         "Authorization: Bearer {token}",
         None,
         '{\n  "members": [\n    {"id": 1, "name": "김회원", "phone": "010-1234-5678"}\n  ]\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "데이터를 불러올 수 없습니다"\n}',
         "Y","SCR-002","FN-002","개발완료"),
        ("API-005","미납 회원 조회","GET","/api/v1/dashboard/unpaid-members","미납 내역이 있는 회원 목록 조회",
         "Authorization: Bearer {token}",
         None,
         '{\n  "members": [\n    {"id": 2, "name": "이회원", "unpaidAmount": 150000}\n  ]\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "조회 실패"\n}',
         "Y","SCR-002","FN-002","개발완료"),
        ("API-006","만료예정 회원 조회","GET","/api/v1/dashboard/expiring-members","7일 내 이용권 만료 예정 회원 조회",
         "Authorization: Bearer {token}",
         None,
         '{\n  "members": [\n    {"id": 3, "name": "박회원", "expiryDate": "2026-03-18", "dDay": -7}\n  ]\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "조회 실패"\n}',
         "Y","SCR-002","FN-002","개발완료"),
        ("API-007","회원 목록 조회","GET","/api/v1/members","회원 목록 조회 (검색/필터/페이지네이션)",
         "Authorization: Bearer {token}",
         "Query: ?search=&status=&gender=&manager_id=&page=1&size=20&sort=name&order=asc",
         '{\n  "data": [\n    {"id": 1, "name": "김회원", "phone": "010-1234-5678", "status": "active", "gender": "M"}\n  ],\n  "pagination": {"page": 1, "size": 20, "total": 450}\n}',
         '{\n  "error": "FORBIDDEN",\n  "message": "접근 권한이 없습니다"\n}',
         "Y","SCR-010","FN-003","개발완료"),
        ("API-008","회원 상세 조회","GET","/api/v1/members/:id","회원 상세 정보 조회 (프로필+이용권+출석+결제+체성분+메모)",
         "Authorization: Bearer {token}",
         None,
         '{\n  "profile": {"id": 1, "name": "김회원", ...},\n  "tickets": [...],\n  "attendance": [...],\n  "payments": [...],\n  "bodyComposition": [...],\n  "memos": [...]\n}',
         '{\n  "error": "NOT_FOUND",\n  "message": "회원을 찾을 수 없습니다"\n}',
         "Y","SCR-011","FN-006","개발완료"),
        ("API-009","회원 등록","POST","/api/v1/members","신규 회원 등록",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "name": "string",\n  "gender": "M|F",\n  "birthDate": "YYYY-MM-DD",\n  "phone": "string",\n  "email": "string",\n  "address": "string",\n  "memo": "string",\n  "managerId": "number"\n}',
         '{\n  "id": 1,\n  "name": "김회원",\n  "status": "active",\n  "createdAt": "2026-03-11T00:00:00Z"\n}',
         '{\n  "error": "DUPLICATE_PHONE",\n  "message": "이미 등록된 연락처입니다"\n}',
         "Y","SCR-012","FN-004","개발완료"),
        ("API-010","회원 수정","PUT","/api/v1/members/:id","기존 회원 정보 수정",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "name": "string",\n  "phone": "string",\n  "email": "string",\n  "address": "string",\n  "memo": "string",\n  "managerId": "number"\n}',
         '{\n  "id": 1,\n  "name": "김회원(수정)",\n  "updatedAt": "2026-03-11T00:00:00Z"\n}',
         '{\n  "error": "NOT_FOUND",\n  "message": "회원을 찾을 수 없습니다"\n}',
         "Y","SCR-013","FN-005","개발완료"),
        ("API-011","회원 삭제","DELETE","/api/v1/members/:id","회원 삭제 (소프트 삭제)",
         "Authorization: Bearer {token}",
         None,
         '{\n  "message": "회원이 삭제되었습니다"\n}',
         '{\n  "error": "FORBIDDEN",\n  "message": "삭제 권한이 없습니다"\n}',
         "Y","SCR-011","FN-005","개발완료"),
        ("API-012","회원 통계 요약","GET","/api/v1/members/stats/summary","회원 상태별 통계 요약",
         "Authorization: Bearer {token}",
         None,
         '{\n  "total": 450,\n  "active": 320,\n  "expired": 105,\n  "suspended": 15,\n  "withdrawn": 10\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "통계 조회 실패"\n}',
         "Y","SCR-010","FN-003","개발완료"),
        ("API-013","출석 기록 조회","GET","/api/v1/attendance","회원 출석 기록 조회",
         "Authorization: Bearer {token}",
         "Query: ?startDate=&endDate=&type=&page=1&size=50",
         '{\n  "data": [\n    {"id": 1, "memberId": 1, "memberName": "김회원", "type": "normal", "inTime": "2026-03-11T09:00:00"}\n  ],\n  "pagination": {...}\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "조회 실패"\n}',
         "Y","SCR-020","FN-008","개발완료"),
        ("API-014","출석 등록","POST","/api/v1/attendance","출석 체크인/체크아웃 등록",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "memberId": "number",\n  "type": "normal|pt|gx",\n  "status": "checkIn|checkOut"\n}',
         '{\n  "id": 1,\n  "memberId": 1,\n  "inTime": "2026-03-11T09:00:00",\n  "visitCount": 45\n}',
         '{\n  "error": "ALREADY_CHECKED_IN",\n  "message": "이미 출석 처리되었습니다"\n}',
         "Y","SCR-020","FN-008","개발완료"),
        ("API-015","매출 조회","GET","/api/v1/sales","매출 내역 조회 (기간/유형/상태 필터)",
         "Authorization: Bearer {token}",
         "Query: ?startDate=&endDate=&type=&status=&page=1&size=20",
         '{\n  "data": [\n    {"id": 1, "date": "2026-03-11", "productName": "3개월 이용권", "amount": 300000, "paymentMethod": "card"}\n  ],\n  "pagination": {...}\n}',
         '{\n  "error": "FORBIDDEN",\n  "message": "매출 조회 권한이 없습니다"\n}',
         "Y","SCR-030","FN-010","개발완료"),
        ("API-016","매출 등록 (결제)","POST","/api/v1/sales","POS 결제 처리 및 매출 등록",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "memberId": "number",\n  "items": [{"productId": 1, "quantity": 1}],\n  "paymentMethod": "cash|card|mileage|mixed",\n  "cashAmount": 0,\n  "cardAmount": 300000,\n  "mileageAmount": 0\n}',
         '{\n  "saleId": 1,\n  "totalAmount": 300000,\n  "receiptNo": "R20260311001"\n}',
         '{\n  "error": "PAYMENT_FAILED",\n  "message": "결제에 실패했습니다"\n}',
         "Y","SCR-032","FN-012","개발완료"),
        ("API-017","매출 통계 요약","GET","/api/v1/sales/stats/summary","매출 통계 요약 (총매출/현금/카드/마일리지)",
         "Authorization: Bearer {token}",
         "Query: ?startDate=&endDate=",
         '{\n  "totalSales": 15000000,\n  "cashSales": 3000000,\n  "cardSales": 11000000,\n  "mileageSales": 1000000,\n  "refundTotal": 500000\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "통계 조회 실패"\n}',
         "Y","SCR-030","FN-010","개발완료"),
        ("API-018","상품 목록 조회","GET","/api/v1/products","상품 목록 조회 (카테고리/상태 필터)",
         "Authorization: Bearer {token}",
         "Query: ?category=&status=",
         '{\n  "data": [\n    {"id": 1, "name": "3개월 이용권", "category": "membership", "cashPrice": 280000, "cardPrice": 300000, "status": "active"}\n  ]\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "조회 실패"\n}',
         "Y","SCR-040","FN-013","개발완료"),
        ("API-019","상품 등록","POST","/api/v1/products","신규 상품 등록",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "category": "membership|pt|gx|etc",\n  "name": "string",\n  "cashPrice": "number",\n  "cardPrice": "number",\n  "period": "number (days)",\n  "kioskExposure": "boolean"\n}',
         '{\n  "id": 1,\n  "name": "3개월 이용권",\n  "createdAt": "2026-03-11"\n}',
         '{\n  "error": "DUPLICATE_NAME",\n  "message": "동일 카테고리에 같은 상품명이 있습니다"\n}',
         "Y","SCR-041","FN-014","개발완료"),
        ("API-020","락커 목록 조회","GET","/api/v1/lockers","구역별 락커 현황 조회",
         "Authorization: Bearer {token}",
         "Query: ?area=A",
         '{\n  "data": [\n    {"id": 1, "number": "A-001", "area": "A", "status": "available", "memberId": null}\n  ]\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "조회 실패"\n}',
         "Y","SCR-050","FN-015","개발완료"),
        ("API-021","락커 배정/해제","PATCH","/api/v1/lockers/:id","락커 배정 또는 해제",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "action": "assign|release",\n  "memberId": "number",\n  "expiryDate": "YYYY-MM-DD"\n}',
         '{\n  "id": 1,\n  "number": "A-001",\n  "status": "occupied",\n  "memberId": 1\n}',
         '{\n  "error": "ALREADY_ASSIGNED",\n  "message": "이미 사용중인 락커입니다"\n}',
         "Y","SCR-051","FN-016","개발완료"),
        ("API-022","직원 목록 조회","GET","/api/v1/staff","직원 목록 조회 (역할/상태 필터)",
         "Authorization: Bearer {token}",
         "Query: ?search=&role=&status=",
         '{\n  "data": [\n    {"id": 1, "name": "박트레이너", "role": "fc", "status": "active", "contact": "010-9999-8888"}\n  ]\n}',
         '{\n  "error": "FORBIDDEN",\n  "message": "직원 조회 권한이 없습니다"\n}',
         "Y","SCR-060","FN-019","개발완료"),
        ("API-023","직원 등록","POST","/api/v1/staff","신규 직원 등록",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "name": "string",\n  "gender": "M|F",\n  "contact": "string",\n  "role": "center_manager|manager|fc|staff",\n  "joinDate": "YYYY-MM-DD",\n  "workType": "full|part"\n}',
         '{\n  "id": 1,\n  "name": "박트레이너",\n  "adminId": "auto_generated_id"\n}',
         '{\n  "error": "DUPLICATE_EMAIL",\n  "message": "이미 등록된 이메일입니다"\n}',
         "Y","SCR-061","FN-020","개발완료"),
        ("API-024","직원 수정","PUT","/api/v1/staff/:id","직원 정보 수정",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "name": "string",\n  "contact": "string",\n  "role": "string",\n  "status": "active|leave|resigned"\n}',
         '{\n  "id": 1,\n  "name": "박트레이너(수정)",\n  "updatedAt": "2026-03-11"\n}',
         '{\n  "error": "NOT_FOUND",\n  "message": "직원을 찾을 수 없습니다"\n}',
         "Y","SCR-060","FN-020","개발완료"),
        ("API-025","직원 삭제","DELETE","/api/v1/staff/:id","직원 삭제 (소프트 삭제)",
         "Authorization: Bearer {token}",
         None,
         '{\n  "message": "직원이 삭제되었습니다"\n}',
         '{\n  "error": "FORBIDDEN",\n  "message": "삭제 권한이 없습니다"\n}',
         "Y","SCR-060","FN-020","개발완료"),
        ("API-026","직원 출근부 조회","GET","/api/v1/staff/attendance","직원 출퇴근 기록 조회",
         "Authorization: Bearer {token}",
         "Query: ?staffId=&yearMonth=2026-03",
         '{\n  "data": [\n    {"date": "2026-03-11", "staffId": 1, "checkIn": "09:00", "checkOut": "18:00", "status": "present"}\n  ]\n}',
         '{\n  "error": "SERVER_ERROR",\n  "message": "조회 실패"\n}',
         "Y","SCR-062","FN-021","개발완료"),
        ("API-027","시스템 초기화","POST","/api/v1/system/reset","지점 데이터 초기화 (개발/테스트용)",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{\n  "branchId": "number",\n  "confirmPassword": "string",\n  "targets": ["members","sales","attendance"]\n}',
         '{\n  "message": "데이터가 초기화되었습니다",\n  "resetTargets": ["members","sales","attendance"]\n}',
         '{\n  "error": "INVALID_PASSWORD",\n  "message": "비밀번호가 올바르지 않습니다"\n}',
         "Y","SCR-094","FN-033","개발완료"),
    ]

    for ri, api in enumerate(apis):
        row = ri + 2
        api_id, api_name, method, endpoint, desc, req_headers, req_body, res_ok, res_fail, auth, scr_id, fn_id, status = api
        ws.cell(row, 1, api_id)
        ws.cell(row, 2, api_name)
        ws.cell(row, 3, method)
        ws.cell(row, 4, endpoint)
        ws.cell(row, 5, desc)
        ws.cell(row, 6, req_headers)
        ws.cell(row, 7, req_body)
        ws.cell(row, 8, res_ok)
        ws.cell(row, 9, res_fail)
        ws.cell(row, 10, auth)
        ws.cell(row, 11, scr_id)
        ws.cell(row, 12, fn_id)
        ws.cell(row, 13, status)
        ws.cell(row, 14, None)  # 담당자
        # [자동] 수식
        ws.cell(row, 15, f"=IFERROR(VLOOKUP(K{row},'IA(정보구조)'!D:E,2,FALSE()),\"\")")
        ws.cell(row, 16, f'=IFERROR(VLOOKUP(L{row},기능명세서!A:B,2,FALSE()),"")')
        ws.cell(row, 17, None)  # 비고
        ws.cell(row, 18, f"=IFERROR(_xlfn.TEXTJOIN(\", \",TRUE(),IF('UI 요소 상세'!H$1:H$500=A{row},'UI 요소 상세'!C$1:C$500,\"\")),\"\")")

        for ci in range(1, 19):
            c = ws.cell(row, ci)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # 드롭다운
    add_dv(ws, "C", "GET,POST,PUT,PATCH,DELETE")
    add_dv(ws, "J", "Y,N")
    add_dv(ws, "K", SCR_IDS_STR)
    add_dv(ws, "L", FN_IDS_STR)
    add_dv(ws, "M", "설계중,개발중,개발완료,테스트중,배포완료")
    add_dv(ws, "N", 담당자_LIST)

    ws.auto_filter.ref = "A1:Q1"
    return ws


# ============================================================
# 7. 데이터 정의서 시트
# ============================================================
def create_data_dict_sheet(wb):
    ws = wb.create_sheet("데이터 정의서")

    headers = ["테이블명","컬럼명","데이터 타입","길이","NULL 허용","기본값","PK","FK","인덱스","설명","관련 API ID","상태","비고"]
    col_widths = {"A":16,"B":18,"C":14,"D":8,"E":10,"F":12,"G":6,"H":16,"I":10,"J":30,"K":12,"L":10,"M":20}
    apply_header(ws, headers, col_widths)

    tables = [
        # tenants
        ("tenants","id","BIGINT",None,"N",None,"Y",None,"Y","테넌트 고유 ID (auto increment)","API-003","확정",None),
        ("tenants","name","VARCHAR",100,"N",None,"N",None,"N","테넌트(본사) 이름",None,"확정",None),
        ("tenants","code","VARCHAR",20,"N",None,"N",None,"UNIQUE","테넌트 코드 (고유)",None,"확정",None),
        ("tenants","plan","ENUM",None,"N","basic","N",None,"N","요금제 (basic/standard/premium)",None,"확정",None),
        ("tenants","status","ENUM",None,"N","active","N",None,"Y","상태 (active/inactive/suspended)",None,"확정",None),
        ("tenants","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        ("tenants","updated_at","DATETIME",None,"Y",None,"N",None,"N","수정일시",None,"확정",None),
        # branches
        ("branches","id","BIGINT",None,"N",None,"Y",None,"Y","지점 고유 ID","API-027","확정",None),
        ("branches","tenant_id","BIGINT",None,"N",None,"N","tenants.id","Y","소속 테넌트 FK",None,"확정",None),
        ("branches","name","VARCHAR",100,"N",None,"N",None,"N","지점명",None,"확정",None),
        ("branches","code","VARCHAR",20,"N",None,"N",None,"UNIQUE","지점 코드 (고유)",None,"확정",None),
        ("branches","address","VARCHAR",255,"Y",None,"N",None,"N","지점 주소",None,"확정",None),
        ("branches","phone","VARCHAR",20,"Y",None,"N",None,"N","지점 연락처",None,"확정",None),
        ("branches","status","ENUM",None,"N","active","N",None,"Y","상태 (active/inactive)",None,"확정",None),
        ("branches","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # users
        ("users","id","BIGINT",None,"N",None,"Y",None,"Y","사용자 고유 ID","API-001","확정",None),
        ("users","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","소속 지점 FK",None,"확정",None),
        ("users","email","VARCHAR",100,"N",None,"N",None,"UNIQUE","로그인 이메일",None,"확정",None),
        ("users","password_hash","VARCHAR",255,"N",None,"N",None,"N","비밀번호 해시",None,"확정",None),
        ("users","name","VARCHAR",50,"N",None,"N",None,"Y","사용자 이름",None,"확정",None),
        ("users","role","ENUM",None,"N",None,"N",None,"Y","역할 (super_admin/center_manager/manager/fc/staff/viewer)",None,"확정",None),
        ("users","status","ENUM",None,"N","active","N",None,"Y","상태 (active/inactive/locked)",None,"확정",None),
        ("users","last_login","DATETIME",None,"Y",None,"N",None,"N","마지막 로그인 일시",None,"확정",None),
        ("users","login_fail_count","INT",None,"N","0","N",None,"N","연속 로그인 실패 횟수",None,"확정",None),
        ("users","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # members
        ("members","id","BIGINT",None,"N",None,"Y",None,"Y","회원 고유 ID","API-007","확정",None),
        ("members","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","소속 지점 FK (데이터 격리)",None,"확정",None),
        ("members","name","VARCHAR",50,"N",None,"N",None,"Y","회원 이름",None,"확정",None),
        ("members","gender","ENUM",None,"N",None,"N",None,"Y","성별 (M/F)",None,"확정",None),
        ("members","birth_date","VARCHAR",10,"Y",None,"N",None,"N","생년월일 (YYYY-MM-DD)",None,"확정",None),
        ("members","phone","VARCHAR",20,"N",None,"N",None,"Y","연락처",None,"확정","지점 내 중복 불가"),
        ("members","email","VARCHAR",100,"Y",None,"N",None,"N","이메일",None,"확정",None),
        ("members","address","VARCHAR",255,"Y",None,"N",None,"N","주소",None,"확정",None),
        ("members","status","ENUM",None,"N","active","N",None,"Y","상태 (active/expired/suspended/withdrawn)",None,"확정",None),
        ("members","manager_id","BIGINT",None,"Y",None,"N","users.id","Y","담당 FC ID",None,"확정",None),
        ("members","attendance_no","VARCHAR",20,"Y",None,"N",None,"UNIQUE","출석 번호",None,"확정",None),
        ("members","company","VARCHAR",100,"Y",None,"N",None,"N","직장명",None,"확정",None),
        ("members","memo","TEXT",None,"Y",None,"N",None,"N","관리자 메모",None,"확정",None),
        ("members","first_reg_date","DATETIME",None,"N","NOW()","N",None,"N","최초 등록일",None,"확정",None),
        ("members","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        ("members","updated_at","DATETIME",None,"Y",None,"N",None,"N","수정일시",None,"확정",None),
        # member_tickets
        ("member_tickets","id","BIGINT",None,"N",None,"Y",None,"Y","이용권 고유 ID","API-008","확정",None),
        ("member_tickets","member_id","BIGINT",None,"N",None,"N","members.id","Y","회원 FK",None,"확정",None),
        ("member_tickets","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("member_tickets","product_id","BIGINT",None,"N",None,"N","products.id","Y","상품 FK",None,"확정",None),
        ("member_tickets","type","ENUM",None,"N",None,"N",None,"Y","유형 (period/count/unlimited)",None,"확정",None),
        ("member_tickets","start_date","DATETIME",None,"N",None,"N",None,"Y","시작일",None,"확정",None),
        ("member_tickets","end_date","DATETIME",None,"Y",None,"N",None,"Y","종료일",None,"확정",None),
        ("member_tickets","remaining_count","INT",None,"Y",None,"N",None,"N","잔여 횟수 (횟수제)",None,"확정",None),
        ("member_tickets","status","ENUM",None,"N","active","N",None,"Y","상태 (active/expired/suspended/cancelled)",None,"확정",None),
        ("member_tickets","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # attendance
        ("attendance","id","BIGINT",None,"N",None,"Y",None,"Y","출석 고유 ID","API-013","확정",None),
        ("attendance","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("attendance","member_id","BIGINT",None,"N",None,"N","members.id","Y","회원 FK",None,"확정",None),
        ("attendance","type","ENUM",None,"N","normal","N",None,"Y","출석 유형 (normal/pt/gx)",None,"확정",None),
        ("attendance","status","ENUM",None,"N","checkIn","N",None,"Y","상태 (checkIn/checkOut)",None,"확정",None),
        ("attendance","category","VARCHAR",50,"Y",None,"N",None,"N","카테고리",None,"확정",None),
        ("attendance","door","VARCHAR",50,"Y",None,"N",None,"N","출입문 정보",None,"확정",None),
        ("attendance","in_time","DATETIME",None,"N",None,"N",None,"Y","입장 시간",None,"확정",None),
        ("attendance","out_time","DATETIME",None,"Y",None,"N",None,"N","퇴장 시간",None,"확정",None),
        ("attendance","visit_count","INT",None,"N","1","N",None,"N","누적 방문 횟수",None,"확정",None),
        ("attendance","pass_info","VARCHAR",50,"Y",None,"N",None,"N","출입 수단 정보 (RFID/앱 등)",None,"확정",None),
        ("attendance","is_other_branch","BOOLEAN",None,"N","false","N",None,"Y","타지점 출석 여부",None,"확정",None),
        ("attendance","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # sales
        ("sales","id","BIGINT",None,"N",None,"Y",None,"Y","매출 고유 ID","API-015","확정",None),
        ("sales","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("sales","no","VARCHAR",30,"N",None,"N",None,"UNIQUE","매출 번호 (R + 날짜 + 순번)",None,"확정",None),
        ("sales","purchase_date","DATETIME",None,"N",None,"N",None,"Y","결제 일시",None,"확정",None),
        ("sales","type","VARCHAR",50,"N",None,"N",None,"Y","매출 유형 (membership/pt/gx/product/etc)",None,"확정",None),
        ("sales","product_name","VARCHAR",100,"N",None,"N",None,"N","상품명",None,"확정",None),
        ("sales","manager_id","BIGINT",None,"Y",None,"N","users.id","Y","담당자 FK",None,"확정",None),
        ("sales","buyer_id","BIGINT",None,"Y",None,"N","members.id","Y","구매 회원 FK",None,"확정",None),
        ("sales","quantity","INT",None,"N","1","N",None,"N","수량",None,"확정",None),
        ("sales","original_price","INT",None,"N","0","N",None,"N","원가",None,"확정",None),
        ("sales","sale_price","INT",None,"N","0","N",None,"N","판매가",None,"확정",None),
        ("sales","discount_price","INT",None,"N","0","N",None,"N","할인 금액",None,"확정",None),
        ("sales","payment_method","VARCHAR",20,"N",None,"N",None,"Y","결제수단 (cash/card/mileage/mixed)",None,"확정",None),
        ("sales","cash","INT",None,"N","0","N",None,"N","현금 결제액",None,"확정",None),
        ("sales","card","INT",None,"N","0","N",None,"N","카드 결제액",None,"확정",None),
        ("sales","mileage","INT",None,"N","0","N",None,"N","마일리지 결제액",None,"확정",None),
        ("sales","card_company","VARCHAR",50,"Y",None,"N",None,"N","카드사",None,"확정",None),
        ("sales","card_number","VARCHAR",20,"Y",None,"N",None,"N","카드번호 (마스킹)",None,"확정",None),
        ("sales","approval_no","VARCHAR",30,"Y",None,"N",None,"N","승인번호",None,"확정",None),
        ("sales","unpaid","INT",None,"N","0","N",None,"N","미납 금액",None,"확정",None),
        ("sales","status","ENUM",None,"N","completed","N",None,"Y","상태 (completed/refunded/partial_refund/unpaid)",None,"확정",None),
        ("sales","category","VARCHAR",50,"Y",None,"N",None,"N","매출 카테고리",None,"확정",None),
        ("sales","memo","TEXT",None,"Y",None,"N",None,"N","메모",None,"확정",None),
        ("sales","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # products
        ("products","id","BIGINT",None,"N",None,"Y",None,"Y","상품 고유 ID","API-018","확정",None),
        ("products","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("products","category","VARCHAR",50,"N",None,"N",None,"Y","대분류 카테고리",None,"확정",None),
        ("products","category_key","ENUM",None,"N",None,"N",None,"Y","카테고리 키 (membership/pt/gx/etc)",None,"확정",None),
        ("products","sub_category","VARCHAR",50,"Y",None,"N",None,"N","소분류",None,"확정",None),
        ("products","name","VARCHAR",100,"N",None,"N",None,"Y","상품명",None,"확정",None),
        ("products","cash_price","INT",None,"N","0","N",None,"N","현금가",None,"확정",None),
        ("products","card_price","INT",None,"N","0","N",None,"N","카드가",None,"확정",None),
        ("products","period","INT",None,"Y",None,"N",None,"N","이용 기간 (일수)",None,"확정",None),
        ("products","kiosk_exposure","BOOLEAN",None,"N","true","N",None,"N","키오스크 노출 여부",None,"확정",None),
        ("products","status","ENUM",None,"N","active","N",None,"Y","상태 (active/inactive)",None,"확정",None),
        ("products","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        ("products","updated_at","DATETIME",None,"Y",None,"N",None,"N","수정일시",None,"확정",None),
        # staff
        ("staff","id","BIGINT",None,"N",None,"Y",None,"Y","직원 고유 ID","API-022","확정",None),
        ("staff","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("staff","name","VARCHAR",50,"N",None,"N",None,"Y","직원 이름",None,"확정",None),
        ("staff","gender","ENUM",None,"Y",None,"N",None,"N","성별 (M/F)",None,"확정",None),
        ("staff","contact","VARCHAR",20,"Y",None,"N",None,"N","연락처",None,"확정",None),
        ("staff","role","ENUM",None,"N",None,"N",None,"Y","역할 (center_manager/manager/fc/staff)",None,"확정",None),
        ("staff","job_group","VARCHAR",50,"Y",None,"N",None,"N","직무 그룹",None,"확정",None),
        ("staff","position","VARCHAR",50,"Y",None,"N",None,"N","직위",None,"확정",None),
        ("staff","team","VARCHAR",50,"Y",None,"N",None,"N","소속 팀",None,"확정",None),
        ("staff","join_date","DATETIME",None,"N",None,"N",None,"Y","입사일",None,"확정",None),
        ("staff","admin_id","VARCHAR",50,"Y",None,"N",None,"UNIQUE","관리자 계정 ID",None,"확정",None),
        ("staff","memo","TEXT",None,"Y",None,"N",None,"N","메모",None,"확정",None),
        ("staff","work_type","ENUM",None,"N","full","N",None,"N","근무 유형 (full/part)",None,"확정",None),
        ("staff","status","ENUM",None,"N","active","N",None,"Y","상태 (active/leave/resigned)",None,"확정",None),
        ("staff","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # staff_attendance
        ("staff_attendance","id","BIGINT",None,"N",None,"Y",None,"Y","직원 출근 고유 ID","API-026","확정",None),
        ("staff_attendance","staff_id","BIGINT",None,"N",None,"N","staff.id","Y","직원 FK",None,"확정",None),
        ("staff_attendance","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("staff_attendance","date","VARCHAR",10,"N",None,"N",None,"Y","날짜 (YYYY-MM-DD)",None,"확정",None),
        ("staff_attendance","status","ENUM",None,"N","present","N",None,"Y","상태 (present/absent/late/half_day/vacation)",None,"확정",None),
        ("staff_attendance","check_in","VARCHAR",5,"Y",None,"N",None,"N","출근 시간 (HH:MM)",None,"확정",None),
        ("staff_attendance","check_out","VARCHAR",5,"Y",None,"N",None,"N","퇴근 시간 (HH:MM)",None,"확정",None),
        ("staff_attendance","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # lockers
        ("lockers","id","BIGINT",None,"N",None,"Y",None,"Y","락커 고유 ID","API-020","확정",None),
        ("lockers","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("lockers","number","VARCHAR",10,"N",None,"N",None,"Y","락커 번호 (A-001 등)",None,"확정",None),
        ("lockers","type","ENUM",None,"N","standard","N",None,"N","유형 (standard/large/premium)",None,"확정",None),
        ("lockers","area","VARCHAR",10,"N",None,"N",None,"Y","구역 (A/B/C...)",None,"확정",None),
        ("lockers","status","ENUM",None,"N","available","N",None,"Y","상태 (available/occupied/expiring/broken)",None,"확정",None),
        ("lockers","member_id","BIGINT",None,"Y",None,"N","members.id","Y","배정 회원 FK",None,"확정",None),
        ("lockers","expiry_date","DATETIME",None,"Y",None,"N",None,"Y","만료일",None,"확정",None),
        ("lockers","gender","ENUM",None,"Y",None,"N",None,"N","성별 구분 (M/F/공용)",None,"확정",None),
        ("lockers","last_updated","DATETIME",None,"Y",None,"N",None,"N","최근 변경일시",None,"확정",None),
        ("lockers","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # payroll
        ("payroll","id","BIGINT",None,"N",None,"Y",None,"Y","급여 고유 ID",None,"확정",None),
        ("payroll","staff_id","BIGINT",None,"N",None,"N","staff.id","Y","직원 FK",None,"확정",None),
        ("payroll","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("payroll","year_month","VARCHAR",7,"N",None,"N",None,"Y","급여 연월 (YYYY-MM)",None,"확정",None),
        ("payroll","base_salary","INT",None,"N","0","N",None,"N","기본급",None,"확정",None),
        ("payroll","incentive","INT",None,"N","0","N",None,"N","인센티브",None,"확정",None),
        ("payroll","deduction","INT",None,"N","0","N",None,"N","공제액",None,"확정",None),
        ("payroll","net_salary","INT",None,"N","0","N",None,"N","실지급액",None,"확정",None),
        ("payroll","status","ENUM",None,"N","pending","N",None,"Y","상태 (pending/paid/cancelled)",None,"확정",None),
        ("payroll","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # messages
        ("messages","id","BIGINT",None,"N",None,"Y",None,"Y","메시지 고유 ID",None,"확정",None),
        ("messages","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("messages","sender_id","BIGINT",None,"N",None,"N","users.id","Y","발송자 FK",None,"확정",None),
        ("messages","type","ENUM",None,"N",None,"N",None,"Y","유형 (sms/lms/kakao)",None,"확정",None),
        ("messages","title","VARCHAR",100,"Y",None,"N",None,"N","메시지 제목",None,"확정",None),
        ("messages","content","TEXT",None,"N",None,"N",None,"N","메시지 본문",None,"확정",None),
        ("messages","recipients","JSON",None,"N",None,"N",None,"N","수신자 목록 (JSON)",None,"확정",None),
        ("messages","sent_at","DATETIME",None,"Y",None,"N",None,"Y","발송 일시",None,"확정",None),
        ("messages","status","ENUM",None,"N","pending","N",None,"Y","상태 (pending/sent/failed/cancelled)",None,"확정",None),
        ("messages","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # coupons
        ("coupons","id","BIGINT",None,"N",None,"Y",None,"Y","쿠폰 고유 ID",None,"확정",None),
        ("coupons","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("coupons","name","VARCHAR",100,"N",None,"N",None,"N","쿠폰명",None,"확정",None),
        ("coupons","type","ENUM",None,"N",None,"N",None,"Y","유형 (rate/amount)",None,"확정",None),
        ("coupons","discount_value","INT",None,"N",None,"N",None,"N","할인값 (% 또는 원)",None,"확정",None),
        ("coupons","start_date","DATETIME",None,"N",None,"N",None,"Y","시작일",None,"확정",None),
        ("coupons","end_date","DATETIME",None,"N",None,"N",None,"Y","종료일",None,"확정",None),
        ("coupons","usage_limit","INT",None,"Y",None,"N",None,"N","사용 한도",None,"확정",None),
        ("coupons","used_count","INT",None,"N","0","N",None,"N","사용 횟수",None,"확정",None),
        ("coupons","status","ENUM",None,"N","active","N",None,"Y","상태 (active/expired/depleted)",None,"확정",None),
        ("coupons","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # contracts
        ("contracts","id","BIGINT",None,"N",None,"Y",None,"Y","계약 고유 ID",None,"확정",None),
        ("contracts","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("contracts","member_id","BIGINT",None,"N",None,"N","members.id","Y","회원 FK",None,"확정",None),
        ("contracts","product_id","BIGINT",None,"N",None,"N","products.id","Y","상품 FK",None,"확정",None),
        ("contracts","start_date","DATETIME",None,"N",None,"N",None,"Y","계약 시작일",None,"확정",None),
        ("contracts","end_date","DATETIME",None,"N",None,"N",None,"Y","계약 종료일",None,"확정",None),
        ("contracts","total_amount","INT",None,"N","0","N",None,"N","총 계약 금액",None,"확정",None),
        ("contracts","paid_amount","INT",None,"N","0","N",None,"N","납부 금액",None,"확정",None),
        ("contracts","status","ENUM",None,"N","active","N",None,"Y","상태 (active/completed/cancelled)",None,"확정",None),
        ("contracts","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
        # mileage
        ("mileage","id","BIGINT",None,"N",None,"Y",None,"Y","마일리지 고유 ID",None,"확정",None),
        ("mileage","branch_id","BIGINT",None,"N",None,"N","branches.id","Y","지점 FK",None,"확정",None),
        ("mileage","member_id","BIGINT",None,"N",None,"N","members.id","Y","회원 FK",None,"확정",None),
        ("mileage","type","ENUM",None,"N",None,"N",None,"Y","유형 (earn/use/cancel)",None,"확정",None),
        ("mileage","amount","INT",None,"N","0","N",None,"N","변동 금액",None,"확정",None),
        ("mileage","balance","INT",None,"N","0","N",None,"N","변동 후 잔액",None,"확정",None),
        ("mileage","description","VARCHAR",200,"Y",None,"N",None,"N","변동 사유",None,"확정",None),
        ("mileage","created_at","DATETIME",None,"N","NOW()","N",None,"N","생성일시",None,"확정",None),
    ]

    write_rows(ws, tables)

    # 드롭다운
    add_dv(ws, "C", "VARCHAR,INT,BIGINT,BOOLEAN,TEXT,JSON,DATETIME,FLOAT,DECIMAL,ENUM,BLOB")
    add_dv(ws, "E", "Y,N")
    add_dv(ws, "G", "Y,N")
    add_dv(ws, "I", "Y,N,UNIQUE")
    add_dv(ws, "L", "작성중,검토중,확정,변경필요,삭제")

    ws.auto_filter.ref = "A1:M1"
    return ws


# ============================================================
# 8. QA테스트케이스 시트
# ============================================================
def create_qa_sheet(wb):
    ws = wb.create_sheet("QA테스트케이스")

    headers = ["TC-ID","플랫폼","테스트 영역","하위 영역","테스트 유형","테스트 시나리오",
               "사전 조건","테스트 절차","기대 결과","실제 결과","테스트 상태","우선순위",
               "심각도","담당자","테스트 환경","브라우저/앱","디바이스","OS","테스트 사이클",
               "자동화 여부","화면 ID","기능 ID","API ID","버그 티켓","스크린샷",
               "[자동] 화면명","[자동] 기능명","[자동] API 엔드포인트","비고"]
    col_widths = {"A":12,"B":14,"C":12,"D":12,"E":12,"F":30,"G":25,"H":40,"I":30,"J":14,
                  "K":10,"L":10,"M":10,"N":12,"O":10,"P":14,"Q":12,"R":12,"S":10,"T":12,
                  "U":10,"V":10,"W":10,"X":14,"Y":18,"Z":16,"AA":16,"AB":20,"AC":14}
    apply_header(ws, headers, col_widths)

    tcs = [
        # 인증
        ("TC-001","Web","인증","로그인","기능 테스트","정상 로그인 - 올바른 이메일/비밀번호로 로그인",
         "등록된 계정 존재, 활성 상태",
         "1) 로그인 페이지 접속\n2) 이메일 입력\n3) 비밀번호 입력\n4) 지점 선택\n5) 로그인 버튼 클릭",
         "대시보드(SCR-002)로 이동, JWT 토큰 저장, 사용자 정보 표시",
         None,"미실행","Critical","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-001","FN-001","API-001",None,None),
        ("TC-002","Web","인증","로그인","기능 테스트","로그인 실패 - 잘못된 비밀번호",
         "등록된 계정 존재",
         "1) 로그인 페이지 접속\n2) 올바른 이메일 입력\n3) 잘못된 비밀번호 입력\n4) 로그인 버튼 클릭",
         "'비밀번호가 올바르지 않습니다' 에러 메시지 표시, 로그인 실패 횟수 +1",
         None,"미실행","Critical","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-001","FN-001","API-001",None,None),
        ("TC-003","Web","인증","로그인","기능 테스트","계정 잠금 - 5회 연속 실패",
         "등록된 계정 존재, 실패 횟수 4회",
         "1) 잘못된 비밀번호로 5번째 로그인 시도",
         "'5회 실패로 계정이 잠겼습니다. 30분 후 다시 시도하세요' 메시지 표시",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-001","FN-001","API-001",None,None),
        ("TC-004","Web","인증","로그인","UI/UX 테스트","비밀번호 마스킹 토글",
         "로그인 페이지 접속",
         "1) 비밀번호 입력\n2) 눈 아이콘 클릭\n3) 비밀번호 표시 확인\n4) 다시 클릭하여 마스킹",
         "토글에 따라 비밀번호 표시/마스킹 전환",
         None,"미실행","Medium","Minor","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-001","FN-001",None,None,None),
        # 대시보드
        ("TC-005","Web","대시보드","통계","기능 테스트","대시보드 KPI 카드 정확성",
         "로그인 완료, 테스트 데이터 존재",
         "1) 대시보드 접속\n2) 총회원/활성/만료임박/만료 카드 수치 확인\n3) DB 데이터와 대조",
         "카드 수치가 DB 실제 데이터와 일치",
         None,"미실행","Critical","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-002","FN-002","API-003",None,None),
        ("TC-006","Web","대시보드","통계","기능 테스트","생일 회원 리스트 표시",
         "오늘 생일인 회원 데이터 존재",
         "1) 대시보드 접속\n2) 생일 회원 섹션 확인",
         "오늘 생일인 회원 목록 표시, 이름/연락처 표시",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-002","FN-002","API-004",None,None),
        # 회원관리
        ("TC-007","Web","회원관리","목록","기능 테스트","회원 검색 - 이름으로 검색",
         "회원 데이터 존재",
         "1) 회원 목록 페이지 접속\n2) 검색바에 '김' 입력\n3) 300ms 대기",
         "'김'이 포함된 회원만 필터링되어 표시",
         None,"미실행","Critical","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-010","FN-003","API-007",None,None),
        ("TC-008","Web","회원관리","목록","기능 테스트","회원 필터 - 상태 필터 적용",
         "다양한 상태의 회원 데이터 존재",
         "1) 상태 필터에서 '만료' 선택\n2) 목록 갱신 확인",
         "만료 상태 회원만 표시, 건수 일치",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-010","FN-003","API-007",None,None),
        ("TC-009","Web","회원관리","등록","기능 테스트","신규 회원 정상 등록",
         "로그인 완료, 매니저 이상 권한",
         "1) '회원등록' 버튼 클릭\n2) 이름 입력: 테스트회원\n3) 성별: 남 선택\n4) 연락처: 010-1111-2222\n5) '저장' 클릭",
         "회원 등록 성공 메시지, 회원 상세 페이지로 이동",
         None,"미실행","Critical","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-012","FN-004","API-009",None,None),
        ("TC-010","Web","회원관리","등록","기능 테스트","회원 등록 실패 - 연락처 중복",
         "동일 연락처 회원 이미 존재",
         "1) 회원 등록 폼에 중복 연락처 입력\n2) '저장' 클릭",
         "'이미 등록된 연락처입니다' 에러 메시지 표시",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-012","FN-004","API-009",None,None),
        ("TC-011","Web","회원관리","등록","UI/UX 테스트","필수 항목 유효성 검사",
         "회원 등록 폼 접속",
         "1) 아무것도 입력하지 않고 '저장' 클릭",
         "필수 항목(이름/성별/연락처)에 에러 메시지 표시, 저장 불가",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-012","FN-004",None,None,None),
        ("TC-012","Web","회원관리","상세","기능 테스트","회원 상세 탭 전환",
         "회원 데이터 존재",
         "1) 회원 상세 페이지 접속\n2) 이용권 탭 클릭\n3) 출석 탭 클릭\n4) 결제 탭 클릭\n5) 체성분 탭 클릭\n6) 메모 탭 클릭",
         "각 탭별 데이터 정상 로드, URL 파라미터 유지",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-011","FN-006","API-008",None,None),
        ("TC-013","Web","회원관리","수정","기능 테스트","회원 정보 수정",
         "회원 데이터 존재",
         "1) 회원 상세 → '수정' 클릭\n2) 이름 변경\n3) '저장' 클릭",
         "수정 성공, 변경 이력 기록, 상세 페이지로 이동",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-013","FN-005","API-010",None,None),
        # 매출/POS
        ("TC-014","Web","매출관리","조회","기능 테스트","매출 기간 필터 조회",
         "매출 데이터 존재",
         "1) 매출 조회 페이지 접속\n2) 시작일/종료일 선택\n3) 결과 확인",
         "선택 기간의 매출만 표시, 통계 카드 갱신",
         None,"미실행","Critical","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-030","FN-010","API-015",None,None),
        ("TC-015","Web","매출관리","조회","기능 테스트","매출 엑셀 다운로드",
         "매출 데이터 존재",
         "1) 매출 조회\n2) '엑셀 다운로드' 클릭",
         "xlsx 파일 다운로드, 현재 필터 결과와 동일한 데이터",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-030","FN-010","API-015",None,None),
        ("TC-016","Web","매출관리","POS","기능 테스트","POS 상품 선택 및 장바구니",
         "상품 데이터 존재",
         "1) POS 페이지 접속\n2) 카테고리 선택\n3) 상품 클릭 (장바구니 추가)\n4) 수량 변경\n5) 합계 확인",
         "장바구니에 상품 추가, 수량 반영, 합계 자동 계산",
         None,"미실행","Critical","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-031","FN-011","API-018",None,None),
        ("TC-017","Web","매출관리","POS","기능 테스트","POS 카드 결제 처리",
         "장바구니에 상품 존재",
         "1) '결제' 버튼 클릭\n2) 회원 검색/선택\n3) 결제수단: 카드 선택\n4) '결제 확인' 클릭",
         "결제 성공, 매출 등록, 영수증 표시",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-032","FN-012","API-016",None,None),
        ("TC-018","Web","매출관리","POS","기능 테스트","POS 복합결제 (현금+카드)",
         "장바구니에 상품 존재",
         "1) 결제수단: 복합 선택\n2) 현금 금액 입력\n3) 나머지 카드 자동 계산\n4) '결제 확인'",
         "복합결제 성공, 현금/카드 금액 정확히 분리 기록",
         None,"미실행","High","Critical","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-032","FN-012","API-016",None,None),
        # 상품
        ("TC-019","Web","상품관리","등록","기능 테스트","신규 상품 등록",
         "센터장 이상 권한",
         "1) 상품등록 페이지 접속\n2) 카테고리: 이용권\n3) 이름: 3개월 이용권\n4) 현금가: 280000\n5) 카드가: 300000\n6) '저장'",
         "상품 등록 성공, 목록에 표시",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-041","FN-014","API-019",None,None),
        # 락커
        ("TC-020","Web","시설관리","락커","기능 테스트","락커 현황 그리드 표시",
         "락커 데이터 존재",
         "1) 락커 현황 페이지 접속\n2) 구역 탭 선택\n3) 그리드 색상 확인",
         "구역별 락커 그리드 표시, 상태별 색상 구분 정확",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-050","FN-015","API-020",None,None),
        ("TC-021","Web","시설관리","락커","기능 테스트","락커 배정",
         "빈 락커 존재, 회원 존재",
         "1) 락커 관리 접속\n2) 회원 검색/선택\n3) 빈 락커 클릭\n4) 만료일 설정\n5) '배정' 클릭",
         "락커 배정 성공, 상태 '사용중'으로 변경",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-051","FN-016","API-021",None,None),
        # 직원
        ("TC-022","Web","직원관리","목록","기능 테스트","직원 목록 조회",
         "직원 데이터 존재, 센터장 권한",
         "1) 직원 목록 접속\n2) 역할 필터 선택\n3) 결과 확인",
         "필터에 맞는 직원만 표시",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-060","FN-019","API-022",None,None),
        ("TC-023","Web","직원관리","등록","기능 테스트","신규 직원 등록",
         "센터장 권한",
         "1) 직원등록 접속\n2) 이름/역할/연락처/입사일 입력\n3) '저장'",
         "직원 등록 성공, 관리자 계정 자동 생성",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-061","FN-020","API-023",None,None),
        # 메시지
        ("TC-024","Web","메시지/마케팅","발송","기능 테스트","SMS 메시지 발송",
         "회원 데이터 존재",
         "1) 메시지 발송 접속\n2) 수신자 선택\n3) SMS 선택\n4) 본문 입력 (90자 이내)\n5) '발송'",
         "발송 성공, 발송 이력에 기록",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-070","FN-023",None,None,None),
        ("TC-025","Web","메시지/마케팅","발송","기능 테스트","메시지 발송 - 수신자 미선택",
         "메시지 발송 페이지 접속",
         "1) 수신자 없이 본문 입력\n2) '발송' 클릭",
         "'수신자를 선택해주세요' 에러, 발송 불가",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-070","FN-023",None,None,None),
        # 계약
        ("TC-026","Web","계약/구독","계약","기능 테스트","계약 마법사 전체 플로우",
         "회원/상품 데이터 존재",
         "1) 계약 마법사 접속\n2) 회원 선택\n3) 상품 선택\n4) 기간/금액 설정\n5) 결제 처리\n6) 확인",
         "계약 등록 성공, 이용권 자동 생성, 매출 자동 등록",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-080","FN-027","API-016",None,None),
        ("TC-027","Web","계약/구독","계약","기능 테스트","계약 마법사 - 이전 단계 미완료",
         "계약 마법사 접속",
         "1) 회원 미선택 상태에서 '다음' 클릭",
         "'회원을 선택하세요' 에러, 다음 단계 진행 불가",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-080","FN-027",None,None,None),
        # 설정/권한
        ("TC-028","Web","설정/권한","권한","기능 테스트","권한 매트릭스 설정",
         "최고관리자 권한",
         "1) 권한 설정 접속\n2) '매니저' 탭 선택\n3) '회원 삭제' 권한 체크 해제\n4) 저장",
         "매니저 역할에서 회원 삭제 권한 제거, 즉시 반영",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-091","FN-030",None,None,None),
        ("TC-029","Web","설정/권한","권한","기능 테스트","권한 부족 접근 차단",
         "스태프 역할로 로그인",
         "1) URL 직접 입력으로 /settings/permissions 접근",
         "접근 거부 메시지 또는 메뉴 미표시",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-091","FN-030",None,None,None),
        ("TC-030","Web","설정/권한","지점","기능 테스트","지점 추가",
         "최고관리자 권한",
         "1) 지점 관리 접속\n2) '지점 추가' 클릭\n3) 이름/코드/주소 입력\n4) 저장",
         "지점 추가 성공, 목록에 표시",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-094","FN-033","API-027",None,None),
        # 데이터 격리
        ("TC-031","Web","보안","데이터 격리","보안 테스트","지점 간 데이터 격리 확인",
         "2개 이상 지점 데이터 존재",
         "1) A지점으로 로그인\n2) 회원 목록 확인\n3) B지점으로 전환\n4) 회원 목록 확인",
         "각 지점에 소속된 회원만 표시, 타지점 데이터 미노출",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-010","FN-003","API-007",None,None),
        ("TC-032","Web","보안","인증","보안 테스트","JWT 토큰 만료 처리",
         "로그인 완료",
         "1) JWT 토큰 만료 대기 (또는 수동 만료)\n2) API 호출 시도",
         "401 Unauthorized, 로그인 페이지로 리다이렉트",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-001","FN-001","API-001",None,None),
        # 성능
        ("TC-033","Web","성능","목록","성능 테스트","회원 목록 1000건 로딩 성능",
         "회원 1000건 데이터 존재",
         "1) 회원 목록 접속\n2) 로딩 시간 측정",
         "3초 이내 첫 페이지 로딩 완료",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-010","FN-003","API-007",None,None),
        ("TC-034","Web","성능","대시보드","성능 테스트","대시보드 초기 로딩 성능",
         "실 데이터 환경",
         "1) 대시보드 접속\n2) 모든 위젯 로딩 시간 측정",
         "2초 이내 대시보드 완전 렌더링",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","자동화 예정",
         "SCR-002","FN-002","API-003",None,None),
        # 호환성
        ("TC-035","Web","호환성","브라우저","호환성 테스트","Safari 브라우저 호환성",
         "macOS Safari 최신",
         "1) Safari로 전체 기능 테스트\n2) 레이아웃/기능 확인",
         "Chrome과 동일한 동작, 레이아웃 깨짐 없음",
         None,"미실행","Medium","Major","QA담당자2","Staging","Safari","Desktop","macOS","1차","수동",
         None,None,None,None,None),
        ("TC-036","Web","호환성","브라우저","호환성 테스트","Firefox 브라우저 호환성",
         "Firefox 최신",
         "1) Firefox로 전체 기능 테스트\n2) 레이아웃/기능 확인",
         "Chrome과 동일한 동작",
         None,"미실행","Medium","Major","QA담당자2","Staging","Firefox","Desktop","Windows 11","1차","수동",
         None,None,None,None,None),
        # 출석
        ("TC-037","Web","출석관리","조회","기능 테스트","출석 기록 일별 조회",
         "출석 데이터 존재",
         "1) 출석 기록 접속\n2) 날짜 선택\n3) 결과 확인",
         "선택 날짜의 출석 기록만 표시",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-020","FN-008","API-013",None,None),
        # 접근성
        ("TC-038","Web","접근성","키보드","UI/UX 테스트","키보드 네비게이션 - 로그인",
         "로그인 페이지 접속",
         "1) Tab키로 이메일→비밀번호→지점→버튼 이동\n2) Enter키로 로그인",
         "Tab 순서 정확, Enter로 로그인 실행",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-001","FN-001",None,None,None),
        # 엣지케이스
        ("TC-039","Web","회원관리","등록","기능 테스트","회원 등록 - 특수문자 이름",
         "회원 등록 폼 접속",
         "1) 이름에 특수문자 입력 (!@#$)\n2) '저장' 시도",
         "유효성 검사: '이름은 한글/영문만 입력 가능합니다' 또는 저장 차단",
         None,"미실행","Medium","Minor","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-012","FN-004","API-009",None,None),
        ("TC-040","Web","매출관리","POS","기능 테스트","POS 결제 - 마일리지 잔액 초과",
         "회원 마일리지 1000원, 상품 10000원",
         "1) 마일리지 결제 선택\n2) 10000원 마일리지 사용 시도",
         "'마일리지 잔액이 부족합니다' 에러, 결제 불가",
         None,"미실행","High","Major","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-032","FN-012","API-016",None,None),
        # 급여
        ("TC-041","Web","직원관리","급여","기능 테스트","급여 현황 조회",
         "급여 데이터 존재, 센터장 권한",
         "1) 급여 현황 접속\n2) 월 선택\n3) 테이블 확인",
         "해당 월 급여 테이블 표시, 합계 정확",
         None,"미실행","High","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-062","FN-021",None,None,None),
        ("TC-042","Web","직원관리","급여","기능 테스트","급여명세서 PDF 다운로드",
         "급여 데이터 존재",
         "1) 급여명세서 접속\n2) 직원/월 선택\n3) 'PDF 다운로드' 클릭",
         "PDF 파일 다운로드, 내용 정확",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-063","FN-022",None,None,None),
        # 쿠폰/마일리지
        ("TC-043","Web","메시지/마케팅","쿠폰","기능 테스트","쿠폰 생성",
         "매니저 이상 권한",
         "1) 쿠폰 관리 접속\n2) '쿠폰 생성' 클릭\n3) 이름/유형/할인값/기간 입력\n4) 저장",
         "쿠폰 생성 성공, 목록에 표시",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-072","FN-025",None,None,None),
        ("TC-044","Web","메시지/마케팅","마일리지","기능 테스트","마일리지 적립",
         "회원 데이터 존재",
         "1) 마일리지 관리 접속\n2) 회원 검색\n3) '적립' 클릭\n4) 금액 입력\n5) 확인",
         "마일리지 적립 성공, 잔액 갱신, 이력 기록",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-073","FN-026",None,None,None),
        # API 직접 테스트
        ("TC-045","Web","API","인증","API 테스트","로그인 API - 정상 응답 구조",
         "테스트 계정 존재",
         "1) POST /api/v1/auth/login 호출\n2) 응답 구조 확인",
         "200 OK, token/user 필드 존재, user에 role/branch_id 포함",
         None,"미실행","Critical","Critical","QA담당자1","Staging",None,None,None,"1차","자동화 예정",
         None,"FN-001","API-001",None,None),
        ("TC-046","Web","API","회원","API 테스트","회원 목록 API - 페이지네이션",
         "회원 50건 이상",
         "1) GET /api/v1/members?page=1&size=20\n2) pagination 필드 확인\n3) page=2 호출\n4) 중복 데이터 없음 확인",
         "pagination 정확, 페이지 간 중복 없음",
         None,"미실행","High","Major","QA담당자1","Staging",None,None,None,"1차","자동화 예정",
         None,"FN-003","API-007",None,None),
        # 시설관리 추가
        ("TC-047","Web","시설관리","RFID","기능 테스트","RFID 카드 등록",
         "RFID 리더 연동 환경",
         "1) RFID 관리 접속\n2) 카드 스캔\n3) 회원 매핑\n4) 저장",
         "카드-회원 매핑 성공, 이력 기록",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-052","FN-017",None,None,None),
        ("TC-048","Web","시설관리","룸","기능 테스트","GX룸 등록",
         "매니저 이상 권한",
         "1) 룸 관리 접속\n2) '룸 등록' 클릭\n3) 룸명/유형/수용인원 입력\n4) 저장",
         "룸 등록 성공, 카드 뷰에 표시",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-053","FN-018",None,None,None),
        # IoT
        ("TC-049","Web","설정/권한","IoT","기능 테스트","IoT 기기 연결 테스트",
         "IoT 기기 설정 존재",
         "1) IoT 설정 접속\n2) '연결 테스트' 클릭",
         "연결 성공/실패 결과 표시",
         None,"미실행","Medium","Major","QA담당자2","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-093","FN-032",None,None,None),
        # 데이터 초기화
        ("TC-050","Web","설정/권한","지점","기능 테스트","데이터 초기화 - 2중 확인",
         "최고관리자 권한",
         "1) 지점 관리 → '데이터 초기화'\n2) 첫 번째 확인 모달\n3) 비밀번호 입력\n4) 초기화 실행",
         "선택한 데이터 초기화 완료, 2중 확인 없이 초기화 불가",
         None,"미실행","Critical","Blocker","QA담당자1","Staging","Chrome","Desktop","Windows 11","1차","수동",
         "SCR-094","FN-033","API-027",None,None),
    ]

    for ri, tc in enumerate(tcs):
        row = ri + 2
        for ci, val in enumerate(tc):
            if val is not None:
                ws.cell(row, ci + 1, val)
        # [자동] 수식 Z, AA, AB
        ws.cell(row, 26, f"=IFERROR(VLOOKUP(U{row},'IA(정보구조)'!D:E,2,FALSE()),\"\")")
        ws.cell(row, 27, f'=IFERROR(VLOOKUP(V{row},기능명세서!A:B,2,FALSE()),"")')
        ws.cell(row, 28, f'=IFERROR(VLOOKUP(W{row},API명세서!A:D,4,FALSE()),"")')

        for ci in range(1, 30):
            c = ws.cell(row, ci)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # 드롭다운
    add_dv(ws, "B", "Web,App (iOS),App (Android),Web + App,공통")
    add_dv(ws, "E", "기능 테스트,UI/UX 테스트,API 테스트,성능 테스트,보안 테스트,호환성 테스트,회귀 테스트,탐색적 테스트")
    add_dv(ws, "K", "미실행,Pass,Fail,Block,N/A,보류")
    add_dv(ws, "L", "Critical,High,Medium,Low")
    add_dv(ws, "M", "Blocker,Critical,Major,Minor,Trivial")
    add_dv(ws, "N", 담당자_LIST)
    add_dv(ws, "O", "Production,Staging,Dev,QA")
    add_dv(ws, "P", "Chrome,Safari,Firefox,Edge,Samsung Internet,인앱브라우저")
    add_dv(ws, "Q", "Desktop,iPhone 15,iPhone 14,Galaxy S24,Galaxy S23,iPad,Galaxy Tab,기타")
    add_dv(ws, "R", "Windows 11,Windows 10,macOS,iOS 17,iOS 16,Android 14,Android 13")
    add_dv(ws, "S", "1차,2차,3차,리그레션,핫픽스")
    add_dv(ws, "T", "수동,자동화 완료,자동화 예정,자동화 불가")
    add_dv(ws, "U", SCR_IDS_STR)
    add_dv(ws, "V", FN_IDS_STR)
    add_dv(ws, "W", API_IDS_STR)

    ws.auto_filter.ref = "A1:AC1"
    return ws


# ============================================================
# 9. 대시보드 시트
# ============================================================
def create_dashboard_sheet(wb):
    ws = wb.create_sheet("대시보드")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    # 타이틀
    c = ws.cell(1, 1, "기획 문서 종합 대시보드")
    c.font = DASH_TITLE_FONT

    # 문서 작성 현황
    row = 3
    c = ws.cell(row, 1, "문서 작성 현황")
    c.fill = DASH_SECTION_FILL
    c.font = DASH_SECTION_FONT
    for ci in range(2, 5):
        ws.cell(row, ci).fill = DASH_SECTION_FILL

    row = 4
    for ci, h in enumerate(["시트", "전체 항목", "확정", "작성중"], 1):
        c = ws.cell(row, ci, h)
        c.fill = SECTION_FILL
        c.font = SECTION_FONT
        c.border = THIN_BORDER

    # IA
    ws.cell(5, 1, "IA(정보구조)")
    ws.cell(5, 2, "=COUNTA('IA(정보구조)'!A2:A200)")
    ws.cell(5, 3, '=COUNTIF(\'IA(정보구조)\'!K2:K200,"확정")')
    ws.cell(5, 4, '=COUNTIF(\'IA(정보구조)\'!K2:K200,"작성중")')
    # 화면설계서
    ws.cell(6, 1, "화면설계서")
    ws.cell(6, 2, "=COUNTA(화면설계서!A2:A200)")
    ws.cell(6, 3, '=COUNTIF(화면설계서!Q2:Q200,"확정")')
    ws.cell(6, 4, '=COUNTIF(화면설계서!Q2:Q200,"작성중")')
    # 기능명세서
    ws.cell(7, 1, "기능명세서")
    ws.cell(7, 2, "=COUNTA(기능명세서!A2:A200)")
    ws.cell(7, 3, '=COUNTIF(기능명세서!P2:P200,"확정")')
    ws.cell(7, 4, '=COUNTIF(기능명세서!P2:P200,"작성중")')
    # API명세서
    ws.cell(8, 1, "API명세서")
    ws.cell(8, 2, "=COUNTA(API명세서!A2:A200)")
    ws.cell(8, 3, '=COUNTIF(API명세서!M2:M200,"개발완료")')
    ws.cell(8, 4, '=COUNTIF(API명세서!M2:M200,"설계중")')
    # 데이터 정의서
    ws.cell(9, 1, "데이터 정의서")
    ws.cell(9, 2, "=COUNTA('데이터 정의서'!A2:A200)")
    ws.cell(9, 3, '=COUNTIF(\'데이터 정의서\'!L2:L200,"확정")')
    ws.cell(9, 4, '=COUNTIF(\'데이터 정의서\'!L2:L200,"작성중")')

    for r in range(5, 10):
        for ci in range(1, 5):
            ws.cell(r, ci).border = THIN_BORDER

    # QA 테스트 현황
    row = 12
    c = ws.cell(row, 1, "QA 테스트 현황")
    c.fill = DASH_SECTION_FILL
    c.font = DASH_SECTION_FONT
    ws.cell(row, 2).fill = DASH_SECTION_FILL

    ws.cell(13, 1, "전체 TC 수")
    ws.cell(13, 2, "=COUNTA(QA테스트케이스!A2:A200)")
    ws.cell(14, 1, "Pass")
    ws.cell(14, 2, '=COUNTIF(QA테스트케이스!K2:K200,"Pass")')
    ws.cell(15, 1, "Fail")
    ws.cell(15, 2, '=COUNTIF(QA테스트케이스!K2:K200,"Fail")')
    ws.cell(16, 1, "Block")
    ws.cell(16, 2, '=COUNTIF(QA테스트케이스!K2:K200,"Block")')
    ws.cell(17, 1, "미실행")
    ws.cell(17, 2, '=COUNTIF(QA테스트케이스!K2:K200,"미실행")')
    ws.cell(18, 1, "Pass Rate (%)")
    ws.cell(18, 2, "=IF(B14>0,B14/(B14+B15+B16)*100,0)")

    for r in range(13, 19):
        for ci in range(1, 3):
            ws.cell(r, ci).border = THIN_BORDER

    # 플랫폼별 QA 현황
    row = 21
    c = ws.cell(row, 1, "플랫폼별 QA 현황")
    c.fill = DASH_SECTION_FILL
    c.font = DASH_SECTION_FONT
    for ci in range(2, 5):
        ws.cell(row, ci).fill = DASH_SECTION_FILL

    for ci, h in enumerate(["플랫폼", "총 TC", "Pass", "Fail"], 1):
        c = ws.cell(22, ci, h)
        c.fill = SECTION_FILL
        c.font = SECTION_FONT
        c.border = THIN_BORDER

    platforms = [("Web", 23), ("App (iOS)", 24), ("App (Android)", 25), ("공통", 26)]
    for pname, r in platforms:
        ws.cell(r, 1, pname)
        ws.cell(r, 2, f'=COUNTIF(QA테스트케이스!B2:B200,"{pname}")')
        ws.cell(r, 3, f'=COUNTIFS(QA테스트케이스!B2:B200,"{pname}",QA테스트케이스!K2:K200,"Pass")')
        ws.cell(r, 4, f'=COUNTIFS(QA테스트케이스!B2:B200,"{pname}",QA테스트케이스!K2:K200,"Fail")')
        for ci in range(1, 5):
            ws.cell(r, ci).border = THIN_BORDER

    # 사용자 스토리 현황
    row = 29
    c = ws.cell(row, 1, "사용자 스토리 현황")
    c.fill = DASH_SECTION_FILL
    c.font = DASH_SECTION_FONT
    ws.cell(row, 2).fill = DASH_SECTION_FILL

    statuses = [("Backlog", 30), ("To Do", 31), ("In Progress", 32), ("In Review", 33), ("Done", 34)]
    for sname, r in statuses:
        ws.cell(r, 1, sname)
        ws.cell(r, 2, f'=COUNTIF(\'사용자 스토리\'!J2:J200,"{sname}")')
        for ci in range(1, 3):
            ws.cell(r, ci).border = THIN_BORDER

    # UI 요소 현황
    row = 36
    c = ws.cell(row, 1, "UI 요소 현황")
    c.fill = DASH_SECTION_FILL
    c.font = DASH_SECTION_FONT
    ws.cell(row, 2).fill = DASH_SECTION_FILL

    ws.cell(37, 1, "전체 UI 요소 수")
    ws.cell(37, 2, "=COUNTA('UI 요소 상세'!A2:A500)")
    ws.cell(38, 1, "완료")
    ws.cell(38, 2, '=COUNTIF(\'UI 요소 상세\'!N2:N500,"완료")')
    ws.cell(39, 1, "설계중")
    ws.cell(39, 2, '=COUNTIF(\'UI 요소 상세\'!N2:N500,"설계중")')
    ws.cell(40, 1, "개발중")
    ws.cell(40, 2, '=COUNTIF(\'UI 요소 상세\'!N2:N500,"개발중")')

    for r in range(37, 41):
        for ci in range(1, 3):
            ws.cell(r, ci).border = THIN_BORDER

    return ws


# ============================================================
# main
# ============================================================
def main():
    wb = Workbook()

    print("1/9 IA(정보구조) 시트 생성...")
    create_ia_sheet(wb)

    print("2/9 사용자 스토리 시트 생성...")
    create_story_sheet(wb)

    print("3/9 화면설계서 시트 생성...")
    create_screen_spec_sheet(wb)

    print("4/9 UI 요소 상세 시트 생성...")
    create_ui_detail_sheet(wb)

    print("5/9 기능명세서 시트 생성...")
    create_function_spec_sheet(wb)

    print("6/9 API명세서 시트 생성...")
    create_api_spec_sheet(wb)

    print("7/9 데이터 정의서 시트 생성...")
    create_data_dict_sheet(wb)

    print("8/9 QA테스트케이스 시트 생성...")
    create_qa_sheet(wb)

    print("9/9 대시보드 시트 생성...")
    create_dashboard_sheet(wb)

    # 대시보드를 맨 앞으로 이동하지 않음 (v2 템플릿 순서 유지)

    wb.save(OUTPUT_PATH)
    print(f"\n✅ 파일 생성 완료: {OUTPUT_PATH}")

    # 요약 출력
    print("\n=== 시트별 요약 ===")
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_row}행 x {ws.max_column}열")


if __name__ == "__main__":
    main()
