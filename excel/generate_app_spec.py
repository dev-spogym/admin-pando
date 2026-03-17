#!/usr/bin/env python3
"""
스포짐 회원앱 기획서 엑셀 생성 스크립트
기존 CRM 기획서(스포짐_기획문서_v3.xlsx)와 동일한 시트 구조로 생성합니다.

사용법:
  cd excel && python3 generate_app_spec.py
  → 스포짐_회원앱_기획서_v1.xlsx 생성
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── 스타일 정의 ──────────────────────────────────────────
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
LOW_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
BODY_FONT = Font(name="맑은 고딕", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

PRIORITY_FILLS = {
    "Critical": CRITICAL_FILL,
    "High": HIGH_FILL,
    "Medium": MEDIUM_FILL,
    "Low": LOW_FILL,
}


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    line_max = max(len(line) for line in lines)
                    max_len = max(max_len, min(line_max + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def apply_priority_fill(ws, col_idx, start_row=2):
    """우선순위 컬럼에 색상 적용"""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[cell.value]


# ══════════════════════════════════════════════════════════
# 1. IA(정보구조) 시트
# ══════════════════════════════════════════════════════════
def create_ia_sheet(wb):
    ws = wb.create_sheet("IA(정보구조)")
    headers = [
        "Depth 1", "Depth 2", "Depth 3", "화면 ID", "화면명",
        "플랫폼", "화면 URL/경로", "로그인 필요", "설명", "우선순위", "상태", "비고"
    ]
    ws.append(headers)
    style_header(ws)

    data = [
        ["인증", "로그인", None, "MA-001", "로그인", "App", "/login", "N", "연락처+비밀번호 로그인", "Critical", "기획", None],
        ["인증", "회원가입", None, "MA-002", "앱 연동 (회원가입)", "App", "/register", "N", "센터 등록 회원의 앱 최초 연동 (SMS 인증)", "Critical", "기획", None],
        ["인증", "비밀번호 찾기", None, "MA-003", "비밀번호 찾기", "App", "/forgot-password", "N", "SMS 인증으로 비밀번호 재설정", "High", "기획", None],
        ["홈", "대시보드", None, "MA-010", "홈 대시보드", "App", "/home", "Y", "이용권 D-day, 오늘 예약, 출석 현황, 공지 배너", "Critical", "기획", "QR 체크인 플로팅 버튼"],
        ["출석", "QR 체크인", None, "MA-030", "QR 체크인", "App", "/checkin", "Y", "QR코드 생성/스캔으로 센터 출석 체크", "Critical", "기획", "60초 자동 갱신"],
        ["출석", "출석 이력", None, "MA-031", "출석 이력", "App", "/attendance", "Y", "월별 캘린더 히트맵 + 출석 리스트", "Critical", "기획", None],
        ["예약", "수업 목록", None, "MA-040", "수업 목록", "App", "/classes", "Y", "날짜별 PT/GX 수업 목록 + 잔여석", "Critical", "기획", None],
        ["예약", "수업 상세", None, "MA-041", "수업 상세/예약", "App", "/classes/:id", "Y", "수업 정보 + 예약/취소 처리", "Critical", "기획", None],
        ["예약", "내 예약", None, "MA-042", "내 예약 목록", "App", "/my-reservations", "Y", "예정/완료/취소 예약 리스트", "High", "기획", None],
        ["내정보", "프로필", None, "MA-050", "프로필", "App", "/profile", "Y", "이름/연락처/사진 조회 및 수정", "High", "기획", "이름/연락처 센터에서만 변경"],
        ["내정보", "이용권", None, "MA-051", "이용권 목록", "App", "/memberships", "Y", "활성/만료 이용권 + 만료 D-day + 잔여 횟수", "Critical", "기획", "연장하기 버튼"],
        ["내정보", "체성분", None, "MA-052", "체성분 변화", "App", "/body-data", "Y", "체중/체지방/골격근 시계열 차트", "High", "기획", "읽기 전용 (측정은 센터)"],
        ["내정보", "결제 이력", None, "MA-053", "결제 이력", "App", "/payments", "Y", "결제 내역 리스트 + 영수증 PDF 다운로드", "High", "기획", None],
        ["내정보", "락커", None, "MA-054", "락커 정보", "App", "/locker", "Y", "배정된 락커 번호/위치/만료일", "Medium", "기획", None],
        ["내정보", "쿠폰함", None, "MA-055", "쿠폰함", "App", "/coupons", "Y", "보유 쿠폰 목록 + 사용 가능 여부", "High", "기획", None],
        ["내정보", "마일리지", None, "MA-056", "마일리지", "App", "/mileage", "Y", "잔액 + 적립/사용 이력", "Medium", "기획", None],
        ["구매", "상품 목록", None, "MA-060", "상품 목록", "App", "/shop", "Y", "이용권/PT/GX 상품 카테고리별 조회", "High", "기획", "판매중 + 키오스크 노출 상품만"],
        ["구매", "상품 상세", None, "MA-061", "상품 상세", "App", "/shop/:id", "Y", "상품 정보 + 가격 + 구매 버튼", "High", "기획", None],
        ["구매", "결제", None, "MA-062", "결제/구매", "App", "/shop/checkout", "Y", "쿠폰/마일리지 적용 + PG 결제 + 이용권 활성화", "High", "기획", "토스페이먼츠/아임포트"],
        ["더보기", "알림 목록", None, "MA-070", "알림 목록", "App", "/notifications", "Y", "만료/예약/생일/공지 알림 리스트", "Medium", "기획", None],
        ["더보기", "알림 설정", None, "MA-071", "알림 설정", "App", "/notifications/settings", "Y", "유형별 푸시 알림 ON/OFF", "Medium", "기획", None],
        ["더보기", "센터 정보", None, "MA-072", "센터 정보", "App", "/center-info", "Y", "센터명/주소/영업시간/연락처/지도", "Low", "기획", "네이버/카카오맵 연동"],
        ["더보기", "공지사항", None, "MA-073", "공지사항", "App", "/notices", "Y", "센터 공지사항 목록 + 상세", "Low", "기획", None],
        ["더보기", "1:1 문의", None, "MA-074", "1:1 문의", "App", "/inquiry", "Y", "문의 작성 + 답변 조회", "Low", "기획", None],
        ["더보기", "구독 관리", None, "MA-075", "구독/자동결제", "App", "/subscription", "Y", "구독 상태/결제일/카드 관리/해지", "Medium", "기획", None],
        ["더보기", "앱 설정", None, "MA-076", "앱 설정", "App", "/settings", "Y", "비밀번호 변경, 계정 탈퇴, 버전 정보", "Medium", "기획", None],
        ["더보기", "로그아웃", None, "MA-077", "로그아웃", "App", "-", "Y", "로그아웃 확인 → 로그인 화면 이동", "Critical", "기획", None],
    ]

    for row in data:
        ws.append(row)

    style_body(ws)
    apply_priority_fill(ws, 10)  # 우선순위 컬럼
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 2. 사용자 스토리 시트
# ══════════════════════════════════════════════════════════
def create_story_sheet(wb):
    ws = wb.create_sheet("사용자 스토리")
    headers = [
        "스토리 ID", "에픽", "사용자 유형", "사용자 스토리",
        "인수 조건", "화면 ID", "우선순위", "사이즈",
        "CRM 연관", "상태", "비고"
    ]
    ws.append(headers)
    style_header(ws)

    stories = [
        ["MUS-001", "인증", "회원", "회원으로서 연락처와 비밀번호로 로그인하여 앱 서비스를 이용하고 싶다",
         "1) 연락처+비밀번호 입력으로 로그인\n2) 로그인 성공 시 홈으로 이동\n3) 자동 로그인 옵션\n4) 5회 실패 시 잠금 안내",
         "MA-001", "Critical", "M", "FN-001", "기획", None],
        ["MUS-002", "인증", "신규회원", "센터에서 등록한 회원으로서 앱을 처음 설치하고 내 계정과 연동하고 싶다",
         "1) 이름+연락처 입력\n2) SMS 인증번호 확인\n3) 비밀번호 설정\n4) 연동 완료 시 홈 이동",
         "MA-002", "Critical", "L", "FN-004", "기획", None],
        ["MUS-003", "인증", "회원", "비밀번호를 잊은 회원으로서 SMS 인증으로 비밀번호를 재설정하고 싶다",
         "1) 연락처로 SMS 인증번호 발송\n2) 인증 성공 후 새 비밀번호 설정\n3) 재설정 완료 후 로그인 이동",
         "MA-003", "High", "S", "FN-001", "기획", None],
        ["MUS-004", "홈", "회원", "회원으로서 홈 화면에서 이용권 상태, 오늘 예약, 출석 현황을 한눈에 보고 싶다",
         "1) 이용권 만료 D-day 카드 (7일 이내 빨강)\n2) 오늘 예약된 수업 카드\n3) 이번 달 출석 횟수\n4) 공지사항 배너\n5) QR 체크인 플로팅 버튼",
         "MA-010", "Critical", "L", "FN-002,FN-006", "기획", None],
        ["MUS-005", "출석", "회원", "회원으로서 센터 도착 시 앱의 QR코드를 스캔하여 간편하게 출석 체크하고 싶다",
         "1) 앱에서 QR코드 표시 (60초 유효)\n2) 키오스크에서 스캔 시 출석 처리\n3) 체크인 완료 시 진동+애니메이션\n4) 이용권 없으면 안내",
         "MA-030", "Critical", "M", "FN-008", "기획", None],
        ["MUS-006", "출석", "회원", "회원으로서 내 출석 이력을 캘린더와 리스트로 확인하고 싶다",
         "1) 월별 캘린더 히트맵\n2) 출석 유형별 구분 (일반/PT/GX)\n3) 월별 출석 횟수 통계\n4) 체크인/아웃 시간 표시",
         "MA-031", "Critical", "M", "FN-008", "기획", None],
        ["MUS-007", "예약", "회원", "회원으로서 예약 가능한 PT/GX 수업 목록을 날짜별로 확인하고 싶다",
         "1) 주간 날짜 스크롤\n2) PT/GX 탭 필터\n3) 트레이너별 필터\n4) 잔여석/마감 표시\n5) 예약됨 표시",
         "MA-040", "Critical", "M", "FN-009", "기획", None],
        ["MUS-008", "예약", "회원", "회원으로서 원하는 수업을 예약하거나 기존 예약을 취소하고 싶다",
         "1) 수업 상세 정보 표시\n2) 예약하기 버튼 → 확인 팝업\n3) PT 잔여 횟수 차감 표시\n4) 24시간 전 자유 취소\n5) 이용권 잔여 횟수 확인",
         "MA-041", "Critical", "L", "FN-009", "기획", None],
        ["MUS-009", "예약", "회원", "회원으로서 나의 예약 현황을 예정/완료/취소 탭으로 확인하고 싶다",
         "1) 예정/완료/취소 탭\n2) 예정: 날짜/시간/트레이너 + 취소 버튼\n3) 완료: 수강 기록\n4) 취소: 사유/일시",
         "MA-042", "High", "S", "FN-009", "기획", None],
        ["MUS-010", "내정보", "회원", "회원으로서 내 프로필 정보를 확인하고 수정하고 싶다",
         "1) 이름/연락처/이메일/생년월일/사진 표시\n2) 프로필 사진 변경\n3) 이메일/주소 수정 가능\n4) 비밀번호 변경 링크",
         "MA-050", "High", "M", "FN-005,FN-006", "기획", "이름/연락처 센터에서만 변경"],
        ["MUS-011", "내정보", "회원", "회원으로서 보유한 이용권의 상태와 만료일을 확인하고 싶다",
         "1) 활성/만료 탭\n2) 상품명/시작~종료일/남은일수\n3) 만료 7일 이내 빨강 경고\n4) PT/GX 잔여 횟수 프로그레스\n5) 연장하기 버튼",
         "MA-051", "Critical", "M", "FN-006", "기획", None],
        ["MUS-012", "내정보", "회원", "회원으로서 내 체성분 측정 기록의 변화 추이를 차트로 보고 싶다",
         "1) 체중/골격근/체지방률/BMI 차트\n2) 기간 선택 (1/3/6개월/전체)\n3) 최근 측정값 카드\n4) 목표 달성률 게이지",
         "MA-052", "High", "M", "FN-007", "기획", "읽기 전용"],
        ["MUS-013", "내정보", "회원", "회원으로서 결제 이력과 영수증을 확인하고 싶다",
         "1) 결제 내역 리스트\n2) 영수증 상세/PDF 다운로드\n3) 기간별 필터",
         "MA-053", "High", "S", "FN-006", "기획", None],
        ["MUS-014", "내정보", "회원", "회원으로서 나에게 배정된 락커 정보를 확인하고 싶다",
         "1) 락커 번호/구역/만료일 표시\n2) 만료 7일 이내 경고\n3) 미배정 시 안내 메시지",
         "MA-054", "Medium", "S", "FN-015", "기획", None],
        ["MUS-015", "내정보", "회원", "회원으로서 보유한 쿠폰 목록을 확인하고 구매 시 사용하고 싶다",
         "1) 사용가능/사용완료/만료 탭\n2) 쿠폰명/할인내용/유효기간\n3) 구매 시 쿠폰 적용\n4) 만료 임박 상단 배치",
         "MA-055", "High", "M", "FN-025", "기획", None],
        ["MUS-016", "내정보", "회원", "회원으로서 마일리지 잔액과 적립/사용 이력을 확인하고 싶다",
         "1) 현재 잔액 크게 표시\n2) 적립/사용 이력 리스트\n3) 소멸 예정 알림",
         "MA-056", "Medium", "S", "FN-026", "기획", None],
        ["MUS-017", "구매", "회원", "회원으로서 구매 가능한 이용권/PT/GX 상품을 확인하고 싶다",
         "1) 카테고리 탭 (이용권/PT/GX)\n2) 상품명/기간/가격 표시\n3) 판매 중 상품만 노출\n4) 상품 탭 → 상세 이동",
         "MA-060", "High", "M", "FN-013", "기획", None],
        ["MUS-018", "구매", "회원", "회원으로서 상품의 상세 정보를 확인하고 구매를 결정하고 싶다",
         "1) 상품명/카테고리/기간/가격\n2) 상품 설명\n3) 구매하기 버튼",
         "MA-061", "High", "S", "FN-013", "기획", None],
        ["MUS-019", "구매", "회원", "회원으로서 앱에서 이용권을 온라인 결제로 구매하고 싶다",
         "1) 주문 요약\n2) 쿠폰 적용\n3) 마일리지 사용\n4) 결제수단 선택 (카드/간편결제)\n5) PG 결제 처리\n6) 이용권 자동 활성화",
         "MA-062", "High", "XL", "FN-027", "기획", None],
        ["MUS-020", "알림", "회원", "회원으로서 센터에서 보낸 알림을 확인하고 싶다",
         "1) 알림 리스트 (시간순)\n2) 유형별 아이콘\n3) 읽음/안읽음 구분\n4) 탭 시 관련 화면 이동",
         "MA-070", "Medium", "M", "FN-023,FN-024", "기획", None],
        ["MUS-021", "알림", "회원", "회원으로서 수신할 알림 유형을 선택하고 싶다",
         "1) 유형별 푸시 알림 ON/OFF 토글\n2) 유형: 만료/수업/마케팅/공지",
         "MA-071", "Medium", "S", "FN-024", "기획", None],
        ["MUS-022", "더보기", "회원", "회원으로서 센터의 위치, 영업시간, 연락처를 확인하고 싶다",
         "1) 센터명/주소/영업시간/전화번호\n2) 지도 표시\n3) 전화 걸기 버튼",
         "MA-072", "Low", "S", "FN-029", "기획", None],
        ["MUS-023", "더보기", "회원", "회원으로서 센터 공지사항을 확인하고 싶다",
         "1) 공지 리스트 (제목/날짜)\n2) 상세 보기 (본문/이미지)\n3) 중요 공지 상단 고정",
         "MA-073", "Low", "S", "신규", "기획", None],
        ["MUS-024", "더보기", "회원", "회원으로서 센터에 문의 사항을 남기고 답변을 확인하고 싶다",
         "1) 카테고리 선택\n2) 제목+본문 작성\n3) 이미지 첨부\n4) 문의 이력+답변 확인\n5) 답변 시 푸시 알림",
         "MA-074", "Low", "M", "신규", "기획", None],
        ["MUS-025", "더보기", "회원", "회원으로서 정기 구독 상태를 확인하고 자동결제를 관리하고 싶다",
         "1) 구독 상품+상태\n2) 다음 결제일 표시\n3) 카드 변경\n4) 해지 (즉시/월말)\n5) 환불 예상금액",
         "MA-075", "Medium", "M", "FN-028", "기획", None],
        ["MUS-026", "설정", "회원", "회원으로서 앱 비밀번호를 변경하고 싶다",
         "1) 현재 비밀번호 입력\n2) 새 비밀번호+확인\n3) 규칙 실시간 검증\n4) 변경 완료 시 재로그인",
         "MA-076", "Medium", "S", "FN-001", "기획", None],
    ]

    for row in stories:
        ws.append(row)

    style_body(ws)
    apply_priority_fill(ws, 7)  # 우선순위 컬럼
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 3. 기능명세서 시트
# ══════════════════════════════════════════════════════════
def create_fn_sheet(wb):
    ws = wb.create_sheet("기능명세서")
    headers = [
        "기능 ID", "기능명", "기능 분류", "플랫폼", "우선순위",
        "기능 설명", "상세 요구사항", "화면 ID",
        "입력 항목", "출력 항목", "비즈니스 규칙", "예외 처리",
        "App 구현 사항", "CRM 연관 기능", "상태", "비고"
    ]
    ws.append(headers)
    style_header(ws)

    functions = [
        ["MFN-001", "회원 로그인", "인증", "App", "Critical",
         "회원이 연락처+비밀번호로 로그인. Supabase Auth 사용, 자동 로그인 지원.",
         "1) 연락처(010-xxxx-xxxx) + 비밀번호 입력\n2) 자동 로그인 토글 (30일 유지)\n3) 로그인 성공 시 홈 이동\n4) 5회 실패 시 잠금\n5) 비밀번호 찾기/앱 연동 링크",
         "MA-001",
         "연락처(String, 필수), 비밀번호(String, 필수), 자동로그인(Boolean, 선택)",
         "액세스 토큰(JWT), 회원 정보({id, name, phone, branchId})",
         "1) 연락처 기반 인증 (phone@member.spogym.local)\n2) 5회 실패 시 30분 잠금\n3) 자동 로그인: refresh token 30일\n4) 회원 role만 허용\n5) 비활성/탈퇴 계정 차단",
         "잘못된 자격증명: '연락처 또는 비밀번호를 확인해주세요'\n계정 잠금: '30분 후 재시도하거나 비밀번호 찾기'\n미연동: '앱 연동이 필요합니다'\n탈퇴: '탈퇴 처리된 계정입니다'",
         "Supabase Auth signInWithPassword\nSecureStorage 토큰 저장\n자동 로그인: 앱 시작 시 토큰 검증",
         "FN-001", "기획", None],
        ["MFN-002", "앱 연동 (최초 가입)", "인증", "App", "Critical",
         "센터 등록 회원이 앱 최초 설치 후 SMS 인증으로 본인확인하고 비밀번호 설정.",
         "1) Step 1: 이름+연락처 입력\n2) Step 2: SMS 인증 (6자리, 3분 유효)\n3) Step 3: 비밀번호 설정 (8자 이상)\n4) Step 4: 연동 완료 → 홈 이동",
         "MA-002",
         "이름(String, 필수), 연락처(String, 필수), 인증번호(String, 필수), 비밀번호(String, 필수)",
         "연동 결과, 회원 정보",
         "1) CRM 등록 이름+연락처 일치 확인\n2) SMS 6자리, 3분 유효, 5회/일 제한\n3) 이미 연동된 회원 → 로그인 안내\n4) CRM 미등록 → 센터 문의 안내\n5) Supabase Auth 계정 자동 생성",
         "불일치: '등록된 회원 정보와 일치하지 않습니다'\n인증 만료: '인증번호가 만료되었습니다'\n이미 연동: '이미 앱에 등록된 회원입니다'",
         "SMS API (CoolSMS)\nSupabase Auth createUser\n단계별 위자드 UI",
         "FN-004", "기획", None],
        ["MFN-003", "비밀번호 찾기/재설정", "인증", "App", "High",
         "SMS 인증으로 비밀번호를 재설정합니다.",
         "1) 등록된 연락처 입력\n2) SMS 인증번호 발송/확인\n3) 새 비밀번호 설정\n4) 재설정 완료 → 로그인 이동",
         "MA-003",
         "연락처(String, 필수), 인증번호(String, 필수), 새 비밀번호(String, 필수)",
         "재설정 결과",
         "1) 연락처로 회원 존재 확인\n2) SMS 인증 후 비밀번호 변경\n3) 이전 3개 비밀번호 재사용 금지",
         "미등록: '등록된 회원 정보가 없습니다'\n인증 실패: '인증번호가 올바르지 않습니다'",
         "Supabase Auth updateUser\nSMS API",
         "FN-001", "기획", None],
        ["MFN-004", "홈 대시보드 조회", "홈", "App", "Critical",
         "회원의 핵심 정보(이용권/예약/출석/공지)를 홈 화면에 요약 표시.",
         "1) 인사말: '안녕하세요, {이름}님!'\n2) 이용권 상태 카드 (D-day)\n3) 오늘 예약 카드\n4) 이번 달 출석 횟수\n5) 공지사항 배너\n6) QR 체크인 플로팅 버튼",
         "MA-010",
         "회원 ID(자동)",
         "활성 이용권, 오늘 예약, 월간 출석, 최신 공지, 쿠폰 수, 마일리지 잔액",
         "1) 만료 D-7 이내: 빨강 강조\n2) 이용권 없음: '구매하러 가기' 버튼\n3) 오늘 예약 없음: 회색 카드\n4) 본인 데이터만 조회\n5) Pull-to-refresh",
         "데이터 로드 실패: '새로고침을 시도해주세요'",
         "병렬 API 호출 (이용권+예약+출석+공지)\nPull-to-refresh\nSkeleton UI",
         "FN-002,FN-006", "기획", None],
        ["MFN-005", "QR 체크인", "출석", "App", "Critical",
         "회원이 앱에서 QR코드를 생성하고 센터에서 스캔하여 출석 체크.",
         "1) QR코드 생성 (member_id+timestamp+HMAC)\n2) 60초마다 자동 갱신\n3) 화면 밝기 자동 최대\n4) 체크인 완료 시 진동+애니메이션\n5) 또는 센터 QR 스캔 (역방향)",
         "MA-030",
         "회원 ID(자동), 지점 ID(자동)",
         "QR코드 데이터, 체크인 결과({success, time, type})",
         "1) 활성 이용권 있어야 QR 생성\n2) QR 60초 유효\n3) 당일 중복 체크인 허용\n4) 서버 HMAC 서명 검증",
         "이용권 없음: '활성 이용권이 없습니다'\n이용권 만료: '연장해주세요'\nQR 실패: '다시 시도해주세요'",
         "QR 라이브러리 (react-native-qrcode / qr_flutter)\n화면 밝기 API\nHaptic feedback",
         "FN-008", "기획", None],
        ["MFN-006", "출석 이력 조회", "출석", "App", "Critical",
         "월별 출석 캘린더 히트맵과 출석 리스트를 조회.",
         "1) 월별 캘린더 히트맵 (출석일 색상)\n2) 유형별 색상 (일반:파랑/PT:초록/GX:보라)\n3) 월별 출석 횟수 통계\n4) 체크인/아웃 시간 표시\n5) 좌우 스와이프 월 이동",
         "MA-031",
         "회원 ID(자동), 월(Date, 선택)",
         "출석 목록([{date, time, type}]), 월별 통계",
         "1) 본인 출석만 조회\n2) branch_id 기준 필터\n3) 타지점 출석 별도 표시",
         "출석 없음: '이번 달 출석 기록이 없습니다'",
         "캘린더 히트맵 커스텀 컴포넌트\n무한 스크롤 (과거 월)",
         "FN-008", "기획", None],
        ["MFN-007", "수업 목록 조회", "예약", "App", "Critical",
         "날짜별 예약 가능한 PT/GX 수업 목록을 조회.",
         "1) 주간 날짜 스크롤\n2) PT/GX 탭 필터\n3) 트레이너별 필터\n4) 수업 카드: 시간/수업명/트레이너/잔여석\n5) 마감 뱃지\n6) 예약됨 표시",
         "MA-040",
         "날짜(Date, 필수), 수업유형(String, 선택), 트레이너ID(UUID, 선택)",
         "수업 목록([{id, title, type, time, trainer, room, 잔여석, is_reserved}])",
         "1) 과거 수업: 회색 처리, 예약 불가\n2) PT: 담당 트레이너 수업 강조\n3) 예약 마감: 수업 1시간 전\n4) 이용권 유형 매칭",
         "수업 없음: '예약 가능한 수업이 없습니다'",
         "주간 날짜 스크롤 (horizontal FlatList)\nReact Query 캐싱",
         "FN-009", "기획", None],
        ["MFN-008", "수업 예약", "예약", "App", "Critical",
         "회원이 PT/GX 수업을 예약합니다.",
         "1) 수업 상세 표시\n2) 예약하기 → 확인 팝업 → 완료\n3) PT 잔여 횟수 차감 표시\n4) GX 잔여석 확인\n5) 예약 완료 시 캘린더 등록 제안",
         "MA-041",
         "수업 ID(UUID, 필수), 회원 ID(자동)",
         "예약 결과, 잔여 횟수",
         "1) PT 잔여 횟수 > 0\n2) GX 잔여석 > 0\n3) 동일 시간 중복 예약 불가\n4) 수업 1시간 전까지만 예약\n5) CRM 일정 자동 반영",
         "횟수 부족: '잔여 횟수가 부족합니다'\n정원 초과: '정원이 마감되었습니다'\n시간 중복: '같은 시간에 이미 예약이 있습니다'",
         "확인 BottomSheet\n캘린더 연동 (device calendar)",
         "FN-009", "기획", None],
        ["MFN-009", "수업 예약 취소", "예약", "App", "Critical",
         "예약한 수업을 취소합니다.",
         "1) 24시간 전: 자유 취소 (PT 횟수 복원)\n2) 24시간 이내: 경고 팝업\n3) 수업 시작 후: 취소 불가\n4) 취소 사유 선택 (선택)\n5) CRM 자동 반영",
         "MA-041",
         "예약 ID(UUID, 필수), 취소 사유(String, 선택)",
         "취소 결과",
         "1) 24시간 전 자유 취소\n2) 24시간 이내 경고 (PT 횟수 차감 가능)\n3) 시작 후 취소 불가\n4) CRM 일정 자동 반영",
         "취소 불가: '수업이 이미 시작되었습니다'\n경고: '24시간 이내 취소는 PT 횟수가 차감될 수 있습니다'",
         "확인 Dialog\n취소 사유 ActionSheet",
         "FN-009", "기획", None],
        ["MFN-010", "내 예약 목록 조회", "예약", "App", "High",
         "예정/완료/취소 예약 리스트를 조회.",
         "1) 예정/완료/취소 탭\n2) 예정: 날짜/시간/트레이너/룸 + 취소 버튼\n3) 완료: 수강 기록\n4) 취소: 취소 사유/일시",
         "MA-042",
         "회원 ID(자동), 상태 탭(String, 선택)",
         "예약 목록",
         "1) 본인 예약만 조회\n2) 예정: 날짜 가까운 순\n3) 완료/취소: 최근순",
         "예약 없음: '예약 내역이 없습니다'",
         "탭 전환 (SegmentedControl)\nFlatList",
         "FN-009", "기획", None],
        ["MFN-011", "프로필 조회", "내정보", "App", "High",
         "회원 프로필 정보를 조회.",
         "1) 이름, 연락처, 이메일, 생년월일, 프로필 사진\n2) 회원번호, 가입일, 담당 FC 표시\n3) 수정 버튼 → 수정 모드",
         "MA-050",
         "회원 ID(자동)",
         "회원 정보 전체",
         "1) 본인 정보만 조회\n2) 이름/연락처 읽기 전용 (센터에서만 변경)",
         "조회 실패: '정보를 불러올 수 없습니다'",
         "Supabase Storage 프로필 이미지",
         "FN-006", "기획", None],
        ["MFN-012", "프로필 수정", "내정보", "App", "High",
         "이메일, 주소, 프로필 사진을 수정.",
         "1) 프로필 사진 변경 (카메라/갤러리)\n2) 이메일 수정\n3) 주소 수정\n4) 저장 → 변경 반영",
         "MA-050",
         "이메일(String, 선택), 주소(String, 선택), 사진(File, 선택)",
         "수정된 회원 정보",
         "1) 이름/연락처 변경 불가 (센터 문의 안내)\n2) 이메일 형식 검증\n3) 사진: 5MB 이하, JPG/PNG",
         "사진 용량 초과: '5MB 이하 이미지를 선택해주세요'\n저장 실패: '다시 시도해주세요'",
         "ImagePicker (카메라/갤러리)\nSupabase Storage 업로드",
         "FN-005", "기획", None],
        ["MFN-013", "이용권 목록 조회", "내정보", "App", "Critical",
         "활성/만료 이용권 목록 + D-day + 잔여 횟수 조회.",
         "1) 활성/만료 탭\n2) 이용권 카드: 상품명/유형/시작~종료/남은일수\n3) D-day 뱃지 (D-30/D-7/D-1)\n4) PT/GX 잔여 횟수 프로그레스\n5) 만료 7일: 빨강 + 연장하기\n6) 만료: 회색",
         "MA-051",
         "회원 ID(자동)",
         "이용권 목록([{id, product_name, type, start, end, days_left, status, total, used}])",
         "1) 본인 이용권만 조회\n2) 활성 상단 (만료 임박 순)\n3) 만료: 최근 6개월만\n4) 연장하기: 구매(MA-062) 이동",
         "이용권 없음: '보유한 이용권이 없습니다. 구매하러 가기'",
         "카드 UI + 프로그레스 바\nD-day 뱃지 컴포넌트",
         "FN-006", "기획", None],
        ["MFN-014", "체성분 변화 조회", "내정보", "App", "High",
         "체성분 측정 기록의 시계열 차트 조회 (읽기 전용).",
         "1) 최근 측정값 카드\n2) 시계열 라인 차트 (항목 토글)\n3) 기간: 1/3/6개월/전체\n4) 목표 달성률 게이지\n5) 측정 기록 리스트",
         "MA-052",
         "회원 ID(자동), 기간(String, 선택)",
         "최근 측정값, 차트 데이터, 달성률, 측정 이력",
         "1) 읽기 전용 (입력은 센터 FC가 CRM에서)\n2) 최소 2건 이상 시 차트 표시\n3) 기록 없음: 안내 메시지",
         "기록 없음: '아직 측정 기록이 없습니다. 센터에서 측정해보세요!'",
         "차트: Victory Native / fl_chart\n게이지: circular progress",
         "FN-007", "기획", None],
        ["MFN-015", "결제 이력 조회", "내정보", "App", "High",
         "결제 내역 리스트와 영수증 조회.",
         "1) 결제 리스트: 날짜/상품명/금액/결제수단\n2) 영수증 상세/PDF 다운로드\n3) 기간별 필터",
         "MA-053",
         "회원 ID(자동), 기간(DateRange, 선택)",
         "결제 이력 목록, 영수증 PDF",
         "1) 본인 결제만 조회\n2) 최근 1년 기본 표시\n3) PDF 공유 가능",
         "결제 없음: '결제 내역이 없습니다'",
         "FlatList + 기간 필터\nPDF 뷰어/공유",
         "FN-006", "기획", None],
        ["MFN-016", "락커 정보 조회", "내정보", "App", "Medium",
         "배정된 락커 번호, 위치, 만료일 조회.",
         "1) 락커 번호, 구역, 만료일\n2) 만료 7일 이내 경고\n3) 미배정 시 안내",
         "MA-054",
         "회원 ID(자동)",
         "락커 정보({number, zone, expire_date})",
         "1) 본인 배정 락커만\n2) 만료 7일 이내 경고 표시",
         "미배정: '배정된 락커가 없습니다'",
         "단순 카드 UI",
         "FN-015", "기획", None],
        ["MFN-017", "쿠폰함 조회", "내정보", "App", "High",
         "보유 쿠폰 목록 조회 + 구매 시 적용.",
         "1) 사용가능/사용완료/만료 탭\n2) 쿠폰 카드: 이름/할인/유효기간\n3) 만료 임박 상단\n4) 바로 사용 버튼\n5) 상세 모달: 적용 조건",
         "MA-055",
         "회원 ID(자동)",
         "쿠폰 목록([{id, name, discount_type, discount_value, valid_from, valid_to, status}])",
         "1) 본인 발급 쿠폰만\n2) 만료 자동 처리\n3) 구매(MA-062)에서 선택 가능",
         "쿠폰 없음: '보유한 쿠폰이 없습니다'",
         "탭 전환 + 쿠폰 카드 디자인\nBottomSheet 상세",
         "FN-025", "기획", None],
        ["MFN-018", "마일리지 조회", "내정보", "App", "Medium",
         "마일리지 잔액 + 적립/사용 이력 조회.",
         "1) 현재 잔액 크게 표시\n2) 적립/사용 이력 리스트\n3) 소멸 예정 마일리지 알림",
         "MA-056",
         "회원 ID(자동)",
         "잔액(Number), 이력([{date, type, amount, balance}])",
         "1) 본인 마일리지만\n2) 잔액 읽기 전용 (조정은 센터에서)\n3) 유효기간 있을 경우 소멸 안내",
         "이력 없음: '마일리지 이력이 없습니다'",
         "잔액 카드 + FlatList",
         "FN-026", "기획", None],
        ["MFN-019", "상품 목록 조회", "구매", "App", "High",
         "이용권/PT/GX 상품을 카테고리별로 조회.",
         "1) 카테고리 탭 (이용권/PT/GX)\n2) 상품 카드: 이름/기간/가격\n3) 판매중 + 앱노출 상품만\n4) 탭 → 상세 이동",
         "MA-060",
         "카테고리(String, 선택)",
         "상품 목록([{id, name, category, price, duration}])",
         "1) branch_id 기준\n2) status='판매중' + kiosk_visible=true\n3) 카드가 기준 표시",
         "상품 없음: '등록된 상품이 없습니다'",
         "카테고리 탭 + 상품 카드 그리드",
         "FN-013", "기획", None],
        ["MFN-020", "상품 상세 조회", "구매", "App", "High",
         "상품 상세 정보 + 구매 버튼.",
         "1) 상품명/카테고리/기간/가격\n2) 상품 설명\n3) 구매하기 버튼 → 결제 이동",
         "MA-061",
         "상품 ID(UUID, 필수)",
         "상품 상세 정보",
         "1) 카드가 기준 가격 표시\n2) 판매 중지 상품 접근 시 안내",
         "판매 중지: '현재 판매하지 않는 상품입니다'",
         "상세 화면 + CTA 버튼",
         "FN-013", "기획", None],
        ["MFN-021", "온라인 결제 (이용권 구매)", "구매", "App", "High",
         "앱에서 쿠폰/마일리지 적용 후 PG 결제로 이용권 구매.",
         "1) 주문 요약 (상품명/기간/원가)\n2) 쿠폰 적용\n3) 마일리지 사용\n4) 최종 결제금액 계산\n5) 결제수단 선택\n6) PG 결제 처리\n7) 이용권 자동 활성화 + 영수증",
         "MA-062",
         "상품ID(UUID), 쿠폰ID(UUID, 선택), 마일리지(Number, 선택), 결제수단(String)",
         "결제 결과, 이용권 정보, 영수증",
         "1) 최종금액 >= 0\n2) 쿠폰: 유효+미사용만\n3) 마일리지: 잔액 이하\n4) 결제 완료 시 이용권 활성화+매출 등록+마일리지 적립\n5) 실패 시 이용권 미생성",
         "결제 실패: '다시 시도해주세요'\n쿠폰 만료: '선택한 쿠폰이 만료되었습니다'\n마일리지 부족: '잔액이 부족합니다'",
         "PG SDK (토스페이먼츠/아임포트)\n결제 결과 콜백 처리\n트랜잭션 보장",
         "FN-027", "기획", None],
        ["MFN-022", "알림 목록 조회", "알림", "App", "Medium",
         "센터 발송 알림 리스트 조회.",
         "1) 알림 리스트 (시간순)\n2) 유형별 아이콘\n3) 읽음/안읽음\n4) 탭 시 관련 화면 이동",
         "MA-070",
         "회원 ID(자동)",
         "알림 목록([{id, type, title, message, read, created_at, deeplink}])",
         "1) 본인 알림만\n2) 최근 3개월 표시\n3) 탭 시 딥링크 이동",
         "알림 없음: '알림이 없습니다'",
         "FlatList + 딥링크 처리",
         "FN-023,FN-024", "기획", None],
        ["MFN-023", "알림 설정", "알림", "App", "Medium",
         "유형별 푸시 알림 ON/OFF 설정.",
         "1) 유형별 토글: 이용권 만료/수업 리마인드/마케팅/공지\n2) 전체 ON/OFF",
         "MA-071",
         "알림설정({type: Boolean})",
         "설정 결과",
         "1) 기본값: 전체 ON\n2) 마케팅 수신 동의 별도 관리\n3) 설정은 서버에 저장",
         "저장 실패: '다시 시도해주세요'",
         "Switch 토글 리스트\nFCM topic 구독/해제",
         "FN-024", "기획", None],
        ["MFN-024", "센터 정보 조회", "더보기", "App", "Low",
         "센터 위치, 영업시간, 연락처 조회.",
         "1) 센터명/주소/영업시간/전화번호\n2) 지도 표시\n3) 전화 걸기 버튼",
         "MA-072",
         "지점 ID(자동)",
         "센터 정보({name, address, hours, phone, lat, lng})",
         "1) 로그인한 지점 정보\n2) 지도: 네이버맵/카카오맵 SDK",
         "조회 실패: '정보를 불러올 수 없습니다'",
         "Map SDK + Linking (전화)",
         "FN-029", "기획", None],
        ["MFN-025", "공지사항 조회", "더보기", "App", "Low",
         "센터 공지사항 목록 + 상세 조회.",
         "1) 공지 리스트 (제목/날짜)\n2) 상세 보기 (본문/이미지)\n3) 중요 공지 상단 고정",
         "MA-073",
         "지점 ID(자동)",
         "공지 목록([{id, title, content, pinned, created_at}])",
         "1) 지점별 공지\n2) 중요(pinned) 상단 고정\n3) 최근 6개월 표시",
         "공지 없음: '공지사항이 없습니다'",
         "FlatList + 상세 화면",
         "신규", "기획", None],
        ["MFN-026", "1:1 문의", "더보기", "App", "Low",
         "문의 작성 + 답변 확인.",
         "1) 카테고리 선택 (이용권/예약/시설/결제/기타)\n2) 제목+본문 작성\n3) 이미지 첨부 (선택)\n4) 문의 이력+답변 확인\n5) 답변 시 푸시 알림",
         "MA-074",
         "카테고리(String), 제목(String), 본문(String), 이미지(File[], 선택)",
         "문의 결과, 문의 이력",
         "1) 본인 문의만 조회\n2) 이미지: 3장 이내, 5MB 이하\n3) 답변 등록 시 FCM 푸시",
         "작성 실패: '다시 시도해주세요'",
         "Form + ImagePicker + FlatList",
         "신규", "기획", None],
        ["MFN-027", "구독/자동결제 관리", "더보기", "App", "Medium",
         "정기 구독 상태 조회 + 자동결제 관리 + 해지.",
         "1) 구독 상품+상태 (활성/일시정지)\n2) 다음 결제일\n3) 결제 카드 변경\n4) 해지 (즉시/월말)\n5) 환불 예상금액 표시",
         "MA-075",
         "회원 ID(자동)",
         "구독 정보({product, status, next_payment, card})",
         "1) 본인 구독만\n2) 해지: 즉시(환불)/월말(기간 종료 후)\n3) 환불: 남은 기간 비례\n4) 카드 변경: PG 토큰 갱신",
         "구독 없음: '구독 중인 상품이 없습니다'\n해지 확인: '정말 해지하시겠습니까?'",
         "구독 카드 UI\n해지 BottomSheet\nPG 카드 변경",
         "FN-028", "기획", None],
        ["MFN-028", "비밀번호 변경", "설정", "App", "Medium",
         "현재 비밀번호 확인 후 새 비밀번호로 변경.",
         "1) 현재 비밀번호 입력\n2) 새 비밀번호+확인\n3) 규칙 실시간 검증 (8자/영문/숫자/특수)\n4) 변경 완료 시 재로그인",
         "MA-076",
         "현재비밀번호(String), 새비밀번호(String), 확인(String)",
         "변경 결과",
         "1) 현재 비밀번호 일치 확인\n2) 이전 3개 비밀번호 재사용 금지\n3) 변경 후 세션 종료 → 재로그인",
         "현재 비밀번호 불일치: '현재 비밀번호가 올바르지 않습니다'\n재사용: '이전에 사용한 비밀번호입니다'",
         "Supabase Auth updateUser\nSecureStorage 토큰 삭제",
         "FN-001", "기획", None],
        ["MFN-029", "로그아웃", "설정", "App", "Critical",
         "로그아웃 확인 후 세션 종료.",
         "1) 로그아웃 확인 팝업\n2) 토큰 삭제 + 세션 종료\n3) 로그인 화면으로 이동",
         "MA-077",
         "없음",
         "없음",
         "1) 로컬 토큰 삭제\n2) Supabase Auth signOut\n3) FCM 토큰 해제",
         "없음",
         "Supabase signOut\nSecureStorage clear\nFCM unregister",
         "FN-001", "기획", None],
        ["MFN-030", "푸시 알림 수신", "시스템", "App", "Medium",
         "CRM 자동 알림 → 회원 앱 푸시 알림 전송.",
         "유형:\n1) 이용권 만료 D-7/D-1\n2) 수업 1시간 전 리마인드\n3) 생일 축하 (쿠폰 포함)\n4) 장기 미출석 (30일)\n5) 결제 완료\n6) 예약 확정\n7) 문의 답변",
         "-",
         "FCM 토큰(자동)",
         "푸시 알림 페이로드({title, body, type, deeplink})",
         "1) FCM 서버 → 클라이언트\n2) 알림 유형별 딥링크\n3) 알림 설정(MFN-023) 기준 필터\n4) 마케팅 수신 동의 확인",
         "FCM 전송 실패: 서버 로그 기록",
         "FCM (Firebase Cloud Messaging)\n딥링크 처리 (navigation)",
         "FN-024", "기획", None],
    ]

    for row in functions:
        ws.append(row)

    style_body(ws)
    apply_priority_fill(ws, 5)  # 우선순위 컬럼
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 4. API명세서 시트
# ══════════════════════════════════════════════════════════
def create_api_sheet(wb):
    ws = wb.create_sheet("API명세서")
    headers = [
        "API ID", "API명", "Method", "엔드포인트", "설명",
        "Request Headers", "Request Body", "Response (성공)", "Response (실패)",
        "인증 필요", "화면 ID", "기능 ID", "상태", "비고"
    ]
    ws.append(headers)
    style_header(ws)

    apis = [
        ["MAPI-001", "회원 로그인", "POST", "/api/v1/member/auth/login",
         "연락처+비밀번호 회원 로그인",
         "Content-Type: application/json",
         '{"phone": "01012345678", "password": "string"}',
         '{"token": "jwt", "member": {id, name, phone, branchId}}',
         '{"error": "INVALID_CREDENTIALS", "message": "연락처 또는 비밀번호를 확인해주세요"}',
         "N", "MA-001", "MFN-001", "기획", None],
        ["MAPI-002", "앱 연동 (SMS 인증 요청)", "POST", "/api/v1/member/auth/verify-request",
         "회원 본인확인 SMS 인증번호 발송",
         "Content-Type: application/json",
         '{"name": "홍길동", "phone": "01012345678"}',
         '{"success": true, "expires_in": 180}',
         '{"error": "MEMBER_NOT_FOUND", "message": "등록된 회원 정보가 없습니다"}',
         "N", "MA-002", "MFN-002", "기획", None],
        ["MAPI-003", "앱 연동 (SMS 인증 확인)", "POST", "/api/v1/member/auth/verify-confirm",
         "SMS 인증번호 확인 + 비밀번호 설정",
         "Content-Type: application/json",
         '{"phone": "01012345678", "code": "123456", "password": "newPassword1!"}',
         '{"success": true, "token": "jwt", "member": {id, name}}',
         '{"error": "INVALID_CODE", "message": "인증번호가 올바르지 않습니다"}',
         "N", "MA-002", "MFN-002", "기획", None],
        ["MAPI-004", "홈 대시보드", "GET", "/api/v1/member/dashboard",
         "회원 홈 대시보드 데이터 (이용권/예약/출석/공지)",
         "Authorization: Bearer {token}",
         None,
         '{"memberships": [...], "today_reservations": [...], "monthly_attendance": 12, "notice": {...}, "coupon_count": 3, "mileage": 5000}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-010", "MFN-004", "기획", "병렬 집계 API"],
        ["MAPI-005", "QR 체크인 토큰 생성", "POST", "/api/v1/member/checkin/qr-token",
         "QR 체크인용 1회성 토큰 생성 (60초 유효)",
         "Authorization: Bearer {token}",
         None,
         '{"qr_data": "encrypted_string", "expires_at": "2026-03-17T10:01:00Z"}',
         '{"error": "NO_ACTIVE_MEMBERSHIP", "message": "활성 이용권이 없습니다"}',
         "Y", "MA-030", "MFN-005", "기획", "HMAC 서명 포함"],
        ["MAPI-006", "출석 이력 조회", "GET", "/api/v1/member/attendance?month=2026-03",
         "월별 출석 이력 + 통계",
         "Authorization: Bearer {token}",
         None,
         '{"attendances": [{date, time, type}], "stats": {total, avg_interval}}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-031", "MFN-006", "기획", None],
        ["MAPI-007", "수업 목록 조회", "GET", "/api/v1/member/classes?date=2026-03-17&type=PT",
         "날짜별 예약 가능 수업 목록",
         "Authorization: Bearer {token}",
         None,
         '{"classes": [{id, title, type, time, trainer, room, capacity, count, is_reserved}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-040", "MFN-007", "기획", None],
        ["MAPI-008", "수업 예약", "POST", "/api/v1/member/reservations",
         "수업 예약 생성",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{"class_id": "uuid"}',
         '{"reservation": {id, class_id, status}, "remaining_count": 8}',
         '{"error": "CAPACITY_FULL", "message": "정원이 마감되었습니다"}',
         "Y", "MA-041", "MFN-008", "기획", None],
        ["MAPI-009", "수업 예약 취소", "DELETE", "/api/v1/member/reservations/:id",
         "예약 취소",
         "Authorization: Bearer {token}",
         '{"cancel_reason": "개인사정"}',
         '{"success": true, "remaining_count": 9}',
         '{"error": "CANCEL_NOT_ALLOWED", "message": "수업이 이미 시작되었습니다"}',
         "Y", "MA-041", "MFN-009", "기획", None],
        ["MAPI-010", "내 예약 목록", "GET", "/api/v1/member/reservations?status=upcoming",
         "예정/완료/취소 예약 조회",
         "Authorization: Bearer {token}",
         None,
         '{"reservations": [{id, class_name, date, time, trainer, status}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-042", "MFN-010", "기획", None],
        ["MAPI-011", "프로필 조회", "GET", "/api/v1/member/profile",
         "회원 프로필 정보 조회",
         "Authorization: Bearer {token}",
         None,
         '{"member": {id, name, phone, email, birthday, photo_url, fc_name, member_no, joined_at}}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-050", "MFN-011", "기획", None],
        ["MAPI-012", "프로필 수정", "PATCH", "/api/v1/member/profile",
         "이메일/주소/사진 수정",
         "Authorization: Bearer {token}\nContent-Type: multipart/form-data",
         '{"email": "new@email.com", "address": "서울시...", "photo": File}',
         '{"member": {updated fields}}',
         '{"error": "INVALID_EMAIL"}',
         "Y", "MA-050", "MFN-012", "기획", None],
        ["MAPI-013", "이용권 목록", "GET", "/api/v1/member/memberships",
         "활성/만료 이용권 조회",
         "Authorization: Bearer {token}",
         None,
         '{"memberships": [{id, product_name, type, start, end, days_left, status, total, used}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-051", "MFN-013", "기획", None],
        ["MAPI-014", "체성분 이력", "GET", "/api/v1/member/body-data?period=3m",
         "체성분 측정 기록 + 차트 데이터",
         "Authorization: Bearer {token}",
         None,
         '{"latest": {weight, muscle, fat, bmi}, "history": [{date, ...}], "goal_rate": 65}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-052", "MFN-014", "기획", None],
        ["MAPI-015", "결제 이력", "GET", "/api/v1/member/payments",
         "결제 내역 + 영수증",
         "Authorization: Bearer {token}",
         None,
         '{"payments": [{id, date, product, amount, method, receipt_url}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-053", "MFN-015", "기획", None],
        ["MAPI-016", "락커 정보", "GET", "/api/v1/member/locker",
         "배정된 락커 조회",
         "Authorization: Bearer {token}",
         None,
         '{"locker": {number, zone, expire_date} | null}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-054", "MFN-016", "기획", None],
        ["MAPI-017", "쿠폰 목록", "GET", "/api/v1/member/coupons",
         "보유 쿠폰 조회",
         "Authorization: Bearer {token}",
         None,
         '{"coupons": [{id, name, discount_type, value, valid_from, valid_to, status}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-055", "MFN-017", "기획", None],
        ["MAPI-018", "마일리지 조회", "GET", "/api/v1/member/mileage",
         "잔액 + 이력",
         "Authorization: Bearer {token}",
         None,
         '{"balance": 5000, "history": [{date, type, amount, balance}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-056", "MFN-018", "기획", None],
        ["MAPI-019", "상품 목록 (앱용)", "GET", "/api/v1/member/products?category=이용권",
         "판매 중 상품 목록 (앱 노출용)",
         "Authorization: Bearer {token}",
         None,
         '{"products": [{id, name, category, price, duration, description}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-060", "MFN-019", "기획", None],
        ["MAPI-020", "온라인 결제 (주문 생성)", "POST", "/api/v1/member/orders",
         "주문 생성 + PG 결제 요청",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{"product_id": "uuid", "coupon_id": "uuid|null", "mileage": 0, "payment_method": "card"}',
         '{"order_id": "uuid", "pg_payment_key": "...", "amount": 95000}',
         '{"error": "COUPON_EXPIRED", "message": "쿠폰이 만료되었습니다"}',
         "Y", "MA-062", "MFN-021", "기획", None],
        ["MAPI-021", "결제 완료 확인", "POST", "/api/v1/member/orders/:id/confirm",
         "PG 결제 완료 → 이용권 활성화",
         "Authorization: Bearer {token}\nContent-Type: application/json",
         '{"pg_payment_key": "...", "pg_order_id": "..."}',
         '{"success": true, "membership": {id, start, end}, "receipt": {...}}',
         '{"error": "PAYMENT_FAILED"}',
         "Y", "MA-062", "MFN-021", "기획", "트랜잭션 보장"],
        ["MAPI-022", "알림 목록", "GET", "/api/v1/member/notifications",
         "알림 리스트",
         "Authorization: Bearer {token}",
         None,
         '{"notifications": [{id, type, title, message, read, created_at, deeplink}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-070", "MFN-022", "기획", None],
        ["MAPI-023", "알림 설정 조회/수정", "GET/PATCH", "/api/v1/member/notification-settings",
         "알림 유형별 ON/OFF",
         "Authorization: Bearer {token}",
         '{"membership_expire": true, "class_remind": true, "marketing": false, "notice": true}',
         '{"settings": {...}}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-071", "MFN-023", "기획", None],
        ["MAPI-024", "공지사항 목록", "GET", "/api/v1/member/notices",
         "센터 공지사항",
         "Authorization: Bearer {token}",
         None,
         '{"notices": [{id, title, content, pinned, created_at}]}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-073", "MFN-025", "기획", None],
        ["MAPI-025", "1:1 문의 CRUD", "GET/POST", "/api/v1/member/inquiries",
         "문의 작성 + 이력 조회",
         "Authorization: Bearer {token}\nContent-Type: multipart/form-data",
         '{"category": "이용권", "title": "...", "content": "...", "images": [File]}',
         '{"inquiry": {id, status, answer}}',
         '{"error": "UNAUTHORIZED"}',
         "Y", "MA-074", "MFN-026", "기획", None],
        ["MAPI-026", "구독 관리", "GET/PATCH/DELETE", "/api/v1/member/subscriptions",
         "구독 조회/카드 변경/해지",
         "Authorization: Bearer {token}",
         '{"action": "cancel", "cancel_type": "immediate|end_of_period"}',
         '{"subscription": {product, status, next_payment, refund_amount}}',
         '{"error": "NO_SUBSCRIPTION"}',
         "Y", "MA-075", "MFN-027", "기획", None],
    ]

    for row in apis:
        ws.append(row)

    style_body(ws)
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 5. 비즈니스정책 시트
# ══════════════════════════════════════════════════════════
def create_policy_sheet(wb):
    ws = wb.create_sheet("비즈니스정책")
    headers = [
        "정책 ID", "정책 영역", "정책명", "정책 설명",
        "관련 화면", "관련 기능", "우선순위", "상태"
    ]
    ws.append(headers)
    style_header(ws)

    policies = [
        ["MBP-001", "인증", "회원 앱 비밀번호 정책",
         "최소 8자, 영문+숫자+특수문자 조합 필수. 이전 3개 비밀번호 재사용 금지.",
         "MA-001,MA-002,MA-076", "MFN-001,MFN-002,MFN-028", "Critical", "확정"],
        ["MBP-002", "인증", "계정 잠금 정책",
         "5회 연속 로그인 실패 시 30분 잠금. SMS 인증으로 잠금 해제 가능.",
         "MA-001", "MFN-001", "Critical", "확정"],
        ["MBP-003", "인증", "앱 연동 정책",
         "센터 CRM에 등록된 회원만 앱 연동 가능. 이름+연락처+SMS 인증으로 본인확인. 1회원 1계정.",
         "MA-002", "MFN-002", "Critical", "확정"],
        ["MBP-004", "출석", "QR 체크인 정책",
         "QR코드 60초 유효. 활성 이용권 필수. HMAC 서명 검증. 당일 중복 체크인 허용.",
         "MA-030", "MFN-005", "Critical", "확정"],
        ["MBP-005", "예약", "수업 예약 정책",
         "수업 시작 1시간 전까지 예약 가능. 동일 시간 중복 예약 불가. PT 잔여 횟수 차감.",
         "MA-040,MA-041", "MFN-007,MFN-008", "Critical", "확정"],
        ["MBP-006", "예약", "예약 취소 정책",
         "24시간 전: 자유 취소 (PT 횟수 복원). 24시간 이내: 경고 후 취소 가능 (PT 횟수 차감 가능). 수업 시작 후: 취소 불가.",
         "MA-041", "MFN-009", "Critical", "확정"],
        ["MBP-007", "결제", "온라인 결제 정책",
         "PG사(토스페이먼츠/아임포트) 연동. 쿠폰+마일리지 복합 할인 가능. 결제 완료 시 이용권 즉시 활성화. 결제 실패 시 이용권 미생성.",
         "MA-062", "MFN-021", "High", "확정"],
        ["MBP-008", "결제", "마일리지 적립 정책",
         "결제 완료 시 결제액의 1% 자동 적립 (센터 설정에 따라). 마일리지 유효기간: 적립 후 1년.",
         "MA-056,MA-062", "MFN-018,MFN-021", "Medium", "확정"],
        ["MBP-009", "구독", "구독 해지/환불 정책",
         "즉시 해지: 남은 기간 비례 환불. 월말 해지: 현재 기간 종료 후 해지 (환불 없음). 전액 환불: 결제 후 7일 이내+미사용 시.",
         "MA-075", "MFN-027", "Medium", "확정"],
        ["MBP-010", "데이터", "회원 데이터 접근 정책",
         "회원은 본인 데이터만 접근 가능 (RLS: member_id = auth.uid()). 관리자 데이터(매출/직원/설정) 접근 불가. 이름/연락처 변경은 센터에서만.",
         "전체", "전체", "Critical", "확정"],
        ["MBP-011", "알림", "푸시 알림 정책",
         "이용권 만료 D-7/D-1, 수업 1시간 전 리마인드, 생일 축하, 장기 미출석 30일. 마케팅 수신 동의 별도 관리.",
         "MA-070,MA-071", "MFN-022,MFN-023,MFN-030", "Medium", "확정"],
        ["MBP-012", "프로필", "프로필 수정 제한 정책",
         "이름/연락처는 센터에서만 변경 가능 (회원 앱에서 읽기 전용). 이메일/주소/사진은 앱에서 수정 가능. 프로필 사진: 5MB 이하, JPG/PNG.",
         "MA-050", "MFN-011,MFN-012", "High", "확정"],
    ]

    for row in policies:
        ws.append(row)

    style_body(ws)
    apply_priority_fill(ws, 7)  # 우선순위 컬럼
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 6. 대시보드 시트
# ══════════════════════════════════════════════════════════
def create_dashboard_sheet(wb):
    ws = wb.create_sheet("대시보드")
    ws.append(["스포짐 회원앱 기획 문서 대시보드"])
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=16, color="2563EB")

    ws.append([])
    ws.append(["문서 작성 현황"])
    ws["A3"].font = Font(name="맑은 고딕", bold=True, size=12)

    ws.append([])
    stats_headers = ["항목", "개수", "상태", "비고"]
    ws.append(stats_headers)
    style_header(ws, row=5)

    stats = [
        ["화면 수", 27, "기획 완료", "MA-001 ~ MA-077"],
        ["기능 수", 30, "기획 완료", "MFN-001 ~ MFN-030"],
        ["사용자 스토리 수", 26, "기획 완료", "MUS-001 ~ MUS-026"],
        ["API 엔드포인트 수", 26, "기획 완료", "MAPI-001 ~ MAPI-026"],
        ["비즈니스 정책 수", 12, "확정", "MBP-001 ~ MBP-012"],
    ]

    for row in stats:
        ws.append(row)

    style_body(ws, start_row=6)

    ws.append([])
    ws.append(["우선순위별 현황"])
    ws["A12"].font = Font(name="맑은 고딕", bold=True, size=12)
    ws.append([])

    pri_headers = ["우선순위", "화면 수", "기능 수", "스토리 수"]
    ws.append(pri_headers)
    style_header(ws, row=14)

    priorities = [
        ["Critical", 10, 10, 8],
        ["High", 9, 12, 12],
        ["Medium", 6, 6, 7],
        ["Low", 3, 3, 3],
    ]
    for row in priorities:
        r = ws.append(row)

    style_body(ws, start_row=15)

    ws.append([])
    ws.append(["개발 로드맵"])
    ws["A20"].font = Font(name="맑은 고딕", bold=True, size=12)
    ws.append([])

    road_headers = ["Phase", "범위", "화면 수", "기능 수", "우선순위"]
    ws.append(road_headers)
    style_header(ws, row=22)

    roadmap = [
        ["Phase 1: MVP", "로그인/홈/출석/예약/이용권", 10, 10, "Critical"],
        ["Phase 2: 확장", "프로필/체성분/결제/구매/쿠폰", 9, 12, "High"],
        ["Phase 3: 부가", "마일리지/알림/구독/설정", 6, 6, "Medium"],
        ["Phase 4: 추가", "센터정보/공지/1:1문의", 3, 3, "Low"],
    ]
    for row in roadmap:
        ws.append(row)

    style_body(ws, start_row=23)
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 7. UI 요소 상세 시트
# ══════════════════════════════════════════════════════════
def create_ui_sheet(wb):
    ws = wb.create_sheet("UI 요소 상세")
    headers = [
        "요소 ID", "화면 ID", "요소명", "요소 유형", "요소 설명",
        "인터랙션 정의", "기능 ID", "API ID",
        "Web 구현", "App 구현", "유효성 검사", "에러 메시지",
        "접근성", "상태", "담당자",
        "[자동] 화면명", "[자동] 기능명", "[자동] API 엔드포인트",
        "비고", "내부 구성", "크기/간격", "상태 목록",
        "호버/클릭 피드백", "소속 섹션", "정렬 순서", "조건부 표시", "관련 디자인토큰"
    ]
    ws.append(headers)
    style_header(ws)

    ui = [
        # ── MA-001 로그인 ──
        ["▼ MA-001", "로그인 (5개 요소)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["MUI-001", "MA-001", "연락처 입력", "Input", "회원 연락처(전화번호) 입력 필드",
         "포커스 시 키패드 (숫자), 자동 하이픈 포매팅 (010-1234-5678)", "MFN-001", "MAPI-001",
         None, "TextInput keyboardType=phone-pad, autoFormat",
         "전화번호 형식 검증 (010-xxxx-xxxx)", "'올바른 연락처를 입력해주세요'",
         "aria-label='연락처'", "기획", "앱개발팀", "로그인", "회원 로그인", "/api/v1/member/auth/login",
         None, "라벨('연락처') + Input(phone) + Phone아이콘", "w:100%, h:48px, p:12px", "default / focus(border-primary) / error(border-red) / filled",
         "focus: border-primary, blur: 에러 시 border-red", "폼 영역", "1", "항상", "MTK-300,MTK-304"],
        ["MUI-002", "MA-001", "비밀번호 입력", "Input", "비밀번호 입력 필드 + 표시/숨김 토글",
         "Eye 아이콘 탭으로 마스킹 토글", "MFN-001", "MAPI-001",
         None, "TextInput secureTextEntry + Eye 토글 버튼",
         "8자 이상", "'비밀번호를 입력해주세요'",
         "aria-label='비밀번호'", "기획", "앱개발팀", "로그인", "회원 로그인", "/api/v1/member/auth/login",
         None, "라벨('비밀번호') + Input(password) + Lock아이콘 + Eye토글", "w:100%, h:48px", "default / focus / error / filled",
         "Eye: 탭 시 secureTextEntry 토글", "폼 영역", "2", "항상", "MTK-300,MTK-304"],
        ["MUI-003", "MA-001", "자동 로그인 체크", "Checkbox", "자동 로그인 ON/OFF 체크박스",
         "체크 시 refresh token 30일 유지", "MFN-001", None,
         None, "Checkbox + 라벨('자동 로그인')",
         None, None,
         "aria-label='자동 로그인'", "기획", "앱개발팀", "로그인", "회원 로그인", None,
         None, "체크박스 + '자동 로그인' 텍스트", "h:20px", "unchecked / checked",
         "탭: 토글", "폼 영역", "3", "항상", None],
        ["MUI-004", "MA-001", "로그인 버튼", "Button", "로그인 실행 CTA 버튼",
         "탭 → API 호출 → 성공 시 홈 이동, 실패 시 에러 표시", "MFN-001", "MAPI-001",
         None, "TouchableOpacity + ActivityIndicator (로딩 시)",
         "연락처+비밀번호 모두 입력 시 활성화", None,
         "aria-label='로그인'", "기획", "앱개발팀", "로그인", "회원 로그인", "/api/v1/member/auth/login",
         None, "'로그인' 텍스트 (Bold, White)", "w:100%, h:52px, radius:10px", "default(primary-500) / pressed(primary-600) / disabled(gray-300) / loading",
         "pressed: 배경 어둡게 + Haptic, loading: ActivityIndicator", "폼 영역", "4", "항상", "MTK-301,MTK-305,MTK-003"],
        ["MUI-005", "MA-001", "하단 링크", "Link", "회원가입 | 비밀번호 찾기 텍스트 링크",
         "탭 → 각 화면 이동", "MFN-002,MFN-003", None,
         None, "TouchableOpacity + Text (primary-500)",
         None, None,
         "aria-role='link'", "기획", "앱개발팀", "로그인", None, None,
         None, "'회원가입' | '비밀번호 찾기' 구분자", "h:20px", "default / pressed(underline)",
         "pressed: 밑줄 표시", "하단 영역", "5", "항상", "MTK-003"],
        # ── MA-010 홈 대시보드 ──
        ["▼ MA-010", "홈 대시보드 (8개 요소)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["MUI-017", "MA-010", "인사말 카드", "Card", "회원 이름 + 프로필 사진 인사말",
         "프로필 사진 탭 → MA-050 이동", "MFN-004", "MAPI-004",
         None, "View + Image(avatar) + Text",
         None, None,
         "aria-label='인사말'", "기획", "앱개발팀", "홈 대시보드", "홈 대시보드 조회", "/api/v1/member/dashboard",
         None, "Avatar(40px) + '안녕하세요, {이름}님!' (heading-xl)", "w:100%, p:16px", "default",
         "사진 탭: MA-050 이동", "인사말 섹션", "1", "항상", "MTK-100,MTK-308"],
        ["MUI-018", "MA-010", "이용권 D-day 카드", "Card", "활성 이용권 + 만료 D-day 뱃지",
         "카드 탭 → MA-051 이동", "MFN-004", "MAPI-004",
         None, "TouchableOpacity + Card 컴포넌트",
         None, None,
         "이용권 상태 전달", "기획", "앱개발팀", "홈 대시보드", "홈 대시보드 조회", "/api/v1/member/dashboard",
         None, "상품명 + D-day 뱃지 (D-30:초록/D-7:주황/D-1:빨강)\n시작~종료일 텍스트\nPT: 잔여횟수 프로그레스", "w:100%, h:auto, p:16px, radius:12px", "default / pressed / 이용권없음",
         "탭: MA-051 이동", "이용권 섹션", "2", "항상", "MTK-303,MTK-010,MTK-011,MTK-012"],
        ["MUI-019", "MA-010", "오늘 예약 카드", "Card", "오늘 예약된 PT/GX 수업 요약",
         "카드 탭 → MA-042 이동", "MFN-004", "MAPI-004",
         None, "TouchableOpacity + Card",
         None, None,
         "예약 정보 전달", "기획", "앱개발팀", "홈 대시보드", "홈 대시보드 조회", "/api/v1/member/dashboard",
         None, "수업명 + 시간 + 트레이너명\n복수 예약: 스크롤 가능 리스트", "w:100%, p:16px, radius:12px", "default / pressed / 예약없음(회색)",
         "탭: MA-042 이동", "예약 섹션", "3", "항상", "MTK-303"],
        ["MUI-020", "MA-010", "출석 현황 카드", "Card", "이번 달 출석 횟수 + 미니 캘린더",
         "카드 탭 → MA-031 이동", "MFN-004", "MAPI-004",
         None, "TouchableOpacity + Card + 미니 캘린더 그리드",
         None, None,
         "출석 횟수 전달", "기획", "앱개발팀", "홈 대시보드", "홈 대시보드 조회", "/api/v1/member/dashboard",
         None, "'이번 달 N회 출석' + 7×5 미니 캘린더 도트", "w:100%, p:16px, radius:12px", "default / pressed / 출석없음",
         "탭: MA-031 이동", "출석 섹션", "4", "항상", "MTK-303,MTK-020,MTK-021,MTK-022"],
        ["MUI-021", "MA-010", "공지 배너", "Banner", "최신 공지사항 1건 슬라이드 배너",
         "탭 → MA-073 이동", "MFN-004", "MAPI-004",
         None, "TouchableOpacity + 배너 카드",
         None, None,
         None, "기획", "앱개발팀", "홈 대시보드", "홈 대시보드 조회", "/api/v1/member/dashboard",
         None, "공지 제목 (1줄 ellipsis) + 날짜", "w:100%, h:60px, radius:8px", "default / pressed / 공지없음(미표시)",
         "탭: MA-073 이동", "공지 섹션", "5", "공지 있을 때만", "MTK-303"],
        ["MUI-022", "MA-010", "혜택 요약", "Card", "쿠폰 N장 | 마일리지 N원 요약 바",
         "쿠폰 탭 → MA-055, 마일리지 탭 → MA-056", "MFN-004", "MAPI-004",
         None, "View + 두 영역 분할",
         None, None,
         None, "기획", "앱개발팀", "홈 대시보드", "홈 대시보드 조회", "/api/v1/member/dashboard",
         None, "'쿠폰 N장' | '마일리지 N원' 2분할", "w:100%, h:48px", "default / pressed",
         "각 영역 탭: 해당 화면 이동", "혜택 섹션", "6", "항상", None],
        ["MUI-023", "MA-010", "QR FAB 버튼", "FAB", "QR 체크인 플로팅 액션 버튼",
         "탭 → MA-030 이동", "MFN-005", None,
         None, "TouchableOpacity + QR아이콘 (absolute 하단 우측)",
         None, None,
         "aria-label='QR 체크인'", "기획", "앱개발팀", "홈 대시보드", "QR 체크인", None,
         None, "QR아이콘 (White) + 원형 배경 (primary-500)", "56×56px, radius:28px, shadow-md", "default / pressed(primary-600)",
         "pressed: scale(0.95) + Haptic", "플로팅", "7", "항상", "MTK-306,MTK-003,MTK-401"],
        ["MUI-024", "MA-010", "Skeleton UI", "Skeleton", "데이터 로딩 중 Skeleton 플레이스홀더",
         "데이터 로드 완료 시 실제 컨텐츠로 교체", "MFN-004", None,
         None, "Skeleton 컴포넌트 (shimmer 애니메이션)",
         None, None,
         "aria-hidden=true", "기획", "앱개발팀", "홈 대시보드", "홈 대시보드 조회", None,
         None, "각 카드 영역 Skeleton (회색 블록 + shimmer)", "카드별 동일 크기", "loading → loaded",
         None, "전체", "0", "로딩 중에만", "MTK-031,MTK-501"],
        # ── MA-030 QR 체크인 ──
        ["▼ MA-030", "QR 체크인 (4개 요소)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["MUI-025", "MA-030", "QR코드 뷰", "QRCode", "회원 식별 QR코드 (60초 자동 갱신)",
         "60초마다 자동 갱신, 화면 밝기 자동 최대", "MFN-005", "MAPI-005",
         None, "QRCode 라이브러리 + 밝기 API",
         None, None,
         "aria-label='QR 체크인 코드'", "기획", "앱개발팀", "QR 체크인", "QR 체크인", "/api/v1/member/checkin/qr-token",
         None, "QR코드 이미지 (240×240) + 하단 타이머", "240×240px, 중앙 정렬", "generating / active / expired(자동갱신)",
         None, "QR 영역", "1", "이용권 있을 때", "MTK-309"],
        ["MUI-026", "MA-030", "갱신 타이머", "Timer", "QR 유효시간 카운트다운 (60초)",
         "0초 도달 시 QR 자동 갱신", "MFN-005", "MAPI-005",
         None, "Text + setInterval (1초)",
         None, None,
         "aria-live='polite'", "기획", "앱개발팀", "QR 체크인", "QR 체크인", None,
         None, "원형 프로그레스 + 남은 초 텍스트", "64×64px", "counting / refreshing",
         None, "QR 영역", "2", "이용권 있을 때", "MTK-105"],
        ["MUI-027", "MA-030", "체크인 완료 오버레이", "Overlay", "체크인 성공 애니메이션",
         "체크 아이콘 애니메이션 + 진동 + 시간 표시", "MFN-005", None,
         None, "Modal + Lottie/Animated + Haptic",
         None, None,
         "aria-label='출석 완료'", "기획", "앱개발팀", "QR 체크인", "QR 체크인", None,
         None, "체크마크 애니메이션 + '출석 완료!' + HH:MM + 확인 버튼", "전체 오버레이", "hidden / visible(애니메이션)",
         "확인 버튼: 오버레이 닫기", "오버레이", "3", "체크인 성공 시", "MTK-010,MTK-502"],
        ["MUI-028", "MA-030", "이용권 없음 안내", "EmptyState", "활성 이용권 없을 때 안내 카드",
         "'이용권 구매' 버튼 → MA-060 이동", "MFN-005", None,
         None, "View + 일러스트 + Text + Button",
         None, None,
         None, "기획", "앱개발팀", "QR 체크인", "QR 체크인", None,
         None, "일러스트 + '활성 이용권이 없습니다' + '이용권 구매하기' 버튼", "w:100%, p:24px", "default",
         "버튼 탭: MA-060 이동", "QR 영역", "1", "이용권 없을 때", "MTK-012"],
        # ── MA-040 수업 목록 ──
        ["▼ MA-040", "수업 목록 (6개 요소)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["MUI-033", "MA-040", "주간 날짜 스크롤", "DateScroller", "가로 스크롤 주간 날짜 선택기",
         "좌우 스와이프, 오늘 강조 (primary), 선택 일 하이라이트", "MFN-007", "MAPI-007",
         None, "FlatList horizontal + 날짜 칩",
         None, None,
         "aria-label='날짜 선택'", "기획", "앱개발팀", "수업 목록", "수업 목록 조회", "/api/v1/member/classes",
         None, "요일(월) + 날짜(17) 칩 × 7일 가로 스크롤", "h:64px, 칩:48×56px", "default / selected(primary) / today(primary-100)",
         "탭: 해당 날짜 수업 로드", "날짜 영역", "1", "항상", "MTK-003,MTK-001"],
        ["MUI-034", "MA-040", "PT/GX 필터", "SegmentedControl", "PT/GX 유형 필터 세그먼트",
         "탭 전환 시 수업 목록 필터링", "MFN-007", "MAPI-007",
         None, "SegmentedControl 컴포넌트",
         None, None,
         "aria-label='수업 유형 필터'", "기획", "앱개발팀", "수업 목록", "수업 목록 조회", None,
         None, "'전체' | 'PT' | 'GX' 세그먼트", "w:100%, h:36px", "전체(default) / PT / GX",
         "탭: 애니메이션 슬라이드 전환", "필터 영역", "2", "항상", "MTK-003"],
        ["MUI-035", "MA-040", "트레이너 필터", "Dropdown", "트레이너별 필터 드롭다운",
         "선택 시 해당 트레이너 수업만 표시", "MFN-007", "MAPI-007",
         None, "ActionSheet / BottomSheet 선택",
         None, None,
         "aria-label='트레이너 선택'", "기획", "앱개발팀", "수업 목록", "수업 목록 조회", None,
         None, "'전체 트레이너 ▼' 드롭다운", "w:auto, h:36px", "closed / open(ActionSheet) / selected",
         "탭: ActionSheet 열기", "필터 영역", "3", "항상", None],
        ["MUI-036", "MA-040", "수업 카드", "Card", "수업 정보 카드 (시간/이름/트레이너/잔여석)",
         "카드 탭 → MA-041 수업 상세 이동", "MFN-007", "MAPI-007",
         None, "TouchableOpacity + Card 컴포넌트",
         None, None,
         "수업 정보 읽기 전달", "기획", "앱개발팀", "수업 목록", "수업 목록 조회", "/api/v1/member/classes",
         None, "시간(HH:MM~HH:MM) + 수업명 + 트레이너 + 룸\n잔여석 프로그레스 (5/20)\n뱃지: '마감'(빨강) / '예약됨'(초록)", "w:100%, p:16px, radius:12px, mb:12px", "default / pressed / 마감(opacity:0.5) / 예약됨(border-green)",
         "탭: MA-041 이동 + scale(0.98)", "수업 리스트", "4", "수업 있을 때", "MTK-303,MTK-010,MTK-012"],
        ["MUI-037", "MA-040", "'내 예약' 아이콘", "IconButton", "헤더 우측 내 예약 바로가기 아이콘",
         "탭 → MA-042 내 예약 목록 이동", "MFN-010", None,
         None, "TouchableOpacity + Calendar 아이콘",
         None, None,
         "aria-label='내 예약'", "기획", "앱개발팀", "수업 목록", "내 예약 목록 조회", None,
         None, "Calendar 아이콘 (24px)", "44×44px (터치 영역)", "default / pressed",
         "탭: MA-042 이동", "헤더", "0", "항상", None],
        ["MUI-038", "MA-040", "수업 없음 빈 상태", "EmptyState", "수업이 없을 때 안내",
         None, "MFN-007", None,
         None, "View + 일러스트 + Text",
         None, None,
         None, "기획", "앱개발팀", "수업 목록", "수업 목록 조회", None,
         None, "일러스트 + '예약 가능한 수업이 없습니다'", "w:100%, p:24px, 중앙", "default",
         None, "수업 리스트", "4", "수업 없을 때", None],
        # ── MA-041 수업 상세/예약 ──
        ["▼ MA-041", "수업 상세/예약 (4개 요소)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["MUI-039", "MA-041", "수업 정보 섹션", "Section", "수업 상세 정보 표시 영역",
         None, "MFN-008", "MAPI-007",
         None, "View + 텍스트 그룹",
         None, None,
         None, "기획", "앱개발팀", "수업 상세/예약", "수업 예약", "/api/v1/member/classes",
         None, "수업명(heading-lg) + 유형 뱃지(PT/GX)\n시간/트레이너/룸/잔여석 정보 행", "w:100%, p:16px", "default",
         None, "수업 정보", "1", "항상", "MTK-101"],
        ["MUI-040", "MA-041", "예약하기 CTA 버튼", "Button", "수업 예약 실행 버튼 (하단 고정)",
         "탭 → 예약 확인 BottomSheet 표시", "MFN-008", "MAPI-008",
         None, "TouchableOpacity + 하단 고정",
         "잔여 횟수/잔여석 확인", None,
         "aria-label='예약하기'", "기획", "앱개발팀", "수업 상세/예약", "수업 예약", "/api/v1/member/reservations",
         None, "'예약하기' (Bold, White, primary-500 배경)", "w:calc(100%-32px), h:52px, 하단 16px", "default / pressed / disabled(마감) / loading",
         "pressed: primary-600 + Haptic", "하단 고정", "4", "미예약 수업", "MTK-301,MTK-305,MTK-003"],
        ["MUI-041", "MA-041", "예약 확인 BottomSheet", "BottomSheet", "예약 전 확인 모달",
         "'확인' → API 호출 → 예약 완료, '취소' → 닫기", "MFN-008", "MAPI-008",
         None, "BottomSheet 컴포넌트 (react-native-bottom-sheet)",
         None, None,
         "aria-modal=true", "기획", "앱개발팀", "수업 상세/예약", "수업 예약", "/api/v1/member/reservations",
         None, "수업명/시간 요약\nPT: '잔여 12회 → 11회' 텍스트\n'예약 확인' + '취소' 버튼", "w:100%, maxH:50%, radius-top:16px", "hidden / visible",
         "'예약 확인': primary, '취소': gray outline", "오버레이", "5", "예약하기 탭 시", "MTK-402"],
        ["MUI-042", "MA-041", "예약 취소 버튼", "Button", "이미 예약된 수업의 취소 버튼",
         "탭 → 취소 확인 Dialog → 취소 처리", "MFN-009", "MAPI-009",
         None, "TouchableOpacity (빨강 outline)",
         None, None,
         "aria-label='예약 취소'", "기획", "앱개발팀", "수업 상세/예약", "수업 예약 취소", "/api/v1/member/reservations/:id",
         None, "'예약 취소' (error-500 텍스트, outline 스타일)", "w:calc(100%-32px), h:52px, 하단 16px", "default / pressed / loading",
         "pressed: error-50 배경", "하단 고정", "4", "예약된 수업만", "MTK-301,MTK-012"],
        # ── MA-051 이용권 목록 ──
        ["▼ MA-051", "이용권 목록 (4개 요소)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["MUI-055", "MA-051", "활성/만료 탭", "SegmentedControl", "활성/만료 이용권 탭 전환",
         "탭 전환 시 리스트 필터링", "MFN-013", "MAPI-013",
         None, "SegmentedControl",
         None, None,
         "aria-label='이용권 상태 필터'", "기획", "앱개발팀", "이용권 목록", "이용권 목록 조회", "/api/v1/member/memberships",
         None, "'활성' | '만료' 세그먼트", "w:100%, h:36px", "활성(default) / 만료",
         "탭: 슬라이드 전환", "탭 영역", "1", "항상", "MTK-003"],
        ["MUI-056", "MA-051", "이용권 카드", "Card", "이용권 상세 카드 (D-day + 잔여 횟수)",
         "카드 자체는 탭 불가 (읽기 전용)", "MFN-013", "MAPI-013",
         None, "View + Card 컴포넌트",
         None, None,
         "이용권 상태 전달", "기획", "앱개발팀", "이용권 목록", "이용권 목록 조회", "/api/v1/member/memberships",
         None, "상품명 + 유형 아이콘 + D-day 뱃지\n시작~종료일\nPT/GX: 프로그레스 바 (used/total)\n만료 7일: 빨강 보더", "w:100%, p:16px, radius:12px, mb:12px", "활성(normal) / 만료임박(border-red) / 만료(gray, opacity:0.7)",
         None, "이용권 리스트", "2", "이용권 있을 때", "MTK-303,MTK-010,MTK-011,MTK-012"],
        ["MUI-057", "MA-051", "연장하기 버튼", "Button", "이용권 연장 (구매 이동) 버튼",
         "탭 → MA-062 결제 화면 (동일 상품 선택 상태)", "MFN-021", None,
         None, "TouchableOpacity (primary outline)",
         None, None,
         "aria-label='이용권 연장하기'", "기획", "앱개발팀", "이용권 목록", "온라인 결제", None,
         None, "'연장하기' 텍스트 (primary-500)", "w:auto, h:36px, px:16px, radius:8px", "default / pressed",
         "탭: MA-062 이동", "이용권 카드 내", "3", "만료 7일 이내 활성 이용권", "MTK-302,MTK-003"],
        ["MUI-058", "MA-051", "이용권 없음 빈 상태", "EmptyState", "이용권이 없을 때 안내",
         "'구매하러 가기' 버튼 → MA-060", "MFN-013", None,
         None, "View + 일러스트 + Text + Button",
         None, None,
         None, "기획", "앱개발팀", "이용권 목록", "이용권 목록 조회", None,
         None, "일러스트 + '보유한 이용권이 없습니다' + '구매하러 가기' 버튼", "w:100%, p:24px, 중앙", "default",
         "버튼: MA-060 이동", "이용권 리스트", "2", "이용권 없을 때", None],
        # ── MA-062 결제 ──
        ["▼ MA-062", "결제/구매 (6개 요소)", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["MUI-085", "MA-062", "주문 요약 섹션", "Section", "선택 상품 정보 + 원가 표시",
         None, "MFN-021", "MAPI-020",
         None, "View + 텍스트",
         None, None,
         None, "기획", "앱개발팀", "결제/구매", "온라인 결제", "/api/v1/member/orders",
         None, "상품명 + 카테고리 + 기간\n원가 (heading-md)", "w:100%, p:16px", "default",
         None, "주문 요약", "1", "항상", "MTK-102"],
        ["MUI-086", "MA-062", "쿠폰 선택 버튼", "Button", "쿠폰 적용 선택 (BottomSheet 열기)",
         "탭 → 쿠폰 선택 BottomSheet → 할인 반영", "MFN-021", None,
         None, "TouchableOpacity → BottomSheet (쿠폰 목록)",
         None, None,
         "aria-label='쿠폰 선택'", "기획", "앱개발팀", "결제/구매", "온라인 결제", None,
         None, "'쿠폰 선택' + 선택된 쿠폰명 표시 + '>' 아이콘", "w:100%, h:48px", "미선택 / 선택됨(쿠폰명 표시)",
         "탭: BottomSheet 열기", "할인 영역", "2", "항상", None],
        ["MUI-087", "MA-062", "마일리지 입력", "Input", "마일리지 사용 금액 입력",
         "잔액 표시 + 전액 사용 버튼", "MFN-021", None,
         None, "TextInput number + '전액 사용' 버튼",
         "0 이상, 잔액 이하", "'마일리지 잔액이 부족합니다'",
         "aria-label='마일리지 사용'", "기획", "앱개발팀", "결제/구매", "온라인 결제", None,
         None, "'마일리지 사용' + 잔액 표시 + NumberInput + '전액 사용'", "w:100%, h:48px", "default / focus / error(잔액초과)",
         "전액 사용: 잔액 자동 입력", "할인 영역", "3", "마일리지 > 0일 때", "MTK-300"],
        ["MUI-088", "MA-062", "결제 금액 표시", "Section", "원가 - 쿠폰 - 마일리지 = 최종금액",
         "할인 항목 변경 시 실시간 재계산", "MFN-021", None,
         None, "View + 항목별 금액 텍스트",
         None, None,
         "금액 정보 전달", "기획", "앱개발팀", "결제/구매", "온라인 결제", None,
         None, "원가 / -쿠폰할인 / -마일리지 / ────── / 최종금액(heading-lg, primary)", "w:100%, p:16px", "default",
         None, "금액 영역", "4", "항상", "MTK-101,MTK-003"],
        ["MUI-089", "MA-062", "결제수단 선택", "RadioGroup", "카드/간편결제 선택",
         "선택 변경 시 결제수단 업데이트", "MFN-021", None,
         None, "RadioGroup (카드/카카오페이/네이버페이/토스)",
         "1개 필수 선택", None,
         "aria-label='결제수단 선택'", "기획", "앱개발팀", "결제/구매", "온라인 결제", None,
         None, "각 결제수단 아이콘 + 이름 라디오 카드", "w:100%, 카드:h:56px", "선택됨(border-primary) / 미선택",
         "탭: border-primary 활성화", "결제수단 영역", "5", "항상", "MTK-003"],
        ["MUI-090", "MA-062", "결제하기 CTA 버튼", "Button", "결제 실행 버튼 (하단 고정)",
         "탭 → PG 모듈 호출 → 결제 처리", "MFN-021", "MAPI-020",
         None, "TouchableOpacity + ActivityIndicator",
         "결제수단 선택 필수", None,
         "aria-label='결제하기'", "기획", "앱개발팀", "결제/구매", "온라인 결제", "/api/v1/member/orders",
         None, "'결제하기 {금액}원' (Bold, White)", "w:calc(100%-32px), h:52px, 하단 16px", "default(primary) / pressed / disabled(미선택) / loading(전체 오버레이)",
         "pressed: primary-600, loading: 오버레이 스피너", "하단 고정", "6", "항상", "MTK-301,MTK-003"],
    ]

    for row in ui:
        ws.append(row)

    style_body(ws)
    # 구분 행(▼) 스타일
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value and str(row[0].value).startswith("▼"):
            for cell in row:
                cell.fill = SUBHEADER_FILL
                cell.font = Font(name="맑은 고딕", bold=True, size=10)
    auto_width(ws, max_width=40)


# ══════════════════════════════════════════════════════════
# 8. 화면설계서 시트
# ══════════════════════════════════════════════════════════
def create_screen_sheet(wb):
    ws = wb.create_sheet("화면설계서")
    headers = [
        "화면 ID", "화면명", "플랫폼", "화면 경로", "화면 설명",
        "[자동] UI 요소 목록", "인터랙션 정의", "Web 특이사항", "App 특이사항",
        "진입 경로", "이동 가능 화면", "에러 화면 정의",
        "접근성 체크", "반응형 체크 (Web)", "가로/세로 모드 (App)",
        "디자인 시안 링크", "상태", "담당자",
        "[자동] 관련 기능 수", "[자동] 관련 TC 수", "[자동] 관련 API 수", "비고",
        "[자동] UI 요소 수", "레이아웃 구조", "섹션 순서", "화면 유형",
        "진입 트리거", "Loading 상태", "Empty 상태", "반응형 상세", "고정 요소"
    ]
    ws.append(headers)
    style_header(ws)

    screens = [
        ["MA-001", "로그인", "App", "/login", "연락처+비밀번호 회원 로그인 화면",
         "MUI-001~005", "비밀번호 마스킹 토글, 자동 로그인 체크", None,
         "SecureStorage 토큰 저장, 생체인증 지원 (추후)",
         "앱 최초 실행 / 로그아웃 후 / 토큰 만료 시", "MA-010 (홈), MA-002 (회원가입), MA-003 (비밀번호 찾기)",
         "잘못된 자격증명: '연락처 또는 비밀번호를 확인해주세요'\n계정 잠금: '30분 후 재시도하세요'\n네트워크 오류: '인터넷 연결을 확인해주세요'",
         "VoiceOver/TalkBack 지원, 입력필드 aria-label", None, "세로 모드 고정",
         None, "기획", "앱개발팀", 2, None, 1, None, 5,
         "[배경] 그라데이션 전체 화면\n[로고] 상단 중앙 (스포짐 로고+텍스트)\n[폼 영역] 중앙 카드\n  - 연락처 입력 (Phone 아이콘)\n  - 비밀번호 입력 (Lock 아이콘) + Eye토글\n  - 자동 로그인 체크\n  - 로그인 버튼\n[하단] 회원가입 | 비밀번호 찾기 링크",
         "1:로고 → 2:연락처입력 → 3:비밀번호입력 → 4:자동로그인 → 5:로그인버튼 → 6:하단링크",
         "page (전체화면, 탭바 없음)", "앱 시작 / 로그아웃 / 토큰 만료",
         "로그인 버튼: ActivityIndicator + '로그인 중...'",
         None, "세로 고정, SafeArea 적용", "없음 (단일 화면)"],

        ["MA-002", "앱 연동 (회원가입)", "App", "/register", "센터 등록 회원의 앱 최초 연동 (SMS 인증 3단계 위자드)",
         "MUI-006~012", "단계별 진행 (이름+연락처 → SMS 인증 → 비밀번호 설정)", None,
         "SMS API 연동 (CoolSMS), 인증번호 타이머 표시",
         "MA-001 (로그인 화면) → '회원가입' 링크", "MA-010 (홈, 연동 완료 시), MA-001 (취소 시)",
         "미등록: '등록된 회원 정보가 없습니다'\n이미 연동: '이미 앱에 등록된 회원입니다'\n인증 만료: '인증번호가 만료되었습니다'",
         "VoiceOver/TalkBack, 단계 진행 안내", None, "세로 모드 고정",
         None, "기획", "앱개발팀", 1, None, 2, None, 7,
         "[헤더] 뒤로가기 + '회원가입' 타이틀\n[Stepper] 3단계 진행 인디케이터\n[Step 1] 이름+연락처 입력 폼\n[Step 2] SMS 인증번호 입력 + 타이머(3:00) + 재발송\n[Step 3] 비밀번호 설정 + 규칙 체크리스트\n[하단] 다음/완료 버튼",
         "1:헤더 → 2:Stepper → 3:단계별 폼 → 4:하단 버튼",
         "page (전체, 탭바 없음)", "'회원가입' 링크 탭",
         "SMS 발송: '인증번호 발송 중...'\n연동: ActivityIndicator",
         "없음", "세로 고정, 키보드 어보이던스", "없음"],

        ["MA-003", "비밀번호 찾기", "App", "/forgot-password", "SMS 인증으로 비밀번호 재설정",
         "MUI-013~016", "연락처 입력 → SMS 인증 → 새 비밀번호 설정", None,
         "SMS API, 비밀번호 규칙 실시간 검증",
         "MA-001 → '비밀번호 찾기' 링크", "MA-001 (재설정 완료 후)",
         "미등록: '등록된 정보가 없습니다'\n인증 실패: '인증번호가 올바르지 않습니다'",
         "VoiceOver/TalkBack", None, "세로 고정",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '비밀번호 찾기'\n[폼] 연락처 → SMS 인증 → 새 비밀번호\n[하단] 확인 버튼",
         "1:헤더 → 2:폼 → 3:확인버튼",
         "page", "'비밀번호 찾기' 링크 탭",
         "SMS 발송 중 스피너", None, "세로 고정", "없음"],

        ["MA-010", "홈 대시보드", "App", "/home", "이용권 D-day, 오늘 예약, 출석 현황, 공지 배너, QR 플로팅 버튼",
         "MUI-017~024", "Pull-to-refresh, 카드 탭 시 상세 이동, QR 플로팅 버튼", None,
         "병렬 API 호출 (이용권+예약+출석+공지), Skeleton UI",
         "로그인 성공 후 / 하단탭 '홈' 탭", "MA-030 (QR), MA-051 (이용권), MA-040 (예약), MA-073 (공지)",
         "데이터 로드 실패: '새로고침을 시도해주세요' + 재시도 버튼",
         "카드별 의미 전달, 대비 색상", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 8,
         "[인사말] '안녕하세요, {이름}님!' + 프로필 사진\n[이용권 카드] 상품명 + D-day 뱃지 (빨강/주황/초록)\n[오늘 예약 카드] 수업명/시간/트레이너\n[출석 현황] 이번 달 출석 횟수 + 미니 캘린더\n[공지 배너] 최신 1건 슬라이드\n[혜택 요약] 쿠폰 N장 | 마일리지 N원\n[QR 플로팅] 하단 우측 고정 FAB 버튼",
         "1:인사말 → 2:이용권카드 → 3:오늘예약 → 4:출석현황 → 5:공지배너 → 6:혜택요약",
         "tab (하단 탭 '홈')", "로그인 성공 / 탭 전환",
         "각 섹션: Skeleton 카드 (3개)\nQR 버튼: 항상 표시",
         "이용권 없음: '이용권이 없습니다. 구매하러 가기'\n예약 없음: '오늘 예약이 없습니다'\n출석 없음: '이번 달 출석 기록이 없습니다'",
         "세로 스크롤, SafeArea", "하단 탭바 + QR FAB 버튼"],

        ["MA-030", "QR 체크인", "App", "/checkin", "QR코드 생성/표시로 센터 출석 체크",
         "MUI-025~028", "QR 자동 갱신 (60초), 밝기 자동 최대, 체크인 완료 시 진동+애니메이션", None,
         "QR 라이브러리, 화면 밝기 API, Haptic feedback",
         "홈 QR FAB 버튼 / 하단탭 '출석' / 출석 이력 → 체크인",
         "MA-031 (출석 이력), MA-060 (이용권 구매, 이용권 없을 때)",
         "이용권 없음: '활성 이용권이 없습니다' + 구매 버튼\n이용권 만료: '이용권이 만료되었습니다' + 연장 버튼",
         "QR 코드 대체 텍스트", None, "세로 고정",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] '출석 체크인' 타이틀\n[QR 영역] 큰 QR코드 (중앙) + 갱신 타이머\n[안내] '센터 리더기에 QR을 보여주세요'\n[하단] '출석 이력 보기' 텍스트 버튼\n\n[체크인 완료 오버레이]\n  - 체크마크 애니메이션\n  - '출석 완료!' + 시간 표시\n  - 확인 버튼",
         "1:헤더 → 2:QR코드 → 3:안내문구 → 4:출석이력링크",
         "tab (하단 탭 '출석')", "FAB 탭 / 탭 전환",
         "QR 생성 중: 스피너",
         "이용권 없음: 안내 카드 + 구매 버튼",
         "세로 고정, 화면 밝기 최대", "없음"],

        ["MA-031", "출석 이력", "App", "/attendance", "월별 캘린더 히트맵 + 출석 리스트",
         "MUI-029~032", "좌우 스와이프 월 이동, 날짜 탭 시 상세 리스트, 유형별 색상", None,
         "커스텀 캘린더 히트맵, 무한 스크롤",
         "MA-030 (체크인) → '출석 이력 보기' / 홈 출석 카드 탭", None,
         "데이터 로드 실패: '새로고침을 시도해주세요'",
         "색상 구분 범례 제공", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] '출석 이력' + 월 선택\n[통계 카드] 이번 달 N회 출석\n[캘린더] 월별 히트맵 (일반:파랑/PT:초록/GX:보라)\n[리스트] 선택 날짜 출석 상세 (시간/유형/체크인·아웃)",
         "1:헤더 → 2:통계카드 → 3:캘린더히트맵 → 4:출석리스트",
         "stack (출석 탭 내)", "탭 전환 / 링크 이동",
         "캘린더: Skeleton 그리드\n리스트: Skeleton 행 3개",
         "출석 없음: '이번 달 출석 기록이 없습니다'",
         "세로 스크롤", "헤더 고정"],

        ["MA-040", "수업 목록", "App", "/classes", "날짜별 PT/GX 수업 목록 + 잔여석 + 예약 상태",
         "MUI-033~038", "주간 날짜 스크롤, PT/GX 탭 필터, 수업 카드 탭 → 상세", None,
         "주간 날짜 horizontal FlatList, React Query 캐싱",
         "하단탭 '예약' 탭", "MA-041 (수업 상세), MA-042 (내 예약)",
         "수업 없음: '예약 가능한 수업이 없습니다'\n로드 실패: '다시 시도해주세요'",
         "수업 카드 정보 전달", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 6,
         "[헤더] '수업 예약' + 내 예약 아이콘 버튼\n[날짜 스크롤] 주간 가로 스크롤 (오늘 강조)\n[필터] PT/GX 세그먼트 + 트레이너 필터\n[수업 리스트] 카드형\n  - 시간 / 수업명 / 트레이너 / 룸\n  - 잔여석 프로그레스\n  - '마감' 뱃지 / '예약됨' 뱃지",
         "1:헤더 → 2:날짜스크롤 → 3:필터 → 4:수업리스트",
         "tab (하단 탭 '예약')", "탭 전환",
         "수업 리스트: Skeleton 카드 3개",
         "수업 없음: '예약 가능한 수업이 없습니다' 일러스트",
         "세로 스크롤, 날짜 가로 스크롤", "헤더+날짜 고정, 탭바"],

        ["MA-041", "수업 상세/예약", "App", "/classes/:id", "수업 정보 + 예약/취소 처리",
         "MUI-039~044", "예약하기 → BottomSheet 확인 → 완료, 취소 → 경고 팝업", None,
         "예약 확인 BottomSheet, 디바이스 캘린더 연동",
         "MA-040 → 수업 카드 탭", "MA-042 (내 예약, 예약 후)",
         "횟수 부족: '잔여 횟수가 부족합니다'\n정원 마감: '정원이 마감되었습니다'\n시간 중복: '같은 시간에 이미 예약이 있습니다'",
         "버튼 상태 구분, 에러 안내", None, "세로 고정",
         None, "기획", "앱개발팀", 2, None, 2, None, 6,
         "[헤더] 뒤로가기 + '수업 상세'\n[수업 정보] 수업명/유형/시간/트레이너/룸/잔여석\n[트레이너 프로필] 사진+이름\n[안내] 취소 정책 안내 텍스트\n[하단 고정] '예약하기' 버튼 (또는 '예약 취소')\n\n[예약 확인 BottomSheet]\n  - 수업명/시간\n  - PT: '잔여 12회 → 11회' 표시\n  - 확인/취소 버튼",
         "1:헤더 → 2:수업정보 → 3:트레이너 → 4:안내 → 5:하단버튼",
         "stack", "수업 카드 탭",
         "예약 처리: ActivityIndicator",
         None, "세로 고정", "하단 CTA 버튼 고정"],

        ["MA-042", "내 예약 목록", "App", "/my-reservations", "예정/완료/취소 예약 리스트",
         "MUI-045~048", "탭 전환 (SegmentedControl), 예정 카드에서 취소 가능", None,
         "SegmentedControl, FlatList",
         "MA-040 → 헤더 '내 예약' 아이콘 / 예약 완료 후", None,
         "로드 실패: '다시 시도해주세요'",
         "탭 전환 안내", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '내 예약'\n[탭] 예정 | 완료 | 취소 (SegmentedControl)\n[리스트] 예약 카드\n  - 예정: 날짜/시간/트레이너/룸 + 취소 버튼\n  - 완료: 날짜/시간/트레이너 (회색)\n  - 취소: 날짜/시간/취소사유 (취소선)",
         "1:헤더 → 2:탭 → 3:예약리스트",
         "stack", "'내 예약' 아이콘 탭",
         "리스트: Skeleton 카드 3개",
         "예약 없음: '예약 내역이 없습니다'",
         "세로 스크롤", "헤더+탭 고정"],

        ["MA-050", "프로필", "App", "/profile", "이름/연락처/사진 조회 + 이메일/주소/사진 수정",
         "MUI-049~054", "프로필 사진 탭 → 카메라/갤러리 ActionSheet, 수정 모드 토글", None,
         "ImagePicker (카메라/갤러리), Supabase Storage 업로드",
         "하단탭 '내정보' 탭", "MA-051~056 (내정보 하위), MA-076 (비밀번호 변경)",
         "저장 실패: '다시 시도해주세요'\n사진 용량 초과: '5MB 이하를 선택해주세요'",
         "프로필 사진 대체 텍스트", None, "세로 우선",
         None, "기획", "앱개발팀", 2, None, 1, None, 6,
         "[헤더] '내 정보'\n[프로필 상단] 큰 프로필 사진 + 카메라 아이콘\n  이름, 회원번호, 담당FC\n[메뉴 리스트]\n  - 이용권 → MA-051\n  - 체성분 → MA-052\n  - 결제 이력 → MA-053\n  - 락커 → MA-054\n  - 쿠폰함 → MA-055\n  - 마일리지 → MA-056\n[하단] 수정/비밀번호변경 링크",
         "1:헤더 → 2:프로필상단 → 3:메뉴리스트 → 4:하단링크",
         "tab (하단 탭 '내정보')", "탭 전환",
         "프로필: Skeleton 카드",
         None, "세로 스크롤", "탭바"],

        ["MA-051", "이용권 목록", "App", "/memberships", "활성/만료 이용권 + D-day + 잔여 횟수",
         "MUI-055~058", "활성/만료 탭 전환, 연장하기 → 구매 이동, PT 프로그레스바", None,
         "카드 UI, 프로그레스 바, D-day 뱃지",
         "MA-050 → '이용권' 메뉴 / 홈 이용권 카드 탭", "MA-062 (구매/연장)",
         "로드 실패: '다시 시도해주세요'",
         "D-day 색상 의미 전달", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '이용권'\n[탭] 활성 | 만료 (SegmentedControl)\n[이용권 카드]\n  - 상품명 + 유형 아이콘\n  - 시작일~종료일\n  - D-day 뱃지 (D-30:초록/D-7:주황/D-1:빨강)\n  - PT/GX: 잔여횟수 프로그레스 (8/12)\n  - '연장하기' 버튼 (만료 7일 이내)",
         "1:헤더 → 2:탭 → 3:이용권카드리스트",
         "stack", "메뉴 탭 / 홈 카드 탭",
         "카드 Skeleton 3개",
         "이용권 없음: '보유한 이용권이 없습니다' + '구매하러 가기' 버튼",
         "세로 스크롤", "헤더+탭 고정"],

        ["MA-052", "체성분 변화", "App", "/body-data", "체중/체지방/골격근 시계열 차트 (읽기 전용)",
         "MUI-059~062", "기간 선택 탭, 차트 항목 토글, 가로 스크롤 차트", None,
         "차트 라이브러리 (Victory Native / fl_chart), 게이지 컴포넌트",
         "MA-050 → '체성분' 메뉴", None,
         "로드 실패: '다시 시도해주세요'",
         "차트 대체 텍스트, 색상 범례", None, "세로+가로 차트",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '체성분 변화'\n[최근 측정 카드] 날짜 + 체중/골격근/체지방/BMI\n[기간 탭] 1개월|3개월|6개월|전체\n[차트] 시계열 라인 차트 (항목별 ON/OFF 토글)\n[목표 게이지] 원형 프로그레스 (달성률 %)\n[측정 이력] 리스트 (날짜순)",
         "1:헤더 → 2:최근측정 → 3:기간탭 → 4:차트 → 5:목표게이지 → 6:이력리스트",
         "stack", "메뉴 탭",
         "차트: Skeleton 영역\n이력: Skeleton 행",
         "기록 없음: '아직 측정 기록이 없습니다. 센터에서 측정해보세요!'",
         "세로 스크롤, 차트 가로 스크롤", "헤더 고정"],

        ["MA-053", "결제 이력", "App", "/payments", "결제 내역 리스트 + 영수증 PDF 다운로드",
         "MUI-063~066", "기간 필터, 영수증 탭 → 상세/PDF 공유", None,
         "FlatList, PDF 뷰어/공유",
         "MA-050 → '결제 이력' 메뉴", None,
         "로드 실패: '다시 시도해주세요'",
         "금액 읽기 지원", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '결제 이력'\n[필터] 기간 선택 (1개월/3개월/6개월/1년)\n[결제 리스트] 카드형\n  - 날짜 / 상품명 / 금액 / 결제수단 아이콘\n  - 탭 → 영수증 상세\n[영수증 상세 모달] 상세 내역 + PDF 공유 버튼",
         "1:헤더 → 2:기간필터 → 3:결제리스트",
         "stack", "메뉴 탭",
         "리스트: Skeleton 카드 3개",
         "결제 없음: '결제 내역이 없습니다'",
         "세로 스크롤", "헤더+필터 고정"],

        ["MA-054", "락커 정보", "App", "/locker", "배정된 락커 번호/위치/만료일",
         "MUI-067~068", "단순 조회 화면", None,
         "단순 카드 UI",
         "MA-050 → '락커' 메뉴", None,
         "로드 실패: '다시 시도해주세요'",
         None, None, "세로 고정",
         None, "기획", "앱개발팀", 1, None, 1, None, 2,
         "[헤더] 뒤로가기 + '락커 정보'\n[락커 카드] 락커 번호 (큰 텍스트) + 구역 + 만료일\n  만료 7일 이내: 빨강 경고 뱃지",
         "1:헤더 → 2:락커카드",
         "stack", "메뉴 탭",
         "카드 Skeleton",
         "미배정: '배정된 락커가 없습니다'",
         "세로 고정", "없음"],

        ["MA-055", "쿠폰함", "App", "/coupons", "보유 쿠폰 목록 + 사용 가능 여부",
         "MUI-069~072", "사용가능/완료/만료 탭, '바로 사용' → 상품 목록, 쿠폰 상세 BottomSheet", None,
         "탭 전환, BottomSheet, 쿠폰 카드 디자인",
         "MA-050 → '쿠폰함' 메뉴 / MA-062 (결제 시 쿠폰 선택)", "MA-060 (상품 목록, '바로 사용' 시)",
         "로드 실패: '다시 시도해주세요'",
         "쿠폰 상태 텍스트 표시", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '쿠폰함'\n[탭] 사용가능 | 사용완료 | 만료\n[쿠폰 카드] 티켓 스타일\n  - 쿠폰명 / 할인 내용 (5,000원 할인 / 10% 할인)\n  - 유효기간\n  - '바로 사용' 버튼\n  - 만료 임박: 빨강 뱃지",
         "1:헤더 → 2:탭 → 3:쿠폰리스트",
         "stack", "메뉴 탭 / 결제 화면 쿠폰 선택",
         "카드 Skeleton 3개",
         "쿠폰 없음: '보유한 쿠폰이 없습니다'",
         "세로 스크롤", "헤더+탭 고정"],

        ["MA-056", "마일리지", "App", "/mileage", "잔액 + 적립/사용 이력",
         "MUI-073~076", "단순 조회, 이력 무한 스크롤", None,
         "잔액 카드 + FlatList",
         "MA-050 → '마일리지' 메뉴", None,
         "로드 실패: '다시 시도해주세요'",
         "금액 읽기 지원", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '마일리지'\n[잔액 카드] 큰 숫자 '5,000P' + 소멸 예정 안내\n[이력 리스트]\n  - 날짜 / 유형 (적립:파랑/사용:빨강) / 금액 / 잔액\n  - 사유 텍스트",
         "1:헤더 → 2:잔액카드 → 3:이력리스트",
         "stack", "메뉴 탭",
         "잔액 Skeleton + 리스트 Skeleton",
         "이력 없음: '마일리지 이력이 없습니다'",
         "세로 스크롤", "헤더+잔액카드 고정"],

        ["MA-060", "상품 목록", "App", "/shop", "이용권/PT/GX 상품 카테고리별 조회",
         "MUI-077~080", "카테고리 탭 전환, 상품 카드 탭 → 상세", None,
         "카테고리 탭 + 상품 카드 그리드",
         "하단탭 없음, MA-055 → '바로 사용' / MA-051 → '연장하기' / 홈 → '구매하러 가기'",
         "MA-061 (상품 상세)",
         "로드 실패: '다시 시도해주세요'",
         "상품 정보 전달", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '상품'\n[카테고리 탭] 이용권 | PT | GX\n[상품 리스트] 카드형\n  - 상품명 / 기간 / 가격 (카드가 기준)\n  - 설명 미리보기",
         "1:헤더 → 2:카테고리탭 → 3:상품리스트",
         "stack", "링크/버튼 이동",
         "카드 Skeleton 3개",
         "상품 없음: '등록된 상품이 없습니다'",
         "세로 스크롤", "헤더+탭 고정"],

        ["MA-061", "상품 상세", "App", "/shop/:id", "상품 정보 + 가격 + 구매 버튼",
         "MUI-081~084", "구매하기 → 결제 화면 이동", None,
         "상세 화면 + CTA 버튼",
         "MA-060 → 상품 카드 탭", "MA-062 (결제)",
         "판매 중지: '현재 판매하지 않는 상품입니다'\n로드 실패: '다시 시도해주세요'",
         None, None, "세로 고정",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + 상품명\n[상품 정보] 카테고리 뱃지 + 상품명 + 기간\n[가격] 카드가 크게 표시\n[설명] 상품 상세 설명\n[하단 고정] '구매하기' CTA 버튼",
         "1:헤더 → 2:상품정보 → 3:가격 → 4:설명 → 5:구매버튼",
         "stack", "상품 카드 탭",
         "상품 Skeleton",
         None, "세로 스크롤", "하단 CTA 고정"],

        ["MA-062", "결제/구매", "App", "/shop/checkout", "쿠폰/마일리지 적용 + PG 결제 + 이용권 활성화",
         "MUI-085~092", "쿠폰 선택 BottomSheet, 마일리지 입력, PG 모듈 호출, 결제 완료 화면", None,
         "PG SDK (토스페이먼츠/아임포트), 트랜잭션 보장",
         "MA-061 → '구매하기' / MA-051 → '연장하기'", "MA-051 (이용권 목록, 결제 완료 후)",
         "결제 실패: '다시 시도해주세요'\n쿠폰 만료: '선택한 쿠폰이 만료되었습니다'\n마일리지 부족: '잔액이 부족합니다'",
         "금액 정보 명확 전달", None, "세로 고정",
         None, "기획", "앱개발팀", 2, None, 2, None, 8,
         "[헤더] 뒤로가기 + '결제'\n[주문 요약] 상품명/기간/원가\n[쿠폰] 선택 버튼 → BottomSheet → 할인액 반영\n[마일리지] 잔액 표시 + 사용 금액 입력\n[결제 금액] 원가 - 쿠폰 - 마일리지 = 최종금액 (강조)\n[결제수단] 카드/카카오페이/네이버페이/토스 선택\n[하단 고정] '결제하기 {금액}원' CTA 버튼\n\n[결제 완료 화면]\n  - 체크 애니메이션\n  - '결제가 완료되었습니다'\n  - 이용권 정보\n  - 영수증 다운로드 버튼\n  - '확인' → 이용권 목록",
         "1:주문요약 → 2:쿠폰 → 3:마일리지 → 4:결제금액 → 5:결제수단 → 6:결제버튼",
         "stack", "구매/연장 버튼 탭",
         "결제 처리: 전체 오버레이 스피너 '결제 처리 중...'",
         None, "세로 고정", "하단 CTA 고정"],

        ["MA-070", "알림 목록", "App", "/notifications", "만료/예약/생일/공지 알림 리스트",
         "MUI-093~096", "알림 탭 → 딥링크 이동, 읽음/안읽음 구분", None,
         "FlatList, 딥링크 처리",
         "하단탭 '더보기' → '알림' / 푸시 알림 탭", None,
         "로드 실패: '다시 시도해주세요'",
         "알림 유형 아이콘 의미 전달", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] '알림'\n[알림 리스트]\n  - 유형 아이콘 (이용권:달력/예약:시계/생일:케이크/결제:카드/공지:메가폰)\n  - 제목 + 메시지 미리보기\n  - 시간 (n분 전/n시간 전/날짜)\n  - 안읽음: 파란 점 표시",
         "1:헤더 → 2:알림리스트",
         "stack", "탭 / 푸시 알림",
         "리스트 Skeleton 5개",
         "알림 없음: '알림이 없습니다'",
         "세로 스크롤", "헤더 고정"],

        ["MA-071", "알림 설정", "App", "/notifications/settings", "유형별 푸시 알림 ON/OFF",
         "MUI-097~098", "Switch 토글", None,
         "FCM topic 구독/해제",
         "MA-070 → 설정 아이콘 / MA-076 → 알림 설정", None,
         "저장 실패: '다시 시도해주세요'",
         "토글 상태 읽기 지원", None, "세로 고정",
         None, "기획", "앱개발팀", 1, None, 1, None, 2,
         "[헤더] 뒤로가기 + '알림 설정'\n[설정 리스트]\n  - 이용권 만료 알림 [토글]\n  - 수업 리마인드 [토글]\n  - 마케팅 메시지 [토글]\n  - 공지사항 [토글]",
         "1:헤더 → 2:설정리스트",
         "stack", "설정 아이콘 탭",
         None, None, "세로 고정", "없음"],

        ["MA-072", "센터 정보", "App", "/center-info", "센터명/주소/영업시간/연락처/지도",
         "MUI-099~102", "지도 탭 → 네이버맵, 전화 걸기 버튼", None,
         "Map SDK (네이버맵/카카오맵), Linking (전화)",
         "MA-076 → '센터 정보' / 더보기 메뉴", None,
         "로드 실패: '다시 시도해주세요'",
         "주소/전화 읽기 지원", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 4,
         "[헤더] 뒤로가기 + '센터 정보'\n[센터 카드] 센터명 + 로고\n[지도] 네이버맵/카카오맵 (센터 위치 마커)\n[정보] 주소 / 영업시간 / 전화번호\n[버튼] 전화 걸기 / 길 찾기",
         "1:헤더 → 2:센터카드 → 3:지도 → 4:정보 → 5:버튼",
         "stack", "메뉴 탭",
         "지도 로딩 스피너",
         None, "세로 스크롤", "없음"],

        ["MA-073", "공지사항", "App", "/notices", "센터 공지사항 목록 + 상세",
         "MUI-103~104", "목록 → 상세 이동, 중요 공지 상단 고정", None,
         "FlatList + 상세 화면",
         "홈 공지 배너 탭 / 더보기 → '공지사항'", None,
         "로드 실패: '다시 시도해주세요'",
         None, None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 2,
         "[헤더] 뒤로가기 + '공지사항'\n[공지 리스트] 제목 + 날짜 + 중요 뱃지\n\n[공지 상세] 제목 + 날짜 + 본문 + 이미지",
         "1:헤더 → 2:공지리스트 (또는 상세)",
         "stack", "배너 탭 / 메뉴 탭",
         "리스트 Skeleton 5개",
         "공지 없음: '공지사항이 없습니다'",
         "세로 스크롤", "헤더 고정"],

        ["MA-074", "1:1 문의", "App", "/inquiry", "문의 작성 + 답변 조회",
         "MUI-105~110", "카테고리 선택, 이미지 첨부, 문의 이력 탭", None,
         "Form + ImagePicker + FlatList",
         "더보기 → '1:1 문의'", None,
         "작성 실패: '다시 시도해주세요'",
         "폼 접근성", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 6,
         "[헤더] 뒤로가기 + '1:1 문의'\n[탭] 문의하기 | 문의 이력\n[문의하기 폼]\n  - 카테고리 선택 (드롭다운)\n  - 제목 입력\n  - 본문 입력 (Textarea)\n  - 이미지 첨부 (최대 3장)\n  - 제출 버튼\n[문의 이력] 리스트 (상태 뱃지: 대기/답변완료)",
         "1:헤더 → 2:탭 → 3:폼 또는 이력리스트",
         "stack", "메뉴 탭",
         "제출: ActivityIndicator",
         "이력 없음: '문의 내역이 없습니다'",
         "세로 스크롤, 키보드 어보이던스", "헤더+탭 고정"],

        ["MA-075", "구독/자동결제", "App", "/subscription", "구독 상태/결제일/카드 관리/해지",
         "MUI-111~116", "카드 변경, 해지 BottomSheet (즉시/월말), 환불 금액 표시", None,
         "구독 카드 UI, 해지 BottomSheet, PG 카드 변경",
         "더보기 → '구독 관리'", None,
         "구독 없음: '구독 중인 상품이 없습니다'\n해지 확인: '정말 해지하시겠습니까?'",
         "구독 상태 텍스트 전달", None, "세로 우선",
         None, "기획", "앱개발팀", 1, None, 1, None, 6,
         "[헤더] 뒤로가기 + '구독 관리'\n[구독 카드]\n  - 상품명 / 상태 뱃지 (활성:초록/일시정지:주황)\n  - 다음 결제일\n  - 결제 금액\n  - 등록 카드 정보 (****1234)\n[버튼] 카드 변경 / 구독 해지\n\n[해지 BottomSheet]\n  - 즉시 해지: 환불 예상금액 표시\n  - 월말 해지: 기간 종료 후 해지\n  - 해지 사유 선택",
         "1:헤더 → 2:구독카드 → 3:버튼",
         "stack", "메뉴 탭",
         "카드 Skeleton",
         "구독 없음: '구독 중인 상품이 없습니다'",
         "세로 스크롤", "없음"],

        ["MA-076", "앱 설정", "App", "/settings", "비밀번호 변경, 계정 탈퇴, 버전 정보",
         "MUI-117~120", "비밀번호 변경 → 현재+새+확인 폼, 계정 탈퇴 → 확인 팝업", None,
         "Supabase Auth updateUser, 앱 버전 조회",
         "MA-050 → '설정' / 더보기 → '설정'", "MA-001 (비밀번호 변경 후 재로그인)",
         "현재 비밀번호 불일치: '올바르지 않습니다'\n재사용: '이전에 사용한 비밀번호입니다'",
         "폼 접근성", None, "세로 고정",
         None, "기획", "앱개발팀", 2, None, 0, None, 4,
         "[헤더] 뒤로가기 + '설정'\n[메뉴 리스트]\n  - 비밀번호 변경 → 비밀번호 폼 화면\n  - 알림 설정 → MA-071\n  - 센터 정보 → MA-072\n  - 버전 정보 (현재 버전 표시)\n  - 로그아웃 (빨강)\n  - 계정 탈퇴 (회색, 하단)",
         "1:헤더 → 2:메뉴리스트",
         "stack / tab (더보기 탭)", "메뉴 탭",
         None, None, "세로 고정", "없음"],
    ]

    for row in screens:
        ws.append(row)

    style_body(ws)
    auto_width(ws, max_width=40)


# ══════════════════════════════════════════════════════════
# 8. 화면흐름도 시트
# ══════════════════════════════════════════════════════════
def create_flow_sheet(wb):
    ws = wb.create_sheet("화면흐름도")
    headers = [
        "흐름 ID", "출발 화면", "트리거 요소", "트리거 액션",
        "도착 화면", "전환 방식", "전달 데이터",
        "뒤로가기 동작", "조건", "비고"
    ]
    ws.append(headers)
    style_header(ws)

    flows = [
        # 인증 플로우
        ["MFLW-001", "MA-001 (로그인)", "로그인 버튼", "tap", "MA-010 (홈)", "replace", "JWT토큰, 회원정보", "로그인으로 돌아감", "연락처+비밀번호 유효 시", None],
        ["MFLW-002", "MA-001 (로그인)", "'회원가입' 링크", "tap", "MA-002 (앱 연동)", "push", "없음", "로그인으로 돌아감", None, None],
        ["MFLW-003", "MA-001 (로그인)", "'비밀번호 찾기' 링크", "tap", "MA-003 (비밀번호 찾기)", "push", "없음", "로그인으로 돌아감", None, None],
        ["MFLW-004", "MA-002 (앱 연동)", "연동 완료", "auto", "MA-010 (홈)", "replace", "JWT토큰, 회원정보", "N/A", "SMS 인증+비밀번호 설정 완료 시", None],
        ["MFLW-005", "MA-003 (비밀번호 찾기)", "재설정 완료", "auto", "MA-001 (로그인)", "replace", "없음", "N/A", "비밀번호 재설정 완료 시", None],
        ["MFLW-006", "전체 화면", "토큰 만료", "auto", "MA-001 (로그인)", "replace", "없음", "N/A", "JWT 만료 + refresh 실패 시", None],
        # 홈 플로우
        ["MFLW-007", "MA-010 (홈)", "QR FAB 버튼", "tap", "MA-030 (QR 체크인)", "push", "없음", "홈으로 돌아감", None, None],
        ["MFLW-008", "MA-010 (홈)", "이용권 카드", "tap", "MA-051 (이용권)", "push", "없음", "홈으로 돌아감", None, None],
        ["MFLW-009", "MA-010 (홈)", "오늘 예약 카드", "tap", "MA-042 (내 예약)", "push", "없음", "홈으로 돌아감", None, None],
        ["MFLW-010", "MA-010 (홈)", "공지 배너", "tap", "MA-073 (공지사항)", "push", "공지 ID", "홈으로 돌아감", None, None],
        ["MFLW-011", "MA-010 (홈)", "'구매하러 가기'", "tap", "MA-060 (상품 목록)", "push", "없음", "홈으로 돌아감", "이용권 없을 때만 표시", None],
        # 출석 플로우
        ["MFLW-012", "MA-030 (QR 체크인)", "'출석 이력 보기'", "tap", "MA-031 (출석 이력)", "push", "없음", "QR로 돌아감", None, None],
        ["MFLW-013", "MA-030 (QR 체크인)", "'이용권 구매'", "tap", "MA-060 (상품 목록)", "push", "없음", "QR로 돌아감", "이용권 없을 때만", None],
        # 예약 플로우
        ["MFLW-014", "MA-040 (수업 목록)", "수업 카드", "tap", "MA-041 (수업 상세)", "push", "수업 ID", "목록으로 돌아감", None, None],
        ["MFLW-015", "MA-040 (수업 목록)", "'내 예약' 아이콘", "tap", "MA-042 (내 예약)", "push", "없음", "목록으로 돌아감", None, None],
        ["MFLW-016", "MA-041 (수업 상세)", "예약 완료", "auto", "MA-042 (내 예약)", "replace", "예약 정보", "N/A", "예약 성공 시", None],
        # 내정보 플로우
        ["MFLW-017", "MA-050 (프로필)", "'이용권' 메뉴", "tap", "MA-051 (이용권)", "push", "없음", "프로필로 돌아감", None, None],
        ["MFLW-018", "MA-050 (프로필)", "'체성분' 메뉴", "tap", "MA-052 (체성분)", "push", "없음", "프로필로 돌아감", None, None],
        ["MFLW-019", "MA-050 (프로필)", "'결제 이력' 메뉴", "tap", "MA-053 (결제 이력)", "push", "없음", "프로필로 돌아감", None, None],
        ["MFLW-020", "MA-050 (프로필)", "'락커' 메뉴", "tap", "MA-054 (락커)", "push", "없음", "프로필로 돌아감", None, None],
        ["MFLW-021", "MA-050 (프로필)", "'쿠폰함' 메뉴", "tap", "MA-055 (쿠폰함)", "push", "없음", "프로필로 돌아감", None, None],
        ["MFLW-022", "MA-050 (프로필)", "'마일리지' 메뉴", "tap", "MA-056 (마일리지)", "push", "없음", "프로필로 돌아감", None, None],
        # 구매 플로우
        ["MFLW-023", "MA-051 (이용권)", "'연장하기' 버튼", "tap", "MA-062 (결제)", "push", "상품 ID", "이용권으로 돌아감", "만료 7일 이내 이용권", None],
        ["MFLW-024", "MA-060 (상품 목록)", "상품 카드", "tap", "MA-061 (상품 상세)", "push", "상품 ID", "목록으로 돌아감", None, None],
        ["MFLW-025", "MA-061 (상품 상세)", "'구매하기' 버튼", "tap", "MA-062 (결제)", "push", "상품 ID", "상세로 돌아감", None, None],
        ["MFLW-026", "MA-062 (결제)", "결제 완료", "auto", "MA-051 (이용권)", "replace", "이용권 정보", "N/A", "PG 결제 성공 시", None],
        ["MFLW-027", "MA-062 (결제)", "'쿠폰 선택'", "tap", "MA-055 (쿠폰함) BottomSheet", "modal", "선택된 쿠폰 ID", "BottomSheet 닫기", None, "BottomSheet 모달"],
        # 더보기 플로우
        ["MFLW-028", "MA-076 (설정)", "'로그아웃'", "tap+confirm", "MA-001 (로그인)", "replace", "없음", "N/A", "확인 팝업 승인 시", None],
        ["MFLW-029", "MA-076 (설정)", "'비밀번호 변경'", "tap", "비밀번호 변경 폼", "push", "없음", "설정으로 돌아감", None, None],
        ["MFLW-030", "푸시 알림", "알림 탭", "tap", "딥링크 대상 화면", "push", "type, target_id", "N/A", "알림 유형별 딥링크", "만료→MA-051, 예약→MA-042 등"],
    ]

    for row in flows:
        ws.append(row)

    style_body(ws)
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 9. 디자인토큰 시트
# ══════════════════════════════════════════════════════════
def create_token_sheet(wb):
    ws = wb.create_sheet("디자인토큰")
    headers = ["토큰 ID", "카테고리", "토큰명", "값", "용도", "비고"]
    ws.append(headers)
    style_header(ws)

    tokens = [
        # 색상 - Primary
        ["MTK-001", "color", "primary-50", "#EFF6FF", "Primary 배경 (매우 연한)", "CRM과 동일"],
        ["MTK-002", "color", "primary-100", "#DBEAFE", "Primary light 배경", None],
        ["MTK-003", "color", "primary-500", "#3B82F6", "Primary 메인 색상 (버튼, 링크)", None],
        ["MTK-004", "color", "primary-600", "#2563EB", "Primary hover/pressed", None],
        ["MTK-005", "color", "primary-700", "#1D4ED8", "Primary dark", None],
        # 색상 - Semantic
        ["MTK-010", "color", "success-500", "#22C55E", "성공/활성 상태 (출석완료, 활성 이용권)", None],
        ["MTK-011", "color", "warning-500", "#F59E0B", "경고/주의 (D-7 만료, 일시정지)", None],
        ["MTK-012", "color", "error-500", "#EF4444", "오류/위험 (D-1 만료, 취소, 에러)", None],
        ["MTK-013", "color", "info-500", "#3B82F6", "정보/안내", None],
        # 색상 - 출석 유형
        ["MTK-020", "color", "attendance-normal", "#3B82F6", "일반 출석 (파랑)", None],
        ["MTK-021", "color", "attendance-pt", "#22C55E", "PT 출석 (초록)", None],
        ["MTK-022", "color", "attendance-gx", "#8B5CF6", "GX 출석 (보라)", None],
        # 색상 - 그레이
        ["MTK-030", "color", "gray-50", "#F9FAFB", "배경", None],
        ["MTK-031", "color", "gray-100", "#F3F4F6", "카드 배경, 구분선", None],
        ["MTK-032", "color", "gray-300", "#D1D5DB", "비활성 테두리", None],
        ["MTK-033", "color", "gray-500", "#6B7280", "보조 텍스트", None],
        ["MTK-034", "color", "gray-700", "#374151", "본문 텍스트", None],
        ["MTK-035", "color", "gray-900", "#111827", "제목 텍스트", None],
        # 타이포그래피
        ["MTK-100", "typography", "heading-xl", "24px / Bold / 1.3", "화면 제목 (홈 인사말)", None],
        ["MTK-101", "typography", "heading-lg", "20px / Bold / 1.3", "섹션 제목", None],
        ["MTK-102", "typography", "heading-md", "18px / SemiBold / 1.4", "카드 제목", None],
        ["MTK-103", "typography", "body-lg", "16px / Regular / 1.5", "본문 기본", None],
        ["MTK-104", "typography", "body-md", "14px / Regular / 1.5", "보조 텍스트, 설명", None],
        ["MTK-105", "typography", "body-sm", "12px / Regular / 1.4", "캡션, 타임스탬프", None],
        ["MTK-106", "typography", "label", "14px / Medium / 1.4", "입력 필드 라벨", None],
        ["MTK-107", "typography", "button", "16px / SemiBold / 1.0", "버튼 텍스트", None],
        ["MTK-108", "typography", "d-day", "32px / Bold / 1.0", "D-day 큰 숫자", "이용권 D-day 뱃지"],
        ["MTK-109", "typography", "mileage", "28px / Bold / 1.0", "마일리지 잔액", None],
        # 간격
        ["MTK-200", "spacing", "space-xs", "4px", "아이콘-텍스트 간격", None],
        ["MTK-201", "spacing", "space-sm", "8px", "요소 내부 간격", None],
        ["MTK-202", "spacing", "space-md", "12px", "카드 내부 패딩", None],
        ["MTK-203", "spacing", "space-lg", "16px", "섹션 간격, 카드 패딩", None],
        ["MTK-204", "spacing", "space-xl", "24px", "큰 섹션 간격", None],
        ["MTK-205", "spacing", "space-2xl", "32px", "화면 상하 여백", None],
        ["MTK-206", "spacing", "safe-area-bottom", "34px", "하단 SafeArea (iPhone)", "디바이스별 동적"],
        # 크기
        ["MTK-300", "size", "input-height", "48px", "입력 필드 높이", "모바일 터치 영역"],
        ["MTK-301", "size", "button-height", "52px", "CTA 버튼 높이", None],
        ["MTK-302", "size", "button-sm-height", "40px", "작은 버튼 높이", None],
        ["MTK-303", "size", "card-radius", "12px", "카드 모서리 라운드", None],
        ["MTK-304", "size", "input-radius", "8px", "입력 필드 라운드", None],
        ["MTK-305", "size", "button-radius", "10px", "버튼 라운드", None],
        ["MTK-306", "size", "fab-size", "56px", "QR FAB 버튼 크기", None],
        ["MTK-307", "size", "avatar-lg", "80px", "프로필 사진 (큰)", None],
        ["MTK-308", "size", "avatar-sm", "40px", "프로필 사진 (작은)", None],
        ["MTK-309", "size", "qr-size", "240px", "QR코드 크기", None],
        ["MTK-310", "size", "tab-bar-height", "56px", "하단 탭바 높이", None],
        # 그림자
        ["MTK-400", "shadow", "shadow-sm", "0 1px 2px rgba(0,0,0,0.05)", "카드 기본 그림자", None],
        ["MTK-401", "shadow", "shadow-md", "0 4px 6px rgba(0,0,0,0.1)", "플로팅 요소", None],
        ["MTK-402", "shadow", "shadow-lg", "0 10px 15px rgba(0,0,0,0.15)", "BottomSheet", None],
        # 애니메이션
        ["MTK-500", "animation", "duration-fast", "150ms", "버튼 피드백, 토글", None],
        ["MTK-501", "animation", "duration-normal", "300ms", "화면 전환, 모달", None],
        ["MTK-502", "animation", "duration-slow", "500ms", "성공 애니메이션", "체크인 완료"],
        ["MTK-503", "animation", "easing-default", "ease-in-out", "기본 이징", None],
    ]

    for row in tokens:
        ws.append(row)

    style_body(ws)
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 10. QA테스트케이스 시트
# ══════════════════════════════════════════════════════════
def create_tc_sheet(wb):
    ws = wb.create_sheet("QA테스트케이스")
    headers = [
        "TC-ID", "플랫폼", "테스트 영역", "하위 영역", "테스트 유형",
        "테스트 시나리오", "사전 조건", "테스트 절차", "기대 결과",
        "실제 결과", "테스트 상태", "우선순위", "심각도", "담당자",
        "테스트 환경", "디바이스", "OS",
        "화면 ID", "기능 ID", "API ID", "비고"
    ]
    ws.append(headers)
    style_header(ws)

    tcs = [
        # 인증
        ["MTC-001", "App", "인증", "로그인", "기능", "정상 로그인 - 올바른 연락처/비밀번호",
         "앱 연동 완료된 회원 계정", "1) 앱 실행\n2) 연락처 입력\n3) 비밀번호 입력\n4) 로그인 버튼 탭",
         "홈(MA-010)으로 이동, 인사말에 회원명 표시", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-001", "MFN-001", "MAPI-001", None],
        ["MTC-002", "App", "인증", "로그인", "기능", "로그인 실패 - 잘못된 비밀번호",
         "앱 연동 완료된 회원 계정", "1) 올바른 연락처 입력\n2) 잘못된 비밀번호 입력\n3) 로그인 버튼 탭",
         "'연락처 또는 비밀번호를 확인해주세요' 에러 메시지", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-001", "MFN-001", "MAPI-001", None],
        ["MTC-003", "App", "인증", "로그인", "기능", "계정 잠금 - 5회 연속 실패",
         "앱 연동 완료된 회원 계정", "1) 잘못된 비밀번호로 5회 연속 로그인 시도",
         "'30분 후 재시도하거나 비밀번호를 찾아주세요' 메시지 + 로그인 버튼 비활성화", None, "대기", "Critical", "Critical", "QA",
         "Staging", "Galaxy S24", "Android 14",
         "MA-001", "MFN-001", "MAPI-001", None],
        ["MTC-004", "App", "인증", "로그인", "기능", "자동 로그인 - 토큰 유지",
         "자동 로그인 체크 후 로그인 완료 상태", "1) 앱 종료\n2) 앱 재실행",
         "로그인 화면 거치지 않고 홈(MA-010)으로 직접 이동", None, "대기", "High", "High", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-001", "MFN-001", "MAPI-001", None],
        ["MTC-005", "App", "인증", "앱 연동", "기능", "정상 앱 연동 (SMS 인증)",
         "CRM에 등록된 회원 (앱 미연동)", "1) 회원가입 탭\n2) 이름+연락처 입력\n3) 인증번호 확인\n4) 비밀번호 설정",
         "연동 완료 → 홈 이동, Supabase Auth 계정 생성", None, "대기", "Critical", "Critical", "QA",
         "Staging", "Galaxy S24", "Android 14",
         "MA-002", "MFN-002", "MAPI-002,MAPI-003", None],
        ["MTC-006", "App", "인증", "앱 연동", "기능", "미등록 회원 연동 시도",
         "CRM에 미등록된 연락처", "1) 회원가입 탭\n2) 이름+미등록 연락처 입력\n3) 다음 버튼 탭",
         "'등록된 회원 정보가 없습니다. 센터에 문의해주세요' 메시지", None, "대기", "High", "High", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-002", "MFN-002", "MAPI-002", None],
        # 홈
        ["MTC-007", "App", "홈", "대시보드", "기능", "홈 대시보드 데이터 표시",
         "로그인 완료, 활성 이용권 보유", "1) 홈 탭 진입",
         "인사말 + 이용권 D-day + 오늘 예약 + 출석 횟수 + 공지 배너 표시", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-010", "MFN-004", "MAPI-004", None],
        ["MTC-008", "App", "홈", "대시보드", "기능", "이용권 없는 회원 홈 표시",
         "로그인 완료, 이용권 없음", "1) 홈 탭 진입",
         "이용권 카드에 '이용권이 없습니다. 구매하러 가기' 표시", None, "대기", "High", "High", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-010", "MFN-004", "MAPI-004", None],
        # 출석
        ["MTC-009", "App", "출석", "QR 체크인", "기능", "QR코드 생성 및 갱신",
         "로그인 + 활성 이용권", "1) QR 체크인 화면 진입\n2) QR코드 표시 확인\n3) 60초 대기",
         "QR코드 표시, 60초 후 자동 갱신, 타이머 카운트다운", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-030", "MFN-005", "MAPI-005", None],
        ["MTC-010", "App", "출석", "QR 체크인", "기능", "이용권 없는 회원 QR 체크인",
         "로그인 완료, 이용권 없음", "1) QR 체크인 화면 진입",
         "'활성 이용권이 없습니다' + 구매 버튼 표시, QR코드 미생성", None, "대기", "Critical", "High", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-030", "MFN-005", "MAPI-005", None],
        ["MTC-011", "App", "출석", "출석 이력", "기능", "월별 출석 이력 캘린더 표시",
         "로그인 + 출석 기록 존재", "1) 출석 이력 화면 진입\n2) 캘린더 확인\n3) 좌우 스와이프",
         "출석일 색상 표시 (일반:파랑/PT:초록/GX:보라), 월 이동 동작", None, "대기", "High", "High", "QA",
         "Staging", "Galaxy S24", "Android 14",
         "MA-031", "MFN-006", "MAPI-006", None],
        # 예약
        ["MTC-012", "App", "예약", "수업 목록", "기능", "날짜별 수업 목록 조회",
         "로그인 + PT/GX 일정 존재", "1) 예약 탭 진입\n2) 날짜 선택\n3) PT/GX 필터 전환",
         "선택 날짜의 수업 목록 표시, 잔여석/마감 뱃지 정상", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-040", "MFN-007", "MAPI-007", None],
        ["MTC-013", "App", "예약", "수업 예약", "기능", "정상 수업 예약",
         "로그인 + PT 이용권 잔여 횟수 > 0", "1) 수업 목록에서 수업 탭\n2) 예약하기 버튼\n3) 확인 팝업 승인",
         "예약 완료 팝업 + 잔여 횟수 1 차감 + 내 예약에 추가", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-041", "MFN-008", "MAPI-008", None],
        ["MTC-014", "App", "예약", "수업 예약", "기능", "정원 마감 수업 예약 시도",
         "로그인 + 정원 초과 수업", "1) 마감된 수업 상세 진입\n2) 예약하기 시도",
         "'정원이 마감되었습니다' 메시지, 예약 불가", None, "대기", "High", "High", "QA",
         "Staging", "Galaxy S24", "Android 14",
         "MA-041", "MFN-008", "MAPI-008", None],
        ["MTC-015", "App", "예약", "예약 취소", "기능", "24시간 전 예약 취소",
         "로그인 + 24시간 이후 수업 예약", "1) 내 예약에서 예약 선택\n2) 취소 버튼\n3) 확인",
         "취소 완료 + PT 잔여 횟수 복원", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-041", "MFN-009", "MAPI-009", None],
        ["MTC-016", "App", "예약", "예약 취소", "기능", "24시간 이내 취소 경고",
         "로그인 + 24시간 이내 수업 예약", "1) 내 예약에서 예약 선택\n2) 취소 버튼",
         "'24시간 이내 취소는 PT 횟수가 차감될 수 있습니다' 경고 팝업", None, "대기", "High", "High", "QA",
         "Staging", "Galaxy S24", "Android 14",
         "MA-041", "MFN-009", "MAPI-009", None],
        # 내정보
        ["MTC-017", "App", "내정보", "이용권", "기능", "이용권 목록 + D-day 표시",
         "로그인 + 활성/만료 이용권", "1) 내정보 → 이용권 진입",
         "활성/만료 탭, D-day 뱃지, PT 잔여 횟수 프로그레스, 연장하기 버튼", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-051", "MFN-013", "MAPI-013", None],
        ["MTC-018", "App", "내정보", "체성분", "기능", "체성분 차트 표시",
         "로그인 + 체성분 측정 기록 2건 이상", "1) 내정보 → 체성분 진입\n2) 기간 변경",
         "시계열 라인 차트 표시, 기간 전환 정상, 최근 측정값 카드", None, "대기", "High", "High", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-052", "MFN-014", "MAPI-014", None],
        ["MTC-019", "App", "내정보", "쿠폰", "기능", "쿠폰함 조회 및 상태 구분",
         "로그인 + 사용가능/사용완료/만료 쿠폰", "1) 내정보 → 쿠폰함 진입\n2) 탭 전환",
         "탭별 쿠폰 분류, 만료 임박 상단 배치, '바로 사용' 버튼", None, "대기", "High", "High", "QA",
         "Staging", "Galaxy S24", "Android 14",
         "MA-055", "MFN-017", "MAPI-017", None],
        # 구매/결제
        ["MTC-020", "App", "구매", "온라인 결제", "기능", "정상 결제 + 이용권 활성화",
         "로그인 + 테스트 PG 환경", "1) 상품 선택 → 구매하기\n2) 쿠폰 적용\n3) 결제수단 선택\n4) 결제 완료",
         "결제 성공 + 이용권 자동 활성화 + 영수증 표시 + 마일리지 적립", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-062", "MFN-021", "MAPI-020,MAPI-021", None],
        ["MTC-021", "App", "구매", "온라인 결제", "기능", "결제 실패 시 이용권 미생성",
         "로그인 + PG 실패 시뮬레이션", "1) 상품 선택 → 구매하기\n2) 결제 실패 유도",
         "'결제에 실패했습니다' 메시지, 이용권 미생성, 쿠폰/마일리지 원복", None, "대기", "Critical", "Critical", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "MA-062", "MFN-021", "MAPI-020", None],
        # 알림
        ["MTC-022", "App", "알림", "푸시 알림", "기능", "이용권 만료 D-7 푸시 알림",
         "활성 이용권 만료 7일 전", "1) 서버 자동 알림 발송\n2) 앱에서 푸시 수신 확인",
         "푸시 알림 수신 + 알림 목록에 추가 + 탭 시 이용권 화면 이동", None, "대기", "Medium", "Medium", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-070", "MFN-030", None, "FCM 연동 필요"],
        # 성능/UX
        ["MTC-023", "App", "성능", "로딩", "성능", "홈 대시보드 로딩 속도",
         "로그인 완료", "1) 홈 탭 진입\n2) 로딩 완료까지 시간 측정",
         "Skeleton UI 즉시 표시, 데이터 로드 2초 이내 완료", None, "대기", "High", "Medium", "QA",
         "Staging", "iPhone 15 / Galaxy S24", "iOS 18 / Android 14",
         "MA-010", "MFN-004", "MAPI-004", None],
        ["MTC-024", "App", "성능", "오프라인", "기능", "네트워크 미연결 시 동작",
         "로그인 완료 후 네트워크 끊기", "1) 비행기 모드 활성화\n2) 화면 전환 시도",
         "'인터넷 연결을 확인해주세요' 배너 표시, 캐시 데이터 표시", None, "대기", "Medium", "Medium", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "전체", "전체", None, None],
        ["MTC-025", "App", "접근성", "VoiceOver", "접근성", "주요 화면 VoiceOver 동작",
         "VoiceOver 활성화", "1) 로그인 → 홈 → 예약 → 내정보 이동\n2) 각 화면 요소 읽기 확인",
         "모든 버튼/입력/카드에 적절한 라벨 읽기, 순서 논리적", None, "대기", "Medium", "Medium", "QA",
         "Staging", "iPhone 15", "iOS 18",
         "전체", "전체", None, None],
    ]

    for row in tcs:
        ws.append(row)

    style_body(ws)
    apply_priority_fill(ws, 12)  # 우선순위 컬럼
    auto_width(ws, max_width=40)


# ══════════════════════════════════════════════════════════
# 11. 데이터 정의서 시트
# ══════════════════════════════════════════════════════════
def create_data_sheet(wb):
    ws = wb.create_sheet("데이터 정의서")
    headers = [
        "테이블명", "컬럼명", "데이터 타입", "길이", "NULL 허용",
        "기본값", "PK", "FK", "인덱스", "설명",
        "관련 API ID", "상태", "비고"
    ]
    ws.append(headers)
    style_header(ws)

    # 회원앱 전용 테이블 + CRM 기존 테이블 확장 컬럼
    tables = [
        # member_app_accounts (앱 연동 계정)
        ["member_app_accounts", "id", "BIGINT", None, "N", None, "Y", None, "Y", "앱 계정 고유 ID", "MAPI-001", "기획", None],
        ["member_app_accounts", "member_id", "BIGINT", None, "N", None, "N", "members.id", "Y", "CRM 회원 ID (FK)", "MAPI-001", "기획", None],
        ["member_app_accounts", "auth_uid", "UUID", None, "N", None, "N", None, "Y", "Supabase Auth UID", "MAPI-001", "기획", None],
        ["member_app_accounts", "phone", "VARCHAR", 20, "N", None, "N", None, "Y", "연락처 (로그인 ID)", "MAPI-001", "기획", None],
        ["member_app_accounts", "fcm_token", "VARCHAR", 255, "Y", None, "N", None, "N", "FCM 푸시 토큰", "MAPI-022", "기획", None],
        ["member_app_accounts", "auto_login", "BOOLEAN", None, "N", "false", "N", None, "N", "자동 로그인 설정", "MAPI-001", "기획", None],
        ["member_app_accounts", "last_login_at", "TIMESTAMP", None, "Y", None, "N", None, "N", "마지막 로그인 시각", "MAPI-001", "기획", None],
        ["member_app_accounts", "login_fail_count", "INT", None, "N", "0", "N", None, "N", "연속 로그인 실패 횟수", "MAPI-001", "기획", None],
        ["member_app_accounts", "locked_until", "TIMESTAMP", None, "Y", None, "N", None, "N", "계정 잠금 해제 시각", "MAPI-001", "기획", None],
        ["member_app_accounts", "created_at", "TIMESTAMP", None, "N", "now()", "N", None, "N", "앱 연동 일시", "MAPI-003", "기획", None],
        # member_reservations (수업 예약)
        ["member_reservations", "id", "BIGINT", None, "N", None, "Y", None, "Y", "예약 고유 ID", "MAPI-008", "기획", None],
        ["member_reservations", "member_id", "BIGINT", None, "N", None, "N", "members.id", "Y", "회원 ID", "MAPI-008", "기획", None],
        ["member_reservations", "schedule_id", "BIGINT", None, "N", None, "N", "schedules.id", "Y", "수업 일정 ID", "MAPI-008", "기획", None],
        ["member_reservations", "branch_id", "BIGINT", None, "N", None, "N", "branches.id", "Y", "지점 ID (멀티테넌트)", "MAPI-008", "기획", None],
        ["member_reservations", "status", "VARCHAR", 20, "N", "'upcoming'", "N", None, "Y", "상태 (upcoming/completed/cancelled)", "MAPI-010", "기획", None],
        ["member_reservations", "cancel_reason", "VARCHAR", 200, "Y", None, "N", None, "N", "취소 사유", "MAPI-009", "기획", None],
        ["member_reservations", "cancelled_at", "TIMESTAMP", None, "Y", None, "N", None, "N", "취소 일시", "MAPI-009", "기획", None],
        ["member_reservations", "created_at", "TIMESTAMP", None, "N", "now()", "N", None, "N", "예약 생성 일시", "MAPI-008", "기획", None],
        # member_notifications (알림)
        ["member_notifications", "id", "BIGINT", None, "N", None, "Y", None, "Y", "알림 고유 ID", "MAPI-022", "기획", None],
        ["member_notifications", "member_id", "BIGINT", None, "N", None, "N", "members.id", "Y", "수신 회원 ID", "MAPI-022", "기획", None],
        ["member_notifications", "branch_id", "BIGINT", None, "N", None, "N", "branches.id", "Y", "지점 ID", "MAPI-022", "기획", None],
        ["member_notifications", "type", "VARCHAR", 30, "N", None, "N", None, "Y", "유형 (expire/reservation/birthday/payment/notice)", "MAPI-022", "기획", None],
        ["member_notifications", "title", "VARCHAR", 100, "N", None, "N", None, "N", "알림 제목", "MAPI-022", "기획", None],
        ["member_notifications", "message", "TEXT", None, "N", None, "N", None, "N", "알림 본문", "MAPI-022", "기획", None],
        ["member_notifications", "deeplink", "VARCHAR", 200, "Y", None, "N", None, "N", "딥링크 경로 (예: /memberships)", "MAPI-022", "기획", None],
        ["member_notifications", "is_read", "BOOLEAN", None, "N", "false", "N", None, "Y", "읽음 여부", "MAPI-022", "기획", None],
        ["member_notifications", "created_at", "TIMESTAMP", None, "N", "now()", "N", None, "Y", "생성 일시", "MAPI-022", "기획", None],
        # member_notification_settings (알림 설정)
        ["member_notification_settings", "id", "BIGINT", None, "N", None, "Y", None, "Y", "설정 ID", "MAPI-023", "기획", None],
        ["member_notification_settings", "member_id", "BIGINT", None, "N", None, "N", "members.id", "Y", "회원 ID", "MAPI-023", "기획", None],
        ["member_notification_settings", "expire_alert", "BOOLEAN", None, "N", "true", "N", None, "N", "이용권 만료 알림", "MAPI-023", "기획", None],
        ["member_notification_settings", "class_remind", "BOOLEAN", None, "N", "true", "N", None, "N", "수업 리마인드 알림", "MAPI-023", "기획", None],
        ["member_notification_settings", "marketing", "BOOLEAN", None, "N", "false", "N", None, "N", "마케팅 메시지 수신", "MAPI-023", "기획", None],
        ["member_notification_settings", "notice_alert", "BOOLEAN", None, "N", "true", "N", None, "N", "공지사항 알림", "MAPI-023", "기획", None],
        # notices (공지사항)
        ["notices", "id", "BIGINT", None, "N", None, "Y", None, "Y", "공지 고유 ID", "MAPI-024", "기획", None],
        ["notices", "branch_id", "BIGINT", None, "N", None, "N", "branches.id", "Y", "지점 ID", "MAPI-024", "기획", None],
        ["notices", "title", "VARCHAR", 200, "N", None, "N", None, "N", "공지 제목", "MAPI-024", "기획", None],
        ["notices", "content", "TEXT", None, "N", None, "N", None, "N", "공지 본문", "MAPI-024", "기획", None],
        ["notices", "is_pinned", "BOOLEAN", None, "N", "false", "N", None, "N", "상단 고정 여부", "MAPI-024", "기획", None],
        ["notices", "created_at", "TIMESTAMP", None, "N", "now()", "N", None, "Y", "작성 일시", "MAPI-024", "기획", None],
        # member_inquiries (1:1 문의)
        ["member_inquiries", "id", "BIGINT", None, "N", None, "Y", None, "Y", "문의 고유 ID", "MAPI-025", "기획", None],
        ["member_inquiries", "member_id", "BIGINT", None, "N", None, "N", "members.id", "Y", "문의 회원 ID", "MAPI-025", "기획", None],
        ["member_inquiries", "branch_id", "BIGINT", None, "N", None, "N", "branches.id", "Y", "지점 ID", "MAPI-025", "기획", None],
        ["member_inquiries", "category", "VARCHAR", 30, "N", None, "N", None, "Y", "카테고리 (이용권/예약/시설/결제/기타)", "MAPI-025", "기획", None],
        ["member_inquiries", "title", "VARCHAR", 200, "N", None, "N", None, "N", "제목", "MAPI-025", "기획", None],
        ["member_inquiries", "content", "TEXT", None, "N", None, "N", None, "N", "본문", "MAPI-025", "기획", None],
        ["member_inquiries", "images", "JSONB", None, "Y", "[]", "N", None, "N", "첨부 이미지 URL 배열", "MAPI-025", "기획", None],
        ["member_inquiries", "status", "VARCHAR", 20, "N", "'pending'", "N", None, "Y", "상태 (pending/answered)", "MAPI-025", "기획", None],
        ["member_inquiries", "answer", "TEXT", None, "Y", None, "N", None, "N", "답변 내용", "MAPI-025", "기획", None],
        ["member_inquiries", "answered_at", "TIMESTAMP", None, "Y", None, "N", None, "N", "답변 일시", "MAPI-025", "기획", None],
        ["member_inquiries", "answered_by", "BIGINT", None, "Y", None, "N", "staff.id", "N", "답변자 (직원 ID)", "MAPI-025", "기획", None],
        ["member_inquiries", "created_at", "TIMESTAMP", None, "N", "now()", "N", None, "Y", "문의 생성 일시", "MAPI-025", "기획", None],
        # member_orders (앱 주문/결제)
        ["member_orders", "id", "BIGINT", None, "N", None, "Y", None, "Y", "주문 고유 ID", "MAPI-020", "기획", None],
        ["member_orders", "member_id", "BIGINT", None, "N", None, "N", "members.id", "Y", "주문 회원 ID", "MAPI-020", "기획", None],
        ["member_orders", "branch_id", "BIGINT", None, "N", None, "N", "branches.id", "Y", "지점 ID", "MAPI-020", "기획", None],
        ["member_orders", "product_id", "BIGINT", None, "N", None, "N", "products.id", "Y", "상품 ID", "MAPI-020", "기획", None],
        ["member_orders", "coupon_id", "BIGINT", None, "Y", None, "N", "coupons.id", "N", "적용 쿠폰 ID", "MAPI-020", "기획", None],
        ["member_orders", "original_amount", "DECIMAL", "10,0", "N", None, "N", None, "N", "원가", "MAPI-020", "기획", None],
        ["member_orders", "coupon_discount", "DECIMAL", "10,0", "N", "0", "N", None, "N", "쿠폰 할인액", "MAPI-020", "기획", None],
        ["member_orders", "mileage_used", "DECIMAL", "10,0", "N", "0", "N", None, "N", "마일리지 사용액", "MAPI-020", "기획", None],
        ["member_orders", "final_amount", "DECIMAL", "10,0", "N", None, "N", None, "N", "최종 결제금액", "MAPI-020", "기획", None],
        ["member_orders", "payment_method", "VARCHAR", 30, "N", None, "N", None, "N", "결제수단 (card/kakaopay/naverpay/tosspay)", "MAPI-020", "기획", None],
        ["member_orders", "pg_payment_key", "VARCHAR", 100, "Y", None, "N", None, "Y", "PG 결제 키", "MAPI-021", "기획", None],
        ["member_orders", "status", "VARCHAR", 20, "N", "'pending'", "N", None, "Y", "상태 (pending/paid/failed/refunded)", "MAPI-021", "기획", None],
        ["member_orders", "membership_id", "BIGINT", None, "Y", None, "N", "memberships.id", "N", "생성된 이용권 ID", "MAPI-021", "기획", None],
        ["member_orders", "created_at", "TIMESTAMP", None, "N", "now()", "N", None, "Y", "주문 생성 일시", "MAPI-020", "기획", None],
        ["member_orders", "paid_at", "TIMESTAMP", None, "Y", None, "N", None, "N", "결제 완료 일시", "MAPI-021", "기획", None],
    ]

    for row in tables:
        ws.append(row)

    style_body(ws)
    auto_width(ws)


# ══════════════════════════════════════════════════════════
# 메인 실행
# ══════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # CRM과 동일한 시트 구조 (순서 맞춤)
    create_dashboard_sheet(wb)
    create_ia_sheet(wb)
    create_story_sheet(wb)
    create_screen_sheet(wb)         # 추가: 화면설계서
    create_ui_sheet(wb)             # 추가: UI 요소 상세
    create_fn_sheet(wb)
    create_api_sheet(wb)
    create_data_sheet(wb)           # 추가: 데이터 정의서
    create_tc_sheet(wb)             # 추가: QA테스트케이스
    create_policy_sheet(wb)
    create_flow_sheet(wb)           # 추가: 화면흐름도
    create_token_sheet(wb)          # 추가: 디자인토큰

    output_path = "스포짐_회원앱_기획서_v1.xlsx"
    wb.save(output_path)
    print(f"[완료] {output_path} 생성됨")
    print(f"  - 시트: {', '.join(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"    {name}: {ws.max_row - 1}건")
    print(f"  - 생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
