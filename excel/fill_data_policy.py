#!/usr/bin/env python3
"""
두 가지 작업 수행:
1. 데이터 정의서 시트의 빈 셀 채우기 (길이, 기본값, FK, 관련 API ID)
2. '비즈니스정책' 새 시트 추가
"""

import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 다른 스크립트가 먼저 실행 중이므로 10초 대기
print("다른 스크립트 완료 대기 중... (10초)")
time.sleep(10)
print("대기 완료. 작업 시작.")

EXCEL_PATH = "/Users/simjaehyeong/Desktop/pando/pando-beta/excel/스포짐_기획문서_v2.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)

# ─────────────────────────────────────────────
# 작업 1: 데이터 정의서 빈 셀 채우기
# ─────────────────────────────────────────────
print("\n[작업 1] 데이터 정의서 빈 셀 채우기...")

ws = wb["데이터 정의서"]

# 헤더: 테이블명(A/0), 컬럼명(B/1), 데이터타입(C/2), 길이(D/3),
#        NULL허용(E/4), 기본값(F/5), PK(G/6), FK(H/7),
#        인덱스(I/8), 설명(J/9), 관련API ID(K/10), 상태(L/11), 비고(M/12)

COL_TABLE   = 1   # A
COL_COLUMN  = 2   # B
COL_TYPE    = 3   # C
COL_LEN     = 4   # D
COL_NULL    = 5   # E
COL_DEFAULT = 6   # F
COL_PK      = 7   # G
COL_FK      = 8   # H
COL_INDEX   = 9   # I
COL_DESC    = 10  # J
COL_API     = 11  # K
COL_STATUS  = 12  # L

# ── FK 참조 규칙 ──
FK_RULES = {
    "tenant_id":   "tenants.id",
    "branch_id":   "branches.id",
    "member_id":   "members.id",
    "product_id":  "products.id",
    "staff_id":    "staff.id",
    "user_id":     "users.id",
    "category_id": "product_categories.id",
    "locker_id":   "lockers.id",
    "room_id":     "rooms.id",
    "rfid_id":     "rfid_cards.id",
    "coupon_id":   "coupons.id",
    "contract_id": "contracts.id",
    "payment_id":  "payments.id",
    "schedule_id": "schedules.id",
    "attendance_id": "attendances.id",
    "message_id":  "messages.id",
    "plan_id":     "plans.id",
    "ticket_id":   "tickets.id",
}

# ── 관련 API ID 매핑 (테이블명 기준) ──
API_MAP = {
    "tenants":              "API-001",
    "branches":             "API-027, API-028, API-029",
    "members":              "API-004, API-005, API-006, API-007",
    "member_tickets":       "API-008, API-009",
    "member_holds":         "API-010",
    "member_body_data":     "API-011",
    "attendances":          "API-012, API-013",
    "products":             "API-016, API-017, API-018",
    "product_categories":   "API-015",
    "lockers":              "API-019, API-020",
    "locker_assignments":   "API-019, API-020",
    "rfid_cards":           "API-021",
    "rooms":                "API-022",
    "schedules":            "API-023, API-024",
    "schedule_bookings":    "API-025",
    "staff":                "API-030, API-031, API-032",
    "staff_attendance":     "API-033",
    "payroll":              "API-034, API-035",
    "sales":                "API-036, API-037",
    "payments":             "API-038",
    "refunds":              "API-039",
    "contracts":            "API-040, API-041",
    "messages":             "API-042, API-043",
    "coupons":              "API-044",
    "coupon_issues":        "API-044",
    "mileage":              "API-045",
    "mileage_history":      "API-045",
    "users":                "API-002, API-003",
    "permissions":          "API-046",
    "settings":             "API-047",
}

def get_length(col_name: str, data_type: str, desc: str) -> str:
    """데이터 타입과 컬럼명/설명에 따른 길이 반환"""
    dt = data_type.upper()

    # 길이 불필요한 타입
    if dt in ("BIGINT", "INT", "SMALLINT", "TINYINT", "BOOLEAN", "BOOL",
              "DATE", "DATETIME", "TIMESTAMP", "TEXT", "LONGTEXT",
              "MEDIUMTEXT", "FLOAT", "DOUBLE", "ENUM", "JSON"):
        return ""

    if dt == "DECIMAL":
        return "(10,2)"

    if dt == "VARCHAR":
        col = col_name.lower()
        desc_lower = desc.lower() if desc else ""

        # 이름/명칭
        if col in ("name", "first_name", "last_name") or "이름" in desc_lower or "명칭" in desc_lower:
            return "50"
        # 이메일
        if "email" in col or "이메일" in desc_lower:
            return "100"
        # 전화/연락처
        if "phone" in col or "tel" in col or "연락처" in desc_lower or "전화" in desc_lower:
            return "20"
        # 코드
        if col.endswith("_code") or col == "code" or "코드" in desc_lower:
            return "20"
        # 비밀번호 해시
        if "password" in col or "비밀번호" in desc_lower:
            return "255"
        # URL/경로
        if "url" in col or "path" in col or "image" in col or "photo" in col:
            return "500"
        # 주소
        if "address" in col or "주소" in desc_lower:
            return "255"
        # 토큰
        if "token" in col or "토큰" in desc_lower:
            return "500"
        # 제목/상품명
        if col in ("title", "subject") or "상품명" in desc_lower or "제목" in desc_lower:
            return "100"
        # 메모/설명
        if col in ("memo", "note", "description", "reason") or "설명" in desc_lower or "메모" in desc_lower or "사유" in desc_lower:
            return "500"
        # 상태값/유형
        if col in ("status", "type", "gender", "role", "method", "category"):
            return "20"
        # 지점명/회사명
        if "지점" in desc_lower or "본사" in desc_lower:
            return "100"
        # 기본
        return "100"

    if dt == "CHAR":
        col = col_name.lower()
        if "code" in col:
            return "10"
        return "10"

    return ""


def get_default(col_name: str, data_type: str, desc: str) -> str:
    """컬럼명과 데이터 타입에 따른 기본값 반환"""
    col = col_name.lower()
    dt = data_type.upper()
    desc_lower = desc.lower() if desc else ""

    # 타임스탬프
    if col == "created_at":
        return "CURRENT_TIMESTAMP"
    if col == "updated_at":
        return "CURRENT_TIMESTAMP"
    if col == "deleted_at":
        return "NULL"

    # BOOLEAN 계열
    if dt in ("BOOLEAN", "BOOL") or dt == "TINYINT":
        if col.startswith("is_") or col.startswith("has_") or col.startswith("use_"):
            # 활성 여부 류
            if "active" in col or "enabled" in col or "show" in col or "display" in col or "visible" in col:
                return "true"
            return "false"

    # status 컬럼
    if col == "status":
        if "active" in desc_lower or "활성" in desc_lower:
            return "'active'"
        if "pending" in desc_lower or "대기" in desc_lower:
            return "'pending'"
        return "'active'"

    # 숫자 기본 0
    if dt in ("INT", "BIGINT", "SMALLINT") and col in (
        "count", "quantity", "amount", "balance", "mileage",
        "penalty_count", "hold_count", "remaining_count",
        "total_amount", "paid_amount", "discount_amount"
    ):
        return "0"

    # DECIMAL 금액
    if dt == "DECIMAL" and ("amount" in col or "price" in col or "fee" in col):
        return "0.00"

    return ""


def get_fk(col_name: str) -> str:
    """컬럼명에 따른 FK 참조 반환"""
    col = col_name.lower()
    return FK_RULES.get(col, "")


def get_api_id(table_name: str, existing: str) -> str:
    """테이블명에 따른 관련 API ID 반환"""
    if existing and existing.strip():
        return existing
    return API_MAP.get(table_name.lower().strip(), "")


filled_len = 0
filled_default = 0
filled_fk = 0
filled_api = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    table_name = row[COL_TABLE - 1].value or ""
    col_name   = row[COL_COLUMN - 1].value or ""
    data_type  = row[COL_TYPE - 1].value or ""
    desc       = row[COL_DESC - 1].value or ""

    if not col_name:
        continue

    # 길이 채우기
    cell_len = row[COL_LEN - 1]
    if not cell_len.value:
        val = get_length(col_name, data_type, desc)
        if val:
            cell_len.value = val
            filled_len += 1

    # 기본값 채우기
    cell_default = row[COL_DEFAULT - 1]
    if not cell_default.value:
        val = get_default(col_name, data_type, desc)
        if val:
            cell_default.value = val
            filled_default += 1

    # FK 채우기
    cell_fk = row[COL_FK - 1]
    if not cell_fk.value:
        val = get_fk(col_name)
        if val:
            cell_fk.value = val
            filled_fk += 1

    # 관련 API ID 채우기
    cell_api = row[COL_API - 1]
    existing_api = cell_api.value or ""
    val = get_api_id(table_name, existing_api)
    if val and not existing_api.strip():
        cell_api.value = val
        filled_api += 1

print(f"  길이 채움: {filled_len}개")
print(f"  기본값 채움: {filled_default}개")
print(f"  FK 채움: {filled_fk}개")
print(f"  관련 API ID 채움: {filled_api}개")


# ─────────────────────────────────────────────
# 작업 2: '비즈니스정책' 새 시트 추가
# ─────────────────────────────────────────────
print("\n[작업 2] '비즈니스정책' 시트 추가...")

if "비즈니스정책" in wb.sheetnames:
    del wb["비즈니스정책"]

ws_bp = wb.create_sheet("비즈니스정책")

# 헤더
headers = ["정책 ID", "정책 영역", "정책명", "정책 설명", "관련 화면", "관련 기능", "우선순위", "상태"]

# 헤더 스타일
header_fill   = PatternFill("solid", fgColor="1F4E79")
header_font   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_side     = Side(style="thin", color="AAAAAA")
thin_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

for ci, h in enumerate(headers, start=1):
    cell = ws_bp.cell(row=1, column=ci, value=h)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = header_align
    cell.border = thin_border

# 열 너비
col_widths = [10, 14, 24, 70, 16, 16, 12, 10]
for ci, w in enumerate(col_widths, start=1):
    ws_bp.column_dimensions[get_column_letter(ci)].width = w

ws_bp.row_dimensions[1].height = 30

# ── 정책 데이터 (42개) ──
policies = [
    # ── 인증 (5개) ──
    ("BP-001", "인증", "비밀번호 정책",
     "최소 8자, 영문+숫자+특수문자 조합 필수. 이전 3개 비밀번호 재사용 금지. 90일 경과 시 변경 권장 팝업(강제 변경 or 건너뛰기 선택 가능).",
     "SCR-001", "FN-001", "Critical", "확정"),

    ("BP-002", "인증", "계정 잠금 정책",
     "로그인 5회 연속 실패 시 30분 잠금. 잠금 시 남은 시간 표시. 잠금 해제 방법: 비밀번호 찾기(이메일 인증) 또는 관리자 수동 해제.",
     "SCR-001", "FN-001", "Critical", "확정"),

    ("BP-003", "인증", "세션 관리 정책",
     "동시 로그인 1개 세션만 허용. 새 로그인 시 기존 세션 강제 종료. JWT 토큰 만료: 2시간(기본). 로그인 유지 선택 시 Refresh Token 7일 유효.",
     "SCR-001", "FN-001", "Critical", "확정"),

    ("BP-004", "인증", "지점 전환 정책",
     "로그인 시 선택한 지점의 데이터만 접근 가능. 헤더 드롭다운으로 재로그인 없이 지점 전환 허용(최고관리자/센터장 한정). 전환 즉시 해당 지점 데이터 리로드.",
     "SCR-001", "FN-002", "High", "확정"),

    ("BP-005", "인증", "권한 매트릭스 정책",
     "역할별 메뉴/기능 접근 제한. 최고관리자>센터장>매니저>FC>스태프>조회전용 계층. 메뉴별 CRUD 권한 세분화. 역할 변경 시 즉시 적용.",
     "SCR-093", "FN-003", "Critical", "확정"),

    # ── 회원 (5개) ──
    ("BP-006", "회원", "회원 상태 정의",
     "활성: 유효한 이용권 1개 이상 보유. 만료: 모든 이용권 만료(마지막 만료일 기준). 예정: 시작일이 미래인 이용권 보유. 임박: 만료 D-30 이내. 홀딩: 이용권 정지 중. 미등록: 이용권 없음.",
     "SCR-010", "FN-010", "Critical", "확정"),

    ("BP-007", "회원", "회원 등록 규칙",
     "이름(2~20자 한글/영문), 성별, 연락처(010-XXXX-XXXX 형식) 필수. 연락처는 동일 지점 내 중복 불가. 회원번호 자동 생성: 지점코드-YYMMDD-순번. 출석번호 4자리 랜덤(중복 체크).",
     "SCR-012", "FN-012", "High", "확정"),

    ("BP-008", "회원", "회원 수정 규칙",
     "수정 시 변경 전/후 값 자동 audit_log 기록(변경자+변경일시). 연락처 변경 시 재중복 체크 필수. 변경 이력은 회원 상세 이력 탭에서 조회 가능.",
     "SCR-013", "FN-013", "High", "확정"),

    ("BP-009", "회원", "회원 삭제 규칙",
     "소프트 삭제(is_deleted 플래그). 미납 건 존재 시 삭제 불가(환불/정산 선행). 유효 이용권 존재 시 삭제 불가. 삭제 후 30일간 복구 가능. 복구 불가 시 영구 삭제 처리.",
     "SCR-011", "FN-011", "High", "확정"),

    ("BP-010", "회원", "회원 검색 정책",
     "이름: 부분일치(2글자 이상). 연락처: 숫자만 추출 후 부분일치. 회원번호: 완전일치. 검색 debounce 300ms 적용. 모든 필터는 AND 조합. 삭제 회원은 기본 미표시(옵션 선택 시 표시).",
     "SCR-010", "FN-010", "Medium", "확정"),

    # ── 매출 (5개) ──
    ("BP-011", "매출", "매출 집계 규칙",
     "총매출 = 현금 + 카드 + 마일리지 - 환불. 순매출 = 총매출 - 부가세. 일/주/월/연 단위 집계. 지점별 완전 독립 집계. 환불 건은 마이너스로 반영.",
     "SCR-030", "FN-030", "Critical", "확정"),

    ("BP-012", "매출", "환불 정책",
     "결제 후 7일 이내: 전액 환불. 7일 초과~이용 시작 전: 위약금 10% 공제. 이용 시작 후: 잔여일수 비례 일할 계산 환불. 환불 시 원래 결제 수단으로 반환. 마일리지 환불 시 잔여 마일리지 복원.",
     "SCR-030", "FN-031", "Critical", "확정"),

    ("BP-013", "매출", "미수금 관리 정책",
     "미납 발생 즉시 회원 상태에 미수금 표시. 7일 경과: SMS 자동 발송(설정 시). 30일 경과: 장기 미납 분류. 이용권 정지 여부는 센터 설정에 따름.",
     "SCR-030", "FN-032", "High", "확정"),

    ("BP-014", "매출", "결제 수단 정책",
     "현금: 거스름돈 계산 + 현금영수증 발행 옵션. 카드: 단말기 연동 또는 수기 입력. 마일리지: 잔액 초과 불가, 최소 사용 1,000원, 결제금액의 50% 한도. 혼합결제: 각 수단 합계 = 총 결제금액 일치 필수.",
     "SCR-032", "FN-033", "High", "확정"),

    ("BP-015", "매출", "비회원 결제 정책",
     "회원 미선택 상태에서 비회원 결제 진행 가능. 마일리지 적립/사용 불가. 영수증에 '비회원' 표시. 비회원 결제 데이터는 별도 통계 집계.",
     "SCR-031", "FN-034", "Medium", "확정"),

    # ── 상품 (4개) ──
    ("BP-016", "상품", "상품 카테고리 체계",
     "시설이용권(헬스/수영/골프 등), 수강권PT(1:1 PT/그룹 PT), 수강권GX(요가/필라테스/사이클 등), 대여(락커/운동복/수건), 기타(음료/보충제/용품). 동일 카테고리 내 상품명 중복 불가.",
     "SCR-040", "FN-040", "High", "확정"),

    ("BP-017", "상품", "가격 정책",
     "현금가 ≤ 카드가 권장(강제 아님). 가격 최소 0원(무료 상품 허용), 최대 99,999,999원. 부가세 포함/별도 선택 가능. 가격 변경 시 기존 계약 가격은 유지(소급 미적용).",
     "SCR-040", "FN-040", "High", "확정"),

    ("BP-018", "상품", "이용 기간 정책",
     "시작일: 구매일 또는 지정일 선택. 홀딩 최대 일수/횟수 상품별 설정. 홀딩 시 잔여일 정지(소진 아님). 만료 후 재활성화 불가(재구매 필요). 홀딩 중 만료일 = 원래 만료일 + 홀딩일수.",
     "SCR-041", "FN-041", "High", "확정"),

    ("BP-019", "상품", "키오스크 노출 정책",
     "노출 ON: 키오스크 판매 화면 표시. 노출 OFF: 키오스크 숨김, POS에서만 판매. 판매 중지: POS/키오스크 모두 비활성화. 재고 0 실물 상품: 자동 비활성화.",
     "SCR-041", "FN-042", "Medium", "확정"),

    # ── 시설 (4개) ──
    ("BP-020", "시설", "락커 배정 정책",
     "동일 유형 내 1회원 1락커 원칙. 배정 시 만료일 필수(상품 기간 연동 또는 수동 입력). 이미 배정된 락커 재배정 불가. 배정 취소는 배정일 당일만 가능.",
     "SCR-050", "FN-050", "High", "확정"),

    ("BP-021", "시설", "락커 만료 처리 정책",
     "D-3: 만료임박 표시(노란색). D-0: 만료 표시(빨간색). D+3: 자동 해제 여부 센터 설정으로 결정. 일괄 해제: 만료+N일 초과 건 일괄 처리 기능 제공.",
     "SCR-050", "FN-050", "Medium", "확정"),

    ("BP-022", "시설", "RFID 카드 정책",
     "1회원 1카드(활성 카드 기준). 카드 번호: 10자리 영숫자. 분실 신고 시 기존 카드 즉시 비활성화 후 신규 발급. 카드 반납 시 초기화 후 재사용 가능. 미등록 카드 태그 시 '등록되지 않은 카드' 표시.",
     "SCR-052", "FN-052", "High", "확정"),

    ("BP-023", "시설", "룸 관리 정책",
     "룸명: 고유, 2~20자. 수용 인원: 1~100명. 유형: PT룸/GX룸/다목적. 미래 예약 존재 시 삭제 불가. 비활성화 시 신규 예약 차단(기존 예약 유지). 예약 가능 시간 요일별 설정.",
     "SCR-053", "FN-053", "Medium", "확정"),

    # ── 출석/일정 (4개) ──
    ("BP-024", "출석/일정", "출석 유형 정의",
     "일반: 키오스크/앱 체크인. PT: PT 수업 참석. GX: GX 수업 참석. 수동: 관리자 등록. 타지점 출석은 전지점 이용권 보유 시 허용, 원지점에 알림 발송, 실제 방문 지점에 기록.",
     "SCR-020", "FN-020", "High", "확정"),

    ("BP-025", "출석/일정", "출석 차단 정책",
     "이용권 만료 회원: 기본 차단(센터 설정으로 허용 가능). 미납 회원: 센터 설정에 따라 허용/경고/차단 중 선택. 차단 시 키오스크 화면에 '관리자에게 문의' 메시지 표시.",
     "SCR-020", "FN-021", "High", "확정"),

    ("BP-026", "출석/일정", "일정 충돌 규칙",
     "동일 시간대+동일 트레이너 배정 시 충돌 경고. 동일 시간대+동일 룸 사용 시 충돌 경고. 1:1 PT는 트레이너당 1건만. GX는 룸 수용 인원 초과 불가. 과거 일정 수정/삭제 불가.",
     "SCR-021", "FN-022", "High", "확정"),

    ("BP-027", "출석/일정", "예약 취소 정책",
     "PT 예약 취소: 수업 24시간 전까지 무료 취소. 당일 취소: 회차 1회 차감(페널티). 노쇼: 회차 1회 차감+패널티 기록. 페널티 3회 누적 시 관리자 알림 발송.",
     "SCR-021", "FN-023", "High", "확정"),

    # ── 직원/급여 (4개) ──
    ("BP-028", "직원/급여", "직원 역할 체계",
     "최고관리자(전체 시스템), 센터장(지점 전체), 매니저(회원/매출/상품), FC(회원 관리/PT 일정), 스태프(출석/키오스크), 조회전용(조회만). 역할 선택 시 기본 권한 자동 배정. 커스텀 권한 추가 설정 가능.",
     "SCR-060", "FN-060", "Critical", "확정"),

    ("BP-029", "직원/급여", "퇴사 처리 정책",
     "퇴사 처리 시 관리자 계정 즉시 비활성화. 담당 회원 재배정 필요(알림 발송). 미지급 급여 정산 선행. 퇴사 후 데이터: 센터 설정에 따라 6개월 보관 또는 영구 보관.",
     "SCR-060", "FN-061", "High", "확정"),

    ("BP-030", "직원/급여", "급여 계산 정책",
     "실지급액 = 기본급 + 인센티브 + 수당 - 공제(4대보험+소득세). PT 인센티브: 건당 커미션(예: 30% 설정 가능). 마감: 매월 25일(설정 가능). 확정완료 후 수정 불가(관리자 잠금 해제 필요).",
     "SCR-062", "FN-062", "High", "확정"),

    ("BP-031", "직원/급여", "급여 명세서 정책",
     "급여 확정완료 후 명세서 자동 생성. PDF 다운로드 + 이메일 자동 발송 옵션. 조회 권한: 본인 명세서 또는 센터장 이상. 이의 제기: 직원 → 관리자 재검토 요청 가능.",
     "SCR-063", "FN-063", "Medium", "확정"),

    # ── 메시지/마케팅 (4개) ──
    ("BP-032", "메시지/마케팅", "메시지 발송 정책",
     "SMS: 90자(70원/건). LMS: 2,000자(30원/건). 카카오 알림톡: 1,000자(8원/건). 발송 전 비용 사전 표시. 예약 발송: 최소 5분 후~최대 30일 후. 발송 실패 시 재시도 1회.",
     "SCR-070", "FN-070", "High", "확정"),

    ("BP-033", "메시지/마케팅", "자동 알림 13종 정책",
     "이용권 만료 D-7/D-3/D-1/당일, 생일 축하, 장기 미출석(30일), 신규 환영, PT 전일 리마인드, 미납 안내, 락커 만료, 쿠폰 만료 임박, 재등록 유도(만료 후 7일), 정기 결제 안내. 매일 오전 9시 실행. 동일 회원+동일 규칙 1일 1회 중복 방지.",
     "SCR-071", "FN-071", "High", "확정"),

    ("BP-034", "메시지/마케팅", "쿠폰 정책",
     "유형: 정률(%)/정액(원)/무료이용권. 유효기간: 특정 기간 또는 발급일+N일. 사용 한도: 전체 N회 또는 회원당 1회. 최소 주문 금액 설정 가능. 만료 쿠폰 자동 비활성화. 발급건 존재 시 쿠폰 삭제 불가.",
     "SCR-072", "FN-072", "Medium", "확정"),

    ("BP-035", "메시지/마케팅", "마일리지 정책",
     "기본 적립률: 결제금액의 N%(센터 설정). 유효기간: N개월(기본 12개월). 최소 사용: 1,000원 이상. 최대 사용: 결제금액의 50%(설정 가능). 현금 전환 불가. 결제 취소 시 자동 회수.",
     "SCR-073", "FN-073", "Medium", "확정"),

    # ── 계약 (3개) ──
    ("BP-036", "계약", "계약 프로세스 정책",
     "Step1 회원 선택(필수), Step2 상품 선택(1개 이상), Step3 기간/금액 설정(시작일/할인/서비스일수), Step4 결제(현금/카드/마일리지/혼합), Step5 전자서명(약관 동의+서명). 각 단계 완료 전 다음 단계 진입 불가.",
     "SCR-080", "FN-080", "Critical", "확정"),

    ("BP-037", "계약", "할인 정책",
     "최대 할인율: 50%. 할인 유형: 재등록/신규/이벤트/관리자 재량. 할인 시 원가/할인가/최종가 모두 표시. 할인 사유 필수 입력. 할인 내역은 매출 통계에서 별도 조회 가능.",
     "SCR-080", "FN-081", "High", "확정"),

    ("BP-038", "계약", "계약 완료 후 자동 처리",
     "이용권 자동 생성(시작일 기준). 매출 자동 등록. 마일리지 자동 적립(설정 시). 환영 메시지 자동 발송(설정 시). 계약서 PDF 자동 생성 및 저장.",
     "SCR-080", "FN-082", "High", "확정"),

    # ── 지점 (3개) ──
    ("BP-039", "지점", "지점 코드 규칙",
     "지점 코드: 영문 대문자 2~5자, 고유. 자동 생성 또는 수동 입력. 한 번 등록된 코드는 변경 불가. 회원번호/출석번호 생성 시 지점 코드 prefix로 사용.",
     "SCR-094", "FN-090", "High", "확정"),

    ("BP-040", "지점", "지점 데이터 격리 정책",
     "지점별 데이터 완전 격리(멀티테넌트). 모든 쿼리에 branch_id 필터 필수 적용. 지점 비활성화: 소속 데이터 유지, 로그인 차단. 데이터 초기화: 최고관리자+비밀번호 확인+2중 확인 필수.",
     "SCR-094", "FN-091", "Critical", "확정"),

    ("BP-041", "지점", "회원 지점 이동 정책",
     "이동 시 이용권 유지 또는 신규 발급 선택. 이동 전 미납 건 정산 선행. 이동 이력 기록(이동일시/처리자/사유). 원지점/이동지점 모두에 이력 표시. 이동 후 원지점 데이터 조회 불가(이동지점만 접근).",
     "SCR-094", "FN-092", "High", "확정"),

    # ── 추가 정책 (1개) - 대시보드/통계 ──
    ("BP-042", "대시보드", "대시보드 데이터 갱신 정책",
     "통계 카드: 페이지 진입 시 + 5분 주기 자동 갱신. 차트: 페이지 진입 시 1회 로드. 알림 리스트: 실시간(WebSocket 또는 30초 폴링). 만료임박 기준: D-7 이내. D-3 이내: 긴급 표시(빨간색).",
     "SCR-002", "FN-002", "Medium", "확정"),
]

# 영역별 색상
area_colors = {
    "인증":          "D6E4F0",
    "회원":          "D5F5E3",
    "매출":          "FDEBD0",
    "상품":          "E8DAEF",
    "시설":          "FDFBD4",
    "출석/일정":     "D1F2EB",
    "직원/급여":     "FADBD8",
    "메시지/마케팅": "EBF5FB",
    "계약":          "F9EBEA",
    "지점":          "EAF2F8",
    "대시보드":      "F0F3F4",
}

priority_colors = {
    "Critical": "C0392B",
    "High":     "E67E22",
    "Medium":   "27AE60",
    "Low":      "7F8C8D",
}

body_font  = Font(name="맑은 고딕", size=9)
wrap_align = Alignment(vertical="top", wrap_text=True)
left_align = Alignment(vertical="top", horizontal="left", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top")

for ri, policy in enumerate(policies, start=2):
    pid, area, name, desc_text, screen, func, priority, status = policy

    area_color = area_colors.get(area, "FFFFFF")
    area_fill  = PatternFill("solid", fgColor=area_color)
    p_color    = priority_colors.get(priority, "7F8C8D")
    p_font     = Font(name="맑은 고딕", size=9, bold=True, color=p_color)

    values = [pid, area, name, desc_text, screen, func, priority, status]
    aligns = [center_align, center_align, left_align, wrap_align, center_align, center_align, center_align, center_align]
    fonts  = [body_font, body_font, body_font, body_font, body_font, body_font, p_font, body_font]

    for ci, (val, aln, fnt) in enumerate(zip(values, aligns, fonts), start=1):
        cell = ws_bp.cell(row=ri, column=ci, value=val)
        cell.font = fnt
        cell.alignment = aln
        cell.border = thin_border
        if ci in (1, 2):
            cell.fill = area_fill

    ws_bp.row_dimensions[ri].height = 48

# 1행 고정
ws_bp.freeze_panes = "A2"

# 자동 필터
ws_bp.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

print(f"  비즈니스정책 시트에 {len(policies)}개 정책 추가 완료")

# ─────────────────────────────────────────────
# 저장
# ─────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f"\n저장 완료: {EXCEL_PATH}")
print("모든 작업 완료!")
