"""
화면설계서 빈셀 채우기 + 대시보드 시트 통계 업데이트
- 작업 1: 화면설계서 시트의 담당자/반응형/접근성/에러화면 빈셀 채우기
- 작업 2: 대시보드 시트에 실제 통계 수치 추가
"""

import openpyxl
import time

EXCEL_PATH = "excel/스포짐_기획문서_v2.xlsx"

# ── 화면별 반응형/접근성/에러 규칙 매핑 ──────────────────────────────────────

# 화면 특성 분류 (화면 ID → 유형)
SCREEN_TYPE_MAP = {
    "SCR-001": "login",
    "SCR-002": "dashboard",
    "SCR-010": "list",
    "SCR-011": "detail_tab",
    "SCR-012": "form",
    "SCR-013": "form",
    "SCR-014": "chart",
    "SCR-020": "list",
    "SCR-021": "calendar",
    "SCR-030": "list",
    "SCR-031": "pos",
    "SCR-032": "modal",
    "SCR-040": "list",
    "SCR-041": "form",
    "SCR-050": "grid",
    "SCR-051": "modal",
    "SCR-052": "list",
    "SCR-053": "card",
    "SCR-060": "list",
    "SCR-061": "form",
    "SCR-062": "list",
    "SCR-063": "form",
    "SCR-070": "form",
    "SCR-071": "form",
    "SCR-072": "modal",
    "SCR-073": "list",
    "SCR-080": "wizard",
    "SCR-081": "list",
    "SCR-090": "settings",
    "SCR-091": "settings",
    "SCR-092": "settings",
    "SCR-093": "settings",
    "SCR-094": "card",
}

# 유형별 기본 반응형 규칙
RESPONSIVE_DEFAULTS = {
    "login":      "1024px 이상 중앙 정렬 카드, 미만 전체 너비",
    "dashboard":  "1280px 이상 4열 카드, 미만 2열→1열",
    "list":       "테이블 가로스크롤 (1024px 미만)",
    "detail_tab": "탭 세로 배치 (768px 미만), 컨텐츠 1열",
    "form":       "2열→1열 (768px 미만)",
    "chart":      "차트 반응형 (컨테이너 너비 100%)",
    "calendar":   "뷰 자동 전환 (주→일, 768px 미만)",
    "pos":        "장바구니 하단 고정 (768px 미만)",
    "modal":      "최대 너비 640px, 모바일 전체화면",
    "grid":       "그리드 열 수 자동 조정 (xs:4, sm:6, md:8)",
    "card":       "카드 그리드 2열→1열 (768px 미만)",
    "wizard":     "스텝 세로 배치 (768px 미만)",
    "settings":   "폼 1열 배치 (768px 미만)",
}

# 유형별 기본 접근성 규칙
ACCESSIBILITY_DEFAULTS = {
    "login":      "Tab 이동 순서, aria-label, 폼 필드 aria-label, 에러 메시지 aria-live",
    "dashboard":  "차트 alt-text, 통계 카드 aria-label, 스크린리더 수치 안내",
    "list":       "role='table', 컬럼 scope='col', 페이지네이션 aria-label, 스크린리더 페이지 안내",
    "detail_tab": "탭 키보드 네비게이션, role='tablist', aria-selected",
    "form":       "폼 필드 aria-label, aria-required, 에러 메시지 aria-live, aria-describedby",
    "chart":      "차트 alt-text, 데이터 테이블 대체 제공",
    "calendar":   "날짜 선택 키보드 접근, aria-label, role='grid'",
    "pos":        "상품 그리드 키보드 선택, 장바구니 수량 aria-label",
    "modal":      "모달 focus trap, aria-modal, ESC 닫기, aria-labelledby",
    "grid":       "색상 + 패턴으로 상태 구분 (색약 대응), aria-label",
    "card":       "카드 키보드 접근, aria-label, focus 표시",
    "wizard":     "스텝 진행 상태 aria-current, 단계 안내 aria-live",
    "settings":   "탭 접근성, 폼 필드 aria-label, 저장 버튼 aria-label",
}

# 유형별 기본 에러 화면 정의
ERROR_DEFAULTS = {
    "login":      "이메일 미입력: '이메일을 입력하세요'\n비밀번호 오류: '비밀번호가 올바르지 않습니다'\n서버 오류: 토스트 알림",
    "dashboard":  "데이터 로딩 실패: 스켈레톤 + 재시도 버튼\n서버 오류: 에러 메시지 + 재시도",
    "list":       "검색 결과 없음: '조건에 맞는 항목이 없습니다'\n로딩 실패: 스켈레톤 + 재시도\n서버 오류: 에러 메시지",
    "detail_tab": "데이터 없음: 404 페이지\n탭 로딩 실패: 탭별 에러 표시 + 재시도",
    "form":       "필수 항목 미입력: 필드 하단 에러 메시지\n서버 오류: 토스트 알림\n저장 실패: 재시도 안내",
    "chart":      "측정 데이터 없음: '기록이 없습니다'\n로딩 실패: 스켈레톤 + 재시도",
    "calendar":   "데이터 로딩 실패: 스켈레톤 + 재시도 버튼\n서버 오류: 에러 메시지 + 재시도",
    "pos":        "장바구니 비어있음: '상품을 선택해주세요'\n재고 없음: 상품 비활성화\n서버 오류: 토스트 알림",
    "modal":      "데이터 로딩 실패: 스켈레톤 + 재시도 버튼\n서버 오류: 에러 메시지 + 재시도\n404: 페이지 없음 안내",
    "grid":       "데이터 없음: '등록된 항목이 없습니다'\n로딩 실패: 스켈레톤 + 재시도",
    "card":       "데이터 없음: '등록된 항목이 없습니다'\n삭제 실패: 토스트 알림",
    "wizard":     "필수 단계 미완료: '이전 단계를 완료해주세요'\n서버 오류: 에러 메시지 + 재시도",
    "settings":   "저장 실패: 토스트 에러\n서버 오류: 에러 메시지 + 재시도",
}


def get_screen_type(screen_id: str) -> str:
    return SCREEN_TYPE_MAP.get(screen_id, "list")


def fill_cell_if_empty(ws, row: int, col: int, value: str) -> bool:
    """셀이 비어있으면 값을 채우고 True 반환, 이미 값이 있으면 False 반환"""
    cell = ws.cell(row=row, column=col)
    if cell.value is None or str(cell.value).strip() == "":
        cell.value = value
        return True
    return False


def fill_화면설계서(wb: openpyxl.Workbook) -> dict:
    ws = wb["화면설계서"]
    stats = {"담당자": 0, "반응형": 0, "접근성": 0, "에러화면": 0}

    # 컬럼 인덱스 (1-based)
    COL_SCREEN_ID  = 1   # 화면 ID
    COL_ERROR      = 12  # 에러 화면 정의 (L)
    COL_ACCESS     = 13  # 접근성 체크 (M)
    COL_RESPONSIVE = 14  # 반응형 체크 (Web) (N)
    COL_MANAGER    = 18  # 담당자 (R)

    for row in range(2, ws.max_row + 1):
        screen_id = ws.cell(row=row, column=COL_SCREEN_ID).value
        if not screen_id:
            continue

        screen_type = get_screen_type(str(screen_id).strip())

        # 1. 담당자 채우기
        if fill_cell_if_empty(ws, row, COL_MANAGER, "개발팀"):
            stats["담당자"] += 1

        # 2. 반응형 체크 (Web) 채우기
        responsive_val = RESPONSIVE_DEFAULTS.get(screen_type, "테이블 가로스크롤 (1024px 미만)")
        if fill_cell_if_empty(ws, row, COL_RESPONSIVE, responsive_val):
            stats["반응형"] += 1

        # 3. 접근성 체크 채우기
        access_val = ACCESSIBILITY_DEFAULTS.get(screen_type, "Tab 이동 순서, aria-label, 스크린리더 안내")
        if fill_cell_if_empty(ws, row, COL_ACCESS, access_val):
            stats["접근성"] += 1

        # 4. 에러 화면 정의 채우기
        error_val = ERROR_DEFAULTS.get(screen_type,
            "데이터 로딩 실패: 스켈레톤 + 재시도 버튼\n서버 오류: 에러 메시지 + 재시도\n404: 페이지 없음 안내")
        if fill_cell_if_empty(ws, row, COL_ERROR, error_val):
            stats["에러화면"] += 1

    return stats


def update_대시보드(wb: openpyxl.Workbook) -> None:
    """
    대시보드 시트는 대부분 수식으로 자동 계산됨.
    수식이 커버하지 않는 GAP 분석 요약 섹션을 추가.
    기존 셀은 건드리지 않음.
    """
    ws = wb["대시보드"]

    # 현재 마지막 데이터 행 파악
    last_row = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                last_row = cell.row

    start_row = last_row + 2  # 빈 줄 하나 띄우고 시작

    # GAP 분석 기반 실제 현황 섹션 추가
    summary_data = [
        ("GAP 분석 기반 실제 현황 (2026-03-11 기준)", None, None, None),
        ("항목", "계획/기획", "구현 완료", "구현율"),
        ("전체 화면 수", 34, 33, "97%"),
        ("전체 기능 수", 33, 33, "100%"),
        ("전체 API 수", 27, 0, "0% (전체 Mock)"),
        ("전체 UI 요소 수", 129, 129, "100%"),
        ("전체 TC 수", 50, 0, "0% (미실행)"),
        ("전체 DB 테이블 수", 16, 0, "0% (설계 완료)"),
        ("전체 DB 컬럼 수", 184, 0, "0% (설계 완료)"),
        ("비즈니스 정책 수", 41, 4, "~10%"),
        (None, None, None, None),
        ("주요 미구현 항목", None, None, None),
        ("API 연동", "전체 Mock 데이터", "실 API 미연동", "0%"),
        ("유효성 검사", "전체 화면", "Login/Form 일부", "15%"),
        ("에러 처리", "전체 화면", "Login만 기본", "5%"),
        ("비즈니스 규칙", "41개 정책", "POS 혼합결제 등 극소수", "10%"),
    ]

    for i, row_data in enumerate(summary_data):
        r = start_row + i
        for j, val in enumerate(row_data):
            if val is not None:
                # 해당 셀이 비어있을 때만 기입
                cell = ws.cell(row=r, column=j + 1)
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = val


def main():
    print("=" * 55)
    print("  화면설계서 빈셀 채우기 + 대시보드 업데이트 시작")
    print("=" * 55)

    print("\n[1/3] 20초 대기 중 (다른 스크립트 작업 완료 대기)...")
    time.sleep(20)

    print("[2/3] 엑셀 파일 로드...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # ── 작업 1: 화면설계서 빈셀 채우기 ─────────────────────
    print("\n[작업 1] 화면설계서 빈셀 채우기...")
    stats = fill_화면설계서(wb)
    print(f"  - 담당자   채운 셀: {stats['담당자']}개")
    print(f"  - 반응형   채운 셀: {stats['반응형']}개")
    print(f"  - 접근성   채운 셀: {stats['접근성']}개")
    print(f"  - 에러화면 채운 셀: {stats['에러화면']}개")

    # ── 작업 2: 대시보드 GAP 요약 추가 ───────────────────────
    print("\n[작업 2] 대시보드 시트 GAP 분석 요약 추가...")
    update_대시보드(wb)
    print("  - GAP 분석 기반 실제 현황 섹션 추가 완료")

    # ── 저장 ─────────────────────────────────────────────────
    print("\n[3/3] 저장 중...")
    wb.save(EXCEL_PATH)
    print(f"\n완료: {EXCEL_PATH} 저장 완료")
    print("=" * 55)


if __name__ == "__main__":
    main()
