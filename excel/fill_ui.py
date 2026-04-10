"""
UI 요소 상세 시트의 빈 셀을 채우는 스크립트
- 유효성 검사 (col K, index 11)
- 에러 메시지 (col L, index 12)
- Web 구현 (col I, index 9)
- 접근성 (col M, index 13)
"""

import openpyxl
import time
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '스포짐_기획문서_v2.xlsx')
SHEET_NAME = 'UI 요소 상세'

# 컬럼 인덱스 (1-based)
COL_요소ID = 1
COL_화면ID = 2
COL_요소명 = 3
COL_유형 = 4
COL_설명 = 5
COL_Web구현 = 9
COL_유효성 = 11
COL_에러메시지 = 12
COL_접근성 = 13


def get_validation(name: str, type_: str) -> str:
    """요소명과 유형에 따른 유효성 검사 규칙 반환"""
    if not type_:
        return ""

    t = type_.strip()
    n = name.strip() if name else ""

    # Button, Card, List, Table, Nav, Chart 등은 비워둠
    if t in ("Button", "Card", "List", "Nav", "Table", "Chart", "Modal", "Link", "Badge", "Toggle", "Tabs", "Tab", "Label"):
        return ""

    if t == "Select":
        return "1개 이상 선택 필수"

    if t == "Radio":
        return "필수 선택"

    if t == "Textarea":
        if "본문" in n or "메시지" in n:
            return "SMS 90자 이하 / LMS 2000자 이하"
        return "최대 500자 이하"

    if t == "Input":
        n_lower = n.lower()
        if "이메일" in n or "email" in n_lower:
            return "이메일 형식 검증 (regex: /^[\\w.-]+@[\\w.-]+\\.[a-z]{2,}$/i)"
        elif "비밀번호" in n or "password" in n_lower:
            return "최소 8자, 영문+숫자+특수문자 포함"
        elif "이름" in n or "성명" in n:
            return "최소 2자 이상, 한글/영문만 허용"
        elif "연락처" in n or "전화" in n or "휴대폰" in n or "핸드폰" in n:
            return "숫자 10~11자리 (하이픈 제외)"
        elif "생년월일" in n or "생일" in n or "날짜" in n or "일자" in n:
            return "YYYY-MM-DD 형식, 유효한 날짜"
        elif "검색" in n:
            return "최소 1자 이상"
        elif "금액" in n or "가격" in n or "요금" in n or "비용" in n:
            return "숫자만 입력, 0 이상"
        elif "횟수" in n or "수량" in n or "개수" in n:
            return "숫자만 입력, 1 이상"
        elif "메모" in n or "내용" in n or "설명" in n or "사유" in n:
            return "최대 500자 이하"
        elif "주소" in n:
            return "최소 5자 이상"
        elif "코드" in n:
            return "영문+숫자 조합"
        else:
            return "필수 입력"

    return ""


def get_error_message(name: str, type_: str, validation: str) -> str:
    """요소명과 유형에 따른 에러 메시지 반환"""
    if not type_:
        return ""

    t = type_.strip()
    n = name.strip() if name else ""

    # Button, Card, List, Table, Nav, Chart 등은 비워둠
    if t in ("Button", "Card", "List", "Nav", "Table", "Chart", "Link", "Badge", "Toggle", "Tabs", "Tab", "Label", "Modal"):
        return ""

    if t == "Select":
        return f"{n}을(를) 선택하세요"

    if t == "Radio":
        return f"{n}을(를) 선택하세요"

    if t == "Textarea":
        return f"{n}을(를) 입력하세요"

    if t == "Input":
        n_lower = n.lower()
        if "이메일" in n or "email" in n_lower:
            return "올바른 이메일 주소를 입력하세요"
        elif "비밀번호" in n or "password" in n_lower:
            return "비밀번호는 8자 이상, 영문+숫자+특수문자를 포함해야 합니다"
        elif "이름" in n or "성명" in n:
            return "이름은 2자 이상 입력하세요"
        elif "연락처" in n or "전화" in n or "휴대폰" in n or "핸드폰" in n:
            return "올바른 연락처를 입력하세요 (숫자 10~11자리)"
        elif "생년월일" in n or "생일" in n:
            return "올바른 생년월일을 입력하세요 (YYYY-MM-DD)"
        elif "날짜" in n or "일자" in n:
            return "올바른 날짜를 입력하세요"
        elif "검색" in n:
            return "검색어를 입력하세요"
        elif "금액" in n or "가격" in n or "요금" in n or "비용" in n:
            return "올바른 금액을 입력하세요"
        elif "횟수" in n or "수량" in n or "개수" in n:
            return "올바른 수량을 입력하세요"
        elif "메모" in n or "내용" in n or "설명" in n:
            return f"{n}을(를) 입력하세요"
        elif "사유" in n:
            return "사유를 입력하세요"
        elif "주소" in n:
            return "주소를 입력하세요"
        elif "코드" in n:
            return "올바른 코드를 입력하세요"
        else:
            return f"{n}을(를) 입력하세요"

    return ""


def get_web_impl(name: str, type_: str) -> str:
    """요소 유형에 따른 Web 구현 힌트 반환"""
    if not type_:
        return ""

    t = type_.strip()
    n = name.strip() if name else ""

    if t == "Textarea":
        return "<textarea> + 글자 수 카운터"

    if t == "Input":
        n_lower = n.lower()
        if "이메일" in n or "email" in n_lower:
            return "<input type='email'> + 실시간 validation"
        elif "비밀번호" in n or "password" in n_lower:
            return "<input type='password'> + visibility toggle"
        elif "검색" in n:
            return "<input type='search'> + debounce(300ms)"
        elif "날짜" in n or "일자" in n or "생년월일" in n:
            return "<input type='date'> + DatePicker 컴포넌트"
        elif "금액" in n or "가격" in n or "요금" in n or "비용" in n:
            return "<input type='number'> + 숫자 포맷 (콤마)"
        elif "전화" in n or "연락처" in n or "휴대폰" in n:
            return "<input type='tel'> + 자동 하이픈 포맷"
        elif "메모" in n or "내용" in n or "설명" in n or "사유" in n:
            return "<textarea> + 글자 수 카운터"
        else:
            return "<input type='text'> + validation"

    elif t == "Select":
        if "멀티" in n or "다중" in n or "필터" in n:
            return "MultiSelect 컴포넌트 + options mapping"
        else:
            return "<select> + options mapping"

    elif t == "Button":
        if "삭제" in n or "취소" in n:
            return "<button type='button'> + 확인 다이얼로그"
        elif "저장" in n or "등록" in n or "완료" in n or "확인" in n:
            return "<button type='submit'> + loading state"
        elif "다운로드" in n or "엑셀" in n:
            return "<button> + 파일 다운로드 핸들러"
        elif "검색" in n:
            return "<button type='submit'> + onClick handler"
        else:
            return "<button type='button'> + onClick handler"

    elif t == "List":
        if "테이블" in n or "목록" in n or "리스트" in n:
            return "DataTable 컴포넌트 + 정렬/페이지네이션"
        else:
            return "List 컴포넌트 + 아이템 렌더링"

    elif t == "Table":
        return "DataTable 컴포넌트 + 정렬/페이지네이션"

    elif t == "Card":
        if "차트" in n or "그래프" in n:
            return "Recharts 라이브러리 + 반응형"
        elif "통계" in n or "매출" in n or "수익" in n:
            return "StatCard 컴포넌트 + 숫자 애니메이션"
        else:
            return "Card 컴포넌트"

    elif t == "Modal":
        return "Modal 컴포넌트 + 상태 관리 (useState)"

    elif t == "Nav":
        if "페이지" in n or "페이지네이션" in n:
            return "Pagination 컴포넌트"
        elif "탭" in n:
            return "Tabs 컴포넌트"
        else:
            return "Nav 컴포넌트"

    elif t == "Chart":
        if "파이" in n or "원형" in n:
            return "Recharts PieChart + 범례"
        elif "막대" in n or "바" in n:
            return "Recharts BarChart + 반응형"
        elif "라인" in n or "선형" in n:
            return "Recharts LineChart + 툴팁"
        else:
            return "Recharts 라이브러리"

    elif t == "Toggle":
        return "<input type='checkbox'> + 커스텀 Toggle 스타일"

    elif t == "Tabs":
        return "Tabs 컴포넌트 + 활성 탭 상태 관리"

    elif t == "Link":
        return "<a> + React Router Link"

    elif t == "Badge":
        return "Badge 컴포넌트 + 상태별 색상"

    return f"{t} 컴포넌트"


def get_accessibility(name: str, type_: str) -> str:
    """요소 유형에 따른 접근성 속성 반환"""
    if not type_:
        return ""

    t = type_.strip()
    n = name.strip() if name else ""

    if t in ("Tab", "Label"):
        return ""

    if t == "Input":
        if "메모" in n or "내용" in n or "설명" in n or "사유" in n:
            return f"aria-label='{n}', role='textbox', tabindex 순서"
        else:
            return f"aria-label='{n}', tabindex 순서"

    elif t == "Textarea":
        return f"aria-label='{n}', role='textbox', aria-multiline='true'"

    elif t == "Radio":
        return f"role='radiogroup', aria-label='{n}', 각 옵션 role='radio'"

    elif t == "Select":
        return f"aria-label='{n}', role='combobox'"

    elif t == "Button":
        if "삭제" in n:
            return f"aria-label='{n}', role='button', aria-describedby='삭제 확인 안내'"
        else:
            return f"aria-label='{n}', role='button'"

    elif t in ("List", "Table"):
        return f"role='table', aria-label='{n}', 컬럼 헤더 scope='col'"

    elif t == "Card":
        if "차트" in n or "그래프" in n:
            return f"role='img', aria-label='{n}', 데이터 테이블 대체 제공"
        else:
            return f"aria-label='{n}', role='region'"

    elif t == "Modal":
        return f"role='dialog', aria-modal='true', aria-label='{n}', 포커스 트랩"

    elif t == "Nav":
        if "페이지" in n or "페이지네이션" in n:
            return f"role='navigation', aria-label='{n}'"
        else:
            return f"role='navigation', aria-label='{n}'"

    elif t == "Chart":
        return f"role='img', aria-label='{n}', 데이터 테이블 대체 제공"

    elif t == "Toggle":
        return f"aria-label='{n}', role='switch', aria-checked"

    elif t == "Tabs":
        return f"role='tablist', aria-label='{n}', 각 탭 role='tab'"

    elif t == "Link":
        return f"aria-label='{n}', role='link'"

    elif t == "Badge":
        return f"aria-label='{n}', role='status'"

    return f"aria-label='{n}'"


def fill_ui_sheet():
    print("3초 대기 중 (다른 프로세스 완료 대기)...")
    time.sleep(3)

    max_retry = 3
    for attempt in range(1, max_retry + 1):
        try:
            print(f"엑셀 파일 열기 시도 {attempt}/{max_retry}: {EXCEL_PATH}")
            wb = openpyxl.load_workbook(EXCEL_PATH)
            break
        except PermissionError as e:
            if attempt == max_retry:
                raise
            print(f"파일 잠금 감지, 5초 후 재시도... ({e})")
            time.sleep(5)

    ws = wb[SHEET_NAME]
    print(f"시트 '{SHEET_NAME}' 로드 완료. 총 {ws.max_row}행")

    updated = {"유효성": 0, "에러메시지": 0, "Web구현": 0, "접근성": 0}

    for row in ws.iter_rows(min_row=2):
        요소id = row[COL_요소ID - 1].value
        요소명_val = row[COL_요소명 - 1].value
        유형_val = row[COL_유형 - 1].value

        # 그룹 헤더 행 건너뜀 (▼ SCR-XXX)
        if not 요소id or str(요소id).startswith("▼"):
            continue

        # UI-XXX 형식인지 확인
        if not str(요소id).startswith("UI-"):
            continue

        name = str(요소명_val).strip() if 요소명_val else ""
        type_ = str(유형_val).strip() if 유형_val else ""

        # 유효성 검사 (col K = index 10, 1-based=11)
        cell_valid = row[COL_유효성 - 1]
        if not cell_valid.value:
            val = get_validation(name, type_)
            if val:
                cell_valid.value = val
                updated["유효성"] += 1

        # 에러 메시지 (col L = index 11, 1-based=12)
        cell_err = row[COL_에러메시지 - 1]
        if not cell_err.value:
            val = get_error_message(name, type_, cell_valid.value or "")
            if val:
                cell_err.value = val
                updated["에러메시지"] += 1

        # Web 구현 (col I = index 8, 1-based=9)
        cell_web = row[COL_Web구현 - 1]
        if not cell_web.value:
            val = get_web_impl(name, type_)
            if val:
                cell_web.value = val
                updated["Web구현"] += 1

        # 접근성 (col M = index 12, 1-based=13)
        cell_acc = row[COL_접근성 - 1]
        if not cell_acc.value:
            val = get_accessibility(name, type_)
            if val:
                cell_acc.value = val
                updated["접근성"] += 1

    print("\n업데이트 현황:")
    for k, v in updated.items():
        print(f"  {k}: {v}셀 채움")

    wb.save(EXCEL_PATH)
    print(f"\n저장 완료: {EXCEL_PATH}")


if __name__ == "__main__":
    fill_ui_sheet()
